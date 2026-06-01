import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import IMP_CON_CRE_INT

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = IMP_CON_CRE_INT
    form = CrearForm
    template_name = 'lista_con_int_cre.html'
    ordering = ['cliente', 'cod_imp', 'categoria']

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = IMP_CON_CRE_INT
    form = CrearForm
    template_name = 'detalles_con_int_cre.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = IMP_CON_CRE_INT
    form = CrearForm
    fields = '__all__'
    # widgets = {'oficina': forms.TextInput(attrs={'class': 'form-control rounded-pill'}),
    #            'codigo': forms.NumberInput(attrs={'class': 'form-control rounded-pill'}),
    #            'centro_costo': forms.TextInput(attrs={'class': 'form-control rounded-pill'})
    #            }
    template_name = 'crear_con_int_cre.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_interes_creditos')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = IMP_CON_CRE_INT
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_con_int_cre.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_interes_creditos')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = IMP_CON_CRE_INT
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_interes_creditos')

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        contabilizacion = IMP_CON_CRE_INT.objects.all().order_by('cliente', 'cod_con')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="IMP_CON_CRE_INT.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)


        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in IMP_CON_CRE_INT:
        # p.drawString(80, 800, f"Cliente: {dato.cliente}")
        # p.drawString(80, 780, f"Código: {dato.cod_imp}")
        # p.drawString(80, 740, f"Categoria: {dato.categoria}")
        # p.drawString(80, 720, f"Cuenta Contable: {dato.kcta_con}")
        # p.drawString(80, 700, f"Cuenta Provisión Individual: {dato.kcta_pro_ind}")
        # p.drawString(80, 680, f"Porcentaje: {dato.kporcentaje}")
        # p.drawString(80, 660, f"Cuenta Intereses: {dato.cta_int}")
        # p.drawString(80, 640, f"Cuenta de Orden Intereses: {dato.cta_ord_int}")

        datos_tabla = [["cliente", "cod_imp", "categoria", "kcta_con","kcta_pro_ind", "kporcentaje", "cta_int", "cta_ord_int"]]

        for dato in contabilizacion:
            datos_tabla.append([dato.cliente, dato.cod_imp, dato.categoria, dato.kcta_con, dato.kcta_pro_ind, dato.kporcentaje, dato.cta_int,dato.cta_ord_int])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = IMP_CON_CRE_INT.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="IMP_CON_CRE_INT.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cliente: {dato.cliente}")
        p.drawString(80, 780, f"Código: {dato.cod_imp}")
        p.drawString(80, 740, f"Categoria: {dato.categoria}")
        p.drawString(80, 720, f"Cuenta Contable: {dato.kcta_con}")
        p.drawString(80, 700, f"Cuenta Provisión Individual: {dato.kcta_pro_ind}")
        p.drawString(80, 680, f"Porcentaje: {dato.kporcentaje}")
        p.drawString(80, 660, f"Cuenta Intereses: {dato.cta_int}")
        p.drawString(80, 640, f"Cuenta de Orden Intereses: {dato.cta_ord_int}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        contabilizacion = IMP_CON_CRE_INT.objects.all()

        p = canvas.Canvas(response)

        for dato in contabilizacion:
            p.drawString(80, 800, f"Cliente: {dato.cliente}")
            p.drawString(80, 780, f"Código: {dato.cod_imp}")
            p.drawString(80, 740, f"Categoria: {dato.categoria}")
            p.drawString(80, 720, f"Cuenta Contable: {dato.kcta_con}")
            p.drawString(80, 700, f"Cuenta Provisión Individual: {dato.kcta_pro_ind}")
            p.drawString(80, 680, f"Porcentaje: {dato.kporcentaje}")
            p.drawString(80, 660, f"Cuenta Intereses: {dato.cta_int}")
            p.drawString(80, 640, f"Cuenta de Orden Intereses: {dato.cta_ord_int}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in IMP_CON_CRE_INT._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        contabilizacion = IMP_CON_CRE_INT.objects.all()
        for row_num, data in enumerate(contabilizacion, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        contabilizacion = IMP_CON_CRE_INT.objects.all()
        for data in contabilizacion:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
