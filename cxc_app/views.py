from django.shortcuts import render
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Max, IntegerField, Sum
from django.db.models.functions import Cast
from .models import CTAS_X_COBRAR,CXC_DET
from django.views.generic import CreateView, UpdateView, DeleteView
from django.views import View
from django.contrib import messages
from django.shortcuts import redirect
from datetime import datetime
from django.http import JsonResponse
from django.views.generic import TemplateView
from terceros_app.models import TERCEROS
from conceptos_app.models import CONCEPTOS
from asociados_app.models import ASOCIADOS
from .forms import CtasXCobrarForm,BusquedaForm,ImportarCxcForm,EliminarCxcForm
from django.http import Http404
from django.views.generic.edit import FormView
from django.http import JsonResponse
from .forms import ImportarCxcForm
import openpyxl
from django.http import JsonResponse
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.urls import reverse
import os
from django.conf import settings


class CtasXCobrarListView(ListView):
    model = CTAS_X_COBRAR
    template_name = 'listar_cxc.html'
    context_object_name = 'ctas_cxc'
    paginate_by = 10  # Puedes ajustar el número de resultados por página

    def get_queryset(self):
        query = self.request.GET.get('q',None)
        est_cta = self.request.GET.get('est_cta',None)
        concepto = self.request.GET.get('concepto', None)
        object_list = CTAS_X_COBRAR.objects.all()
        if concepto:
            object_list = object_list.filter(concepto_id=concepto) 
        if query :
            object_list = object_list.filter(
                Q(tercero__nombre__icontains=query) |
                Q(tercero__doc_ide=query)
            )
        if est_cta:
            object_list = object_list.filter(aplicado=est_cta) 
            object_list = object_list.order_by('cod_cxc')  # Cambia 'num_cta' por el campo que prefieras
        return object_list
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = BusquedaForm(cliente_id=self.request.session.get('cliente_id')) 
        ctas_cxc = self.get_queryset()
        context['resultados'] = [] 
        for cta_cxc in ctas_cxc:
            concep = CONCEPTOS.objects.filter(id = cta_cxc.concepto_id).first()
            resultado = {
                'id' : cta_cxc.id,
                'cod_cxc' : cta_cxc.cod_cxc,
                'nombre' :  cta_cxc.tercero.nombre,
                'concepto' : concep.cod_con,
                'fecha_des' : cta_cxc.fecha_des,
                'valor' : cta_cxc.valor,
                'aplicado' : cta_cxc.aplicado,
            }
            context['resultados'].append(resultado)
        return context
    
from django.views.generic import TemplateView, CreateView
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.utils import timezone
from django.contrib import messages
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from datetime import date

class CtasXCobrarBaseView(View):
    template_name = 'listar_cxc.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = self.kwargs.get('pk') if 'pk' in self.kwargs else None
        form = self.get_form()
        if pk:  # Modo de actualización
            ctas_x_cobrar = CTAS_X_COBRAR.objects.get(id=pk)
            fecha_des_form = ctas_x_cobrar.fecha_des.strftime('%d/%m/%Y') if ctas_x_cobrar.fecha_des else None
            fecha_exi_form = ctas_x_cobrar.fecha_exi.strftime('%d/%m/%Y') if ctas_x_cobrar.fecha_exi else None           
            context.update({
                'cod_cxc': ctas_x_cobrar.cod_cxc,
                'cod_aso': ctas_x_cobrar.tercero.doc_ide if ctas_x_cobrar.tercero else 'No disponible',
                'valor': ctas_x_cobrar.valor,
                'fecha_des': fecha_des_form,
                'fecha_exi': fecha_exi_form,
                'asociado_nombre': ctas_x_cobrar.tercero.nombre if ctas_x_cobrar.tercero else 'Desconocido',
                'concepto': ctas_x_cobrar.concepto,
                'aplicado': ctas_x_cobrar.aplicado if ctas_x_cobrar.aplicado else 'X',
                'operation': 'update',
                'button_text': 'Modificar',
            })
        else:  # Modo de creación
            context['operation'] = 'create'
            context['button_text'] = 'Guardar'
            max_value_str = self.get_form_kwargs().get('max_value_str')
            form.fields['cod_cxc'].initial = max_value_str  # O cualquier otra lógica que decidas

        context['form'] = form  # Asegúrate de asignar el formulario al contexto
        return context

    def get_form(self, form_class=None):
        if form_class is None:
            form_class = self.get_form_class()
        return form_class(**self.get_form_kwargs())
    
    def validate_models(self, form):
        errors = {}
        fecha_des = form.cleaned_data.get('fecha_des')
        if not self.is_valid_date_format(fecha_des):
            errors['fecha_des'] = 'La fecha debe estar en formato dd/mm/yyyy.'
        fecha_exi = form.cleaned_data.get('fecha_exi')
        if not self.is_valid_date_format(fecha_exi):
            errors['fecha_exi'] = 'La fecha debe estar en formato dd/mm/yyyy.'
        cod_aso = form.cleaned_data.get('cod_aso')
        if not cod_aso :
            errors['cod_aso'] = 'Debe Existir el documento del asociado Titular de la Cuenta.'
        concepto = form.cleaned_data.get('concepto')
        if not concepto:
            errors['concepto'] = 'Debe seleccionar concepto'
        try:
            get_object_or_404(CONCEPTOS, pk=concepto.id)
        except ValidationError:
            errors['concepto'] = 'El concepto No es Valido'
        return errors

    def is_valid_date_format(self,date_value):
        # Verificar si date_value es una cadena
        if isinstance(date_value, str):
            try:
                datetime.strptime(date_value, '%d/%m/%Y')
                return True
            except ValueError:
                return False
        # Verificar si date_value es un objeto datetime
        elif isinstance(date_value, (datetime, date)):
            return True
        else:
            return False


    def handle_errors(self, errors):
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return None

class CtasXCobrarCreateView(CtasXCobrarBaseView, CreateView):
    model = CTAS_X_COBRAR
    form_class = CtasXCobrarForm
    template_name = 'ctas_x_cobrar_form.html'
    success_url = reverse_lazy('listar_cxc')  # Asegúrate que este es el nombre correcto
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        kwargs['cliente_id'] = cliente_id
        kwargs['oficina_id'] = oficina_id
        kwargs['operation'] = 'create'
        max_value = CTAS_X_COBRAR.objects.filter(oficina_id=oficina_id).annotate(
            cod_cxc_int=Cast('cod_cxc', IntegerField())
        ).aggregate(max_value=Max('cod_cxc_int'))['max_value']
        max_value_str = str(max_value + 1).zfill(8) if max_value else '00000001'
        kwargs['max_value_str'] = max_value_str
        return kwargs

    def post(self, request):
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        form = CtasXCobrarForm(request.POST)
        post_data = request.POST.copy()
        cod_aso = post_data.get('cod_aso')
        if 'fecha_des' in post_data:
            fecha_str = post_data.get('fecha_des')
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                post_data['fecha_des'] = fecha_obj  # Sobrescribir con el objeto de fecha
            except ValueError:
                pass  # Manejar si falla la conversión, si quieres manejar errores aquí
        if 'fecha_exi' in post_data:
            fecha_str = post_data.get('fecha_exi')
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                post_data['fecha_exi'] = fecha_obj  # Sobrescribir con el objeto de fecha
            except ValueError:
                pass  # Manejar si falla la conversión, si quieres manejar errores aquí
        form = CtasXCobrarForm(post_data)
        asociado = TERCEROS.objects.filter(cliente_id = cliente_id ,doc_ide = cod_aso).first()
        if asociado:
            form.instance.tercero = asociado  
        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
            cta_cxc = CTAS_X_COBRAR.objects.filter(oficina_id = oficina_id,cod_cxc = request.POST.get('cod_cxc')).first()
            if cta_cxc != None:
                return JsonResponse({'success': False, 'errors': 'No se puede crear una cuenta que ya existe'}, status=200)
            cta_cxc = CTAS_X_COBRAR.objects.create(oficina_id = oficina_id, 
                cod_cxc = post_data.get('cod_cxc'),
                tercero = asociado,
                concepto_id = int(post_data.get('concepto')),
                fecha_des = post_data.get('fecha_des'),
                fecha_exi = post_data.get('fecha_exi'),
                valor = post_data.get('valor'),
                aplicado = post_data.get('aplicado')
            )

            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'cod_cxc': cta_cxc.cod_cxc,
                    'mensaje': 'La cuenta se ha grabado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
        else:
            error_list = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_list.append(f'Error en {field}: {error}')
            print('Lista de errores:', error_list)
            if is_ajax:
                return JsonResponse({'success': False, 'errors': error_list}, status=200)
            else:
                return render(request, 'ctas_x_cobrar_form.html', {'form': form})
            
class CtasCxcUpdateView(CtasXCobrarBaseView,UpdateView):
    model = CTAS_X_COBRAR
    form_class = CtasXCobrarForm
    template_name = 'ctas_x_cobrar_form.html'
    success_url = reverse_lazy('listar_cxc')
    
    def get_object(self, queryset=None):    
        padre_pk = self.kwargs.get('pk')
        cta_cxc = CTAS_X_COBRAR.objects.filter(id=padre_pk).first()
        if cta_cxc is None:
            raise Http404("No se encontró   la Cta por Cobrar ")
        return cta_cxc
        
    def get(self, request, *args, **kwargs):
        self.object = self.get_object()  # Asegúrate de asignar el objeto
        context = self.get_context_data(object=self.object)
        context['pk'] = self.kwargs.get('pk')
        return render(request, self.template_name, context)

    def form_valid(self, form):
        lin_aho = form.cleaned_data['lin_aho']
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        form = CtasXCobrarForm(request.POST)
        post_data = request.POST.copy()
        cod_aso = post_data.get('cod_aso')
        if 'fecha_des' in post_data:
            fecha_str = post_data.get('fecha_des')
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                post_data['fecha_des'] = fecha_obj  # Sobrescribir con el objeto de fecha
            except ValueError:
                pass  # Manejar si falla la conversión, si quieres manejar errores aquí
        if 'fecha_exi' in post_data:
            fecha_str = post_data.get('fecha_exi')
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                post_data['fecha_exi'] = fecha_obj  # Sobrescribir con el objeto de fecha
            except ValueError:
                pass  # Manejar si falla la conversión, si quieres manejar errores aquí
        form = CtasXCobrarForm(post_data)
        asociado = TERCEROS.objects.filter(cliente_id = cliente_id ,doc_ide = cod_aso).first()
        if asociado:
            form.instance.tercero = asociado  
        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
                else:
                    messages.error(request, 'Errores en el formulario.')
                    return render(request, 'ctas_x_cobrar_form.html', {'form': form})
            cta_cxc = CTAS_X_COBRAR.objects.filter(oficina_id = oficina_id,cod_cxc = request.POST.get('cod_cxc')).first()
            if cta_cxc == None:
                return JsonResponse({'success': False, 'errors': 'No Existe cuenta por cobrar'}, status=200)
            cta_cxc.tercero = asociado
            cta_cxc.concepto_id = int(post_data.get('concepto'))
            cta_cxc.fecha_des = post_data.get('fecha_des')
            cta_cxc.fecha_exi = post_data.get('fecha_exi')
            cta_cxc.valor = post_data.get('valor')
            cta_cxc.aplicado = post_data.get('aplicado')
            cta_cxc.save()
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'cod_cxc': cta_cxc.cod_cxc,
                    'mensaje': 'La cuenta se ha moedificado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
        else:
            error_list = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_list.append(f'Error en {field}: {error}')
            if is_ajax:
                return JsonResponse({'success': False, 'errors': error_list}, status=200)
            else:
                return render(request, 'ctas_x_cobrar_form.html', {'form': form})
            
class CtasCxcDeleteView(UpdateView):
    model = CTAS_X_COBRAR
    form_class = CtasXCobrarForm
    template_name = 'listar_cxc.html'
    success_url = reverse_lazy('listar_cxc')
    def get_object(self, queryset=None):
        padre_pk = self.kwargs.get('pk')
        cta_cxc = CTAS_X_COBRAR.objects.filter(id=padre_pk).first()
        if cta_cxc is None:
            raise Http404("No se encontró   la Cta por Cobrar ")
        else:
            cta_cxc.delete()
        return cta_cxc
    
class CtasXCobrarImportar(CreateView):
    model = CTAS_X_COBRAR
    form_class = ImportarCxcForm
    template_name = 'importar_cxc.html'
    success_url = reverse_lazy('listar_cxc')  # Asegúrate que este es el nombre correcto

    def get(self, request, *args, **kwargs):
        form = self.get_form()
        return render(request, self.template_name, {'form': form})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        kwargs['cliente_id'] = cliente_id
        kwargs['oficina_id'] = oficina_id
        return kwargs

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            if form.is_valid():
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
        return super().post(request, *args, **kwargs)
    
    def form_valid(self, form):
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        cod_con = form.cleaned_data.get('concepto')  # Obtener el campo 'concepto'
        fecha_des = form.cleaned_data.get('fecha_des')  # Obtener el campo 'fecha'
        fecha_exi = form.cleaned_data.get('fecha_exi')  # Obtener el campo 'fecha'
        excel_file = form.cleaned_data['excelFile']  # Obtener el archivo subido
        print('cliente id ',cliente_id)
        concepto = CONCEPTOS.objects.filter(cliente_id=cliente_id, cod_con=cod_con).first()
        errors = {}
        if concepto is None:
            errors['concepto'] = 'No Existe Concepto'
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active  # Acceder a la primera hoja de cálculo
            for row in sheet.iter_rows(min_row=2, values_only=True):
                doc_ide = row[0]  # Primer columna (A)
                valor = row[1]  # Segunda columna (B)
                tercero = TERCEROS.objects.filter(cliente_id=cliente_id, doc_ide=doc_ide).first()
                if tercero is None:
                    errors[doc_ide] = 'No existe Doc_ide'
            if len(errors) == 0:
                sheet.cell(row=1, column=3).value = "cod_cxc"
                sheet.cell(row=1, column=4).value = "concepto"
                for row in sheet.iter_rows(min_row=2):  # Sin values_only para modificar celdas
                    doc_ide = row[0].value  # Primer columna (A)
                    valor = row[1].value  # Segunda columna (B)
                    tercero = TERCEROS.objects.filter(cliente_id=cliente_id, doc_ide=doc_ide).first()
                    max_value = CTAS_X_COBRAR.objects.filter(oficina_id=oficina_id).annotate(
                        cod_cxc_int=Cast('cod_cxc', IntegerField())
                    ).aggregate(max_value=Max('cod_cxc_int'))['max_value']
                    max_value_str = str(max_value + 1).zfill(8) if max_value else '00000001'
                    try:
                        # Crear y guardar el registro
                        cta_cxc = CTAS_X_COBRAR.objects.create(
                            oficina_id=oficina_id,
                            cod_cxc=max_value_str,
                            tercero=tercero,
                            concepto=concepto,
                            valor=valor,
                            fecha_des=fecha_des,
                            fecha_exi=fecha_exi,
                            aplicado='X'
                        )
                        cta_cxc.save()
                        sheet.cell(row=row[0].row, column=3, value=cta_cxc.cod_cxc)  # Columna C
                        sheet.cell(row=row[0].row, column=4, value=cta_cxc.concepto.cod_con)  # Columna D
                    except Exception as e:
                        print('Error al guardar el registro:', e)
                file_path = os.path.join(settings.MEDIA_ROOT, 'archivo_procesado.xlsx')
                wb.save(file_path)
                download_url = reverse('descargar_archivo_procesado')    #  a qui se invoca la url que ejecuta 
                return JsonResponse({'success': True,'mensaje': 'Proceso Exitoso ','download_url': download_url})
            else:
                print('Errores ', errors)
                return JsonResponse({'success': False, 'mensaje': 'Errores encontrados', 'errors': errors})
        except Exception as e:
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': str(e)})
            else:
                form.add_error(None, f'Error procesando archivo: {str(e)}')
                return self.form_invalid(form)

    def form_invalid(self, form):
        print('Formulario inválido:', form.errors)  # Imprimir los errores
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors})
        else:
            return super().form_invalid(form)
        
    def dispatch(self, request, *args, **kwargs):
        print('Método de solicitud:', request.method)  # Para verificar el método
        return super().dispatch(request, *args, **kwargs)
    
from django.http import FileResponse

def descargar_archivo_procesado(request):
    file_path = os.path.join(settings.MEDIA_ROOT, 'archivo_procesado.xlsx')
    response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="archivo_procesado.xlsx"'
    return response

class CtasXCobrarEliminar(CreateView):
    model = CTAS_X_COBRAR
    form_class = EliminarCxcForm
    template_name = 'eliminar_importar.html'
    success_url = reverse_lazy('listar_cxc')  # Asegúrate que este es el nombre correcto
    
    def get(self, request, *args, **kwargs):
        form = self.get_form()
        return render(request, self.template_name, {'form': form})
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        kwargs['cliente_id'] = cliente_id
        kwargs['oficina_id'] = oficina_id
        return kwargs

    def post(self, request, *args, **kwargs):
        form = self.get_form()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            if form.is_valid():
                return self.form_valid(form)
            else:
                return self.form_invalid(form)
        return super().post(request, *args, **kwargs)
    
    def form_valid(self, form):
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        excel_file = form.cleaned_data['excelFile']  # Obtener el archivo subido
        print('cliente id ',cliente_id)
        errors = {}
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active  # Acceder a la primera hoja de cálculo
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cod_cxc_str = row[0]  # Primer columna (A)
                if cod_cxc_str == None:
                    continue
                cta_cxc = CTAS_X_COBRAR.objects.filter(oficina_id=oficina_id,cod_cxc=cod_cxc_str).first()
                if cta_cxc == None:
                    errors[cod_cxc_str] = 'No existe cod_cxc'
                else:
                    if cta_cxc.aplicado != 'X':
                        errors[cod_cxc_str] = 'No se puede Eliminar por que esta Activo'
            if len(errors) == 0:
                sheet.cell(row=1, column=2).value = "Eliminado"
                for row in sheet.iter_rows(min_row=2):  
                    cod_cxc_str = row[0].value  # Primer columna (A)
                    if cod_cxc_str == None:
                        continue
                    try:
                        # Crear y guardar el registro
                        cta_cxc = CTAS_X_COBRAR.objects.filter(oficina_id=oficina_id,cod_cxc=cod_cxc_str).first()
                        if cta_cxc != None:
                            cta_cxc.delete()
                        sheet.cell(row=row[0].row, column=2, value='Ok')  # Columna C
                    except Exception as e:
                        print('Error al guardar el registro:', e)
                file_path = os.path.join(settings.MEDIA_ROOT, 'archivo_procesado.xlsx')
                wb.save(file_path)
                download_url = reverse('descargar_archivo_procesado')    #  a qui se invoca la url que ejecuta 
                return JsonResponse({'success': True,'mensaje': 'Proceso Exitoso ','download_url': download_url})
            else:
                return JsonResponse({'success': False, 'mensaje': 'Errores encontrados', 'errors': errors})
        except Exception as e:
            if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': str(e)})
            else:
                form.add_error(None, f'Error procesando archivo: {str(e)}')
                return self.form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'errors': form.errors})
        else:
            return super().form_invalid(form)
        
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    
def lista_cxp(id_socio, fecha_corte):
    try:
        asociado = ASOCIADOS.objects.select_related('tercero').get(id=id_socio)
        tercero = asociado.tercero

        cuentas = CTAS_X_COBRAR.objects.filter(tercero=tercero).select_related('concepto')

        lista_resultados = []

        for cta in cuentas:
            # Sumar solo los movimientos hasta la fecha de corte
            movimientos = CXC_DET.objects.filter(
                cuenta_x_cobrar=cta,
                tip_mov='P',
                fecha__lte=fecha_corte  # 🔥 Importante: Solo movimientos hasta esa fecha
            ).aggregate(total=Sum('valor'))['total'] or 0  # Aquí corregido 
                        
            saldo = (cta.valor or 0) + movimientos
            if saldo > 0:
                datos = {
                    'val_ini': cta.valor or 0,
                    'fec_des': cta.fecha_des,
                    'fec_ven': cta.fecha_exi,
                    'concepto': cta.concepto.descripcion if cta.concepto else '',
                    'saldo': saldo
                }
                lista_resultados.append(datos)

        return lista_resultados

    except ASOCIADOS.DoesNotExist:
        return [] 