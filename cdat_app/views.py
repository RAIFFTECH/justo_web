from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from django import forms
from .forms import CtasCdatsForm,ReporteCdatForm
from ctas_ahorros_app.models import CTAS_AHORRO
from django.http import HttpResponse
from django.views.decorators.http import require_GET
from lineas_ahorro_app.models import LINEAS_AHORRO
from .models import CTA_CDAT
from django.db.models import Max, IntegerField, Func, F
from django.db.models.functions import Cast
from asociados_app.models import ASOCIADOS
from django.http import JsonResponse
from terceros_app.models import TERCEROS
from oficinas_app.models import OFICINAS
from ampliacion_cdat_app.models import CTA_CDAT_AMP
from django.shortcuts import get_object_or_404
from django.http import Http404 
from liquidacion_cdat_app.models import CTA_CDAT_LIQ 
from reportlab.lib.pagesizes import letter
from num2words import num2words

# Asegúrate de tener la función get_max_consecutive_for_line en el mismo archivo o importarla
def get_max_consecutive_for_line(id_lin):   
    lin_aho = LINEAS_AHORRO.objects.filter(id=id_lin).first()
    queryset = CTAS_AHORRO.objects.filter(lin_aho=lin_aho)
    queryset = queryset.annotate(
        consecutive_number=Cast(
            Func(
                F('num_cta'),
                function='SUBSTRING',
                template="SUBSTRING(%(expressions)s FROM 4 FOR 6)",  # Desde el carácter 4 por 6 posiciones
                output_field=IntegerField()
            ),
            output_field=IntegerField()
        )
    )
    max_consecutive = queryset.aggregate(max_consecutive=Max('consecutive_number'))['max_consecutive']+1
    if max_consecutive is None:
        max_consecutive = 0
    max_consecutive_str = str(max_consecutive).zfill(6)
    max_value = f"{lin_aho.cod_lin_aho}-{max_consecutive_str}"
    return max_value

class BuscarCdatshoView(View):   
    def get(self, request):
        num_cta = request.GET.get('num_cta', '')
        nombre = request.GET.get('nombre', '')
        est_cta = request.GET.get('est_cta', 'A')  # Valor por defecto
        resultados = CTA_CDAT.objects.all()
        if num_cta:
            resultados = resultados.filter(num_cta__icontains=num_cta)
        if nombre:
            resultados = resultados.filter(asociado__tercero__nombre__icontains=nombre)
        if est_cta:
            resultados = resultados.filter(est_cta=est_cta)
        context = {
            'resultados': resultados,
            'num_cta': num_cta,
            'nombre': nombre,
            'est_cta': est_cta,
        }
        return render(request, 'lista_cdats.html', context)

from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q

class CtasCdatListView(ListView):
    model = CTAS_AHORRO
    template_name = 'lista_cdats.html'
    context_object_name = 'cdats'
    paginate_by = 10  # Puedes ajustar el número de resultados por página

    def get_queryset(self):
        query = self.request.GET.get('txtbus')
        est_cta = self.request.GET.get('est_cta')
        object_list = CTAS_AHORRO.objects.select_related('asociado__tercero').all()
        if query :
            object_list = object_list.filter(
                Q(num_cta__startswith='04-') & (
                    Q(num_cta__icontains=query) | 
                    Q(asociado__tercero__nombre__icontains=query) |
                    Q(asociado__tercero__doc_ide__icontains=query)
                )
            )
            est_cta = est_cta.strip() if est_cta else None
            if est_cta:
                object_list = object_list.filter(est_cta=est_cta) 
            object_list = object_list.order_by('num_cta')  # Cambia 'num_cta' por el campo que prefieras
            return object_list
        else:
            return CTA_CDAT.objects.none() 
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cdats = self.get_queryset()
        context['resultados'] = [] 
        for cdat in cdats:
            max_ampliacion = CTA_CDAT.objects.filter(cta_aho_id=cdat.id).aggregate(max_ampliacion=Max('ampliacion'))
            cdat0 = CTA_CDAT.objects.filter(cta_aho_id=cdat.id,ampliacion = 0).first()
            if cdat0 == None:
                valor_cdat = 0
            else:
                valor_cdat = cdat0.valor
            max_ampliacion_value = max_ampliacion['max_ampliacion'] or 0
            resultado = {
                'id' : cdat.id,
                'num_cta' : cdat.num_cta,
                'nombre' :  cdat.asociado.tercero.nombre,
                'fecha_apertura' : cdat.fec_apertura,
                'valor' : valor_cdat,
                'estado' : cdat.est_cta,
                'ampliaciones' : max_ampliacion_value
            }
            context['resultados'].append(resultado)
        return context
        
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView
from .models import CTAS_AHORRO
from django.shortcuts import get_object_or_404
from django.core.exceptions import ValidationError
from django.db import transaction
from datetime import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views import View
from django.core.paginator import Paginator
from .models import CTAS_AHORRO, CTA_CDAT
from contabilizacion_lineas_ahorros_app.models import IMP_CON_LIN_AHO
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from datetime import date

class CtasCdatsBaseView(View):
    template_name = 'lista_cdats.html'

    def get_context_data(self, **kwargs):
        print('get_context data base')
        oficina_id = self.request.session.get('oficina_id')
        context = super().get_context_data(**kwargs)        
        pk = self.kwargs.get('pk') if 'pk' in self.kwargs else None
        if pk:
            # Modo actualización
            cta_aho = CTAS_AHORRO.objects.filter(id=pk).first()
            cta_cdat = CTA_CDAT.objects.filter(cta_aho=cta_aho, ampliacion=0).first()
            fecha_form = cta_aho.fec_apertura.strftime('%d/%m/%Y') if cta_cdat.fecha else None
            if cta_cdat is None:
                return context  # Manejar caso donde no existe cta_cdat
            ampliaciones = CTA_CDAT.objects.filter(cta_aho=cta_aho).order_by('ampliacion')
            xNuevaAmpl = -1
            xUltFec = cta_aho.fec_apertura
            xNuavaTiea = 0
            for ampliacion in ampliaciones:
                cdat_amp = CTA_CDAT_AMP.objects.filter(cta_aho_id = cta_aho.id ,cta_amp_id = ampliacion.id).first()
                if cdat_amp != None:
                    if ampliacion.aplicado == 'S': 
                        xNuevaAmpl = ampliacion.ampliacion
                        xUltFec = cdat_amp.fecha
                        xNuavaTiea = ampliacion.tiae

            fecha_amp_form = xUltFec.strftime('%d/%m/%Y') 
            print('xUltFec ------------> ',xUltFec)
            form = CtasCdatsForm(
                cliente_id=cta_aho.asociado.cod_aso, 
                oficina_id=oficina_id,
                operation='update', 
                instance=cta_cdat,
                fecha_inicial=fecha_form,
                fecha_amp_inicial = fecha_amp_form,
            #    tiae_amp_inicial = xUltAmp.tiae if xUltAmp != None or xUltFec else cta_cdat.tiae
            )
            form.fields['num_cta'].initial = cta_aho.num_cta
            form.fields['aplicado'].initial = cta_cdat.aplicado 
            form.fields['aplicado'].widget = forms.Select(choices=[('S', 'Sí'), ('N', 'No')])
            context['num_cta'] = cta_aho.num_cta
            context['cod_aso'] = cta_aho.asociado.cod_aso if cta_aho.asociado else 'No disponible'
            context['fecha'] = fecha_form
            context['asociado_nombre'] = cta_aho.asociado.tercero.nombre if cta_aho.asociado and cta_aho.asociado.tercero else 'Desconocido'
            context['imp_con'] = cta_cdat.imp_con
            context['aplicado'] = 'S' if cta_cdat.aplicado else 'N'
            context['ampliacion'] = xNuevaAmpl + 1
            context['fecha_amp'] = xUltFec 
            context['tiae_amp'] = xNuavaTiea
            context['operation'] = 'update'
            context['button_text'] = 'Modificar'
            
            paginator = Paginator(ampliaciones, 6)  # Mostrar 6 detalles por página
            page_number = self.request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            context['form'] = form
            context['ampliaciones'] = page_obj
            context['paginator'] = paginator
        else:
            form = self.form_class()
            context['ampliacion'] = 0
            context['operation'] = 'create'
            context['button_text'] = 'Guardar'
            context['form'] = form  # Asigna el formulario al contexto
        return context

    def validate_models(self, form):
        errors = {}
        fec_apertura = form.cleaned_data.get('fecha')
        print('Fecha apertura -------------> ',fec_apertura,'   datos  ',type(fec_apertura ))
        if not self.is_valid_date_format(fec_apertura):
            errors['fec_apertura'] = 'La fecha debe estar en formato dd/mm/yyyy.'
        fecha_amp = form.cleaned_data.get('fecha_amp')
        print('Fecha amp -------------> ',fecha_amp,'   datos  ',type(fecha_amp ))
        if not self.is_valid_date_format(fecha_amp):
            errors['fecha_amp'] = 'La fecha debe estar en formato dd/mm/yyyy.'
        cod_aso = form.cleaned_data.get('cod_aso')
        if not cod_aso :
            errors['cod_aso'] = 'Debe Existir el documento del asociado Titular de la Cuenta.'
        imp_con = form.cleaned_data.get('imp_con')
        if not imp_con:
            errors['imp_con'] = 'Debe seleccionar un codigo de imputacion.'
        else:
            try:
                get_object_or_404(IMP_CON_LIN_AHO, pk=imp_con.id)
            except ValidationError:
                errors['li'] = 'la Liea de Ahorro no es valida.'
        print('sale de validate models')
        return errors

    def is_valid_date_format(self,date_value):
        # Verificar si date_value es una cadena
        if isinstance(date_value, str):
            try:
                datetime.strptime(date_value, '%d/%m/%Y')
                return True
            except ValueError:
                return False
        # Verificar si date_value es un objeto datetime
        elif isinstance(date_value, (datetime, date)):
            return True
        else:
            return False

    def save_models(self, form):
        print('entra aqui ?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=?=')
        try:
            with transaction.atomic():
                # Guarda el modelo principal HechoEcono
                CTAS_AHORRO = form.save(commit=False)
                CTAS_AHORRO.save()
                self.save_or_update_details(CTAS_AHORRO.id, form)
                self.save_or_update_accounting_details(CTAS_AHORRO, form)
        except Exception as e:
            # Maneja cualquier error que pueda ocurrir durante el guardado
            raise e

    def handle_errors(self, errors):
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return None
    
from django.db.models import Max, IntegerField
from django.db.models.functions import Cast, Substr
                           
class CtasCdatsCreateView(CtasCdatsBaseView,CreateView):
    model = CTA_CDAT
    form_class = CtasCdatsForm
    template_name = 'ctas_cdat_forms.html'
    success_url = reverse_lazy('ctas_ahorro_list')

    def get_form_kwargs(self):
        kwargs = super(CtasCdatsCreateView, self).get_form_kwargs()
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        kwargs['cliente_id'] = cliente_id
        kwargs['oficina_id'] = oficina_id
        kwargs['operation'] = 'create'
        max_value = CTAS_AHORRO.objects.filter(
            num_cta__startswith='04-',oficina_id = oficina_id
        ).annotate(
        num_cta_int=Cast(Substr('num_cta', 4, 6), IntegerField())
        ).aggregate(
            max_value=Max('num_cta_int')
        )['max_value']
        if max_value is not None:
            max_value_str = '04-'+str(max_value+1).zfill(6)
        else:
            max_value_str = '04-'+'000000'  # Manejo de casos donde no haya registros coincidentes
        kwargs['initial'] = {'num_cta': max_value_str}
        return kwargs
    
    def get_context_data(self, **kwargs):
        print('context data create')
        context = super(CtasCdatsCreateView, self).get_context_data(**kwargs)
        form = self.get_form()  # O self.form_class(initial=self.get_form_kwargs().get('initial'))
        max_value_str = self.get_form_kwargs()['initial'].get('num_cta', None)
        context['num_cta_initial'] = max_value_str
        context['form'] = form  # Asigna el formulario al contexto
        return context

    def form_valid(self, form):
        lin_aho = form.cleaned_data['lin_aho']
        ultimo_num = CTAS_AHORRO.objects.filter(num_cta__startswith=f"{lin_aho.cod_lin_aho}-").count()
        form.instance.num_cta = f"{lin_aho.cod_lin_aho}-{ultimo_num + 1}"
        return super().form_valid(form)

    def post(self, request):
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        oficina_id = self.request.session.get('oficina_id')
        form = CtasCdatsForm(request.POST)
        post_data = request.POST.copy()
        cod_aso = post_data.get('cod_aso')
        if 'fecha' in post_data:
            fecha_str = post_data.get('fecha')
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                post_data['fecha'] = fecha_obj  # Sobrescribir con el objeto de fecha
            except ValueError:
                pass  # Manejar si falla la conversión, si quieres manejar errores aquí
        form = CtasCdatsForm(post_data)
        asociado = ASOCIADOS.objects.filter(oficina_id = oficina_id,cod_aso = cod_aso).first()
        if asociado:
            form.instance.asociado = asociado  
        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    print('Hay Errores y es ajax')
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
                else:
                    print('Hay Errores y NO es ajax')
                    messages.error(request, 'Errores en el formulario.')
                    return render(request, 'ctas_cdat_forms.html', {'form': form})
            cta_aho = CTAS_AHORRO.objects.filter(oficina_id=oficina_id,num_cta=request.POST.get('num_cta')).first()
            if cta_aho != None:
                return JsonResponse({'success': False, 'errors': 'No se puede crear una cuenta que ya existe'}, status=200)
            cta_aho = CTAS_AHORRO.objects.create(oficina_id=oficina_id, 
                lin_aho_id = 2,
                asociado = asociado,
                num_cta = post_data.get('num_cta'),
                est_cta = 'A',
                fec_apertura = post_data.get('fecha'),
                exc_tas_mil = 'S',
                cod_imp = '04'
            )

            print('fecha ',post_data.get('fecha'),'   Tipo ',type(post_data.get('fecha')))
            cdat = CTA_CDAT.objects.create(
                cta_aho = cta_aho,
                imp_con_id = post_data.get('imp_con'),
                ampliacion = 0,
                valor = float(post_data.get('valor')),
                fecha =  post_data.get('fecha'),
                plazo_mes = int(post_data.get('plazo_mes')),
                tiae = float(post_data.get('tiae')),
                Periodicidad = int(post_data.get('Periodicidad')),
                cta_int_ret = post_data.get('cta_int_ret'),
                aplicado = 'N'
            )
            print('tasa int anual ',cdat.tiae,'    otro  ',type(cdat.tiae))
            xTIAE = cdat.tiae/100
            xNumMes = cdat.plazo_mes / cdat.Periodicidad
            xPerAno = 12/cdat.plazo_mes*cdat.Periodicidad
            xTIAN = ((1 + xTIAE) ** (1 / xPerAno) - 1) * xPerAno
            xIntGan = int(cdat.valor * (xTIAN / 12 * xNumMes))
            print('xtIAN  ---',xTIAN,'   ------------> ',xIntGan)
            fecha_int = cdat.fecha + relativedelta(months=xNumMes)
            cdat_amp = CTA_CDAT_AMP.objects.create(
                cta_aho = cta_aho,
                cta_amp = cdat,
                fecha = fecha_int,
                num_liq = 1,
                valor = xIntGan,
                cta_aho_afe = cdat.cta_int_ret,
                aplicado = 'N'
            )

            if is_ajax:
                print('Es Ajax')
                return JsonResponse({
                    'success': True,
                    'num_cta': cta_aho.num_cta,
                    'mensaje': 'La cuenta se ha grabado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
            else:
                print('NO Es Ajax')
                messages.success(request, 'Cuenta de ahorro guardada correctamente. Num_cta -->'+cta_aho.num_cta)
                return redirect('ctas_cdat_forms.html')  # Redirige a otra vista si no es una solicitud AJAX
        else:
            error_list = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_list.append(f'Error en {field}: {error}')
            print('Lista de errores:', error_list)
            if is_ajax:
                return JsonResponse({'success': False, 'errors': error_list}, status=200)
            else:
                return render(request, 'ctas_cdat_forms.html', {'form': form})
        
class CtasCdatsUpdateView(CtasCdatsBaseView,UpdateView):
    model = CTA_CDAT
    form_class = CtasCdatsForm
    template_name = 'ctas_cdat_forms.html'
    success_url = reverse_lazy('ctas_ahorro_list')
    
    def get_object(self, queryset=None): 
        print('Ingresa a Get_objets ....')   
        padre_pk = self.kwargs.get('pk')
        cdat = CTA_CDAT.objects.filter(cta_aho_id=padre_pk, ampliacion=0).first()
        if cdat is None:
            cta = CTAS_AHORRO.objects.filter(id = padre_pk).first()
            cdat = CTA_CDAT.objects.create(cta_aho_id=padre_pk, ampliacion=0)
            cdat.fecha = cta.fec_apertura
            cdat.save()
        return cdat
        
    def get(self, request, *args, **kwargs):
        print('Ingresa a Get ....')   
        self.object = self.get_object()  # Asegúrate de asignar el objeto
        context = self.get_context_data(object=self.object)
        context['pk'] = self.kwargs.get('pk')
        print('Entra por Update ---> ',context)
        return render(request, self.template_name, context)

    def form_valid(self, form):
        lin_aho = form.cleaned_data['lin_aho']
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        print('Entra a Post')
        self.object = self.get_object()
        pk = self.kwargs.get('pk')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        oficina_id = self.request.session.get('oficina_id')
        form = CtasCdatsForm(request.POST)
        post_data = request.POST.copy()
        cod_aso = post_data.get('cod_aso')
        if 'fecha' in post_data:
            fecha_str = post_data.get('fecha')
            fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
            post_data['fecha'] = fecha_obj  # Sobrescribir con el objeto de fecha
        if 'fecha_amp' in post_data:
            fecha_amp_str = post_data.get('fecha_amp')
            fecha_amp_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
            post_data['fecha_amp'] = fecha_amp_obj  # Sobrescribir con el objeto de fecha
        form = CtasCdatsForm(post_data)
        print('paso 1')
        cod_aso = request.POST.get('cod_aso')
        asociado = ASOCIADOS.objects.filter(oficina_id = oficina_id,cod_aso = cod_aso).first()
        if asociado:
            form.instance.asociado = asociado  
        print('paso 2')
        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
                else:
                    messages.error(request, 'Errores en el formulario.')
                    return render(request, 'hecho_econo_form.html', {'form': form})  
            fecha_str = request.POST.get('fecha')
            fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
            cta_aho = CTAS_AHORRO.objects.filter(id = pk).first()
            PorAplicar = post_data.get('sinaplicar')
            if PorAplicar == 'S':
                xAmp = request.POST.get('ampliacion') 
                fecha_amp_str = request.POST.get('fecha_amp')
                fecha_amp = datetime.strptime(fecha_amp_str, '%d/%m/%Y').date()
                print('Ampliacion --->',xAmp,'  Fecha Ampl ',fecha_amp,'  Valor ',float(post_data.get('valor')))
                cta_aho = CTAS_AHORRO.objects.filter(id = pk).first()   
                cta_cdat = CTA_CDAT.objects.filter(cta_aho_id = pk,ampliacion = xAmp,cta_aho = cta_aho).first()
                if cta_cdat == None:
                    cta_cdat = CTA_CDAT.objects.create(cta_aho_id = pk,ampliacion = xAmp,cta_aho = cta_aho)
                #cta_cdat.imp_con = post_data.get('imp_con'),
                cta_cdat.valor = float(post_data.get('valor'))
                cta_cdat.fecha =  fecha_amp
                cta_cdat.plazo_mes = int(post_data.get('plazo_mes'))
                cta_cdat.tiae = float(post_data.get('tiae_amp'))
                cta_cdat.Periodicidad = int(post_data.get('Periodicidad'))
                cta_cdat.cta_int_ret = post_data.get('cta_int_ret')
                cta_cdat.aplicado = 'N'
                cta_cdat.save()
                xTIAE = cta_cdat.tiae/100
                xNumMes = cta_cdat.plazo_mes / cta_cdat.Periodicidad
                xPerAno = 12/cta_cdat.plazo_mes*cta_cdat.Periodicidad
                xTIAN = ((1 + xTIAE) ** (1 / xPerAno) - 1) * xPerAno
                xIntGan = int(cta_cdat.valor * (xTIAN / 12 * xNumMes))
                fecha_int = fecha_amp + relativedelta(months=xNumMes)
                cta_cdat_amp = CTA_CDAT_AMP.objects.filter(cta_aho = cta_aho,cta_amp = cta_cdat,fecha = fecha_int).first()
                if cta_cdat_amp  == None:
                    cta_cdat_amp = CTA_CDAT_AMP.objects.create(cta_aho = cta_aho,cta_amp = cta_cdat,fecha = fecha_int)
                cta_cdat_amp.num_liq = 1
                cta_cdat_amp.valor = xIntGan
                cta_cdat_amp.cta_aho_afe = cta_cdat.cta_int_ret
                cta_cdat_amp.aplicado = 'N'            
                cta_cdat_amp.save()
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'num_cta': cta_aho.num_cta,
                    'mensaje': 'La cuenta se ha grabado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
           
            else:
                messages.success(request, 'Cuenta de ahorro guardada correctamente. Num_cta -->'+cta_ahorro.num_cta)
                return redirect('nombre_de_la_vista')  # Redirige a otra vista si no es una solicitud AJAX
        else:
            return JsonResponse({'success': False, 'errors': ['Método no permitido']}, status=405)

class CtasAhorroDeleteView(DeleteView):
    model = CTAS_AHORRO
    template_name = 'ctas_ahorro_confirm_delete.html'
    success_url = reverse_lazy('ctas_ahorro_list')

class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cdats = CTA_CDAT.objects.all().order_by(
            'cta_aho', 'cta_amp', 'fecha')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cdats.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in cdats:
        # p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
        # p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
        # p.drawString(80, 760, f"Fecha: {dato.fecha}")
        # p.drawString(80, 740, f"Número Liquidación: {dato.num_liq}")
        # p.drawString(80, 720, f"Valor: {dato.valor}")
        # p.drawString(80, 700, f"Cuenta de Ahorro Afectada: {dato.cta_aho_afe}")
        # p.drawString(80, 680, f"Número Documento: {dato.docto}")
        # p.drawString(80, 660, f"Clase: {dato.clase}")
        # p.drawString(80, 640, f"Documento: {dato.documento}")
        # p.drawString(80, 620, f"Aplicado?: {dato.aplicado}")

        datos_tabla = [["Cuenta de Ahorro", "Cuenta Ampliación", "Fecha", "Número Liquidación",
                        "Valor", "Cuenta de Ahorro Afectada", "Número Documento", "Clase", "Documento", "Aplicado?"]]

        for dato in cdats:
            datos_tabla.append([dato.cta_aho, dato.cta_amp, dato.fecha, dato.num_liq, dato.valor,
                               dato.cta_aho_afe, dato.docto, dato.clase, dato.documento, dato.aplicado])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        padre = CTA_CDAT.objects.get(pk=pk)
        padre0 = CTA_CDAT.objects.filter(cta_aho = padre.cta_aho,ampliacion = 0).first()
        hijo = CTA_CDAT_AMP.objects.filter(cta_amp_id = pk).first()
        nietos = CTA_CDAT_LIQ.objects.filter(cta_amp_id = hijo.id).order_by('fecha')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cdats.pdf"'

        p = canvas.Canvas(response)
        val_cdat_for = "${:,.2f}".format(padre0.valor)
        p.drawString(80, 800, f"Cuenta Aho.: {padre0.cta_aho.num_cta} Ampliación: {padre.ampliacion} Fecha Amp.: {padre.fecha} Valor Cdat: {val_cdat_for}")
        p.drawString(80, 780, f"Cuenta Aho pago    {padre0.cta_int_ret}   Tasa IEA : {padre.tiae}    Fecha de pago ampliacion: {hijo.fecha}")
        p.drawString(80, 750, "Fecha          Tipo                          Valor_int                Valor_ret   ")
        col = 720
        for nieto in nietos :
            val_int_for = "${:,.2f}".format(nieto.val_int)
            val_ret_for = "${:,.2f}".format(nieto.val_ret)
            if nieto.tip_liq == 'D':
                tipo = "Caus. Fin de Mes"
            elif nieto.tip_liq == 'C':
                tipo = "Caus. Vencimiento"
            elif nieto.tip_liq == 'P':
                tipo = "Traslado a Cuenta"
            p.drawString(80,col, f"{nieto.fecha}")
            p.drawString(150,col, tipo)
            p.drawString(260,col, f"{val_int_for}")
            p.drawString(360,col, f"{val_ret_for}")
            col-=20

        p.showPage()
        p.save()

        return response

class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cdats = CTA_CDAT.objects.all()

        p = canvas.Canvas(response)

        for dato in cdats:
            p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
            p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
            p.drawString(80, 760, f"Fecha: {dato.fecha}")
            p.drawString(80, 740, f"Número Liquidación: {dato.num_liq}")
            p.drawString(80, 720, f"Valor: {dato.valor}")
            p.drawString(
                80, 700, f"Cuenta de Ahorro Afectada: {dato.cta_aho_afe}")
            p.drawString(80, 680, f"Número Documento: {dato.docto}")
            p.drawString(80, 660, f"Clase: {dato.clase}")
            p.drawString(80, 640, f"Documento: {dato.documento}")
            p.drawString(80, 620, f"Aplicado?: {dato.aplicado}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CTA_CDAT._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cdats = CTA_CDAT.objects.all()
        for row_num, data in enumerate(cdats, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cdats = CTA_CDAT.objects.all()
        for data in cdats:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response

#--------------------------------------------------------------------------------------------------------------------------------
from django.forms.models import model_to_dict

def buscar_cta_int(request, codigo):
    if len(codigo) < 4:
        return JsonResponse({'message': 'Sin Datos'}, status=404)
    asociado = ASOCIADOS.objects.filter(oficina_id = 1,cod_aso=codigo).first()
    if asociado == None:
        return JsonResponse({'message': 'Sin Datos', 'results': []}, status=200)
    results = []
    ctas_ahorro = CTAS_AHORRO.objects.filter(oficina_id=1, asociado=asociado).exclude(num_cta__startswith='04-')
    for cta_aho in ctas_ahorro:
        results.append({
            'num_cta': cta_aho.num_cta,
            'fec_apertura': cta_aho.fec_apertura,
            'estado': cta_aho.est_cta,
        })
    if not results:
        return JsonResponse({'message': 'Sin Datos', 'results': []}, status=200)
    return JsonResponse(results, safe=False)


from django.shortcuts import render
from django.http import HttpResponse

import os

def generar_reportes_cdat(request):
    if request.method == 'POST':
        form = ReporteCdatForm(request.POST)
        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            directorio = form.cleaned_data['directorio']
            nombre_archivo = form.cleaned_data['nombre_archivo']
            accion = request.POST.get('accion')
            # Validar que el directorio exista
            if not os.path.exists(directorio):
                return HttpResponse(f"Error: El directorio '{directorio}' no existe.", status=400)
            ruta_completa = os.path.join(directorio, nombre_archivo)
            # Simulación de la generación del reporte
            if accion == "vencidos":
                mensaje = f"Reporte de vencidos generado en: {ruta_completa}"
            elif accion == "nuevos":
                mensaje = f"Reporte de nuevos generado en: {ruta_completa}"
            else:
                mensaje = "Acción no válida."
            return HttpResponse(mensaje)
    else:
        form = ReporteCdatForm()

    return render(request, 'reportes_cdat.html', {'form': form})


class ImprimeReporteCdatPDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cdats.pdf"'

        p = canvas.Canvas(response)
        col = 720

        p.showPage()
        p.save()

        return response
    
from django.http import JsonResponse
from django.shortcuts import render

def imprimirRepCdat(request):
    if request.method == "POST":
        directorio = request.POST.get("directorio")
        nombre_archivo = request.POST.get("nombre_archivo")
        if not directorio or not nombre_archivo:
            return JsonResponse({"error": "Debes ingresar un directorio y un nombre de archivo"}, status=400)

        return JsonResponse({"mensaje": f"Reporte guardado en {directorio}/{nombre_archivo}"})

    return render(request, "reportes.html")

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt  # Para pruebas, aunque lo mejor es usar CSRF en producción
def guardar_directorio(request):
    if request.method == "POST":
        data = json.loads(request.body)
        request.session["directorio_guardado"] = data.get("directorio")  # Guardar en sesión
        return JsonResponse({"mensaje": "Directorio guardado correctamente"})
    return JsonResponse({"error": "Método no permitido"}, status=405)


import os
import pandas as pd
from django.http import JsonResponse

import os

def repCdatsVecidos(request):
    if request.method == "POST":
        fecha_inicio_str = request.POST.get("fecha_inicial", "").strip()
        fecha_fin_str = request.POST.get("fecha_final", "").strip()
        print(f"Fecha inicio recibida: {fecha_inicio_str} tipo {type(fecha_inicio_str)}  ")  
        print(f"Fecha fin recibida: {fecha_fin_str}")  # Depuración
        vencidos = CTA_CDAT_AMP.objects.filter(aplicado = 'S',fecha__range=(fecha_inicio_str, fecha_fin_str))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cdats.pdf"'
        p = canvas.Canvas(response)
        #p.setFont("Helvetica", 8)  # Reducir tamaño de fuente
        p.setFont("Courier", 10)
        p.drawString(180, 800, f"Reporte de Cdats Vencidos entre {fecha_inicio_str} y {fecha_fin_str}")
        p.drawString(180, 780, f"====== == ===== ========   ===== ============== ===")
        p.drawString(10, 750, "FECHA              CDAT    AMPL.  DOC_IDE        NOMBRE                                                            TIEA              VALOR          VAL. INTER")
        
        col = 720
        for vencido in vencidos:
            val_tiea = "{:,.2f}".format(vencido.cta_amp.tiae)
            cdat = CTA_CDAT.objects.filter(cta_aho=vencido.cta_aho, ampliacion=0).first()
            val_cdat = "${:,.0f}".format(cdat.valor)
            val_int = "${:,.0f}".format(vencido.valor)
            nombre = vencido.cta_aho.asociado.tercero.nombre[:32]  # Limitar a 30 caracteres
            p.drawString(10, col, f"{vencido.fecha}")
            p.drawString(70, col, f"{vencido.cta_aho.num_cta}")
            p.drawString(114, col, f"{vencido.cta_amp.ampliacion}")
            p.drawString(130, col, f"{vencido.cta_aho.asociado.tercero.doc_ide}")
            p.drawString(180, col, f"{nombre}")
            p.drawRightString(410, col, f"{val_tiea}")   # Ajustado a la derecha en X=410
            p.drawRightString(490, col, f"{val_cdat}")   # Ajustado a la derecha en X=470
            p.drawRightString(560, col, f"{val_int}")    # Ajustado a la derecha en X=560  
            col -= 12  # Reducir el espacio entre filas
            if col < 50:  
                p.showPage()  # Guardar la página actual
                p.setFont("Helvetica", 8)  # Restablecer fuente
                col = 750  # Reiniciar la posición en la nueva página

        p.showPage()
        p.save()
    return response
    
        #directorio_base = request.POST.get("directorio", "").strip()
        #nombre_archivo = request.POST.get("nombre_archivo", "").strip()
        #formato = request.POST.get("formato", "").strip()
        #if not directorio_base or not nombre_archivo:
        #    return JsonResponse({"error": "Debes ingresar un nombre de archivo o una ruta válida."})
        #extension = ".pdf" if formato == "pdf" else ".xlsx"
        #if not nombre_archivo.endswith(extension):
        #    nombre_archivo += extension
        #ruta_completa = os.path.join(directorio_base, nombre_archivo)
        #directorio_final = os.path.dirname(ruta_completa)
        #formato = request.POST.get("formato")  # Obtiene "pdf" o "xlsx"
        #if not formato:
        #    formato = "pdf"  # Valor por defecto si no se selecciona nada
        #print(f"Formato seleccionado: {formato}")
        #if formato == "pdf":
        #    generar_pdf()
        #elif formato == "xlsx":
        #    generar_excel()


#-------------------------------------------CORIFIJO----------------------------------------------------


class ImprimirTitulo(View):
    def get(self, request, *args, **kwargs):
        id_ofi = request.session.get('oficina_id')
        oficina = OFICINAS.objects.filter(id=id_ofi).first()
        
        pk = kwargs.get("pk")
        padre = CTA_CDAT.objects.get(pk=pk)
        padre0 = CTA_CDAT.objects.filter(cta_aho = padre.cta_aho,ampliacion = 0).first()
        hijo = CTA_CDAT_AMP.objects.filter(cta_amp_id = pk).first()
        nietos = CTA_CDAT_LIQ.objects.filter(cta_amp_id = hijo).order_by('fecha')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cdats.pdf"'
        p = canvas.Canvas(response, pagesize= letter)
        p.setFont("Courier-Bold", 11)
        val_cdat_for = "${:,.2f}".format(padre0.valor)
        total_intereses = 0
        total_retencion = 0
        ultima_fecha = None
        col = 720

        for nieto in nietos:
    
            if nieto.tip_liq == 'D':
                tipo = "Caus. Fin de Mes"
            elif nieto.tip_liq == 'C':
                tipo = "Caus. Vencimiento"
            elif nieto.tip_liq == 'P':
                tipo = "Traslado a Cuenta"

            total_intereses += float(nieto.val_int)
            total_retencion += float(nieto.val_ret)
            ultima_fecha = nieto.fecha
            col -= 20

        # # Mostrar totales y última fecha
        # val_int = "${:,.2f}".format(total_intereses)
        # val_ret = "${:,.2f}".format(total_retencion)


        p.drawString(119.07, 439.425, f"{-nieto.val_int:,} ")
        p.drawString(286.335, 439.425, f"{-nieto.val_ret:,}")
        p.drawString(414.745, 439.425, f"{ultima_fecha.strftime('%d  %m  %Y') }")

        #Fila 1
        p.drawString(85.05,654.885, f"{oficina.ciudad.nombre.strip()}")#####pendiente
        p.drawString(201.285,654.885, f"{oficina.nombre_oficina.upper()}")#####pendiente
        p.drawString(294.84,654.885, f"{padre0.cta_aho.num_cta} ")
        p.drawString(410.800,654.885, f"{padre.fecha.strftime('%d  %m  %Y') }")
        #Fila 2
        p.drawString(155.925,620.865, f"{padre0.cta_aho.asociado.tercero.nombre}")
        p.drawString(435.59,620.865, f"{padre0.cta_aho.asociado.tercero.doc_ide}")
        #Fila 3
        
        # Dibujar valor en letras 
        pesos = ''
        if padre.valor == 0:
            pesos = 'PESOS M/CTE.'
        elif padre.valor % 1_000_000 == 0:
            pesos = 'DE PESOS M/CTE.'
        else:
            pesos = 'PESOS M/CTE.'

        numero_en_letras = num2words(padre.valor, lang='es').upper()
        num_let_formateado = f"{numero_en_letras} {pesos}"  

        p.drawString(95,547.155, f"{num_let_formateado}")
        #Fila 4
        p.drawString(95,525.475, "***************************")
        p.drawString(445.095,524.475, f"{ padre.valor:,}")
        #Fila 5
        mes = padre0.plazo_mes
        mes= mes*30
        p.drawString(85.05,479.115, f"{mes}")
        p.drawString(195.625,479.115, "Tasa nominal")##Formula    
        p.drawString(294.84,479.115, f"{padre.tiae}")
        p.drawString(385.56,479.115, "vencido el plazo")
    
        p.showPage()
        p.save()

        return response




