import csv, os
from django.db import connection
from django.utils.timezone import make_aware
import pandas as pd
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth
from math import ceil
from reportlab.lib.pagesizes import landscape, letter, legal
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.db import models
from django.db.models import Sum, Max, Min, Value, FloatField, IntegerField, DateField, Case, When, F, DecimalField, Count, ExpressionWrapper, Subquery, OuterRef, Q
from django.db.models.functions import Cast, Coalesce, Now, ExtractYear, ExtractMonth, TruncDate, ExtractDay
from django.core.paginator import Paginator
from datetime import date, timedelta, datetime
from itertools import chain
from django import forms
from dateutil.relativedelta import relativedelta  
from pagadores_app.models import PAGADORES

from .forms import CrearForm
from .models import PLAN_APORTES
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from asociados_app.models import ASOCIADOS
from terceros_app.models import TERCEROS
from detalle_producto_app.models import DETALLE_PROD
from hecho_economico_app.models import HECHO_ECONO
from creditos_app.models import CREDITOS
from estados_financieros_app.models import ESTADOS_FIN
from localidades_app.models import LOCALIDADES
from recla_carte_app.models import CARTE_CAT_HIS
from conceptos_app.models import CONCEPTOS
from justo_app.opciones import OPC_EST_SOCIO, OPC_EDUCACION, OPC_EST_CIV
from justo_app.funciones_principales import formato_fecha, formatear_cod_aso

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = PLAN_APORTES
    form = CrearForm
    template_name = 'lista_aportes.html'
    ordering = ['oficina','-agno']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        for obj in queryset:
            # Calculamos el aporte mensual
            try:
                obj.aporte_mensual = obj.totadu / obj.meses if obj.meses else 0
            except (TypeError, ZeroDivisionError):
                obj.aporte_mensual = 0
        return queryset

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = PLAN_APORTES
    form = CrearForm
    template_name = 'detalles_aporte.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = PLAN_APORTES
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_aporte.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_aporte')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = PLAN_APORTES
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_aporte.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_aporte')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = PLAN_APORTES
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_aporte')

# Primera consulta: Saldo Anterior
def ejecutar_consulta_orm(oficina_id,subcuenta,fecha_ini,fecha_fin):
    # Primera parte de la consulta
    saldo_anterior = DETALLE_PROD.objects.filter(
        producto = 'AP',
        subcuenta = subcuenta,
        hecho_econo__fecha__lt = fecha_ini - timedelta(days = 1),
        hecho_econo__docto_conta__oficina_id=oficina_id
    ).aggregate(
        Saldo=Coalesce(-Sum('valor'), 0, output_field=FloatField())
    )
    resultado_saldo_anterior = {
        'Fecha' : fecha_ini - timedelta(days = 1),
        'Comprobante' : 'Saldo Anterior',
        'Numero' : 0,
        'Concepto': '',
        'Aporte' : 0,
        'Retiro' : 0,
        'Saldo' : saldo_anterior['Saldo']
    }
    
    # Segunda parte de la consulta
    movimientos = DETALLE_PROD.objects.filter(
        producto='AP',
        subcuenta=subcuenta,
        hecho_econo__fecha__range=(fecha_ini, fecha_fin),
        hecho_econo__docto_conta__oficina_id=oficina_id
    ).annotate(
        Fecha=F('hecho_econo__fecha'),
        Comprobante=F('hecho_econo__docto_conta__nombre'),
        Numero=F('hecho_econo__numero'),
        Concepto=F('concepto'),
        Aporte=Case(
            When(valor__lt=0, then=-F('valor')),
            default=Value(0),
            output_field=FloatField()
        ),
        Retiro=Case(
            When(valor__gt=0, then=F('valor')),
            default=Value(0),
            output_field=FloatField()
        ),
        Saldo=Value(0, output_field=FloatField())
    ).values(
        'Fecha', 'Comprobante', 'Numero', 'Concepto', 'Aporte', 'Retiro', 'Saldo'
    ).order_by('Fecha')

    # Unir las dos partes
    resultados = [resultado_saldo_anterior] + list(movimientos)    
    return resultados

def movtos_aporte_socio(request, cliente_id=None, oficina_id=None):
    if request.method == 'GET':
        accion = request.GET.get("accion")
        filtro_socio = request.GET.get('filtro_cod_socio', '').strip()
        fecha_ini = request.GET.get('fecha_inicio', None)
        fecha_fin = request.GET.get('fecha_final', None)
        from datetime import datetime
        date_format = "%Y-%m-%d"
        fecha_actual = datetime.now()
        if not fecha_ini:
            fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
        else:
            fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
        if not fecha_fin:
            fecha_final = fecha_actual.date()
        else:
            fecha_final = datetime.strptime(fecha_fin, date_format).date()
        oficina_id = int(request.GET.get('oficina_id', 1))
        if len(filtro_socio) > 3:
            resultados = ejecutar_consulta_orm(oficina_id, filtro_socio, fecha_inicio, fecha_final)
        else:
            resultados = []
        sal_acu = 0
        rows = []
        for row in resultados:
            sal_acu += float(row['Aporte']) - float(row['Retiro']) + float(row['Saldo'])
            upd_row = row
            upd_row['Saldo'] = sal_acu
            rows.append(upd_row)
        print('filtro_socio  ',filtro_socio)
        saldo_aporte = saldo_aporte_socio_fecha(1,filtro_socio,fecha_final)
        saldo_esperado = aporte_esperado_fecha(filtro_socio,fecha_final)
        from django.core.paginator import Paginator
        paginator = Paginator(rows, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        if accion == "exportar":
            return exportar(request, rows)
        elif accion == "csv":
            return exportar_csv(request, rows)
        elif accion == 'imprimir':
            return imprimir(request, rows)

        return render(request, 'movtos_apo_soc.html', {
            'context': page_obj,
            'page_obj': page_obj,
            'filtro_cod_socio': filtro_socio,
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_final': fecha_final.strftime('%Y-%m-%d'),
            'saldo_aporte': saldo_aporte,
            'aporte_esperado': saldo_esperado,
        })
    return HttpResponse("Método no permitido", status=405)

def exportar(request, rows):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    filtro_socio = request.GET.get('filtro_cod_socio', '').strip()
    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    print('Filtro  ',filtro_socio,'    fecha_ini ',fecha_ini,'   fecha_fin ',fecha_fin)
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))
  
    # Lógica para exportar Crear el libro de Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimiento Aportes" 

    entidad = CLIENTES.objects.filter(id=id_cli).first()
    oficina = OFICINAS.objects.filter(id=id_ofi).first()
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()

    # Llama a la función para obtener los datos
    resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)

    if len(filtro_socio) > 3: 
        resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        sal_acu = sal_acu + float(row['Aporte']) - float(row['Retiro']) + float(row['Saldo'])
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)
                    
    if not resultados:
        return HttpResponse("No se encontraron datos para exportar", status=404)
    
    fecha_inicio_formateada = formato_fecha(fecha_inicio)
    fecha_final_formateada = formato_fecha(fecha_final)
                        
    # Añadir datos adicionales encima de los encabezados
    empresa = f"{entidad.nombre.strip()}"
    nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
    reporte = "MOVIMIENTO DE APORTES "
    asociado = f"Asociado: {filtro_socio} {tercero.nombre}"
    detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fecha_inicio_formateada} a {fecha_final_formateada}"
                        
    datos_adicionales = [
        [empresa],
        [nit_empresa],
        [reporte],
        [asociado],
        [detalles]                
    ]

    # Insertar los datos adicionales
    for row_num, fila in enumerate(datos_adicionales, start=1):
        for col_num, valor in enumerate(fila, start=1):
            sheet.cell(row=row_num, column=col_num, value=valor)

    # Determinar la fila donde empiezan los encabezados
    header_row = len(datos_adicionales) + 1

    # Añadir los encabezados
    headers = resultados[0].keys()
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_num, value=header)

    # Añadir los datos
    for row_num, data in enumerate(resultados, start=header_row + 1):
        for col_num, field in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=data[field])

    nombre_archivo = f"mov_aportes_{tercero.nombre.strip()}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    workbook.save(response)
    return response
            
def exportar_csv(request, rows):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    filtro_socio = request.GET.get('filtro_cod_socio', '').strip()
    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    print('Filtro  ',filtro_socio,'    fecha_ini ',fecha_ini,'   fecha_fin ',fecha_fin)
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))
        
    # Llama a la función para obtener los datos
    resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)

    if len(filtro_socio) > 3: 
        resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        sal_acu = sal_acu + float(row['Aporte']) - float(row['Retiro']) + float(row['Saldo'])
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)

    if not resultados:
        return HttpResponse("No se encontraron datos para exportar", status=404)
    
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()         
    # Configurar la respuesta HTTP para un archivo CSV
    nombre_archivo = f"mov_aportes_{tercero.nombre.strip()}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                
    # Crear el escritor CSV
    writer = csv.writer(response)
                
    # Añadir encabezados de las columnas
    headers = resultados[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
    writer.writerow(headers)

    # Añadir los datos
    for fila in resultados:
        writer.writerow([fila[col] for col in headers])
    return response
            
def imprimir(request, rows):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    filtro_socio = request.GET.get('filtro_cod_socio', '').strip()
    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    print('Filtro  ',filtro_socio,'    fecha_ini ',fecha_ini,'   fecha_fin ',fecha_fin)
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    
    fecha_inicio_formateada = formato_fecha(fecha_inicio)
    fecha_final_formateada = formato_fecha(fecha_final)    
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))

    resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)

    if len(filtro_socio) > 3: 
        resultados = ejecutar_consulta_orm(oficina_id,filtro_socio,fecha_inicio,fecha_final)
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        sal_acu = sal_acu + float(row['Aporte']) - float(row['Retiro']) + float(row['Saldo'])
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)
        
    # Lógica para imprimir o generar vista previa  Configuración del PDF
    entidad = CLIENTES.objects.filter(id=id_cli).first()
    oficina = OFICINAS.objects.filter(id=id_ofi).first()
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()

    nombre_archivo = f"mov_aportes_{tercero.nombre.strip()}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    margin_x, margin_y = 50, 60

    # Configuración inicial
    filas_por_pagina = 64 # Número máximo de filas por página
    total_filas = len(resultados)
    total_paginas = ceil(total_filas / filas_por_pagina)

    # Función para dibujar encabezado
    def dibujar_encabezado():
        empresa = f"{entidad.nombre.strip()}"
        texto_empresa = stringWidth(empresa, "Times-Roman", 12)
        p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                    
        nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
        texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
        p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

        reporte = "MOVIMIENTO DE APORTES"
        texto_reporte = stringWidth(reporte, "Times-Roman", 12)
        p.drawString((width - texto_reporte) / 2, height - 60, reporte)

        p.setFont("Helvetica", 10)
        detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fecha_inicio_formateada} a {fecha_final_formateada}"
        texto_detalles = stringWidth(detalles, "Helvetica", 10)
        p.drawString((width - texto_detalles) / 2, height - 75, detalles)

        asociado = f"Asociado: {filtro_socio} {tercero.nombre.strip()}"
        texto_filtro = stringWidth(asociado, "Helvetica", 12)
        p.drawString((width - texto_filtro) / 2, height - 90, asociado)

        p.setFont("Helvetica", 8)
        texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
        p.drawRightString(width - margin_x, height - 90, texto_paginas)

        p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 20, height - margin_y - 10, 60, 60)

    # Función para dibujar pie de página
    def dibujar_pie(pagina_actual, total_paginas):
        line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
        p.line(margin_x-40, line_y, margin_x+552, line_y)  # Dibuja la línea

        p.setFont("Courier", 9)
        texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
        p.drawString(margin_x-20, margin_y - 40, texto_pie)

    # Función para dibujar encabezado de tabla
    def dibujar_encabezado_tabla(y):
        line_y = height - 115  # Ajusta para que esté justo encima del texto
        p.line(margin_x-40, line_y, margin_x+552, line_y)
        p.line(margin_x-40, height - 100, margin_x+552, height - 100)
        columnas = [
            ("Fecha", 50), 
            ("Comprobante", 180), 
            ("Número", 50), 
            ("Concepto", 60), 
            ("        Aporte", 75), 
            ("        Retiro", 75), 
            ("        Saldo", 80)
            ]
        x = margin_x-30
        p.setFont("Helvetica-Bold", 9)
        for col, ancho in columnas:
            p.drawString(x, y, col)
            x += ancho
                        
    # Dibujar contenido
    pagina_actual = 1
    dibujar_encabezado()
    dibujar_pie(pagina_actual, total_paginas)
    dibujar_encabezado_tabla(height - margin_y - 50)
    y = height - margin_y - 65

    def dibujar_fila(y, row):
        x = margin_x-30
        columnas = [
            (row['Fecha'], 50), 
            (row['Comprobante'],180), 
            (row['Numero'], 50),
            (row['Concepto'], 50), 
            (f"{row['Aporte']:,.2f}",70,'right'), 
            (f"{row['Retiro']:,.2f}", 70,'right'), 
            (f"{row['Saldo']:,.2f}", 80, 'right')
            ]

        p.setFont("Helvetica", 9)
        for col in columnas:
            if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                texto, ancho, _ = col
                p.drawRightString(x + ancho, y, texto)
            else:  # Alineación izquierda
                texto, ancho = col[:2]
                p.drawString(x, y, str(texto))
            x += ancho

    for idx, row in enumerate(resultados):
        if y < margin_y - 30:
            p.showPage()
            pagina_actual += 1
            dibujar_encabezado()
            dibujar_encabezado_tabla(height - margin_y - 50)
            dibujar_pie(pagina_actual, total_paginas)
            y = height - margin_y - 65    

        dibujar_fila(y, row)
        y -= 10
    p.showPage()
    p.save()
    return response

def liquidar_aportes(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    if request.method == 'GET':
        return render(request, 'aportes_a_la_fecha.html')  # tu formulario
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_corte')
        try:
            fecha_corte = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha inválida", status=400)
        resultados = []
        print('Hora de Inicio ',datetime.now())
        result= (
            DETALLE_PROD.objects.filter(
                producto = 'AP',
                hecho_econo__fecha__lte = fecha_corte,
                hecho_econo__docto_conta__oficina_id=1
            )
            .values('subcuenta')  # Agrupamos por 'subcuenta'
            .annotate(
                cod_aso=F('subcuenta'),
                Aporte_fecha=-Sum(F('valor')),  # Negamos la suma como en la SQL
                fec_ult_apo=Max('hecho_econo__fecha')
            )
        )
        result_list = list(result)
        for item in result_list:
            if item['Aporte_fecha'] is not None or item['fec_ult_apo'] is not None:
                asociado = ASOCIADOS.objects.filter(oficina_id = 1,cod_aso = item['cod_aso']).first()
                if asociado == None:
                    print('No Hay Asociado con código ',item['cod_aso'])
                    continue
                if asociado.tercero_id == None:
                    print('Error de Integridad cod_aso ',asociado.cod_aso)
                    continue
                tercero = TERCEROS.objects.filter(id = asociado.tercero_id).first()
                if tercero != None:
                    resultados.append({
                        'cod_aso': asociado.cod_aso,
                        'nombre': tercero.nombre,
                        'Aporte_fecha': item['Aporte_fecha'],
                        'fec_ult_apo': item['fec_ult_apo']
                    })

        wb = Workbook()
        ws = wb.active
        ws.title = "Saldos_aportes_a_la_fecha"
        titulo = f"Saldo de aportes a {fecha_corte}"
        ws.merge_cells('A1:E1')
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:E2')
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        headers = ['cod_aso','Nombre','Aporte','fec_ult_pag']
        ws.append([])  # Fila 3 vacía
        ws.append(headers)
        for col_index in range(1, len(headers) + 1):
            cell = ws.cell(row=4, column=col_index)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        # Volcar los resultados en la hoja de Excel
        fila_inicio_datos = 5  # La fila donde empiezan los datos
        fila_actual = fila_inicio_datos
        total_aporte = 0
        for item in resultados:
            if item['Aporte_fecha'] == 0:
                continue
            ws.append([
                item['cod_aso'],
                item['nombre'],
                item['Aporte_fecha'],
                item['fec_ult_apo']
            ])
            # Acumular el total (solo si Aporte_fecha es numérico)
            if isinstance(item['Aporte_fecha'], (int, float)):
                total_aporte += item['Aporte_fecha']
            elif isinstance(item['Aporte_fecha'], str):
                try:
                    total_aporte += float(item['Aporte_fecha'])
                except:
                    pass
            fila_actual += 1
        ws.cell(row=fila_actual, column=2).value = "TOTAL"
        ws.cell(row=fila_actual, column=2).font = Font(bold=True)
        ws.cell(row=fila_actual, column=3).value = total_aporte
        ws.cell(row=fila_actual, column=3).font = Font(bold=True)


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"aportes_a_{fecha_corte}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

def saldo_aportes_fecha(oficina_id, cod_socio, fecha_corte):
    resultado = (
        DETALLE_PROD.objects.filter(
            producto='AP',
            subcuenta=cod_socio,
            hecho_econo__fecha__lte=fecha_corte,
            hecho_econo__docto_conta__oficina_id=oficina_id
        )
        .aggregate(
            Fecha=Max('hecho_econo__fecha'),
            Aporte_fecha=Coalesce(-Sum('valor', output_field=FloatField()), 0.0)
        )
    )
    if resultado['Fecha'] is None:
        return {
            'cod_soc': cod_socio,
            'Fecha': fecha_corte,
            'Aporte_fecha': 0.0
        }
    return {
        'cod_soc': cod_socio,
        'Fecha': resultado['Fecha'],
        'Aporte_fecha': resultado['Aporte_fecha']
    }

def saldo_aporte_socio_fecha(oficina_id, subcuenta, fecha_corte):
    saldo = (
        DETALLE_PROD.objects
            .filter(
                oficina_id = oficina_id,
                producto = "AP",
                subcuenta = subcuenta,
                hecho_econo__fecha__lte = fecha_corte
            )
        .aggregate(saldo=Coalesce(Sum("valor"), Value(0, output_field=FloatField())))
    ) 
    return -saldo["saldo"]

def promedio_aporte_socio_fecha(oficina_id, subcuenta, fecha_corte):
    fecha_inicio = date(fecha_corte.year, 1, 1)
    # fecha_inicio = fecha_corte.replace(year=fecha_corte.year - 1)
    saldo = (
        DETALLE_PROD.objects
            .filter(
                oficina_id=oficina_id,
                producto="AP",
                subcuenta=subcuenta,
                hecho_econo__fecha__gt=fecha_inicio,
                hecho_econo__fecha__lte=fecha_corte,
                valor__lt=0
            )
            .aggregate(saldo=Coalesce(Sum("valor"), Value(0, output_field=FloatField())))
    )
    return -saldo["saldo"]/365

def saldo_aporte_extra_socio_fecha(oficina_id, subcuenta, fecha_corte):
    extra = CONCEPTOS.objects.filter(tip_dev_ap='E').values_list('cod_con', flat=True)
    saldo = (
        DETALLE_PROD.objects
            .filter(
                oficina_id=oficina_id,
                producto="AP",
                subcuenta=subcuenta,
                hecho_econo__fecha__lte=fecha_corte,
                concepto__in=extra
            )
            .aggregate(saldo=Coalesce(Sum("valor"), Value(0, output_field=FloatField())))
    )
    return -saldo["saldo"]

def saldo_aporte_voluntario_socio_fecha(oficina_id, subcuenta, fecha_corte):
    voluntario = CONCEPTOS.objects.filter(tip_dev_ap='V').values_list('cod_con', flat=True)
    saldo = (
        DETALLE_PROD.objects
            .filter(
                oficina_id=oficina_id,
                producto="AP",
                subcuenta=subcuenta,
                hecho_econo__fecha__lte=fecha_corte,
                concepto__in=voluntario
            )
            .aggregate(saldo=Coalesce(Sum("valor"), Value(0, output_field=FloatField())))
    )
    return -saldo["saldo"]

def saldo_revalorizacion_aportes_socio_fecha(oficina_id, subcuenta, fecha_corte):
    revalorizacion = CONCEPTOS.objects.filter(tip_dev_ap='R').values_list('cod_con', flat=True)
    fecha_inicio = date(fecha_corte.year, 1, 1)
    saldo = (
        DETALLE_PROD.objects
            .filter(
                oficina_id=oficina_id,
                producto="AP",
                subcuenta=subcuenta,
                hecho_econo__fecha__gte=fecha_inicio,
                hecho_econo__fecha__lte=fecha_corte,
                concepto__in=revalorizacion
            )
            .aggregate(saldo=Coalesce(Sum("valor"), Value(0, output_field=FloatField())))
    )
    return -saldo["saldo"]

def saldo_aportes(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    if request.method == 'GET':
        return render(request, 'saldo_aportes.html') 
    if request.method == 'POST':
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        accion = request.POST.get("accion")   
      
        fecha_corte = request.POST.get('fecha_corte')
            
        saldos = obtener_reporte(id_cli, id_ofi, fecha_corte)

        if accion == "exportar":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
        
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "SALDO APORTES A UNA FECHA" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos        
            saldos = obtener_reporte(id_cli, id_ofi, fecha_corte)                            
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                      
            fecha_corte_formateada = formato_fecha(fecha_corte)
            
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = f"SALDO DE APORTES A LA FECHA {fecha_corte_formateada}"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
                                
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = saldos[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(saldos, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])

            nombre_archivo = f"saldo_aportes_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
           
        elif accion == "csv":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
                            
            # Llama a la función para obtener los datos
            saldos = obtener_reporte(id_cli, id_ofi, fecha_corte)
           
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"saldo_aportes_{fecha_corte_formateada}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                        
            # Crear el escritor CSV
            writer = csv.writer(response)
                        
            # Añadir encabezados de las columnas
            headers = saldos[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in saldos:
                writer.writerow([fila[col] for col in headers])
            return response
        else:
            return HttpResponse("Acción no permitida", status=405)
                
    return render(request, 'saldo_aportes.html') 
#     return HttpResponse("Método no permitido", status=405)

def score_creditos_por_nit(oficina_id, fecha_corte):
    fecha_limite = fecha_corte - timedelta(days=365*3)
    resultado = (
        CARTE_CAT_HIS.objects
        .filter(oficina_id=oficina_id, fecha__gt=fecha_corte - timedelta(days=365*3))  # Filtro de fecha
        .values("nit")  # Agrupar por `nit`
        .annotate(
            suma_cat_arr=Coalesce(
                Sum(
                    Case(
                        When(cat_arr="A", then=Value(1.0)),  # Si ARRASTRE es "A", contar 1.0
                        default=Value(0.0),  # De lo contrario, 0.0
                        output_field=FloatField(),
                    )
                ),
                Value(0.0)  # Si no hay valores, poner 0.0
            ),
            conteo_total=Coalesce(Count("id", output_field=FloatField()), Value(1.0)),  # Contar total, evitar división por 0
        )
        .annotate(porcentaje=F("suma_cat_arr") / F("conteo_total") * 100)  # Calcular el porcentaje
    )
    #print("🔍 Resultado QuerySet:", list(resultado))
    return {item["nit"]: item["porcentaje"] for item in resultado}

def contar_creditos_socio(oficina_id, fecha_limite):
    # Ejecutar una sola consulta y devolver un diccionario {cod_aso: total}
    resultado = (
        CREDITOS.objects
        .filter(
            oficina_id=oficina_id,
            fec_des__lt=fecha_limite
        )
        .exclude(estado="X")
        .values("socio__cod_aso")  # Agrupar por socio
        .annotate(total=Count("id"))  # Contar los créditos por socio
    )

    # Convertir resultados en un diccionario {cod_aso: total}
    return {item["socio__cod_aso"]: item["total"] for item in resultado}

def obtener_reporte(cliente_id, oficina_id, fecha_corte_str):
   
    NIV_EST_DICT = dict(OPC_EDUCACION)
    EST_CIV_DICT = dict(OPC_EST_CIV)
    EST_SOC_DICT = dict(OPC_EST_SOCIO)

    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    
    creditos_socio = contar_creditos_socio(oficina_id, fecha_corte)
    score_nit = score_creditos_por_nit(oficina_id, fecha_corte)

    asociados = list(ASOCIADOS.objects.filter(
         oficina_id=oficina_id
        ).exclude(
            estado='R', 
            fec_ret__lte=fecha_corte
        ).annotate(
            fecha_maxima=Max('tercero__estados_fin__fec_inf')  # Fecha más reciente por tercero
            ).filter(
            tercero__estados_fin__fec_inf=F('fecha_maxima')
        ).values(
            'tercero__cla_doc', 'cod_aso', 'tercero__nombre', 'fec_afi', 'tercero__cod_ciu_res_id__codigo', 
            'tercero__direccion', 'tercero__celular1', 'estado', 'sexo', 'emp_ent', 'fec_nac', 'tercero__doc_ide', 
            'tercero__email', 'niv_est', 'ocupacion', 'tercero__estados_fin__ing_tot', 'estrato', 'zona', 'cab_fam', 'est_civ', 'per_a_cargo', 'num_hij_may',
            'num_hij_men', 'tip_viv'
    ))

    for asociado in asociados:
        asociado['total_aportes'] = saldo_aporte_socio_fecha(oficina_id,asociado['cod_aso'],fecha_corte)
        asociado['apo_esperado'] = aporte_esperado_fecha(asociado['cod_aso'],fecha_corte)
        asociado['total_creditos'] = creditos_socio.get(asociado['tercero__doc_ide'], 0)
        asociado['score_cartera'] = score_nit.get(asociado['tercero__doc_ide'],0)
        asociado['niv_est'] = NIV_EST_DICT.get(asociado['niv_est'], "Sin Estudio").upper()
        asociado['est_civ'] = EST_CIV_DICT.get(asociado['est_civ'], "Sin Estado Civil").upper()
        asociado['estado'] = EST_SOC_DICT.get(asociado['estado'], "Asociado Sin Estado").upper()
        if asociado['fec_nac']:
            asociado['edad'] = calcular_edad(asociado['fec_nac'], fecha_corte)
        else:
            asociado['edad'] = None  # O un valor predeterminado
        if asociado['fec_afi']:
            asociado['ant_mes'] = calcular_antiguedad(asociado['fec_afi'], fecha_corte)
        else:
            asociado['ant_mes'] = None  # O un valor predeterminado      
       
    orden = [
        'tercero__cla_doc', 'cod_aso', 'tercero__nombre', 'fec_afi', 'tercero__cod_ciu_res_id__codigo',
        'tercero__direccion', 'tercero__celular1', 'total_aportes', 'estado', 'sexo', 
        'emp_ent', 'fec_nac', 'tercero__doc_ide', 'tercero__email', 'edad', 'score_cartera', 
        'ant_mes', 'total_creditos', 'niv_est', 'ocupacion', 'tercero__estados_fin__ing_tot', 'estrato', 'zona', 
        'cab_fam', 'per_a_cargo', 'num_hij_may', 'num_hij_men', 'tip_viv', 'apo_esperado'
        ]
    
    renombrar = {
        'tercero__cla_doc': 'Tip_Doc',
        'cod_aso': 'Cod_Aso',
        'tercero__nombre': 'Nombre Completo',
        'fec_afi': 'Fec_Afi',
        'tercero__cod_ciu_res_id__codigo': 'Ciudad',
        'tercero__direccion': 'Dirección',
        'tercero__celular1': 'Teléfono',
        'total_aportes': 'Total Aportes',
        'estado': 'Estado',
        'sexo': 'Sexo',
        'emp_ent': 'Empleado_Ent',
        'fec_nac': 'Fec_Nac',
        'tercero__doc_ide': 'NIT',
        'tercero__email': 'Email',
        'edad': 'Edad',
        'score_cartera': 'Score_Car',
        'ant_mes': 'Ant_(Meses)',
        'total_creditos': 'Num_Cre',
        'niv_est': 'Niv_Est',
        'ocupacion': 'Ocupacion',
        'tercero__estados_fin__ing_tot': 'Niv_Ingreso',
        'estrato': 'Estrato',
        'zona': 'Zona',
        'cab_fam': 'Cab_Fam',
        'per_a_cargo': 'Per_a_Cargo',
        'num_hij_may': 'Num_Hij_May',
        'num_hij_men': 'Num_Hij_Men',
        'tip_viv': 'Tip_Viv',
        'apo_esperado': 'Apor_al_dia'
        }
    # Reorganizar los diccionarios en el orden deseado
    for asociado in asociados:
        asociado_ordenado = {campo: asociado.get(campo) for campo in orden}
        asociado.clear()
        asociado.update({renombrar.get(k, k): v for k, v in asociado_ordenado.items()})
    
    return asociados

# Funciones auxiliares
def calcular_antiguedad(fecha_afiliacion, fecha_corte):
    ant_mes = (fecha_corte.year - fecha_afiliacion.year)* 12 + (
        (fecha_corte.month)-(fecha_afiliacion.month)
        )
    return ant_mes

def calcular_edad(fecha_nacimiento, fecha_corte):
    edad = fecha_corte.year - fecha_nacimiento.year - (
        (fecha_corte.month, fecha_corte.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
    )
    return edad

def aporte_esperado_fecha(cod_aso, fec_cor):
    if cod_aso is None or not cod_aso:
        return 0
    asociado = ASOCIADOS.objects.select_related('tercero').get(cod_aso=cod_aso, oficina=1)
    tip_ter = asociado.tercero.tip_ter
    inicio = asociado.fec_afi if asociado.fec_afi is not None else date.today()
    corte = fec_cor
    print('Asociado  ',asociado.cod_aso,' corte ',corte,'type ',type(corte),' inicio',type(inicio))
    if corte < inicio:
        return 0
    monthly_by_year = {}
    for year in range(inicio.year, corte.year + 1):
        plan = PLAN_APORTES.objects.get(agno=year)

        if tip_ter == 'J':
            mensual = (plan.totadu or 0) / 12
        else:
            fecha_corte_edad = date(year, 1, 1)
            edad_dias = (fecha_corte_edad - asociado.fec_nac).days
            es_menor = edad_dias < 365 * 18
            if es_menor:
                mensual = (plan.totchi2 or 0) / 12
            else:
                mensual = (plan.totadu or 0) / 12

        monthly_by_year[year] = mensual

    # 2) Acumular por mes, prorrateando por días
    total = 0.0
    y, m = inicio.year, inicio.month
    while (y, m) <= (corte.year, corte.month):
        dias_mes = monthrange(y, m)[1]
        ini_mes = date(y, m, 1)
        fin_mes = date(y, m, dias_mes)

        desde = max(inicio, ini_mes)
        hasta = min(corte, fin_mes)

        if desde <= hasta:
            dias_cubiertos = (hasta - desde).days + 1  # inclusivo
            mensual = monthly_by_year[y]
            total += mensual * dias_cubiertos / dias_mes

        # siguiente mes
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1

    # Si quieres incluir la "cuota de ingreso" (iniadu/inichi2/inijur), súmala aquí 1 sola vez:
    # if tip_ter == 'J':
    #     total += plan_inicial_del_año_de_afiliacion.inijur
    # else:
    #     total += plan_inicial_del_año_de_afiliacion.inichi2 o iniadu según edad
    print('este se supone debe ser el esperado', total)
    return round(total, 2)

def aporte_mensual(oficina_id, cod_aso, fec_cor):
    asociado = ASOCIADOS.objects.filter(cod_aso=cod_aso, oficina=oficina_id).first()
    if not asociado:
        return 0  # Si no existe el asociado, retornar 0
    
    x_tip_soc = asociado.tercero.tip_ter
    plan_apo = PLAN_APORTES.objects.filter(agno=fec_cor.year).first()

    if not plan_apo:
        return 0  # Si no hay plan de aportes para ese año, retornar 0

    # Determinar el aporte anual según el tipo de asociado
    if x_tip_soc == "J":
        aporte_anual = plan_apo.totjur  # Persona jurídica usa totjur
    else:
        x_dia_ing = (fec_cor - asociado.fec_nac).days
        if x_dia_ing < 365 * 18:
            aporte_anual = plan_apo.totchi2  # Chico 1 y 2
        else:
            aporte_anual = plan_apo.totadu  # Adulto

    # Calcular el aporte mensual dividiendo el anual entre 12
    aporte_mensual = aporte_anual / 12

    return round(aporte_mensual, 2)  # Redondear a 2 decimales

def aportes_super(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    if request.method == 'GET':
        return render(request, 'aportes_supersolidaria.html') 
    if request.method == 'POST':
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        accion = request.POST.get("accion")   
      
        fecha_corte = request.POST.get('fecha_corte')
        
        saldos = reporte_super(id_cli, id_ofi, fecha_corte)

        if accion == "exportar":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
        
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "APORTES SUPERSOLIDARIA" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos        
            saldos = reporte_super(id_cli, id_ofi, fecha_corte)                            
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            print('fecha_corte',fecha_corte)          
            
            fecha_corte_formateada = formato_fecha(fecha_corte)
            print('fec_cor_formato', fecha_corte_formateada)                        
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = f"APORTES SUPERSOLIDARIA A LA FECHA {fecha_corte_formateada}"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
                                
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = saldos[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(saldos, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])

            nombre_archivo = f"aportes_supersolidaria_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
           
        elif accion == "csv":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
                            
            # Llama a la función para obtener los datos
            saldos = reporte_super(id_cli, id_ofi, fecha_corte)
           
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"aportes_supersolidaria_{fecha_corte}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                        
            # Crear el escritor CSV
            writer = csv.writer(response)
                        
            # Añadir encabezados de las columnas
            headers = saldos[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in saldos:
                writer.writerow([fila[col] for col in headers])
            return response
        else:
            return HttpResponse("Acción no permitida", status=405)
                
    return render(request, 'aportes_supersolidaria.html')

def reporte_super(cliente_id, oficina_id, fecha_corte_str):
    fec_cor = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()

    # Traer asociados activos con su tercero
    asociados = list(ASOCIADOS.objects.filter(oficina_id=oficina_id)
        .exclude(estado='R', fec_ret__lte=fec_cor)
        .select_related('tercero')
    )

    asociados_dict = []

    for asociado in asociados:
        tercero = asociado.tercero

        # Formatear cod_aso si cla_doc es "N"
        if tercero.cla_doc == "N":
            cod_aso_formateado = formatear_cod_aso(asociado.cod_aso)
            if tercero.dig_ver:
                cod_aso_formateado += f"-{tercero.dig_ver}"
        else:
            cod_aso_formateado = asociado.cod_aso

        # Cálculo de fechas y saldos
        fecha_aporte = saldo_aportes_fecha(oficina_id, asociado.cod_aso, fec_cor)
        aporte_ordinario = saldo_aporte_socio_fecha(oficina_id, asociado.cod_aso, fec_cor)
        revalorizacion = saldo_revalorizacion_aportes_socio_fecha(oficina_id, asociado.cod_aso, fec_cor)
        extra = saldo_aporte_extra_socio_fecha(oficina_id, asociado.cod_aso, fec_cor)
        voluntario = saldo_aporte_voluntario_socio_fecha(oficina_id, asociado.cod_aso, fec_cor)
        promedio_aporte = promedio_aporte_socio_fecha(oficina_id, asociado.cod_aso, fec_cor)

        datos = {
            'tip_doc': "I" if tercero.cla_doc == "T" else (tercero.cla_doc if tercero.cla_doc else ""),
            'cod_aso': cod_aso_formateado,
            'aporte_mensual': aporte_mensual(oficina_id, asociado.cod_aso, fec_cor),
            'aporte_ordinario': aporte_ordinario - revalorizacion - extra - voluntario,
            'aporte_extra_ordinario': extra,
            'revalorizacion': revalorizacion,
            'promedio_aporte': round(promedio_aporte,0), #(asociado.cod_aso, fec_cor),
            'fec_ult_apo': fecha_aporte["Fecha"].strftime('%d/%m/%Y') if fecha_aporte else None,
            'aporte_voluntario': voluntario,
            'fec_nac': asociado.fec_nac,
        }

        datos['total_aportes'] = (
            datos['aporte_ordinario']
            + datos['aporte_extra_ordinario']
            + datos['revalorizacion']
            + datos['aporte_voluntario']
        )

        asociados_dict.append(datos)

    # Orden y renombramiento
    orden = [
        'tip_doc', 'cod_aso', 'total_aportes', 'aporte_mensual',
        'aporte_ordinario', 'aporte_extra_ordinario', 'revalorizacion',
        'promedio_aporte', 'fec_ult_apo', 'aporte_voluntario', 'fec_nac'
    ]

    renombrar = {
        'tip_doc': 'Tip_Doc',
        'cod_aso': 'Cod_Aso',
        'total_aportes': 'Total Aportes',
        'aporte_mensual': 'Aporte Mensual',
        'aporte_ordinario': 'Aporte Ordinario',
        'aporte_extra_ordinario': 'Aporte Extra',
        'revalorizacion': 'Revalorizacion',
        'promedio_aporte': 'Promedio Aporte',
        'fec_ult_apo': 'Fec_Ult_Apo',
        'aporte_voluntario': 'Aporte Voluntario',
        'fec_nac': 'Fec_Nac'
    }

    # Filtrar asociados con aportes > 0
    asociados_filtrados = [a for a in asociados_dict if a['total_aportes'] > 0]

    # Reordenar y renombrar campos
    asociados_renombrados = []
    for a in asociados_filtrados:
        nuevo = {renombrar.get(k, k): a.get(k) for k in orden}
        asociados_renombrados.append(nuevo)

    # Orden final por código de asociado
    asociados_ordenados = sorted(asociados_renombrados, key=lambda x: x.get('Cod_Aso', ''))

    return asociados_ordenados

def riesgo_de_liquidez_aportes(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    print('id_ofi  ',id_ofi)

    if request.method == 'GET':
        return render(request, 'Riesgo_liq_aportes.html')  # tu formulario

    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_corte')
        try:
            fecha_base = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha inválida", status=400)
        accion = request.POST.get('accion')
        # Fecha final: por ejemplo, hoy o una fecha del formulario
        fecha_inicio = (fecha_base - relativedelta(months=12)) + timedelta(days=1)
        # Estructura equivalente al cursor AHORRO:
        ahorro = []  # Lista de dicts con claves: 'DIA', 'M01', ..., 'M12'
        for dia in range(1, 32):
            fila = {'DIA': dia}
            for mes in range(1, 13):
                clave_mes = f"M{mes:02}"  # M01, M02, ..., M12
                fila[clave_mes] = 0
            ahorro.append(fila)
        MESES = [[0 for _ in range(31)] for _ in range(12)]
        pagador_codigo = Subquery(
            ASOCIADOS.objects.filter(
                cod_aso=OuterRef('subcuenta'),oficina_id=id_ofi
            ).filter(id_pag__isnull=False
            ).values('id_pag__codigo')[:1]
        )
        queryset = DETALLE_PROD.objects.filter(
            producto='AP',
            hecho_econo__anulado='N',
            hecho_econo__fecha__year__gt=2014,
            hecho_econo__docto_conta__oficina_id=id_ofi,
        ).annotate(
            fecha=F('hecho_econo__fecha'),
            cod_ent=pagador_codigo
        ).values(
            'fecha','subcuenta','concepto','valor','cod_ent'
        )
        codigos_lib = [
            "005", "013", "999", "016", "017", "018", "019", "021", "022", "025",
            "027", "028", "029", "031", "032", "033", "034", "035", "036", "038", "040", "041"
        ]
        if accion == 'Libranza': 
            resultado = [fila for fila in queryset if fila['cod_ent'] in codigos_lib]
        elif accion == 'Personal':
            resultado = [fila for fila in queryset if fila['cod_ent'] not in codigos_lib]
        elif accion == 'Ingreso Libranza':
            resultado = [fila for fila in queryset if fila['cod_ent'] in codigos_lib and valor < 0 and concepto != 'APREV']
        elif accion == 'Retiro Libranza':
            resultado = [fila for fila in queryset if fila['cod_ent'] in codigos_lib and valor > 0 and concepto != 'APREV']
        elif accion == 'Ingreso Personal':
            resultado = [fila for fila in queryset if fila['cod_ent'] not in codigos_lib and valor < 0 and concepto != 'APREV']
        elif accion == 'Retiro Personal':
            resultado = [fila for fila in queryset if fila['cod_ent'] not in codigos_lib and valor > 0 and concepto != 'APREV']
        else:
            resultado = list(queryset)
        xant_aporte = 0
        if accion != 'activos_1_ano':
            for fila in resultado:
                fecha = fila['fecha']
                subcuenta = fila['subcuenta']
                concepto  = fila['concepto']
                valor = fila['valor']
                if fecha >= fecha_inicio:
                    mes_idx = fecha.month - 1  # 0 basado
                    dia_idx = fecha.day - 1    # 0 basado
                    MESES[mes_idx][dia_idx] -= valor
                else:
                    xant_aporte -= valor
            print('Accion ',accion,'   Valor Anterior ',xant_aporte)
            wb = Workbook()
            ws = wb.active
            ws.title = 'Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")
            ws.append(['Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")])
            ws.append([])
            encabezado = ['Día'] + [f"Mes {m+1:02}" for m in range(12)]
            ws.append(encabezado)
            for dia in range(31):
                fila = [f"Día {dia+1:02}"]  # Primera columna: Día
                for mes in range(12):
                    fila.append(MESES[mes][dia])  # Valor de ese día en cada mes
                ws.append(fila)
            ws.append([])
            ws.append([])
            ws.append(['Aporte Anterior :', xant_aporte ])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            xFecFin = fecha_base  # Suponiendo que ya la tienes como datetime.date
            xFecIni = xFecFin - relativedelta(months=12) + relativedelta(days=1)
            if xFecIni.month == 2 and xFecIni.year % 4 == 0:
                xFecIni = xFecIni + relativedelta(days=1)
            tab_rie_apo = {}
            xFecCor = xFecIni
            while xFecCor <= xFecFin:
                mes_str = f"{xFecCor.year}-{xFecCor.month:02}"
                mes_str = f"{xFecCor.year}-{xFecCor.month:02}"
                tab_rie_apo[mes_str] = {
                    "NumAntIng": 0,
                    "NumAntRet": 0,
                    "AntLibIng": 0.0,
                    "AntLibEgr": 0.0,
                    "AntPerIng": 0.0,
                    "AntPerEgr": 0.0,
                    "NumNueIng": 0,
                    "NumNueRet": 0,
                    "NueLibIng": 0.0,
                    "NueLibEgr": 0.0,
                    "NuePerIng": 0.0,
                    "NuePerEgr": 0.0,
                }
                xFecCor += relativedelta(months=1)
            xIniAntLib = 0
            xIniAntPer = 0
            xIniNueLib = 0
            xIniNuePer = 0
            for fila in resultado:
                fecha = fila['fecha']
                subcuenta = fila['subcuenta']
                concepto  = fila['concepto']
                cod_ent = fila['cod_ent']
                valor = fila['valor']
                mes = f"{fecha.year:04}-{fecha.month:02}"
                if mes in tab_rie_apo:
                    socio = ASOCIADOS.objects.filter(oficina_id = id_ofi,cod_aso = subcuenta).first()
                    if socio == None:
                        continue
                    if fecha >= xFecIni and fecha <= xFecFin and concepto != 'APREV':
                        if (socio.estado != "R" and socio.fec_afi < xFecIni)  or (socio.estado == "R" and socio.fec_ret is not None and socio.fec_ret >= xFecIni):
                            if cod_ent in codigos_lib:  #  Libranza
                                if valor < 0: # Ingreso
                                    tab_rie_apo[mes]["AntLibIng"] -= valor
                                else:
                                    tab_rie_apo[mes]["AntLibEgr"] -= valor
                            else:
                                if valor < 0: # Ingreso
                                    tab_rie_apo[mes]["AntPerIng"] -= valor
                                else:
                                    tab_rie_apo[mes]["AntPerEgr"] -= valor
                        else:
                            if cod_ent in codigos_lib:  #  Libranza
                                if valor < 0: # Ingreso
                                    tab_rie_apo[mes]["NueLibIng"] -= valor
                                else:
                                    tab_rie_apo[mes]["NueLibEgr"] -= valor
                            else:
                                if valor < 0: # Ingreso
                                    tab_rie_apo[mes]["NuePerIng"] -= valor
                                else:
                                    tab_rie_apo[mes]["NuePerEgr"] -= valor
                    else:
                        if fecha < xFecIni:
                            if cod_ent in codigos_lib:
                                xIniAntLib -= valor
                            else:
                                xIniAntPer -= valor
            xNumAntIng = 0
            socios = ASOCIADOS.objects.filter(oficina_id = id_ofi)
            for socio in socios:
                if socio.fec_afi is None:
                    continue
                if socio.fec_afi <= xFecFin:
                    if (socio.estado != 'R' and socio.fec_afi < xFecIni):
                        xNumAntIng=xNumAntIng+1    
                    if (socio.estado == 'R' and socio.fec_ret is not None and socio.fec_ret >= xFecIni and socio.fec_afi < xFecIni):   # antiguos retirados
                        mes = f"{socio.fec_ret.year:04}-{socio.fec_ret.month:02}"
                        if mes in tab_rie_apo:
                            tab_rie_apo[mes]["NumAntRet"] += 1
                    if (socio.estado != 'R' and socio.fec_afi >= xFecIni):   #   Nuevos
                        mes = f"{socio.fec_afi.year:04}-{socio.fec_afi.month:02}"
                        if mes in tab_rie_apo:
                            tab_rie_apo[mes]["NumNueIng"] += 1
                    if (socio.estado == 'R' and socio.fec_afi >= xFecIni and socio.fec_ret <= xFecFin):
                        mes = f"{socio.fec_ret.year:04}-{socio.fec_ret.month:02}"
                        if mes in tab_rie_apo:
                            tab_rie_apo[mes]["NumNueRet"] += 1
            tab_rie_apo["SalIni"] = {
                "NumAntIng": xNumAntIng,
                "NumAntRet": 0,
                "AntLibIng": xIniAntLib,
                "AntLibEgr": 0.0,
                "AntPerIng": xIniAntPer,
                "AntPerEgr": 0.0,
                "NumNueIng": 0,
                "NumNueRet": 0,
                "NueLibIng": xIniNueLib,
                "NueLibEgr": 0.0,
                "NuePerIng": xIniNuePer,
                "NuePerEgr": 0.0,
            }

            wb = Workbook()
            ws = wb.active
            ws.title = 'Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")
            ws.append(['Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")])
            ws.append([])
            headers = [
                "Mes",
                "NumAntIng", "NumAntRet", "AntLibIng", "AntLibEgr", "AntPerIng", "AntPerEgr",
                "NumNueIng", "NumNueRet", "NueLibIng", "NueLibEgr", "NuePerIng", "NuePerEgr"
            ]
            ws.append(headers)
            claves = list(tab_rie_apo.keys())
            if "SalIni" in claves:
                claves.remove("SalIni")
                claves = ["SalIni"] + sorted(claves)
            else:
                claves = sorted(claves)
            for mes in claves:
                fila = [mes]
                fila += [
                    tab_rie_apo[mes].get("NumAntIng", 0),
                    tab_rie_apo[mes].get("NumAntRet", 0),
                    tab_rie_apo[mes].get("AntLibIng", 0.0),
                    tab_rie_apo[mes].get("AntLibEgr", 0.0),
                    tab_rie_apo[mes].get("AntPerIng", 0.0),
                    tab_rie_apo[mes].get("AntPerEgr", 0.0),
                    tab_rie_apo[mes].get("NumNueIng", 0),
                    tab_rie_apo[mes].get("NumNueRet", 0),
                    tab_rie_apo[mes].get("NueLibIng", 0.0),
                    tab_rie_apo[mes].get("NueLibEgr", 0.0),
                    tab_rie_apo[mes].get("NuePerIng", 0.0),
                    tab_rie_apo[mes].get("NuePerEgr", 0.0),
                ]
                ws.append(fila)
            print('Fin Por menos de un año')
        
        nombre_archivo = f"rl_aportes_{accion}_{fecha_base}.xlsx"
        print('Ahora Graba ',nombre_archivo )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(response)
        return response