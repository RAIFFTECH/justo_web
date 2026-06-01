import csv, json
from math import ceil
from io import BytesIO
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse, JsonResponse, FileResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.db.models import Q
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, BaseDocTemplate, PageTemplate, Frame
from .forms import MovCajaFilterForm, MovCajaForm
from django import forms

from .models import MOV_CAJA
from cajeros_app.models import CAJEROS
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from terceros_app.models import TERCEROS
from hecho_economico_app.models import HECHO_ECONO


class MOV_CAJACreateView(SuccessMessageMixin, CreateView):
    model = MOV_CAJA
    form_class = MovCajaForm
    template_name = 'mov_caja_form.html'
    success_url = reverse_lazy('mov_caja_list')
    success_message = "Movimiento de caja creado correctamente."

    def get_context_data(self, **kwargs):
        # Obtén el contexto original de la clase base
        context = super().get_context_data(**kwargs)
        # Agrega las denominaciones al contexto
        context['denominaciones'] = [100000, 50000, 20000, 10000, 5000, 2000, 1000, 500, 200, 100, 50]
        return context
    

    @method_decorator(csrf_exempt)  # Solo usar si no estás usando el CSRF token en el frontend
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)  # Carga los datos del cuerpo de la solicitud
            fecha = data.get('fecha')
            if not fecha:
                return JsonResponse({'error': 'La fecha es obligatoria.'}, status=400)
            try:
                # Intentar parsear la fecha al formato esperado
                fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({'error': 'El formato de la fecha debe ser YYYY-MM-DD.'}, status=400)
            id_usr = data.get('cajero')
            if not id_usr:
                return JsonResponse({'error': 'El cajero es obligatorio.'}, status=400)
            try:
                # Comprobar si el usuario existe
                cajero = CAJEROS.objects.get(id=id_usr)
            except cajero.DoesNotExist:
                return JsonResponse({'error': f'No se encontró un usuario con ID {id_usr}.'}, status=400)

            saldo_ini = data.get('saldo_ini')
            debitos = data.get('debitos')
            creditos = data.get('creditos')
            val_che_dev = data.get('val_che_dev')
            saldo_fin = data.get('saldo_fin')
            diferencia = data.get('diferencia')
            val_cheques = data.get('val_cheques')
            val_vales = data.get('val_vales')
            monedas = data.get('monedas', [])
            # Realiza validaciones necesarias sobre los datos
            if not monedas:
                return JsonResponse({'error': 'No se proporcionaron monedas.'}, status=400)
            
            # Procesar las monedas o realizar operaciones
            for moneda in monedas:
                denominacion = moneda.get('denominacion', 0)
                cantidad = moneda.get('cantidad', 0)
                # Aquí podrías guardar cada moneda en una base de datos o realizar cálculos

            mov_caj_dia = MOV_CAJA.objects.filter(cajero_id = id_usr,fecha = fecha).first()
            if mov_caj_dia == None:
                mov_caj_dia = MOV_CAJA.objects.create(cajero_id = id_usr,fecha = fecha)
                mov_caj_dia.saldo_ini = saldo_ini
                mov_caj_dia.debitos = debitos
                mov_caj_dia.creditos =creditos
                mov_caj_dia.val_che_dev = val_che_dev
                mov_caj_dia.saldo_fin = saldo_fin
                mov_caj_dia.diferencia = diferencia
                mov_caj_dia.val_cheques = val_cheques
                mov_caj_dia.val_vales = val_vales
                mov_caj_dia.monedas = monedas
                mov_caj_dia.save()

            return JsonResponse({'message': 'Datos procesados correctamente.'}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Los datos enviados no son válidos.'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
   
class MOV_CAJAUpdateView(SuccessMessageMixin, UpdateView):
    model = MOV_CAJA
    form_class = MovCajaForm
    template_name = 'mov_caja_form.html'
    success_url = reverse_lazy('mov_caja_list')
    success_message = "Movimiento de caja actualizado correctamente."

class MOV_CAJADetailView(DetailView):
    model = MOV_CAJA
    template_name = 'mov_caja_details.html'
    context_object_name = 'mov_caja'

class MOV_CAJADeleteView(DeleteView):
    model = MOV_CAJA
    template_name = 'mov_caja_confirm_delete.html'
    success_url = reverse_lazy('mov_caja_list')

# Formulario para los filtros
class MovCajaFilterForm(forms.Form):
    fecha_inicio = forms.DateField(required=False, widget=forms.DateInput(attrs={'type': 'date'}))
    fecha_fin = forms.DateField(required=False, widget=forms.DateInput(attrs={'type': 'date'}))
    cajero = forms.ModelChoiceField(queryset=CAJEROS.objects.all(), required=False)

def mov_caja_list(request):
    form = MovCajaFilterForm(request.GET or None)
    mov_cajas = MOV_CAJA.objects.none()  # Inicialmente vacío
    

    if form.is_valid():
        fecha_inicio = form.cleaned_data.get('fecha_inicio')
        fecha_fin = form.cleaned_data.get('fecha_fin')
        cajero = form.cleaned_data.get('cajero')

        # Filtrar según los criterios proporcionados
        filters = Q()
        if fecha_inicio:
            filters &= Q(fecha__gte=fecha_inicio)
        if fecha_fin:
            filters &= Q(fecha__lte=fecha_fin)
        if cajero:
            filters &= Q(cajero=cajero)

        mov_cajas = MOV_CAJA.objects.filter(filters)

    return render(request, 'mov_caja_list.html', {'form': form, 'mov_cajas': mov_cajas})

# # Impresión de reportes
# def movimiento_caja(request):
#     print(request.POST)
#     if request.method == 'GET':
#         return render(request, 'mov_caja_form.html')
#     if request.method == 'POST':
#         accion = request.POST.get("accion")  # Obtener la acción
#         # Recupera los parámetros del formulario
#         id_cli = request.session.get('cliente_id')
#         id_ofi = request.session.get('oficina_id')
#         # id_cajero = request.POST.get('cajero')
#         cajero = request.POST.get('cajero', 'N/A')
#         fecha = request.POST.get('fecha', 'N/A')
#         print('Recibe fecha ---->',fecha)
#         jornada = request.POST.get('jornada', 'N/A')
#         cerrado = request.POST.get('cerrado', 'N/A')
#         saldo_ini = request.POST.get('saldo_ini', '0')
#         debitos = request.POST.get('debitos', '0')
#         creditos = request.POST.get('creditos', '0')
#         val_che_dev = request.POST.get('val_che_dev', '0')
#         saldo_fin = request.POST.get('saldo_fin', '0')
#         diferencia = request.POST.get('diferencia', '0')
#         val_cheques = request.POST.get('val_cheques', '0')
#         val_vales = request.POST.get('val_vales', '0')

#         # Procesar el JSON de monedas si está incluido
#         monedas_json = request.POST.get('monedas', '{}')
#         print('este es el metodo--->', request.method)
#         resultados = saldos_cajero_dia(request).content
#         datos = json.loads(resultados)
#         print('DATOS -- - ->',datos)
#         if not datos.get('Exito'):
#             return JsonResponse({'error': 'No se pudo generar el reporte.'})


#         resultado = datos   
#         total_filas = len(resultado)     
#         print(resultado)
#         print('total filas--->', total_filas)
        
#         if accion == "exportar":
          
#             # Lógica para exportar Crear el libro de Excel
#             workbook = Workbook()
#             sheet = workbook.active
#             sheet.title = "movimiento_caja" 

#             entidad = CLIENTES.objects.filter(id=id_cli).first()
#             oficina = OFICINAS.objects.filter(id=id_ofi).first()

#             # Llama a la función para obtener los datos
#             resultado = saldos_cajero_dia(request)
#             print(resultado)

#             # Obtener los datos relevantes del JSON
#             datos = resultado.get("data", [])
#             # Validar si los datos son una lista de diccionarios
#             if isinstance(datos, list) and all(isinstance(item, dict) for item in datos):
#                 # Añadir los encabezados
#                 headers = datos[0].keys()
#                 header_row = 1
#                 for col_num, header in enumerate(headers, start=1):
#                     sheet.cell(row=header_row, column=col_num, value=header)
#                 # Añadir los datos
#                 for row_num, data in enumerate(datos, start=header_row + 1):
#                     for col_num, field in enumerate(headers, start=1):
#                         sheet.cell(row=row_num, column=col_num, value=data[field])

#                 # Guardar el archivo Excel
#                 workbook.save("resultados.xlsx")
#                 print("Archivo Excel generado exitosamente.")
#             else:
#                 print("El JSON no contiene una lista de diccionarios en 'data'.")

            
#             if not resultado:
#                 return HttpResponse("No se encontraron datos para exportar", status=404)
                  
#             # Añadir datos adicionales encima de los encabezados
#             empresa = f"{entidad.nombre.strip()}"
#             nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
#             reporte = "Movimiento de Caja"
#             detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha: {fecha}"
#             usuario = f"Cajero: {cajero}"
            
#             datos_adicionales = [
#                 [empresa],
#                 [nit_empresa],
#                 [reporte],
#                 [detalles],
#                 [usuario]
#             ]

#             # Insertar los datos adicionales
#             for row_num, fila in enumerate(datos_adicionales, start=1):
#                 for col_num, valor in enumerate(fila, start=1):
#                     sheet.cell(row=row_num, column=col_num, value=valor)

#             # Determinar la fila donde empiezan los encabezados
#             header_row = len(datos_adicionales) + 1

#             # Añadir los encabezados
#             headers = resultado[0].keys()
#             for col_num, header in enumerate(headers, start=1):
#                 sheet.cell(row=header_row, column=col_num, value=header)

#             # Añadir los datos
#             for row_num, data in enumerate(resultado, start=header_row + 1):
#                 for col_num, field in enumerate(headers, start=1):
#                     sheet.cell(row=row_num, column=col_num, value=data[field]) 

#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="movimiento_caja.xlsx"'
#             workbook.save(response)
#             return response

#         elif accion == "exportar_cierre":
#             # Lógica para exportar Crear el libro de Excel
#             workbook = Workbook()
#             sheet = workbook.active
#             sheet.title = "cierre_caja" 

#             entidad = CLIENTES.objects.filter(id=id_cli).first()
#             oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
#             # Llama a la función para obtener los datos
#             # resultado = get_cierre(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
#             if not resultado:
#                 return HttpResponse("No se encontraron datos para exportar", status=404)
                  
#             # Añadir datos adicionales encima de los encabezados
#             empresa = f"{entidad.nombre.strip()}"
#             nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
#             reporte = "Cierre de Caja"
#             detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha: {fecha}"
#             usuario = f"Cajero: {cajero.tercero.nombre}"
            
#             datos_adicionales = [
#                 [empresa],
#                 [nit_empresa],
#                 [reporte],
#                 [detalles],
#                 [usuario]                
#             ]

#             # Insertar los datos adicionales
#             for row_num, fila in enumerate(datos_adicionales, start=1):
#                 for col_num, valor in enumerate(fila, start=1):
#                     sheet.cell(row=row_num, column=col_num, value=valor)

#             # Determinar la fila donde empiezan los encabezados
#             header_row = len(datos_adicionales) + 1

#             # Añadir los encabezados
#             headers = resultado[0].keys()
#             for col_num, header in enumerate(headers, start=1):
#                 sheet.cell(row=header_row, column=col_num, value=header)

#             # Añadir los datos
#             for row_num, data in enumerate(resultado, start=header_row + 1):
#                 for col_num, field in enumerate(headers, start=1):
#                     sheet.cell(row=row_num, column=col_num, value=data[field]) 

#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = 'attachment; filename="cierre_caja.xlsx"'
#             workbook.save(response)
#             return response

#         elif accion == "imprimir":
#             # Lógica para imprimir o generar vista previa  Configuración del PDF
#             entidad = CLIENTES.objects.filter(id=id_cli).first()
#             oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
#             response = HttpResponse(content_type='application/pdf')
#             response['Content-Disposition'] = 'inline; filename="movimiento_caja.pdf"'

#             p = canvas.Canvas(response, pagesize=letter)
#             width, height = letter
#             margin_x, margin_y = 50, 60

#             # Configuración inicial
#             filas_por_pagina = 46 # Número máximo de filas por página
#             total_filas = len(resultado)
#             total_paginas = ceil(total_filas / filas_por_pagina)

#             # Función para dibujar subtotales
#             # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
#             #     x = margin_x - 30
#             #     p.setFont("Helvetica-Bold", 9)
#             #     p.drawString(x, y, "Subtotal: "+cod_cta)
#             #     x += 640  # Posicionar en la columna de débito
#             #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
#             #     x += 80  # Posicionar en la columna de crédito
#             #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

#             # Función para dibujar encabezado
#             def dibujar_encabezado():
#                 empresa = f"{entidad.nombre.strip()}"
#                 texto_empresa = stringWidth(empresa, "Times-Roman", 12)
#                 p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
#                 nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
#                 texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
#                 p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

#                 reporte = "Movimiento de Caja"
#                 texto_reporte = stringWidth(reporte, "Times-Roman", 12)
#                 p.drawString((width - texto_reporte) / 2, height - 60, reporte)

#                 p.setFont("Helvetica", 10)
#                 detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha: {fecha}"
#                 texto_detalles = stringWidth(detalles, "Helvetica", 10)
#                 p.drawString((width - texto_detalles) / 2, height - 75, detalles)

#                 usuario = f"Cajero: {cajero}"
#                 texto_filtro = stringWidth(usuario, "Helvetica", 10)
#                 p.drawString((width - texto_filtro) / 2, height - 90, usuario)

#                 p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

#             # Función para dibujar pie de página
#             def dibujar_pie(pagina_actual, total_paginas):
#                 line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
#                 p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

#                 p.setFont("Courier", 9)
#                 texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
#                 p.drawString(margin_x, margin_y - 40, texto_pie)

#                 texto_paginas = f"Página {pagina_actual} de {total_paginas}"
#                 p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

#             # Función para dibujar encabezado de tabla
#             def dibujar_encabezado_tabla(y):
#                 line_y = height - 115  # Ajusta para que esté justo encima del texto
#                 p.line(margin_x-40, line_y, margin_x+732, line_y)
#                 p.line(margin_x-40, height - 100, margin_x+732, height - 100)
#                 columnas = [ 
#                     ("Documento", 20), 
#                     ("Número", 40), 
#                     ("Fecha", 50), 
#                     ("        Débito", 70), 
#                     ("        Crédito", 70), 
#                     ("        Saldo", 80)
#                     ]
#                 x = margin_x-30
#                 p.setFont("Helvetica-Bold", 9)
#                 for col, ancho in columnas:
#                     p.drawString(x, y, col)
#                     x += ancho
                    
#             # Dibujar contenido
#             pagina_actual = 1
#             dibujar_encabezado()
#             dibujar_pie(pagina_actual, total_paginas)
#             dibujar_encabezado_tabla(height - margin_y - 50)
#             y = height - margin_y - 65

#             def dibujar_fila(y, row):
#                 datos = resultado.content
#                 datos = json.loads(datos)
#                 for row in datos:
#                     # if not all(key in row for key in ['saldo_anterior', 'tot_debitos', 'tot_creditos', 'val_cheques', 'val_vales', 'monedas', 'docto', 'numero', 'debito', 'credito', 'saldo']):
#                     #     raise ValueError("Faltan claves en una de las filas del JSON.")
#                     #     dibujar_fila(y, row)
#                         y -= 10


#                 # p.setFont("Helvetica", 9)
#                 # for col in columnas:
#                 #     if len(col) == 3 and col[2] == 'right':  # Alineación derecha
#                 #         texto, ancho, _ = col
#                 #         p.drawRightString(x + ancho, y, texto)
#                 #     else:  # Alineación izquierda
#                 #         texto, ancho = col[:2]
#                 #         p.drawString(x, y, str(texto))
#                 #     x += ancho

#             for idx, row in enumerate(resultado):
#                 if y < margin_y - 30:
#                     p.showPage()
#                     pagina_actual += 1
#                     dibujar_encabezado()
#                     dibujar_encabezado_tabla(height - margin_y - 50)
#                     dibujar_pie(pagina_actual, total_paginas)
#                     y = height - margin_y - 65    

#                 dibujar_fila(y, row)
#                 y -= 10

#             p.showPage()
#             p.save()
#             return response
        
#         elif accion == "imprimir_cierre":
#             # Lógica para imprimir o generar vista previa  Configuración del PDF
#             entidad = CLIENTES.objects.filter(id=id_cli).first()
#             oficina = OFICINAS.objects.filter(id=id_ofi).first()
#             print(monedas_json)
#             response = HttpResponse(content_type='application/pdf')
#             response['Content-Disposition'] = 'inline; filename="cierre_caja.pdf"'

#             p = canvas.Canvas(response, pagesize=letter)
#             width, height = letter
#             margin_x, margin_y = 50, 60

#             # Configuración inicial
#             filas_por_pagina = 46 # Número máximo de filas por página
#             total_filas = len(resultado)
#             total_paginas = ceil(total_filas / filas_por_pagina)

#             # Función para dibujar subtotales
#             # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
#             #     x = margin_x - 30
#             #     p.setFont("Helvetica-Bold", 9)
#             #     p.drawString(x, y, "Subtotal: "+cod_cta)
#             #     x += 640  # Posicionar en la columna de débito
#             #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
#             #     x += 80  # Posicionar en la columna de crédito
#             #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

#             # Función para dibujar encabezado
#             def dibujar_encabezado():
#                 empresa = f"{entidad.nombre.strip()}"
#                 texto_empresa = stringWidth(empresa, "Times-Roman", 12)
#                 p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
#                 nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
#                 texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
#                 p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

#                 reporte = "Cierre de Caja"
#                 texto_reporte = stringWidth(reporte, "Times-Roman", 12)
#                 p.drawString((width - texto_reporte) / 2, height - 60, reporte)

#                 p.setFont("Helvetica", 10)
#                 detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha: {fecha}"
#                 texto_detalles = stringWidth(detalles, "Helvetica", 10)
#                 p.drawString((width - texto_detalles) / 2, height - 75, detalles)

#                 usuario = f"Cajero: {cajero.tercero.nombre}"
#                 texto_filtro = stringWidth(usuario, "Helvetica", 10)
#                 p.drawString((width - texto_filtro) / 2, height - 90, usuario)

#                 p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

#             # Función para dibujar pie de página
#             def dibujar_pie(pagina_actual, total_paginas):
#                 line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
#                 p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

#                 p.setFont("Courier", 9)
#                 texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
#                 p.drawString(margin_x, margin_y - 40, texto_pie)

#                 texto_paginas = f"Página {pagina_actual} de {total_paginas}"
#                 p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

#             # Función para dibujar encabezado de tabla
#             def dibujar_encabezado_tabla(y):
#                 line_y = height - 115  # Ajusta para que esté justo encima del texto
#                 p.line(margin_x-40, line_y, margin_x+732, line_y)
#                 p.line(margin_x-40, height - 100, margin_x+732, height - 100)
#                 columnas = [
#                     ("Cajero:", cajero),
#                     ("Fecha:", fecha),
#                     ("Jornada:", jornada),
#                     ("Cerrado:", cerrado),
#                     ("Saldo Inicial:", saldo_ini),
#                     ("Débitos:", debitos),
#                     ("Créditos:", creditos),
#                     ("Valor Cheques Devueltos:", val_che_dev),
#                     ("Saldo Final:", saldo_fin),
#                     ("Diferencia:", diferencia),
#                     ("Valor Cheques:", val_cheques),
#                     ("Valor Vales:", val_vales),
#                     ]
#                 x = margin_x-30
#                 p.setFont("Helvetica-Bold", 9)
#                 for col, ancho in columnas:
#                     p.drawString(x, y, f"{col} {ancho}")
#                     x += line_y

#         # for label, value in fields:
#         #     pdf.drawString(50, y, f"{label} {value}")
#         #     y -= line_height

#                 # Agregar tabla de denominaciones (si es necesario)
#                 if monedas_json:
#                     from json import loads
#                     monedas = loads(monedas_json)

#                     p.drawString(50, y, "Detalle de Dinero en Efectivo:")
#                     y -= line_y

#                     p.setFont("Helvetica-Bold", 10)
#                     p.drawString(60, y, "Denominación")
#                     p.drawString(160, y, "Cantidad")
#                     p.drawString(260, y, "Total")
#                     p.setFont("Helvetica", 10)
#                     y -= line_y

#                     for moneda in monedas:
#                         p.drawString(60, y, f"{moneda['denominacion']}")
#                         p.drawString(160, y, f"{moneda['cantidad']}")
#                         p.drawString(260, y, f"{moneda['total']}")
#                         y -= line_y
                    
#             # Dibujar contenido
#             pagina_actual = 1
#             dibujar_encabezado()
#             dibujar_pie(pagina_actual, total_paginas)
#             dibujar_encabezado_tabla(height - margin_y - 50)
#             y = height - margin_y - 65

#             # def dibujar_fila(y, row):
#             #     x = margin_x-30
#             #     columnas = [
#             #         (row['cod_cta'], 50), 
#             #         (row['nom_cta'][:17],100), 
#             #         (row['doc_ide'], 60),
#             #         (row['raz_soc'][:15], 95), 
#             #         (row['docto'],20), 
#             #         (row['numero'], 40),
#             #         (row['fecha'], 50), 
#             #         (f"{row['detalle'][:24]}", 125), 
#             #         (f"{row['debito']:,.2f}",70,'right'), 
#             #         (f"{row['credito']:,.2f}", 70,'right'), 
#             #         (f"{row['sal_acu']:,.2f}", 80, 'right')
#             #         ]

#             #     p.setFont("Helvetica", 9)
#             #     for col in columnas:
#             #         if len(col) == 3 and col[2] == 'right':  # Alineación derecha
#             #             texto, ancho, _ = col
#             #             p.drawRightString(x + ancho, y, texto)
#             #         else:  # Alineación izquierda
#             #             texto, ancho = col[:2]
#             #             p.drawString(x, y, str(texto))
#             #         x += ancho

#             # for idx, row in enumerate(resultado):
#             #     if y < margin_y - 30:
#             #         p.showPage()
#             #         pagina_actual += 1
#             #         dibujar_encabezado()
#             #         dibujar_encabezado_tabla(height - margin_y - 50)
#             #         dibujar_pie(pagina_actual, total_paginas)
#             #         y = height - margin_y - 65    

#             #     dibujar_fila(y, row)
#             #     y -= 10

#             p.showPage()
#             p.save()
#             return response
#         else:
#             return HttpResponse("Acción no válida", status=400)
#     return HttpResponse("Método no permitido", status=405)


def saldos_cajero_dia(request):
    print("Session data:", request.session)
    print("GET data:", request.GET)
    cliente_id = request.session.get('cliente_id')  # Obtiene el valor de 'cliente_id' en la sesión
    oficina_id = request.session.get('oficina_id')  # Obtiene el valor de 'oficina_id' en la sesión
    per_con = request.session.get('per_con') 
    hoy = date.today()  # Obtiene la fecha actual
    hoy_formato = hoy.strftime('%Y-%m-%d') 
    id_cajero = request.GET.get('id_cajero')
    fecha = request.GET.get('fecha')
    mov_cierre = []
    cajero = CAJEROS.objects.filter(oficina_id = oficina_id,id = id_cajero).first()
    if cajero == None:
        response_data = {
            'Exito': False, 
            'mensaje': 'No existe cajero',
        }
        return JsonResponse(response_data)
    tot_debitos = 0
    tot_creditos = 0
    monedas_dia = ''
    mov_dia = MOV_CAJA.objects.filter(fecha = fecha,cajero_id = cajero.user_id).first()
    if mov_dia == None:
        val_cheques = 0
        val_vales = 0
    else:
        val_cheques = mov_dia.val_cheques
        val_vales = mov_dia.val_vales
        monedas_dia = mov_dia.monedas
    mov_dia = MOV_CAJA.objects.filter(fecha = fecha,cajero_id = cajero.user_id).first()
    if mov_dia == None:
        val_cheques = 0
        val_vales = 0
    else:
        val_cheques = mov_dia.val_cheques
        val_vales = mov_dia.val_vales
    if isinstance(fecha, str):
        fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
    xfecha = fecha - timedelta(days=1)
    saldo_anterior = 0
    xconta = 0
    while xconta < 20 and saldo_anterior == 0:
        mov_dia_ant = MOV_CAJA.objects.filter(fecha = xfecha,cajero_id = cajero.user_id).first()
        if mov_dia_ant != None:
            if mov_dia_ant.monedas :
                saldo_anterior = mov_dia_ant.saldo_fin
                #for moneda in mov_dia_ant.monedas:
                #    saldo_anterior = saldo_anterior + moneda['denominacion']*moneda['numero']
        xconta = xconta + 1
        xfecha = xfecha - timedelta(days=1)
    
    mov_cierre.append({'docto': 'Saldo Ant.', 'numero': 0,'credito' : saldo_anterior})
    movs_dia_caj = HECHO_ECONO.objects.filter(docto_conta__oficina_id = oficina_id,user_id = cajero.user,fecha = fecha)    
    saldo = saldo_anterior
    for mov_caj in movs_dia_caj:
        if mov_caj.canal == 'CHE':
            continue
        saldo = saldo + mov_caj.valor
        if mov_caj.valor > 0:
            tot_debitos = tot_debitos + mov_caj.valor
            mov_cierre.append({'docto': mov_caj.docto_conta.nombre , 'numero': mov_caj.numero,'debito' : mov_caj.valor,'credito' : 0,'saldo' : saldo})
        else:
            tot_creditos = tot_creditos - mov_caj.valor
            mov_cierre.append({'docto': mov_caj.docto_conta.nombre , 'numero': mov_caj.numero,'debito' : 0,'credito' : -mov_caj.valor,'saldo' : saldo})
    response_data = {
        'Exito': True, 
        'mensaje': 'Ok',
        'saldo_anterior' : saldo_anterior,
        'tot_debitos': tot_debitos,
        'tot_creditos': tot_creditos,
        'val_cheques' : val_cheques,
        'val_vales' : val_vales,
        'monedas' :  monedas_dia,
        'datos_array': mov_cierre,
    }
    return JsonResponse(response_data)
    


# Función para encabezado y pie de página
def agregar_encabezado_pie(canvas, doc):
    styles = getSampleStyleSheet()
    canvas.saveState()
    oficina = OFICINAS.objects.filter(id=1).first()
    entidad = CLIENTES.objects.filter(id=1).first()

    # Encabezado
    canvas.setFont("Helvetica-Bold", 12)
    empresa = f"{entidad.nombre.strip()}"
    texto_empresa = stringWidth(empresa, "Helvetica-Bold", 12)
    canvas.drawString((650 - texto_empresa)/2, 10.5 * inch, empresa)
    # canvas.drawString(100, 10.5 * inch, "Reporte de Movimientos y Cierre de Caja")
    
    nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
    texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
    canvas.drawString((650 - texto_nit) / 2, 10.3 * inch, nit_empresa)
    canvas.setFont("Helvetica", 10)
    canvas.drawString(100, 10.3 * inch, f"Generado el: {doc.fecha_generacion}")
   
                  

    reporte = "Movimiento de Caja"
    texto_reporte = stringWidth(reporte, "Times-Roman", 12)
    canvas.drawString((inch - texto_reporte) / 2, 720 - 60, reporte)

    canvas.setFont("Helvetica", 10)
    detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
    texto_detalles = stringWidth(detalles, "Helvetica", 10)
    canvas.drawString((inch - texto_detalles) / 2, 720 - 75, detalles)
    

    # usuario = f"Cajero: {cajero}"
    # texto_filtro = stringWidth(usuario, "Helvetica", 10)
    # canvas.drawString((inch - texto_filtro) / 2, inch - 90, usuario)

    canvas.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", inch - 50, 720, 60, 60)

    # Pie de página
    canvas.line(inch-65, 0.5 * inch, inch+530, 0.5 * inch)  # Dibuja la línea
    canvas.setFont("Courier", 9)
    texto_pie = f"{oficina.direccion.strip()}  Tel.: {oficina.celular.strip()}  E-mail: {oficina.email.strip()}  {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}  Pág. {doc.page} / {doc.page_count}"
    canvas.drawString(15, 0.3 * inch, texto_pie)
    canvas.restoreState()




def generar_reporte_caja(response_data, tipo_reporte):
    buffer = BytesIO()
    
    # Configurar documento con espacio para encabezado y pie de página
    doc = BaseDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=inch,
        leftMargin=inch,
        topMargin=inch,
        bottomMargin=inch,
    )

    # Guardar datos adicionales en el documento
    doc.fecha_generacion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    doc.page_count = 0  # Se calculará después
    doc.page = 0

    # Crear el marco principal
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id="normal",
    )
    template = PageTemplate(id="test", frames=[frame])
    doc.addPageTemplates([template])

    elements = []
    styles = getSampleStyleSheet()

    # Títulos
    titulo = "Movimiento de Caja" if tipo_reporte == "movimientos" else "Cierre de Caja"
    elements.append(Paragraph(titulo, styles["Title"]))

    # Decodificar datos JSON
    if isinstance(response_data, bytes):
        response_data = json.loads(response_data.decode("utf-8"))

    if tipo_reporte == "cierre":
        saldo_anterior = f"{float(response_data.get('saldo_anterior', 0)):,.2f}"
        tot_debitos = f"{float(response_data.get('tot_debitos', 0)):,.2f}"
        tot_creditos = f"{float(response_data.get('tot_creditos', 0)):,.2f}"
        val_cheques = f"{float(response_data.get('val_cheques', 0)):,.2f}"
        val_vales = f"{float(response_data.get('val_vales', 0)):,.2f}"

        elements.append(Paragraph(f"Saldo Anterior: {saldo_anterior}", styles["Normal"]))
        elements.append(Paragraph(f"Total Débitos: {tot_debitos}", styles["Normal"]))
        elements.append(Paragraph(f"Total Créditos: {tot_creditos}", styles["Normal"]))
        elements.append(Paragraph(f"Valor Cheques: {val_cheques}", styles["Normal"]))
        elements.append(Paragraph(f"Valor Vales: {val_vales}", styles["Normal"]))
        elements.append(Paragraph(" ", styles["Normal"]))  # Espaciado

        # Agregar campo `monedas` al reporte
        monedas = response_data.get("monedas", {})
        if monedas:
            elements.append(Paragraph("Detalle de Monedas:", styles["Heading2"]))
            for moneda, cantidad in monedas.items():
                elements.append(Paragraph(f"{moneda}: {cantidad}", styles["Normal"]))

        # Datos de la tabla
        encabezados = ["Denominación", "Cantidad", "Valor"]
        datos_array =response_data.get("monedas", [])
        datos_tabla = [encabezados]

        for row in datos_array:
            datos_tabla.append([
                # str(row.get("docto", "")),
                # str(row.get("numero", "")),
                f"{float(row.get('denominacion', 0)):,.2f}",
                f"{float(row.get('cantidad', 0)):,.2f}",
                f"{float(row.get('valor', 0)):,.2f}",
            ])

        # Crear tabla
        tabla = Table(datos_tabla, colWidths=[200, 50, 80, 80, 80])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.transparent),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.transparent),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.transparent),
        ]))

        elements.append(tabla)


    # # Datos generales
    # saldo_anterior = f"{float(response_data.get('saldo_anterior', 0)):,.2f}"
    # tot_debitos = f"{float(response_data.get('tot_debitos', 0)):,.2f}"
    # tot_creditos = f"{float(response_data.get('tot_creditos', 0)):,.2f}"
    # val_cheques = f"{float(response_data.get('val_cheques', 0)):,.2f}"
    # val_vales = f"{float(response_data.get('val_vales', 0)):,.2f}"

    # elements.append(Paragraph(f"Saldo Anterior: {saldo_anterior}", styles["Normal"]))
    # elements.append(Paragraph(f"Total Débitos: {tot_debitos}", styles["Normal"]))
    # elements.append(Paragraph(f"Total Créditos: {tot_creditos}", styles["Normal"]))
    # elements.append(Paragraph(f"Valor Cheques: {val_cheques}", styles["Normal"]))
    # elements.append(Paragraph(f"Valor Vales: {val_vales}", styles["Normal"]))
    # elements.append(Paragraph(" ", styles["Normal"]))  # Espaciado

    else:
        # Datos de la tabla
        encabezados = ["Documento", "Número", "        Débito", "        Crédito", "        Saldo"]
        datos_array = response_data.get("datos_array", [])
        datos_tabla = [encabezados]

        for row in datos_array:
            datos_tabla.append([
                str(row.get("docto", "")),
                str(row.get("numero", "")),
                f"{float(row.get('debito', 0)):,.2f}",
                f"{float(row.get('credito', 0)):,.2f}",
                f"{float(row.get('saldo', 0)):,.2f}",
            ])

        # Crear tabla
        tabla = Table(datos_tabla, colWidths=[200, 50, 80, 80, 80])
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.transparent),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.transparent),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.transparent),
        ]))

        elements.append(tabla)

    # Construir el PDF
    doc.build(elements)
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=False, content_type="application/pdf")



def reporte_movimientos_caja(request):
    data = json.loads(request.body)
    id_cajero = data.get("id_cajero")
    fecha = data.get("fecha")

    print('Ejecuta reporte_movimientos_caja  ',id_cajero,'  fecha -----> ',fecha)

    if not request.GET.get('id_cajero') or not request.GET.get('fecha'):
        request.GET = request.GET.copy()
        request.GET['id_cajero'] = id_cajero
        request.GET['fecha'] = fecha # Reemplaza con una fecha válida           

    response_data = saldos_cajero_dia(request)
    if response_data.status_code != 200:
        return HttpResponse("Error al obtener los datos", status=500)
    
    datos = json.loads(response_data.content.decode('utf-8'))
    tipo_reporte = request.GET.get('tipo_reporte', 'movimientos')

    return generar_reporte_caja(datos, tipo_reporte)



#         # elif accion == "csv":
#         #     # Llama a la función para obtener los datos
#         #     # resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
#         #     if not resultado:
#         #         return HttpResponse("No se encontraron datos para exportar", status=404)
            
#         #     # Configurar la respuesta HTTP para un archivo CSV
#         #     response = HttpResponse(content_type='text/csv')
#         #     response['Content-Disposition'] = 'attachment; filename="movimiento_caja.csv"'
            
#         #     # Crear el escritor CSV
#         #     writer = csv.writer(response)
            
#         #     # Añadir encabezados de las columnas
#         #     headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
#         #     writer.writerow(headers)

#         #     # Añadir los datos
#         #     for fila in resultado:
#         #         writer.writerow([fila[col] for col in headers])
#         #     return response

