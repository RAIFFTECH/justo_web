import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from justo_app.funciones_principales import formato_fecha, formatear_cod_aso

from django import forms
from .forms import CrearForm
from .models import ORIGINACION

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = ORIGINACION
    form = CrearForm
    template_name = 'lista_originacion.html'
    # ordering = ['cliente', 'codigo']

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = ORIGINACION
    form = CrearForm
    template_name = 'detalles_originacion.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = ORIGINACION
    form_class = CrearForm
    template_name = 'crear_originacion.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_originacion')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ORIGINACION
    form_class = CrearForm
    template_name = 'actualizar_originacion.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_originacion')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = ORIGINACION
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_originacion')

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Originación"

        # Crear los encabezados en la primera fila
        sheet.append([
            "Asociado", "Línea de Crédito", "Monto", "Plazo",
            "Garantía Crédito Solidario?", "Línea Crédito Solidario",
            "Modalidad Crédito Solidario", "Modalidad Crédito Solidario",
            "Forma de Pago"
        ])

        # Obtener los datos del modelo ORIGINACION
        originaciones = ORIGINACION.objects.all()

        # Agregar los datos de cada originación a la hoja de cálculo
        for originacion in originaciones:
            sheet.append([
                originacion.asociado.cod_aso if originacion.asociado else '',
                originacion.lin_cre,
                originacion.monto,
                originacion.plazo,
                originacion.gar_cre_sol,
                originacion.lin_cre_sol,
                originacion.mod_cre_sol,
                originacion.sol_cre_edu,
                originacion.for_pag
            ])

        # Preparar la respuesta HTTP para devolver el archivo Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=originacion.xlsx"

        # Guardar el libro de trabajo en la respuesta
        workbook.save(response)

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = ORIGINACION.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="originacion.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Asociado: {dato.asociado}")
        p.drawString(80, 780, f"Línea de Crédito: {dato.lin_cre}")
        p.drawString(80, 760, f"Monto: {dato.monto}")
        p.drawString(80, 740, f"Plazo: {dato.plazo}")
        p.drawString(80, 720, f"Garantía Crédito Solidario?: {dato.gar_cre_sol}")
        p.drawString(80, 700, f"Línea Crédito Solidario: {dato.lin_cre_sol}")
        p.drawString(80, 680, f"Modalidad Crédito Solidario: {dato.mod_cre_sol}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        originacion = ORIGINACION.objects.all()

        p = canvas.Canvas(response)

        for dato in originacion:
            p.drawString(80, 800, f"Asociado: {dato.asociado}")
            p.drawString(80, 780, f"Línea de Crédito: {dato.lin_cre}")
            p.drawString(80, 760, f"Monto: {dato.monto}")
            p.drawString(80, 740, f"Plazo: {dato.plazo}")
            p.drawString(80, 720, f"Garantía Crédito Solidario?: {dato.gar_cre_sol}")
            p.drawString(80, 700, f"Línea Crédito Solidario: {dato.lin_cre_sol}")
            p.drawString(80, 680, f"Modalidad Crédito Solidario: {dato.mod_cre_sol}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in originacion._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        originacion = ORIGINACION.objects.all()
        for row_num, data in enumerate(originacion, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        originacion = ORIGINACION.objects.all()
        for data in originacion:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
    

from creditos_app.models import CREDITOS
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from urllib.parse import quote


def reporte_evaluacion_cartera(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')

    if request.method == 'GET':
        return render(request, 'Evaluacion_cartera.html') 

    if request.method == 'POST':
        print('Generando resultados ---->')
        accion = request.POST.get("accion")   
        fecha_corte = request.POST.get('fecha_corte')
        min_dias = 0
        max_dias = 99999

        saldos = evaluacion_catera(id_cli, id_ofi, fecha_corte, min_dias, max_dias)
        fecha_corte_formateada = formato_fecha(fecha_corte)
        if not saldos:
            return HttpResponse("No se encontraron datos para exportar", status=404)

        if accion == "exportar":
            print('Exportar a Excel...')
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CARTERA A UNA FECHA" 
            headers = list(saldos[0].keys())
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)
        
            for row_num, data in enumerate(saldos, start=2):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])

            nombre_archivo = f"evaluacion_cartera_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            print('Reporte Grabado ..',nombre_archivo)
            return response

        elif accion == "csv":
            nombre_archivo = f"evaluacion_cartera_{fecha_corte_formateada}.csv"
            print('Imprime Archivo CSV......', nombre_archivo)
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            writer = csv.writer(response)
            headers = list(saldos[0].keys())
            writer.writerow(headers)
            for fila in saldos:
                writer.writerow([fila[col] for col in headers])
                
            return response

        else:
            return HttpResponse("Acción no permitida", status=405)

    return render(request, 'cartera_a_una_fecha.html')

from datetime import datetime
from dateutil.relativedelta import relativedelta
from terceros_app.models import TERCEROS
from django.db.models import Count
from justo_app.justo_creditos import Liquida_cre
from recla_carte_app.models import CARTE_CAT_HIS
from creditos_app.models import CREDITOS,GAR_NO_IDONEA
from categorias_creditos_app.models import CAT_DES_DIA_CRE 
from estados_financieros_app.models import ESTADOS_FIN
from django.db.models import Q


def evaluacion_catera(cliente_id, id_oficina, fecha_corte_str,min_dias,max_dias):
    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    tres_anios_atras = fecha_corte - relativedelta(years=3)
    print('fecha Corte y tres años ',fecha_corte,tres_anios_atras,id_oficina)
    resultados = []
    Creditos = CREDITOS.objects.filter(oficina_id = id_oficina, fec_des__lte = fecha_corte 
            ).exclude(
                # Excluir créditos con estado 'C' y fec_ult_pag < self.fecha_ant_rec
                Q(estado='C') & Q(fec_ult_pag__lt = fecha_corte)
            ).exclude(
                # Excluir créditos con estado 'H'
                estado='H'
            ).exclude(
                # Excluir créditos con estado 'H'
                fec_des__gt = fecha_corte
            )
    for credito in Creditos:
        his_num_cre = (
            CREDITOS.objects.filter(
                oficina_id=1,
                fec_des__gte=tres_anios_atras,
                socio__tercero__doc_ide = credito.socio.tercero.doc_ide
            )
            .exclude(estado= 'X')
            .values('socio__tercero__doc_ide')
            .annotate(total=Count('id'))
        )
        if his_num_cre:
            xtot_cre_ter = his_num_cre[0]['total']
        else:
            xtot_cre_ter = 0
        
        base_his = (
            CARTE_CAT_HIS.objects.filter(
                nit = credito.socio.tercero.doc_ide,
                fecha__gte = tres_anios_atras,  # ← filtro por fecha
                nit__in=TERCEROS.objects.filter(cliente_id=1)
                    .values_list('doc_ide', flat=True)
                )
                .values('nit')
                .annotate(
                total=Count('id'),
                buenas=Count('id', filter=Q(cat_arr='A'))
            )
        )

        if  base_his:
            xPorCum = base_his[0]['buenas'] / base_his[0]['total'] * 100 if base_his[0]['total'] > 0 else 100
        else:
            xPorCum = 0

        diferencia = relativedelta(fecha_corte, credito.socio.fec_afi)
        xAntMes = (diferencia.years * 12 + diferencia.months) if credito.socio.estado == 'A' or credito.socio.estado == 'R' else 0

        liq_cre = Liquida_cre(credito.cod_cre,fecha_corte)
        if liq_cre.lista_mov == None:
            credito.estado = 'H'
            credito.save()
            continue 
        liq_cre.liq_al_dia(fecha_corte)
        if liq_cre.sal_cap_tot <= 0 :
            continue 
        liq_cre.calculo_periodo()
        xdias_mor = (liq_cre.fecha_focal-liq_cre.fec_al_dia).days
        xdias_mor = xdias_mor if xdias_mor > 0 else 0
        if xdias_mor < min_dias or xdias_mor > max_dias:
            continue
        aportes_his = CARTE_CAT_HIS.objects.filter(oficina_id = id_oficina,fecha = fecha_corte,nit = credito.socio.tercero.doc_ide)
        xSalApo = 0
        xProApo = 0
        if aportes_his:
            for aporte in aportes_his:
                xSalApo = xSalApo + aporte.aporte 
                if aporte.cod_cre ==  credito.cod_cre:
                    xProApo = aporte.aporte

        micarhis = CARTE_CAT_HIS.objects.filter(oficina_id = id_oficina,fecha = fecha_corte,cod_cre = credito.cod_cre).first()
        xSalMin = 1440000
        xAprobo = ' ' * 10
        xTipGar = ' ' * 10
        if micarhis == None:
            if xdias_mor < 1:
                xCat = 'A'
            else:
                # print(' ord(Credito.cod_des)',Credito.cod_des)
                CatDesDia = CAT_DES_DIA_CRE.objects.filter(cliente_id = cliente_id,codigo = ord(credito.cod_des),
                    minimo_dias__lte=xdias_mor,maximo_dias__gte=xdias_mor).first()
                if CatDesDia == None:
                    xCat = 'F'
                else:
                    xCat = CatDesDia.categoria
            xcat_mor = xCat
            xcat_arr = 'X'
            xcat_eva = 'X'
            xcat_ree = 'X'
            xcat_mod = 'X'
            xcategoria = 'X'
            xProIndKap = 0
        else:
            xcat_mor = micarhis.cat_mor
            xcat_arr = micarhis.cat_arr
            xcat_eva = micarhis.cat_eva
            xcat_ree = micarhis.cat_ree
            xcat_mod = micarhis.cat_mod
            xcategoria = micarhis.categoria
            xProIndKap = micarhis.pro_ind_kap
            xAprobo = ' ' * 10
            if credito.cod_lin_cre.cod_lin_cre == 49:
                xAprobo = 'CONSEJO.ADM'
            elif credito.cod_lin_cre.cod_lin_cre == 50:
                xAprobo = 'GERENCIA'
            elif credito.cod_lin_cre.cod_lin_cre == 51:
                xAprobo = 'GERENCIA'
            elif credito.cod_lin_cre.cod_lin_cre == 52:
                xAprobo = 'COM.CREDITO'
            elif credito.cod_lin_cre.cod_lin_cre == 53:
                xAprobo = 'CONSEJO.ADM'
            elif credito.cod_lin_cre.cod_lin_cre == 54:
                if credito.cap_ini <= xSalMin * 40 :
                    xAprobo = 'GERENCIA'
                elif credito.cap_ini <= xSalMin * 93 :
                    xAprobo = 'COM. CREDITO'
                else:
                    xAprobo = 'CONSEJO.ADM'
        xTipGar = 'Hipotecario' if credito.tip_gar == '2' else ('No Idonea' if credito.tip_gar == '15' else 'Admisible' )
        est_fin = ESTADOS_FIN.objects.filter(cliente_id = cliente_id,tercero = credito.socio.tercero).first()
        codeu = GAR_NO_IDONEA.objects.filter(oficina_id = id_oficina,credito = credito).first()
        if codeu != None:
            cod_ter = TERCEROS.objects.filter(cliente_id = cliente_id,doc_ide = codeu.doc_ide).exclude(doc_ide=credito.socio.tercero.doc_ide).first()
        else:
            cod_ter = None
        xSalProMes = liq_cre.sal_cap_tot/((credito.num_cuo_act-liq_cre.cuo_pag) if credito.num_cuo_act != liq_cre.cuo_pag else 1)
        resultados.append({
            'cod_soc' : credito.socio.cod_aso,
            'nombre': credito.socio.tercero.nombre,
            'cod_cre': credito.cod_cre, 
            'fec_des' : credito.fec_des.strftime('%d/%m/%Y'),
            'fec_ven' : credito.fec_ven,
            'tasint' : credito.tiae_ic_act,
            'numcuo' : credito.num_cuo_act,
            'altura' : liq_cre.altura,
            'CapIni' : credito.cap_ini,
            'ValCuo' : credito.val_cuo_act,
            'CuoPag' : liq_cre.cuo_pag,
            'DiaMor' : xdias_mor,
            'saldo_K' : liq_cre.sal_cap_tot,
            'salpromes' : xSalProMes,
            'AltReal' : 0,
            'tasintactnom' : credito.tian_ic_act,
            'Provision' : xProIndKap,
            'CatIni' : 'A' if credito.cat_nue == ' ' else credito.cat_nue,
            'CalAct' : xcat_arr,  # podria ser categoriA
            'TipCre' : credito.imputacion.descripcion,
            'Aprobo' : xAprobo,
            'TipGar' : xTipGar,
            'TelRes' : credito.socio.tercero.celular1,           
            'TelOfi' :  credito.socio.tel_tra,
            'Celular' : credito.socio.tercero.celular2,
            'email' : credito.socio.tercero.email,
            'nomdescre' : credito.imputacion.descripcion,
            'ValGart' : credito.val_gar_hip,
            'SalMora' : liq_cre.sal_cap_dia+liq_cre.sal_int_dia+liq_cre.sal_int_mor,
            'ConAli' : 'N',
            'FecAli' : ' '*10,
            'numcuoali' : 0,
            'ocupacion' : credito.socio.ocupacion,
            'CiuRes' : credito.socio.ciu_tra.nombre,
            'edad' : (fecha_corte.year - credito.socio.fec_nac.year),
            'Genero' : credito.socio.sexo,
            'nivel_est' : credito.socio.niv_est,
            'estrato' : credito.socio.estrato,
            'ing_mensual ' : est_fin.ing_tot if est_fin != None else 0,
            'GasMen' : est_fin.egr_tot if est_fin != None else 0,
            'Patrimonio' : est_fin.tot_act - est_fin.tot_pas if est_fin != None else 0,
            'EstCiv' : credito.socio.est_civ,
            's01MujCabFam ' : 'NO',
            'Profesion' : credito.socio.profesion,
            'num_hijos' : credito.socio.num_hij_may + credito.socio.num_hij_men,
            'PorCum' : xPorCum,
            'Creditos' : xtot_cre_ter,
            'Nit' : credito.socio.tercero.doc_ide,
            'SalApo' : xSalApo,
            'FecIngC' : credito.socio.fec_afi,
            'antiaso' : int(xAntMes / 12),
            'FecRetC' : credito.socio.fec_ret,
            'FecIngLab' : credito.socio.fec_ing_tra,
            'ValAct' : est_fin.tot_act if est_fin != None else 0,
            'ValPas' : est_fin.tot_pas if est_fin != None else 0,
            'TipCon' : credito.socio.tip_con,
            'ProApo' : xProApo,

        })
    print('Culmina Generacion')
    return resultados


