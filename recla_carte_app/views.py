import csv, calendar, time, redis, os, openpyxl
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
from dateutil.relativedelta import relativedelta
from django.views import View
from django.http import HttpResponse
from django.db.models.functions import Substr, Coalesce, ExtractYear
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import JsonResponse
from django.db.models import Sum, Count, Case, When, IntegerField, F, Value, Q, Prefetch, FloatField, ExpressionWrapper, DecimalField
from io import BytesIO
from .tasks import tarea_lenta
from recla_carte_app.tasks import ejecutar_modelo_task

from reportlab.graphics.shapes import Drawing, Rect
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.textlabels import Label
from reportlab.graphics.widgets.markers import makeMarker
from reportlab.lib.colors import lightblue, navy
from reportlab.graphics import renderPDF

from justo_app.justo_creditos import perdida_esperada, Liquida_cre, Reclasificacion
from categorias_creditos_app.models import CAT_DES_DIA_CRE
from justo_app.funciones_principales import formato_fecha, formatear_cod_aso
from clientes_app.models import CLIENTES
from pagadores_app.models import PAGADORES
from oficinas_app.models import OFICINAS
from terceros_app.models import TERCEROS
from asociados_app.models import ASOCIADOS, ASO_BENEF, ASO_REFERENCIAS
from localidades_app.models import LOCALIDADES
from hecho_economico_app.models import HECHO_ECONO, DOCTO_CONTA
from detalle_producto_app.models import DETALLE_PROD
from cuentas_app.models import PLAN_CTAS
from recla_carte_app.models import PE_CARTE_HIS,RPKI
from creditos_app.models import CREDITOS, CODEUDORES, GAR_NO_IDONEA
from creditos_app.views import lista_creditos_asociado, fecha_ultimo_movimiento
from ctas_ahorros_app.views import tasa_nominal_anual
from .models import CARTE_CAT_HIS
from detalle_economico_app.models import DETALLE_ECONO
from contabilizacion_intereses_creditos_app.models import IMP_CON_CRE_INT
from estados_financieros_app.models import ESTADOS_FIN
from creditos_app.models import GAR_NO_IDONEA

def cifin(request):
    if request.method == 'GET':
        return render(request, 'cifin.html')
    if request.method == 'POST':
        accion = request.POST.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_cor = datetime.strptime(request.POST['fec_cor'], '%Y-%m-%d').date()
        reporte_entrega = request.POST['reporte_entrega']
        print('reporte a entregar', reporte_entrega)
        reporte_entrega_nom = {
            '1': 'Primera Entrega',
            '2': 'Segunda Entrega'
            }
        estado_nombre = reporte_entrega_nom.get(reporte_entrega)
        
        if reporte_entrega == '2':
            fec_fin = fec_cor + timedelta(days=20)
            print('fecha mas 20', fec_fin)
        elif reporte_entrega == '1':
            fec_fin = fec_cor

        print('entrega reporte', reporte_entrega)          
        resultado = listado_cifin(fec_fin, reporte_entrega)
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte_Cifin" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = listado_cifin(fec_ini, reporte_entrega)
            # print('estos resultados--->', resultado)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "REPORTE CENTRAL DE INFORMACIÓN FINANCIERA"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_cor} a {fec_fin}"
            filtro_cuentas = f"Situación: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"cifin_{estado_nombre}_{fec_ini}_{fec_cor}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = listado_cifin(fec_fin, reporte_entrega)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"cifin_{estado_nombre}_{fec_cor}_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            # headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            # writer.writerow(headers)
            
            # with open("cifin.txt", "w", encoding="utf-8") as f:
            # for linea in registros:
            #     f.write(linea)
            
            # with open(nombre_archivo, "w", newline="") as f:
                # writer = csv.writer(f)
                # fila = "Hola mundo"
                # writer.writerow([fila])  # Una sola columna con "Hola mundo"

            # Añadir los datos
            for fila in resultado:
                # writer(fila)
                # writer.writerow([fila[col] for col in headers])
                response.write(fila + "\r\n")
                # writer.writerow([fila])
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)

# def listado_cifin(fec_corte: date, tipo_reporte: str):
#     # Fecha focal según tipo de reporte
#     if tipo_reporte == 1:
#         f_focal = fec_corte
#     else:
#         f_focal = fec_corte + timedelta(days=20)

#     # Lista de líneas del archivo
#     registros = []

#     # Encabezado
#     registros.append(f"121050181          01{f_focal.strftime('%Y%m%d')}\r\n")

#     # Créditos base
#     creditos = CREDITOS.objects.filter(
#         # ~Q(est_jur='P'),
#         # ~Q(com_des_id=''),
#         estado = 'A',
#         rep_cen_rie='S',
#         fec_des__lte=f_focal
#     ).order_by('cod_cre')
#     # print('estos creditos',creditos)
#     for credito in creditos:
#         cod_cre = credito.cod_cre
#         # Busca la categoría del crédito a la fecha
#         categoria = CARTE_CAT_HIS.objects.filter(
#             cod_cre=cod_cre,
#             fecha__lte=f_focal
#         ).order_by('-fecha').first()

#         # Validación si continúa
#         if not categoria and not (
#             credito.fec_des <= f_focal.replace(day=1) - timedelta(days=1)
#             and credito.fec_ult_pag.year == f_focal.year
#             and credito.fec_ult_pag.month == f_focal.month
#         ) and credito.est_jur != "K":
#             continue

#         # Lógica de cálculo del saldo, mora, cuotas, fechas, estado, etc.
#         # Aquí deberías llamar a tus funciones CalculoEdadCr y LiqCreAlDia adaptadas a Python
#         saldo = calcular_saldo_credito(credito, f_focal)
#         if saldo <= 0 and credito.est_jur != "K":
#             continue

#         # Información del titular (Socio y Tercero)
#         asociado = ASOCIADOS.objects.filter(id=credito.socio_id).first()
#         tercero = asociado.tercero if asociado else None
#         pagador = PAGADORES.objects.filter(id=asociado.id_pag_id).first() if asociado else None

#         # Información del codeudor
#         codeudores = GAR_NO_IDONEA.objects.filter(credito_id=credito.id)
#         codeudor_data = []
#         for codeudor in codeudores:
#             tercero_codeudor = TERCEROS.objects.filter(doc_ide=codeudor.doc_ide).first()
#             if tercero_codeudor:
#                 codeudor_data.append(obtener_linea_codeudor(credito, tercero_codeudor, f_focal, categoria))

#         # Construcción de la línea del titular
#         # linea_titular = construir_linea_titular(credito, tercero, categoria, saldo, fec_corte, f_focal, asociado, pagador)
#         linea_titular = construir_linea_titular(credito, tercero, f_focal, categoria)
#         registros.append(linea_titular)

#         # Agregar las líneas de codeudores
#         registros.extend(codeudor_data)

#     # Pie con cantidad de registros
#     registros.append(f"9{str(len(registros)).zfill(8)}\r\n")

#     # # Guardar el archivo
#     # with open(nom_archivo, 'w', encoding='utf-8') as f:
#     #     f.writelines(registros)
#     # print('estos registros cifin', registros)
        
#     return registros #nom_archivo

# # Funciones auxiliares básicas
# def calcular_saldo_credito(credito, f_focal):
#     # Adaptar la lógica de CalculoEdadCr y LiqCreAlDia
#     # Ejemplo simplificado:
#     return credito.cap_ini

# # def construir_linea_titular(credito, tercero, categoria, saldo, fec_cor, f_focal, socio, pagador):
# def construir_linea_titular(credito, tercero, fecha_corte, categoria):
#     # Aquí ensamblas la línea tipo 2, usando los campos que has traído.
#     # Por simplicidad solo mostramos una parte
#     # 🔹 Diccionarios de mapeo
#     mapeo_tip_doc = {'C': '01', 'N': '02', 'E': '03', 'T': '04', 'P': '05','R': '09'}
#     mapeo_tip_ter = {'N': '000', 'J': '002'} 
#     mapeo_categoria = {'A': '01', 'B': '02', 'C': '03', 'D': '04', 'E': '05'}
#     xCuoMor = 0
#     xValMor = "000000000000"
#     if categoria.dias_mor < 30:
#         xEdadMora = "00"
#         xValMor = "000000000000"
#     elif 30 <= categoria.dias_mor < 60:
#         # if credito.not_mor and credito.fec_not_mor < fecha_corte:
#             # xEdadMora = "01"
#         # else:
#             xEdadMora = "00"
#             xCuoMor = 0
#     elif 60 <= categoria.dias_mor < 90:
#         xEdadMora = "02"
#     elif 90 <= categoria.dias_mor < 120:
#         xEdadMora = "03"
#     elif 120 <= categoria.dias_mor < 150:
#         xEdadMora = "04"
#     elif 150 <= categoria.dias_mor < 180:
#         xEdadMora = "05"
#     elif 180 <= categoria.dias_mor < 210:
#         xEdadMora = "06"
#     elif 210 <= categoria.dias_mor < 240:
#         xEdadMora = "07"
#     elif 240 <= categoria.dias_mor < 270:
#         xEdadMora = "08"
#     elif 270 <= categoria.dias_mor < 300:
#         xEdadMora = "09"
#     elif 300 <= categoria.dias_mor < 330:
#         xEdadMora = "10"
#     elif 330 <= categoria.dias_mor < 360:
#         xEdadMora = "11"
#     elif 360 <= categoria.dias_mor < 540:
#         xEdadMora = "12"
#     elif 540 <= categoria.dias_mor < 730:
#         xEdadMora = "13"
#     elif categoria.dias_mor >= 730:
#         xEdadMora = "14"
        
#     xAnosMora = categoria.dias_mor // 360
        
#     tip_doc =  mapeo_tip_doc.get(tercero.cla_doc) if tercero else ''.ljust(2)
#     doc_ide = tercero.doc_ide.ljust(15) if tercero else ''.ljust(15)
#     pri_ape = tercero.pri_ape.ljust(15) if tercero else ''.ljust(15)
#     seg_ape = tercero.seg_ape.ljust(15) if tercero else ''.ljust(15)
#     pri_nom = tercero.pri_nom.strip() if tercero else ''
#     seg_nom = tercero.seg_nom.strip() if tercero else ''
#     nombre_completo = pri_nom+' '+seg_nom
#     nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
#     espacio2 = ''.ljust(2)
#     fec_lim_pag = fecha_corte.strftime('%Y%m%d')
#     cod_cre = credito.cod_cre.ljust(20)
#     cod_age = '000001'
#     deu_pri = 'P'
#     calificacion = mapeo_categoria.get(categoria.cat_arr, '05') if tercero else ''.ljust(2)
#     sit_est_tit = '05' # xSitEstTit+ ;		&& Situacion o Estado del Titular (Tabla 24 Cifin)
#     estado = '01' # xEstado+ ;		&& Estado (Tabla 10 Cifin) xEstado = IIF(S24SALCAP<=1,"07","01") && 07=SALDADO, 01=VIGENTE
#     edad_mora = xEdadMora #		carte_cat_his__dias_mor     && Edad de Mora (Tabla 8 Cifin)
#     anos_mora = xAnosMora # xAnosMora+ ;		&& Anos En Mora
#     fec_act = fecha_corte.strftime('%Y%m%d') # xFecAct+ ;		fec_cor                     && Fecha de Corte  (AAAAMMDD)
#     fec_des = credito.fec_des.strftime('%Y%m%d')#		creditos__fec_des           && Fecha Inicial o de Expedicion (AAAAMMDD)
#     fec_ven = credito.fec_ven.strftime('%Y%m%d')# xFecVen+ ;creditos__fec_ven       && Fecha Terminacion (AAAAMMDD)                 
#     fec_exi = fec_ven # xFecExi+ ;		&& Fecha Exigibilidad (AAAAMMDD)
#     espacio8 = ''.ljust(8)# space(8)+ ;		&& Fecha Prescripcion (AAAAMMDD)
#     fec_ult_pag = credito.fec_ult_pag.strftime('%Y%m%d') # xFecPag+ ;		creditos__fec_ult_pag       && Fecha Pago (AAAAMMDD)
#     mod_ext = '  '# xModExt+ ;		&& Modo Extincion (Tabla 9 Cifin)
#     tip_pago = '01' # "01" + ;			&& Tipo Pago (Tabla 31 Ci"01" --> Volunt"02" -->No voluntario
#     per_pag = credito.per_ano # xPerPag + ;		creditos__per_ano           && Periocidad de Pago (Tabla 6 Cifin)
#     pro_no_pago = 'GYE' # xProbNoPag	+ ;	&& Probabilidad de No Pago (Valor entre 0 y 100)
#     # xCuoCan+ ;		&& Numero Cuotas Pagadas
#     tot_cuo = credito.num_cuo_ini # xTotCuo+ ;		creditos__num_cuo_ini       && Numero Cuotas Pactadas
#     cuo_mor = xCuoMor # xCuoMor+ ;		&& Cuotas en Mora
#     cap_ini = '{:012}'.format(int(categoria.cap_ini / 1000)) # xValIni+ ;		carte_cat_his__cap_ini      && Valor o Cupo (En Miles)
#     val_mor = xValMor # xValMor+ ;		&& Valor de Mora (En Miles)
#     val_saldo = '{:012}'.format(int(categoria.saldo_1 / 1000))# xValSaldo+ ;		carte_cat_his__sal_cap      && Valor Saldo (En Miles)
#     val_cuo = '{:012}'.format(int(credito.val_cuo_ini / 1000)) # xValCuo+ ;		creditos__val_cuo_ini       && Valor de la Cuota (En Miles)
#     espacio12 = ''.ljust(12) # space(12)+ ;		&& Valor de Cargo Fijo ### (No Aplica) ###
#     lin_cre = '008' # "008"+ ;		carte_cat_his__cod_lin_cre  && Linea Credito (Tabla 3 Cifin)
#     espacio3 = ''.ljust(3)# space(3)+ ;		&& Clausula de Permanencia ### (No Aplica) ###
#     # "001"+ ;			&& Tipo de Contrato (Tabla 25 Cifin)
#     # "001"+ ;			&& Estado del Contrato (Tabla 26 Cifin)
#     espacio5 = ''.ljust(5)
#     # space(3)+ ;		&& Termino o Vigencia del Contrato ### (No Aplica) ###
#     # space(2)+ ;		&& Numero Meses Contrato ### (No Aplica) ###
#     naturaleza = mapeo_tip_ter.get(tercero.tip_ter) if tercero else ''.ljust(3)# xNaturaleza+ ;	&& Naturaleza Juridica (Tabla 19 Cifin)
#     mod_cre = '02' # "02"+ ;			&& Modalidad de Credito (Tabla 2 Cifin) -- consumo
#     # xTipMon+ ;		&& Tipo de Moneda (Tabla 5 Cifin)
#     # xTipGar+ ;		&& Tipo Garantia (Tabla 4 Cifin)
#     # "000000000000"+ ;	&& Valor Garantia (En Miles)
#     # "02"+ ;			&& Obligacion Reestructurada (Tabla 32 Cifin)
#     espacio39 = ''.ljust(39)
#     # space(2)+ ;		&& Naturaleza Reestructuracion (Tabla 11 Cifin)
#     # space(3)+ ;		&& Numero Reestructuraciones
#     # space(3)+ ;		&& Clase Tarjeta ### (No Aplica) ### 
#     # space(4)+ ;		&& Numero de Cheques Devueltos ### (No Aplica) ###		
#     # space(2)+ ;		&& Categoria Servicios ### (No Aplica) ###
#     # space(2)+ ;		&& Plazo ### (No Aplica) ### 
#     # space(6)+ ;		&& Dias Cartera ### (No Aplica) ###
#     # space(2)+ ;		&& Tipo Cuenta ### (No Aplica) ###
#     # space(12)+ ;		&& Cupo Sobregiro ### (No Aplica) ###
#     # space(3)+ ;		&& Dias Autorizados ### (No Aplica) ###
#     direccion = tercero.direccion # xDirRes + ; 		tercero__direccion          && Direccion Titular
#     tel_res = tercero.celular1 # xTelRes + ;		tercero__celular1           && Telefono Casa Titular
#     cod_ciu = tercero.cod_ciu_res.codigo # "000001" + ;		tercero__localidades_codigo && Codigo Ciudad Casa
#     ciu_res = tercero.cod_ciu_res.nombre.strip()  # xCiuRes + ;		tercero__localidades_nombre && Ciudad Casa Titular
#     cod_dep = tercero.cod_ciu_res.cod_pos[:2].ljust(3, '0') # "050" + ;	tercero__localidades_cod_pos[:2]    && Codigo Departamento Casa
#     departamento = tercero.cod_ciu_res.departamento.strip() # "META" + ; tercero__localidades_departamento   && Departamento Casa Titular
#     espacio293 = ''.ljust(293)
#     # space(60)+ ;		&& Nombre Empresa
#     # space(60)+ ;		&& Direccion Empresa
#     # space(20)+ ;		&& Telefono Empresa
#     # space(6)+ ;		&& Codigo Ciudad Empresa
#     # space(6)+ ;		&& Codigo Departamento Empresa
#     # space(20)+ ;		&& Departamento Empresa Titular
#     # space(8)+ ;		&& Fecha Inicio Excension GMF
#     # space(8)+ ;		&& Fecha Terminacion Escension GMF
#     # space(2)+ ;		&& Numero Renovacion CDT
#     # space(2)+ ;		&& Cuenta Ahorro Excenta GMF
#     # space(2)+ ;		&& Tipo Identificacion Originaria
#     # space(14)+ ;		&& Numero Identificacion Originaria
#     # space(3)+ ;		&& Tipo Entidad Originaria
#     # space(3)+ ;		&& Codigo Entidad Originaria
#     # space(2)+ ;		&& Tipo Fideicomiso
#     # space(3)+ ;		&& Numero Fideicomiso
#     # space(60)+ ;		&& Nombre Fideicomiso
#     # space(4)+ ;		&& Tipo Deuda Cartera
#     # space(4)+ ;		&& Tipo de Poliza
#     # space(6) +;		&& Codigo de Ramo
    

#     # Debes agregar todos los campos en el orden y tamaño exacto como en tu FoxPro
#     return f"2{tip_doc}{doc_ide}{pri_ape}{seg_ape}{nombre}{espacio2}{fec_lim_pag}{cod_cre}{cod_age}{deu_pri}{calificacion}{sit_est_tit}{estado}{edad_mora}{anos_mora}{fec_act}{fec_des}{fec_ven}{fec_exi}{espacio8}{fec_ult_pag}{mod_ext}{tip_pago}{fec_ult_pag}{per_pag}{pro_no_pago}{tot_cuo}{cuo_mor}{cap_ini}{val_saldo}{val_mor}{val_cuo}{espacio12}{lin_cre}{espacio3}{espacio5}{naturaleza}{mod_cre}{espacio39}{direccion}{tel_res}{cod_ciu}{ciu_res}{cod_dep}{departamento}{espacio293}\r\n"

# def obtener_linea_codeudor(credito, tercero, fecha_corte, categoria):
#     # Similar a construir_linea_titular pero con los datos del codeudor
#     # nit = tercero.doc_ide.ljust(15)
#     # nombre = tercero.nombre.ljust(30)
#     # cod_cre = credito.cod_cre.ljust(20)
    
#     # 🔹 Diccionarios de mapeo
#     mapeo_tip_doc = {'C': '01', 'N': '02', 'E': '03', 'T': '04', 'P': '05','R': '09'}
#     mapeo_tip_ter = {'N': '000', 'J': '002'} 
#     mapeo_categoria = {'A': '01', 'B': '02', 'C': '03', 'D': '04', 'E': '05'}
    
#     if categoria.dias_mor < 30:
#         xEdadMora = "00"
#         xValMor = "000000000000"
#     elif 30 <= categoria.dias_mor < 60:
#         # if credito.not_mor and credito.fec_not_mor < fec_cor:
#             # xEdadMora = "01"
#         # else:
#             xEdadMora = "00"
#             # xCuoMor = 0
#     elif 60 <= categoria.dias_mor < 90:
#         xEdadMora = "02"
#     elif 90 <= categoria.dias_mor < 120:
#         xEdadMora = "03"
#     elif 120 <= categoria.dias_mor < 150:
#         xEdadMora = "04"
#     elif 150 <= categoria.dias_mor < 180:
#         xEdadMora = "05"
#     elif 180 <= categoria.dias_mor < 210:
#         xEdadMora = "06"
#     elif 210 <= categoria.dias_mor < 240:
#         xEdadMora = "07"
#     elif 240 <= categoria.dias_mor < 270:
#         xEdadMora = "08"
#     elif 270 <= categoria.dias_mor < 300:
#         xEdadMora = "09"
#     elif 300 <= categoria.dias_mor < 330:
#         xEdadMora = "10"
#     elif 330 <= categoria.dias_mor < 360:
#         xEdadMora = "11"
#     elif 360 <= categoria.dias_mor < 540:
#         xEdadMora = "12"
#     elif 540 <= categoria.dias_mor < 730:
#         xEdadMora = "13"
#     elif categoria.dias_mor >= 730:
#         xEdadMora = "14"
        
#     xAnosMora = categoria.dias_mor / 360
        
#     tip_doc =  mapeo_tip_doc.get(tercero.cla_doc) if tercero else ''.ljust(2)
#     doc_ide = tercero.doc_ide.ljust(15) if tercero else ''.ljust(15)
#     pri_ape = tercero.pri_ape.ljust(15) if tercero else ''.ljust(15)
#     seg_ape = tercero.seg_ape.ljust(15) if tercero else ''.ljust(15)
#     pri_nom = tercero.pri_nom.strip() if tercero else ''
#     seg_nom = tercero.seg_nom.strip() if tercero else ''
#     nombre_completo = pri_nom+' '+seg_nom
#     nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
#     espacio2 = ''.ljust(2)
#     fec_lim_pag = fecha_corte.strftime('%Y%m%d')
#     cod_cre = credito.cod_cre.ljust(20)
#     cod_age = '000001'
#     deu_pri = 'C'
#     calificacion = mapeo_categoria.get(categoria.cat_arr, '05') if tercero else ''.ljust(2)
#     sit_est_tit = '05' # xSitEstTit+ ;		&& Situacion o Estado del Titular (Tabla 24 Cifin)
#     estado = '01' # xEstado+ ;		&& Estado (Tabla 10 Cifin) xEstado = IIF(S24SALCAP<=1,"07","01") && 07=SALDADO, 01=VIGENTE
#     edad_mora = xEdadMora #		carte_cat_his__dias_mor     && Edad de Mora (Tabla 8 Cifin)
#     anos_mora = xAnosMora # xAnosMora+ ;		&& Anos En Mora
#     fec_act = fecha_corte.strftime('%Y%m%d') # xFecAct+ ;		fec_cor                     && Fecha de Corte  (AAAAMMDD)
#     fec_des = credito.fec_des.strftime('%Y%m%d')#		creditos__fec_des           && Fecha Inicial o de Expedicion (AAAAMMDD)
#     fec_ven = credito.fec_ven.strftime('%Y%m%d')# xFecVen+ ;creditos__fec_ven       && Fecha Terminacion (AAAAMMDD)                 
#     fec_exi = fec_ven # xFecExi+ ;		&& Fecha Exigibilidad (AAAAMMDD)
#     espacio8 = ''.ljust(8)# space(8)+ ;		&& Fecha Prescripcion (AAAAMMDD)
#     fec_ult_pag = credito.fec_ult_pag.strftime('%Y%m%d') # xFecPag+ ;		creditos__fec_ult_pag       && Fecha Pago (AAAAMMDD)
#     mod_ext = '  '# xModExt+ ;		&& Modo Extincion (Tabla 9 Cifin)
#     tip_pago = '01' # "01" + ;			&& Tipo Pago (Tabla 31 Ci"01" --> Volunt"02" -->No voluntario
#     per_pag = credito.per_ano # xPerPag + ;		creditos__per_ano           && Periocidad de Pago (Tabla 6 Cifin)
#     # xProbNoPag	+ ;	&& Probabilidad de No Pago (Valor entre 0 y 100)
#     # xCuoCan+ ;		&& Numero Cuotas Pagadas
#     tot_cuo = credito.num_cuo_ini # xTotCuo+ ;		creditos__num_cuo_ini       && Numero Cuotas Pactadas
#     # xCuoMor+ ;		&& Cuotas en Mora
#     cap_ini = '{:012}'.format(int(categoria.cap_ini / 1000)) # xValIni+ ;		carte_cat_his__cap_ini      && Valor o Cupo (En Miles)
#     # xValMor+ ;		&& Valor de Mora (En Miles)
#     val_saldo = '{:012}'.format(int(categoria.saldo_1 / 1000))# xValSaldo+ ;		carte_cat_his__sal_cap      && Valor Saldo (En Miles)
#     val_cuo = '{:012}'.format(int(credito.val_cuo_ini / 1000)) # xValCuo+ ;		creditos__val_cuo_ini       && Valor de la Cuota (En Miles)
#     espacio12 = ''.ljust(12) # space(12)+ ;		&& Valor de Cargo Fijo ### (No Aplica) ###
#     # "008"+ ;			carte_cat_his__cod_lin_cre  && Linea Credito (Tabla 3 Cifin)
#     espacio3 = ''.ljust(3)# space(3)+ ;		&& Clausula de Permanencia ### (No Aplica) ###
#     # "001"+ ;			&& Tipo de Contrato (Tabla 25 Cifin)
#     # "001"+ ;			&& Estado del Contrato (Tabla 26 Cifin)
#     # space(3)+ ;		&& Termino o Vigencia del Contrato ### (No Aplica) ###
#     # space(2)+ ;		&& Numero Meses Contrato ### (No Aplica) ###
#     naturaleza = mapeo_tip_ter.get(tercero.tip_ter) if tercero else ''.ljust(3)# xNaturaleza+ ;	&& Naturaleza Juridica (Tabla 19 Cifin)
#     mod_cre = '02' # "02"+ ;			&& Modalidad de Credito (Tabla 2 Cifin) -- consumo
#     # xTipMon+ ;		&& Tipo de Moneda (Tabla 5 Cifin)
#     # xTipGar+ ;		&& Tipo Garantia (Tabla 4 Cifin)
#     # "000000000000"+ ;	&& Valor Garantia (En Miles)
#     # "02"+ ;			&& Obligacion Reestructurada (Tabla 32 Cifin)
#     # space(2)+ ;		&& Naturaleza Reestructuracion (Tabla 11 Cifin)
#     # space(3)+ ;		&& Numero Reestructuraciones
#     # space(3)+ ;		&& Clase Tarjeta ### (No Aplica) ### 
#     # space(4)+ ;		&& Numero de Cheques Devueltos ### (No Aplica) ###		
#     # space(2)+ ;		&& Categoria Servicios ### (No Aplica) ###
#     # space(2)+ ;		&& Plazo ### (No Aplica) ### 
#     # space(6)+ ;		&& Dias Cartera ### (No Aplica) ###
#     # space(2)+ ;		&& Tipo Cuenta ### (No Aplica) ###
#     # space(12)+ ;		&& Cupo Sobregiro ### (No Aplica) ###
#     # space(3)+ ;		&& Dias Autorizados ### (No Aplica) ###
#     direccion = tercero.direccion # xDirRes + ; 		tercero__direccion          && Direccion Titular
#     tel_res = tercero.celular1 # xTelRes + ;		tercero__celular1           && Telefono Casa Titular
#     # "000001" + ;		tercero__localidades_codigo && Codigo Ciudad Casa
#     # xCiuRes + ;		tercero__localidades_nombre && Ciudad Casa Titular
#     # "050" + ;			tercero__localidades_cod_pos[:2]    && Codigo Departamento Casa
#     # "META" + ;		tercero__localidades_departamento   && Departamento Casa Titular
#     # space(60)+ ;		&& Nombre Empresa
#     # space(60)+ ;		&& Direccion Empresa
#     # space(20)+ ;		&& Telefono Empresa
#     # space(6)+ ;		&& Codigo Ciudad Empresa
#     # space(6)+ ;		&& Codigo Departamento Empresa
#     # space(20)+ ;		&& Departamento Empresa Titular
#     # space(8)+ ;		&& Fecha Inicio Excension GMF
#     # space(8)+ ;		&& Fecha Terminacion Escension GMF
#     # space(2)+ ;		&& Numero Renovacion CDT
#     # space(2)+ ;		&& Cuenta Ahorro Excenta GMF
#     # space(2)+ ;		&& Tipo Identificacion Originaria
#     # space(14)+ ;		&& Numero Identificacion Originaria
#     # space(3)+ ;		&& Tipo Entidad Originaria
#     # space(3)+ ;		&& Codigo Entidad Originaria
#     # space(2)+ ;		&& Tipo Fideicomiso
#     # space(3)+ ;		&& Numero Fideicomiso
#     # space(60)+ ;		&& Nombre Fideicomiso
#     # space(4)+ ;		&& Tipo Deuda Cartera
#     # space(4)+ ;		&& Tipo de Poliza
#     # space(6) +;		&& Codigo de Ramo

#     return f"2{tip_doc}{doc_ide}{pri_ape}{seg_ape}{nombre}{espacio2}{fec_lim_pag}{cod_cre}{cod_age}{deu_pri}{calificacion}{sit_est_tit}{estado}{edad_mora}{anos_mora}{fec_act}{fec_des}{fec_ven}{fec_exi}{espacio8}{fec_ult_pag}{mod_ext}{tip_pago}{fec_ult_pag}{per_pag}{tot_cuo}{cap_ini}{val_saldo}{val_cuo}{espacio12}{espacio3}{naturaleza}{mod_cre}{direccion}{tel_res}\r\n"


# def listado_cifin(fec_corte: date, tipo_reporte: str):
#     # Fecha focal según tipo de reporte
#     if tipo_reporte == 1:
#         f_focal = fec_corte
#     else:
#         f_focal = fec_corte + timedelta(days=20)

#     # Lista de líneas del archivo
#     registros = []

#     # Encabezado
#     registros.append(f"121050181          01{fec_corte.strftime('%Y%m%d')}\r\n")

#     # Créditos base
#     creditos = CREDITOS.objects.filter(
#         # ~Q(est_jur='P'),
#         # ~Q(com_des_id=''),
#         estado = 'A',
#         rep_cen_rie='S',
#         fec_des__lte=fec_corte
#     ).order_by('cod_cre')
#     print('estos creditos',creditos)
#     for credito in creditos:
#         cod_cre = credito.cod_cre
#         # Busca la categoría del crédito a la fecha
#         categoria = CARTE_CAT_HIS.objects.filter(
#             cod_cre=cod_cre,
#             fecha__lte=fec_corte
#         ).order_by('-fecha').first()

#         # Validación si continúa
#         if not categoria and not (
#             credito.fec_des <= fec_corte.replace(day=1) - timedelta(days=1)
#             and credito.fec_ult_pag.year == f_focal.year
#             and credito.fec_ult_pag.month == f_focal.month
#         ) and credito.est_jur != "K":
#             continue

#         # Lógica de cálculo del saldo, mora, cuotas, fechas, estado, etc.
#         # Aquí deberías llamar a tus funciones CalculoEdadCr y LiqCreAlDia adaptadas a Python
#         saldo = calcular_saldo_credito(credito, f_focal)
#         if saldo <= 0 and credito.est_jur != "K":
#             continue

#         # Información del titular (Socio y Tercero)
#         asociado = ASOCIADOS.objects.filter(id=credito.socio_id).first()
#         tercero = asociado.tercero if asociado else None
#         pagador = PAGADORES.objects.filter(id=asociado.id_pag_id).first() if asociado else None

#         # Información del codeudor
#         codeudores = GAR_NO_IDONEA.objects.filter(credito_id=credito.id)
#         codeudor_data = []
#         for codeudor in codeudores:
#             tercero_codeudor = TERCEROS.objects.filter(doc_ide=codeudor.doc_ide).first()
#             if tercero_codeudor:
#                 codeudor_data.append(obtener_linea_codeudor(credito, tercero_codeudor, fec_corte))

#         # Construcción de la línea del titular
#         # linea_titular = construir_linea_titular(credito, tercero, categoria, saldo, fec_corte, f_focal, asociado, pagador)
#         linea_titular = construir_linea_titular(credito, tercero, fec_corte, categoria)
#         registros.append(linea_titular)

#         # Agregar las líneas de codeudores
#         registros.extend(codeudor_data)

#     # Pie con cantidad de registros
#     registros.append(f"9{str(len(registros)).zfill(8)}\r\n")

#     # # Guardar el archivo
#     # with open(nom_archivo, 'w', encoding='utf-8') as f:
#     #     f.writelines(registros)
#     # print('estos registros cifin', registros)
#     return registros #nom_archivo

# # Funciones auxiliares básicas
# def calcular_saldo_credito(credito, f_focal):
#     # Adaptar la lógica de CalculoEdadCr y LiqCreAlDia
#     # Ejemplo simplificado:
#     return credito.cap_ini

# # def construir_linea_titular(credito, tercero, categoria, saldo, fec_cor, f_focal, socio, pagador):
# def construir_linea_titular(credito, tercero, fecha_corte, categoria):
#     # Aquí ensamblas la línea tipo 2, usando los campos que has traído.
#     # Por simplicidad solo mostramos una parte
#     # 🔹 Diccionarios de mapeo
#     mapeo_tip_doc = {'C': '01', 'N': '02', 'E': '03', 'T': '04', 'P': '05','R': '09'}
#     mapeo_tip_ter = {'N': '000', 'J': '002'} 
#     mapeo_categoria = {'A': '01', 'B': '02', 'C': '03', 'D': '04', 'E': '05'}
    
#     if categoria.dias_mor < 30:
#         xEdadMora = "00"
#         xValMor = "000000000000"
#     elif 30 <= categoria.dias_mor < 60:
#         # if credito.not_mor and credito.fec_not_mor < fec_cor:
#             # xEdadMora = "01"
#         # else:
#             xEdadMora = "00"
#             # xCuoMor = 0
#     elif 60 <= categoria.dias_mor < 90:
#         xEdadMora = "02"
#     elif 90 <= categoria.dias_mor < 120:
#         xEdadMora = "03"
#     elif 120 <= categoria.dias_mor < 150:
#         xEdadMora = "04"
#     elif 150 <= categoria.dias_mor < 180:
#         xEdadMora = "05"
#     elif 180 <= categoria.dias_mor < 210:
#         xEdadMora = "06"
#     elif 210 <= categoria.dias_mor < 240:
#         xEdadMora = "07"
#     elif 240 <= categoria.dias_mor < 270:
#         xEdadMora = "08"
#     elif 270 <= categoria.dias_mor < 300:
#         xEdadMora = "09"
#     elif 300 <= categoria.dias_mor < 330:
#         xEdadMora = "10"
#     elif 330 <= categoria.dias_mor < 360:
#         xEdadMora = "11"
#     elif 360 <= categoria.dias_mor < 540:
#         xEdadMora = "12"
#     elif 540 <= categoria.dias_mor < 730:
#         xEdadMora = "13"
#     elif categoria.dias_mor >= 730:
#         xEdadMora = "14"
        
#     xAnosMora = categoria.dias_mor / 360
        
#     tip_doc =  mapeo_tip_doc.get(tercero.cla_doc) if tercero else ''.ljust(2)
#     doc_ide = tercero.doc_ide.ljust(15) if tercero else ''.ljust(15)
#     pri_ape = tercero.pri_ape.ljust(15) if tercero else ''.ljust(15)
#     seg_ape = tercero.seg_ape.ljust(15) if tercero else ''.ljust(15)
#     pri_nom = tercero.pri_nom.strip() if tercero else ''
#     seg_nom = tercero.seg_nom.strip() if tercero else ''
#     nombre_completo = pri_nom+' '+seg_nom
#     nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
#     espacio2 = ''.ljust(2)
#     fec_ult_pag = credito.fec_ult_pag.strftime('%Y%m%d')
#     cod_cre = credito.cod_cre.ljust(20)
#     cod_age = '000001'
#     deu_pri = 'P'
#     calificacion = mapeo_categoria.get(categoria.cat_arr, '05') if tercero else ''.ljust(2)
#     sit_est_tit = '05' # xSitEstTit+ ;		&& Situacion o Estado del Titular (Tabla 24 Cifin)
#     estado = '01' # xEstado+ ;		&& Estado (Tabla 10 Cifin) xEstado = IIF(S24SALCAP<=1,"07","01") && 07=SALDADO, 01=VIGENTE
#     edad_mora = xEdadMora #		carte_cat_his__dias_mor     && Edad de Mora (Tabla 8 Cifin)
#     anos_mora = xAnosMora # xAnosMora+ ;		&& Anos En Mora
#     fec_act = fecha_corte.strftime('%Y%m%d') # xFecAct+ ;		fec_cor                     && Fecha de Corte  (AAAAMMDD)
#     fec_des = credito.fec_des.strftime('%Y%m%d')#		creditos__fec_des           && Fecha Inicial o de Expedicion (AAAAMMDD)
#     fec_ven = credito.fec_ven.strftime('%Y%m%d')# xFecVen+ ;		creditos__fec_ven           && Fecha Terminacion (AAAAMMDD)                 
#     # xFecExi+ ;		&& Fecha Exigibilidad (AAAAMMDD)
#     espacio8 = ''.ljust(8)# space(8)+ ;		&& Fecha Prescripcion (AAAAMMDD)
#     # xFecPag+ ;		creditos__fec_ult_pag       && Fecha Pago (AAAAMMDD)
#     # xModExt+ ;		&& Modo Extincion (Tabla 9 Cifin)
#     tip_pago = '01' # "01" + ;			&& Tipo Pago (Tabla 31 Ci"01" --> Volunt"02" -->No voluntario
#     per_pag = credito.per_ano # xPerPag + ;		creditos__per_ano           && Periocidad de Pago (Tabla 6 Cifin)
#     # xProbNoPag	+ ;	&& Probabilidad de No Pago (Valor entre 0 y 100)
#     # xCuoCan+ ;		&& Numero Cuotas Pagadas
#     tot_cuo = credito.num_cuo_ini # xTotCuo+ ;		creditos__num_cuo_ini       && Numero Cuotas Pactadas
#     # xCuoMor+ ;		&& Cuotas en Mora
#     cap_ini = '{:012}'.format(int(categoria.cap_ini / 1000)) # xValIni+ ;		carte_cat_his__cap_ini      && Valor o Cupo (En Miles)
#     # xValMor+ ;		&& Valor de Mora (En Miles)
#     val_saldo = '{:012}'.format(int(categoria.saldo_1 / 1000))# xValSaldo+ ;		carte_cat_his__sal_cap      && Valor Saldo (En Miles)
#     val_cuo = '{:012}'.format(int(credito.val_cuo_ini / 1000)) # xValCuo+ ;		creditos__val_cuo_ini       && Valor de la Cuota (En Miles)
#     espacio12 = ''.ljust(12) # space(12)+ ;		&& Valor de Cargo Fijo ### (No Aplica) ###
#     # "008"+ ;			carte_cat_his__cod_lin_cre  && Linea Credito (Tabla 3 Cifin)
#     espacio3 = ''.ljust(3)# space(3)+ ;		&& Clausula de Permanencia ### (No Aplica) ###
#     # "001"+ ;			&& Tipo de Contrato (Tabla 25 Cifin)
#     # "001"+ ;			&& Estado del Contrato (Tabla 26 Cifin)
#     # space(3)+ ;		&& Termino o Vigencia del Contrato ### (No Aplica) ###
#     # space(2)+ ;		&& Numero Meses Contrato ### (No Aplica) ###
#     naturaleza = mapeo_tip_ter.get(tercero.tip_ter) if tercero else ''.ljust(3)# xNaturaleza+ ;	&& Naturaleza Juridica (Tabla 19 Cifin)
#     mod_cre = '02' # "02"+ ;			&& Modalidad de Credito (Tabla 2 Cifin) -- consumo
#     # xTipMon+ ;		&& Tipo de Moneda (Tabla 5 Cifin)
#     # xTipGar+ ;		&& Tipo Garantia (Tabla 4 Cifin)
#     # "000000000000"+ ;	&& Valor Garantia (En Miles)
#     # "02"+ ;			&& Obligacion Reestructurada (Tabla 32 Cifin)
#     # space(2)+ ;		&& Naturaleza Reestructuracion (Tabla 11 Cifin)
#     # space(3)+ ;		&& Numero Reestructuraciones
#     # space(3)+ ;		&& Clase Tarjeta ### (No Aplica) ### 
#     # space(4)+ ;		&& Numero de Cheques Devueltos ### (No Aplica) ###		
#     # space(2)+ ;		&& Categoria Servicios ### (No Aplica) ###
#     # space(2)+ ;		&& Plazo ### (No Aplica) ### 
#     # space(6)+ ;		&& Dias Cartera ### (No Aplica) ###
#     # space(2)+ ;		&& Tipo Cuenta ### (No Aplica) ###
#     # space(12)+ ;		&& Cupo Sobregiro ### (No Aplica) ###
#     # space(3)+ ;		&& Dias Autorizados ### (No Aplica) ###
#     direccion = tercero.direccion # xDirRes + ; 		tercero__direccion          && Direccion Titular
#     tel_res = tercero.celular1 # xTelRes + ;		tercero__celular1           && Telefono Casa Titular
#     # "000001" + ;		tercero__localidades_codigo && Codigo Ciudad Casa
#     # xCiuRes + ;		tercero__localidades_nombre && Ciudad Casa Titular
#     # "050" + ;			tercero__localidades_cod_pos[:2]    && Codigo Departamento Casa
#     # "META" + ;		tercero__localidades_departamento   && Departamento Casa Titular
#     # space(60)+ ;		&& Nombre Empresa
#     # space(60)+ ;		&& Direccion Empresa
#     # space(20)+ ;		&& Telefono Empresa
#     # space(6)+ ;		&& Codigo Ciudad Empresa
#     # space(6)+ ;		&& Codigo Departamento Empresa
#     # space(20)+ ;		&& Departamento Empresa Titular
#     # space(8)+ ;		&& Fecha Inicio Excension GMF
#     # space(8)+ ;		&& Fecha Terminacion Escension GMF
#     # space(2)+ ;		&& Numero Renovacion CDT
#     # space(2)+ ;		&& Cuenta Ahorro Excenta GMF
#     # space(2)+ ;		&& Tipo Identificacion Originaria
#     # space(14)+ ;		&& Numero Identificacion Originaria
#     # space(3)+ ;		&& Tipo Entidad Originaria
#     # space(3)+ ;		&& Codigo Entidad Originaria
#     # space(2)+ ;		&& Tipo Fideicomiso
#     # space(3)+ ;		&& Numero Fideicomiso
#     # space(60)+ ;		&& Nombre Fideicomiso
#     # space(4)+ ;		&& Tipo Deuda Cartera
#     # space(4)+ ;		&& Tipo de Poliza
#     # space(6) +;		&& Codigo de Ramo
    

#     # Debes agregar todos los campos en el orden y tamaño exacto como en tu FoxPro
#     return f"2{tip_doc}{doc_ide}{pri_ape}{seg_ape}{nombre}{espacio2}{fec_ult_pag}{cod_cre}{cod_age}{deu_pri}{calificacion}{sit_est_tit}{estado}{edad_mora}{anos_mora}{fec_act}{fec_des}{fec_ven}{espacio8}{tip_pago}{per_pag}{tot_cuo}{cap_ini}{val_saldo}{val_cuo}{espacio12}{espacio3}{naturaleza}{mod_cre}{direccion}{tel_res}\r\n"

# # def construir_linea_codeuddor(credito, tercero, categoria, saldo, fec_cor, f_focal, socio, pagador):
# def obtener_linea_codeudor(credito, tercero, fec_corte):
#     # Similar a construir_linea_titular pero con los datos del codeudor
#     nit = tercero.doc_ide.ljust(15)
#     nombre = tercero.nombre.ljust(30)
#     cod_cre = credito.cod_cre.ljust(20)

#     return f"2{nit}{nombre}{cod_cre}...resto de campos de codeudor...\r\n"

def datacredito():
    # REPLACE registro WITH dTipIde+;
	# 			dNumIde+;
	# 			dNumCta+;
	# 			dNombre+;
	# 			dSitTit+;
	# 			dFecApe+;
	# 			dFecVen+;
	# 			dCalDeu+;
	# 			dTipObl+;
	# 			dSubHip+;
	# 			dFecSubHip+;
	# 			dTerCon+;
	# 			dForPag+;
	# 			dPerPag+;		
	# 			dNovedad+;
	# 			dEstOriCta+;
	# 			dFecEstOrg+;
	# 			dEstCta+;
	# 			dFecEstCta+;
	# 			dEstPla+;
	# 			dFecEstPla+;
	# 			dAdjetivo+;
	# 			dFecAdj+;
	# 			dClaTar+;
	# 			dFranquicia+;
	# 			dNomMarPri+;
	# 			dTipMon+;
	# 			dTipGar+;
	# 			dCalificacion+;
	# 			dProInc+;
	# 			dEdaMor+;
	# 			dValIni+;
	# 			dSalDeu+;
	# 			dValDis+;
	# 			dValCuoMen+;
	# 			dValSalMor+;
    pass
        

def cartera_fecha(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')

    if request.method == 'GET':
        return render(request, 'cartera_a_una_fecha.html') 

    if request.method == 'POST':
        accion = request.POST.get("accion")   
        fecha_corte = request.POST.get('fecha_corte')
        min_dias = int(request.POST.get('min_dias', 0))
        max_dias = int(request.POST.get('max_dias', 0))

        saldos = report_cartera_a_una_fecha(id_cli, id_ofi, fecha_corte, min_dias, max_dias)

        if not saldos:
            return HttpResponse("No se encontraron datos para exportar", status=404)

        if accion == "exportar":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CARTERA A UNA FECHA" 
            
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            fecha_corte_formateada = formato_fecha(fecha_corte)
                                 
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = f"CARTERA A UNA FECHA {fecha_corte_formateada.upper()}"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
            dias_mora = f"Días de Mora  Mínimo: {min_dias}   Máximo: {max_dias}"
                                
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [dias_mora]
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1
           
            # Escribir encabezados de la tabla en la fila 5
            headers = saldos[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Escribir los datos desde la fila 6
            for row_num, data in enumerate(saldos, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])
            nombre_archivo = f"cartera_a_la_fecha_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            print('Reporte Grabado ..',nombre_archivo)
            return response

        elif accion == "csv":
            nombre_archivo = f"cartera_a_la_fecha_{fecha_corte}.csv"
            response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            writer = csv.writer(response)

            # Encabezado adicional

            headers = list(saldos[0].keys())
            writer.writerow(headers)

            for fila in saldos:
                writer.writerow([fila[col] for col in headers])

            return response

        else:
            return HttpResponse("Acción no permitida", status=405)

    return render(request, 'cartera_a_una_fecha.html')
            
def cartera_super(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    if request.method == 'GET':
        return render(request, 'cartera_supersolidaria.html') 
    if request.method == 'POST':
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        accion = request.POST.get("accion")   
      
        fecha_corte = request.POST.get('fecha_corte')
        
        saldos = reporte_cartera(id_cli, id_ofi, fecha_corte)

        if accion == "exportar":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
        
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "CARTERA SUPERSOLIDARIA" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos        
            saldos = reporte_cartera(id_cli, id_ofi, fecha_corte)                            
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            print('fecha_corte',fecha_corte)          
            
            fecha_corte_formateada = formato_fecha(fecha_corte)
            print('fec_cor_formato', fecha_corte_formateada)                        
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = f"CARTERA SUPERSOLIDARIA A LA FECHA {fecha_corte_formateada}"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
                                
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = saldos[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(saldos, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])

            nombre_archivo = f"cartera_supersolidaria_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
           
        elif accion == "csv":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
                            
            # Llama a la función para obtener los datos
            saldos = reporte_cartera(id_cli, id_ofi, fecha_corte)
           
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"cartera_supersolidaria_{fecha_corte}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                        
            # Crear el escritor CSV
            writer = csv.writer(response)
                        
            # Añadir encabezados de las columnas
            headers = saldos[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in saldos:
                writer.writerow([fila[col] for col in headers])
            return response
        else:
            return HttpResponse("Acción no permitida", status=405)
                
    return render(request, 'cartera_supersolidaria.html')

def report_cartera_a_una_fecha(cliente_id, id_oficina, fecha_corte_str,min_dias,max_dias):
    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    tres_anios_atras = fecha_corte - relativedelta(years=3)
    print('fecha Corte y tres años ',fecha_corte,tres_anios_atras,id_oficina)
    resultados = []
    Creditos = CREDITOS.objects.filter(oficina_id = id_oficina, fec_des__lte = fecha_corte 
            ).exclude(
                # Excluir créditos con estado 'C' y fec_ult_pag < self.fecha_ant_rec
                Q(estado='C') & Q(fec_ult_pag__lt = fecha_corte)
            ).exclude(
                # Excluir créditos con estado 'H'
                estado='H'
            ).exclude(
                # Excluir créditos con estado 'H'
                fec_des__gt = fecha_corte
            )
    for credito in Creditos:
        his_num_cre = (
            CREDITOS.objects.filter(
                oficina_id=1,
                fec_des__gte=tres_anios_atras,
                socio__tercero__doc_ide = credito.socio.tercero.doc_ide
            )
            .exclude(estado= 'X')
            .values('socio__tercero__doc_ide')
            .annotate(total=Count('id'))
        )
        if his_num_cre:
            xtot_cre_ter = his_num_cre[0]['total']
        else:
            xtot_cre_ter = 0
        
        base_his = (
            CARTE_CAT_HIS.objects.filter(
                nit = credito.socio.tercero.doc_ide,
                fecha__gte = tres_anios_atras,  # ← filtro por fecha
                nit__in=TERCEROS.objects.filter(cliente_id=1)
                    .values_list('doc_ide', flat=True)
                )
                .values('nit')
                .annotate(
                total=Count('id'),
                buenas=Count('id', filter=Q(cat_arr='A'))
            )
        )

        if  base_his:
            xPorCum = base_his[0]['buenas'] / base_his[0]['total'] * 100 if base_his[0]['total'] > 0 else 100
        else:
            xPorCum = 0

        diferencia = relativedelta(fecha_corte, credito.socio.fec_afi)
        xAntMes = (diferencia.years * 12 + diferencia.months) if credito.socio.estado == 'A' or credito.socio.estado == 'R' else 0

        liq_cre = Liquida_cre(credito.cod_cre,fecha_corte)
        if liq_cre.lista_mov == None:
            credito.estado = 'H'
            credito.save()
            continue 
        liq_cre.liq_al_dia(fecha_corte)
        if liq_cre.sal_cap_tot <= 0 :
            continue 
        liq_cre.calculo_periodo()
        xdias_mor = (liq_cre.fecha_focal-liq_cre.fec_al_dia).days
        xdias_mor = xdias_mor if xdias_mor > 0 else 0
        if xdias_mor < min_dias or xdias_mor > max_dias:
            continue
        micarhis = CARTE_CAT_HIS.objects.filter(oficina_id = id_oficina,fecha = fecha_corte,cod_cre = credito.cod_cre).first()
        xSalMin = 1440000
        xAprobo = ' ' * 10
        xTipGar = ' ' * 10
        if micarhis == None:
            if xdias_mor < 1:
                xCat = 'A'
            else:
                # print(' ord(Credito.cod_des)',Credito.cod_des)
                CatDesDia = CAT_DES_DIA_CRE.objects.filter(cliente_id = cliente_id,codigo = ord(credito.cod_des),
                    minimo_dias__lte=xdias_mor,maximo_dias__gte=xdias_mor).first()
                if CatDesDia == None:
                    xCat = 'F'
                else:
                    xCat = CatDesDia.categoria
            xcat_mor = xCat
            xcat_arr = 'X'
            xcat_eva = 'X'
            xcat_ree = 'X'
            xcat_mod = 'X'
            xcategoria = 'X'
            xProIndKap = 0
        else:
            xcat_mor = micarhis.cat_mor
            xcat_arr = micarhis.cat_arr
            xcat_eva = micarhis.cat_eva
            xcat_ree = micarhis.cat_ree
            xcat_mod = micarhis.cat_mod
            xcategoria = micarhis.categoria
            xProIndKap = micarhis.pro_ind_kap
            xAprobo = ' ' * 10
            if credito.cod_lin_cre.cod_lin_cre == 49:
                xAprobo = 'CONSEJO.ADM'
            elif credito.cod_lin_cre.cod_lin_cre == 50:
                xAprobo = 'GERENCIA'
            elif credito.cod_lin_cre.cod_lin_cre == 51:
                xAprobo = 'GERENCIA'
            elif credito.cod_lin_cre.cod_lin_cre == 52:
                xAprobo = 'COM.CREDITO'
            elif credito.cod_lin_cre.cod_lin_cre == 53:
                xAprobo = 'CONSEJO.ADM'
            elif credito.cod_lin_cre.cod_lin_cre == 54:
                if credito.cap_ini <= xSalMin * 40 :
                    xAprobo = 'GERENCIA'
                elif credito.cap_ini <= xSalMin * 93 :
                    xAprobo = 'COM. CREDITO'
                else:
                    xAprobo = 'CONSEJO.ADM'
        xTipGar = 'Hipotecario' if credito.tip_gar == '2' else ('No Idonea' if credito.tip_gar == '15' else 'Admisible' )
        est_fin = ESTADOS_FIN.objects.filter(cliente_id = cliente_id,tercero = credito.socio.tercero).first()
        codeu = GAR_NO_IDONEA.objects.filter(oficina_id = id_oficina,credito = credito).first()
        if codeu != None:
            cod_ter = TERCEROS.objects.filter(cliente_id = cliente_id,doc_ide = codeu.doc_ide).exclude(doc_ide=credito.socio.tercero.doc_ide).first()
        else:
            cod_ter = None
        resultados.append({
            'cod_soc' : credito.socio.cod_aso,
            'fec_nac' : credito.socio.fec_nac, 
            'edad' : (fecha_corte.year - credito.socio.fec_nac.year),
            'Nit' : credito.socio.tercero.doc_ide,
            'est_soc' : credito.socio.estado,
            'nombre': credito.socio.tercero.nombre,
            'cod_cre': credito.cod_cre, 
            'codlincre': credito.cod_lin_cre.descripcion,
            'fec_des' : credito.fec_des.strftime('%d/%m/%Y'),
            'fec_ven' : credito.fec_ven,
            'tasint' : credito.tiae_ic_act,
            'numcuo' : credito.num_cuo_act,
            'altura' : liq_cre.altura,
            'CapIni' : credito.cap_ini,
            'ValCuo' : credito.val_cuo_act,
            'CuoPag' : liq_cre.cuo_pag,
            'DiaMor' : xdias_mor,
            'saldo_K' : liq_cre.sal_cap_tot,
            'TasIntIniEfe' : credito.tiae_ic_ini,
            'TasIntActEfe' : credito.tiae_ic_act,
            'TasIntActNom' : credito.tian_ic_act,
            'Provision' : xProIndKap,            
            'cat_mor' : xcat_mor,
            'cat_arr' : xcat_arr,
            'cat_eva' : xcat_eva,
            'cat_ree' : xcat_ree,
            'cat_mod' : xcat_mod,
            'categoria' : xcategoria,
            'FecCal' : fecha_corte,
            'FecCalAnt' : (fecha_corte - relativedelta(months=1)) if credito.cat_nue != ' ' else  None,
            'TipCre' : credito.imputacion.descripcion,
            'Aprobo' : xAprobo,
            'TipGar' : xTipGar,
            'TelRes' : credito.socio.tercero.celular1,           
            'TelOfi' :  credito.socio.tel_tra,
            'Celular' : credito.socio.tercero.celular2,
            'email' : credito.socio.tercero.email,
            'nomdescre' : credito.imputacion.descripcion,
            'PorCum' : xPorCum,
            'AntMes' : xAntMes,
            'Creditos' : xtot_cre_ter,
            'ValGart' : credito.val_gar_hip,
            'SalMora' : liq_cre.sal_cap_dia+liq_cre.sal_int_dia+liq_cre.sal_int_mor,
            'ConAli' : 'N',
            'FecAli' : ' '*10,
            'numcuoali' : 0,
            'sec_eco' : credito.socio.sector_emp,
            'ocupacion' : credito.socio.ocupacion,
            'CiuRes' : credito.socio.ciu_tra.nombre,
            'Genero' : credito.socio.sexo,
            'nivel_est' : credito.socio.niv_est,
            'estrato' : credito.socio.estrato,
            'ing_mensual ' : est_fin.ing_tot if est_fin != None else 0,
            'Activos' : est_fin.tot_act if est_fin != None else 0,
            'Pasivos' : est_fin.tot_pas if est_fin != None else 0,
            'num_hijos' : credito.socio.num_hij_may + credito.socio.num_hij_men,
            'tipo_viv' : credito.socio.tip_viv,
            'AntTra' : (fecha_corte.year - credito.socio.fec_ing_tra.year),
            'TipCon' : credito.socio.tip_con,
            'CreRes' : ' ',
            'ppe' : 'NO',
            'Patrimonio' : est_fin.tot_act - est_fin.tot_pas if est_fin != None else 0,
            'egresos' : est_fin.egr_tot if est_fin != None else 0,
            'EstCiv' : credito.socio.est_civ,
            'NitCod1' : codeu.doc_ide if codeu != None else '',
            'NomCod!' : cod_ter.nombre if cod_ter != None else '',
            'NitCod2' : '',
            'NomCod2' : '',
            'pagare' : credito.pagare,                 
            'tipo_salario' : credito.socio.tip_sal,
        })
    print('riesgos geinner', resultados)
    return resultados
    
def reporte_cartera(cliente_id, id_oficina, fecha_corte_str):
    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    
    creditos = CARTE_CAT_HIS.objects.filter(
        oficina_id=id_oficina,
        fecha=fecha_corte
        ).select_related(
        'credito', 'credito__socio', 'credito__socio__tercero', 'credito__socio__id_pag', 'credito__cod_lin_cre'
    )
    resultados = []   
    for idx, cartera in enumerate(creditos, start=1):
        asociado = cartera.credito.socio
        tercero = asociado.tercero
        nit_patron = asociado.id_pag.nit
        nom_patron = asociado.id_pag.nombre
        if tercero.cla_doc == "N":
            cod_aso_formateado = formatear_cod_aso(asociado.cod_aso) #tercero.doc_ide)
            if tercero.dig_ver:  # Agregar el DV si existe
                cod_aso_formateado += f"-{tercero.dig_ver}"
        else:
            cod_aso_formateado = asociado.cod_aso
            
        # Buscar IMP_CON_CRE_INT por línea crédito y categoria
        imp_con = IMP_CON_CRE_INT.objects.filter(
            cod_imp=cartera.cod_imp_con,
            categoria=cartera.categoria
        ).first()
        
        if imp_con.kcta_con:
            cuenta_contable = imp_con.kcta_con
        else:
            cuenta_contable = 'FALTA LA CUENTA CONTABLE'  # o lo que quieras por defecto
        
        # cuenta_contable = imp_con.Kcta_con
                         
        liq_cre = Liquida_cre(cartera.cod_cre, fecha_corte)
        cuo_pag = liq_cre.cuo_pag
        
        fec_des = cartera.credito.fec_des
        cuotas = cartera.plazo
        fec_ven = fec_des + relativedelta(months=cuotas)
        
        est_jur = 0
        if cartera.credito.est_jur == 'P':
            est_jur = 1
        elif cartera.credito.est_jur == 'J':
            est_jur = 2
        if cartera.credito.termino == 'D':
            cupo_rotativo = 1
        elif cartera.credito.termino == 'R':
            cupo_rotativo = 2
        elif cartera.credito.termino == 'C':
            cupo_rotativo = 3
        fec_ult_pag = fecha_ultimo_movimiento(fecha_corte, cartera.cod_cre)   
        val_cap_mora = cartera.sal_cap_dia if cartera.dias_mor > 0 else 0
        resultados.append({
            'tip_doc': "I" if tercero.cla_doc == "T" else (tercero.cla_doc if tercero.cla_doc else ""),
            'doc_ide': cod_aso_formateado, 
            'cod_con': cuenta_contable[:6],
            'modificacion': 4,  # 2=Reestructurado,  3=Novado, 4=Ninguna, 5=Alivio, 6=Otras modificaciones de las condiciones iniciales
            'cod_cre': cartera.cod_cre,
            'fec_des': fec_des.strftime('%d/%m/%Y'), 
            'fec_ven': fec_ven.strftime('%d/%m/%Y'),
            'dias_mora': cartera.dias_mor,
            'tipo_cuota': 1, #tip_cuo, # 1 = Fija  2 = Variable  3 = Otra (semifija o semivariable). En el caso de variable, la amortización a capital es fija.
            'cuotas_pagadas': cuo_pag,
            'periodo_pago': 30, # dias
            'modalidad': 2, # 1 = Anticipada   2 = Vencida
            'tas_int_efe': cartera.credito.tiae_ic_act,
            'valor_prestamo': cartera.cap_ini,
            'valor_cuota': cartera.credito.val_cuo_ini, # valor_cuota,
            'saldo_capital': cartera.sal_cap_pe,
            'saldo_interes': cartera.sal_cat_int, # Registre el saldo de los intereses causados y no pagados, de cada una de las obligaciones del deudor, que deben estar registrados en la respectiva subcuenta de la cuenta 140000 del Catálogo Único de Información Financiera con fines de supervisión.
            'saldo_poliza': 0, #sal_pol,
            'garantia_1': cartera.val_gar_hip,# val_gar,
            'fec_ult_aval': "", # fec_ult_aval,
            'det_cap': cartera.pro_ind_kap, #Registre el valor del deterioro individual por concepto de capital de acuerdo con las normas vigentes.
            'det_int': cartera.pro_ind_int, # Registre el valor del deterioro causado por concepto de intereses de la cartera de créditos, a la fecha de corte., por cada una de las categorías relacionadas.
            'con_int': cartera.sal_int_contin, #Contingencia Intereses: Registre el saldo de los intereses no pagados, que se registran en deudoras contingentes por suspensión de la causación de acuerdo con las normas vigentes.
            'val_cuo_ext': 0,  # val_cuo_ext, # Registre el valor de los abonos extraordinarios que se hacen al capital de la obligación. En el caso de que existan varios abonos extraordinarios para un solo crédito, separar su valor con "/".
            'mes_cuo_ext': 0,  #mes_cuo_ext, # 34.Meses cuotas extraordinarias: Registre el número al que corresponde el mes en que se hace el abono extraordinario a capital de la obligación. En el caso de que existan varios abonos extraordinarios para un solo crédito, separe los meses en que se efectúa el abono con "/", teniendo en cuenta que los meses se identifican con la siguiente numeración:
            # 1 = Enero
            # 2 = Febrero
            # 3 = Marzo
            # 4 = Abril
            # 5 = Mayo
            # 6 = Junio
            # 7 = Julio 
            # 8 = Agosto
            # 9 = Septiembre
            # 10 = Octubre
            # 11 = Noviembre
            # 12 = Diciembre
            'fec_ult_pag': fec_ult_pag,
            'clase_garantia': cartera.cla_gar, # Clase de garantía 1: Registre la clase de garantía otorgada que respalda el saldo de la obligación, según las siguientes convenciones: 
            # 1 = Garantía no idónea
            # 2 = Hipotecaria
            # 3 = Prendaria (Vehículos y maquinaria)
            # 6 = Contratos de Fiducia
            # 8 = Pignoración de rentas
            # 9 = Otras garantías idóneas (*)
            # 10 = Depósitos de dinero en garantía
            # 11 = Garantía soberana de la nación
            # 12 = Garantías emitidas por fondos de garantías que administren recursos
            # públicos (FNG, FAG)
            # 13 = Derechos de cobro
            # 14 = Fiducia sobre inmuebles
            # 15 = Sin garantía
            'des_cre': 'S', #cartera.credito. des_cre,
            'cod_ofi': 1, # cod_ofi,
            'amor_cap': 30, #Periodicidad amortización de capital (Días): Indique en días, cada cuanto se paga la cuota por concepto de capital.
            'val_cap_mora': val_cap_mora, # val_cap_mora, #Valor del capital de cuotas en mora: Registre la suma total de capital adeudado de todas las cuotas morosas. Nota: Es solo el valor del capital correspondiente a cada cuota.
            'cla_viv': "", # cla_viv, # Clase de vivienda: Aplica solo a créditos de vivienda o microcrédito inmobiliario, y corresponde a la clase de vivienda a financiar, según las siguientes categorías:
            # 1 = Nueva.
            # 2 = Usada.
            # 3 = Mejoramiento.
            # 4 = Lote con servicios.
            # 5 = Construcción en sitio propio. (Este campo solo se diligencia si son créditos de vivienda de lo contrario se debe dejar en blanco).
            'viv_vis': "", # viv_vis, #43.Señal VIS: Aplica solo a créditos de vivienda o microcrédito inmobiliario, e indica si es vivienda de interés social o no. Valores: 1-VIS, 0-No VIS. Si es un microcrédito inmobiliario, su valor es 1-VIS.
            'rango_viv': "", # rango_viv, #44.Tipo o Rango de vivienda: Aplica solo a créditos de vivienda o microcrédito inmobiliario, para clases de vivienda nueva o usada. Si el crédito es VIS (Vivienda de Interés Social) los tipos son los siguientes:
            # 1 = Tipo 1: Cuyo valor de la vivienda sea menor o igual a 50 SMML
            # 2 = Tipo 2: Cuyo valor de la vivienda sea mayor a 50 SMML y menor o igual a 70 SMML
            # 3 = Tipo 3: Cuyo valor de la vivienda sea mayor a 70 SMML y menor o igual a 100 SMML
            # 4 = Tipo 4: Cuyo valor de la vivienda sea mayor a 100 SMML y menor o igual a 135 SMML
            # Para Vivienda No VIS:
            # 1 - Rango 1: Cuyo monto sea mayor a VIS y menor o igual a 643.100 UVR
            # 2 - Rango 2: Cuyo monto sea mayor a 643.100 UVR y menor o igual a 2’411.625 UVR
            # 3 - Rango 3: Cuyo valor sea mayor a 2’411.625 UVR. (Este campo solo se diligencia si son créditos de vivienda de lo contrario se debe dejarlo en blanco).
            'ent_redes': "", # ent_redes, # 46.Entidad de Redescuento (VIS): Si es un créditos VIS, indicar la entidad de redescuento:
            # 0 = Ninguna
            # 1 = FINDETER
            # Nota: Ninguna indica sin redescuento. Este campo solo se diligencia si son créditos de vivienda de lo contrario se debe dejarlo en blanco.
            'margen_redes': "", # margen_redes, # 47.Margen de Redescuento (VIS): Si la entidad de redescuento es FINDETER, se debe registrar un valor entre el 50% y el 90%, correspondiente al monto prestado por FINDETER; de lo contrario 0. Este campo solo se diligencia si son créditos de vivienda de lo contrario se debe dejar en blanco.
            'subsidio': "", # subsidio, # 45.Señal de subsidio:
            # 1 = SI
            # 0 = No
            # Aplica solo a créditos de vivienda o microcrédito inmobiliario, para clases de vivienda nueva, usada, mejoramiento y construcción en sitio propio. Este campo solo se diligencia si son créditos de vivienda de lo contrario se debe dejarlo en blanco.
            'sujeto_desem': 1, # sujeto_desem, # 48.Sujeto de desembolso:
            # 1 = Desembolso directo
            # 2 = Desembolso a constructor
            # 3 = Subrogación.
            # Nota:
            # La señal 3 = Subrogación aplica a créditos de vivienda o microcrédito.
            # La señal 2 = Solo aplica si el tipo de crédito es comercial.
            'moneda': 1, # moneda, # 49.Moneda del crédito: Registre el código de moneda en que está el crédito:
            # 1 = Peso
            # 2 = UVR
            # Nota: Aunque un crédito esté en UVR, todos sus valores (desembolso, capital, intereses, etc.), se reportan en pesos.
            'total_aportes': cartera.aporte, # Registre el valor de los aportes sociales, según la distribución a prorrata, teniendo en cuenta el total de las obligaciones a la fecha de corte del reporte.
            'lin_cre': cartera.cod_lin_cre, # lin_cre, # Línea de Crédito de la entidad: Registre la línea por la cual se desembolsó el crédito. Para el caso de la línea de crédito productivo, de acuerdo con el valor digitado en el campo Valor Préstamo se desplegará el listado de línea de crédito productivo a la cual aplica según las condiciones determinadas en el Decreto 455 de 2023 de acuerdo con las siguientes:
            # 1. Micro - Crédito popular productivo rural: Se define como crédito popular productivo rural el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas rurales y rurales dispersas cuyo monto no exceda de seis (6) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 2. Comercial - Crédito popular productivo rural: Se define como crédito popular productivo rural el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas rurales y rurales dispersas cuyo monto no exceda de seis (6) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 3. Micro - Crédito popular productivo urbano: Se define como crédito popular productivo urbano el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas urbanas cuyo monto no exceda de seis (6) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 4. Comercial - Crédito popular productivo urbano: Se define como crédito popular productivo urbano el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas urbanas cuyo monto no exceda de seis (6) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 5. Micro - Crédito productivo rural: Se define como crédito productivo rural el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas rurales y rurales dispersas cuyo monto sea mayor de seis (6) salarios mínimos legales mensuales vigentes y hasta veinticinco (25) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 6. Comercial - Crédito productivo rural: Se define como crédito productivo rural el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas rurales y rurales dispersas cuyo monto sea mayor de seis (6) salarios mínimos legales mensuales vigentes y hasta veinticinco (25) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 7. Micro - Crédito productivo urbano: Se define como crédito productivo urbano el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas urbanas cuyo monto sea mayor de seis (6) salarios mínimos legales mensuales vigentes y hasta veinticinco (25) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 8. Comercial - Crédito productivo urbano: Se define como crédito productivo urbano el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica en zonas urbanas cuyo monto sea mayor de seis (6) salarios mínimos legales mensuales vigentes y hasta veinticinco (25) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            # 9. Crédito productivo de mayor monto: Se define como crédito de mayor monto el otorgado a personas naturales o jurídicas para el desarrollo de cualquier actividad económica cuyo monto sea mayor a veinticinco (25) salarios mínimos legales mensuales vigentes y hasta ciento veinte (120) salarios mínimos legales mensuales vigentes, al momento de la aprobación de la respectiva operación activa de crédito.
            'num_mod': 0, # Número de MOdificaciones
            'est_cre': est_jur, # 0 = Vigente: El crédito está al día. 
            # 1 = En cobro Pre jurídico: Despliegue de la actividad profesional que efectúa el acreedor en procura de recaudar el valor de la obligación en mora sin que medie un proceso judicial, es decir, se trata de un cobro extraprocesal y persuasivo previo – www.superfinanciera.gov.co.  
            # 2 = En cobro jurídico: Ejercicio de la acción de cobro mediante la instauración de las acciones ejecutivas a través de un proceso judicial ante los jueces de la República - www.superfinanciera.gov.co. 
            # 3 = Deudor Insolvente: Situación jurídica en la que se encuentra una persona física o empresa cuando no puede hacer frente al pago de sus deudas.
            'nit_deu_pat': nit_patron, #Si la entidad diligencia información en los códigos contables 140400(140405, 140410, 140415, 140420 y 140425); 141100 (141105, 141110, 141115,141120, 141125); 144100 (144105, 144110, 144115, 144120, 144125), se habilitará la opción para la captura de información del NIT de la deudora patronal(NIT).
            'nom_deu_pat': nom_patron, #Si la entidad diligencia información en los códigos contables 140400(140405, 140410, 140415, 140420 y 140425); 141100 (141105, 141110, 141115, 141120, 141125); 144100 (144105, 144110, 144115, 144120, 144125), se habilitará la opción para la captura de la razón social de la deudora patronal (Nombre).
            'cupo_rotativo': cupo_rotativo, # 1=No, 2=Cupo Rotativo, 3=Tarjeta de Crédito
            'ent_garantia': 4, # 1=Bancoldex, 2=Finagro, 3=Findeter, 4=Fogacoop, 5=Establecimiento, 6=Otro, 7=Ninguno
            'tas_int_nom': round(tasa_nominal_anual(cartera.credito.tiae_ic_act, 12),3),
            'consecutivo': idx,
            'nombre': tercero.nombre,
            # fecres	
            # catres
            # modificado	
            # tipmod	
            # fecmodcre	
            # caliante	
            # mesesgra
            # activo
        })
    return resultados
    
class CatRecPerEsp(ListView):
    template_name = 'lista_comp_recla.html'
    context_object_name = 'filas'

    def get_queryset(self):
        oficina_id = self.request.session.get('oficina_id')
        oficina = OFICINAS.objects.filter(id = oficina_id).first()
        per_con = self.request.session.get('per_con')
        MESES_ES = {
            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
            5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
            9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
        }
        registros = DETALLE_ECONO.objects.filter(
            cuenta__per_con = per_con,hecho_econo__docto_conta__oficina_id = oficina_id,
            hecho_econo__docto_conta__per_con = per_con,
            hecho_econo__docto_conta__codigo = 34
        ).values(
            'hecho_econo__numero',
            'hecho_econo__fecha',
            'hecho_econo__protegido'
        ).annotate(
            cantidad = Count('id'),
            difer = Coalesce(Sum(F('debito') - F('credito'), output_field=FloatField()), 0.0),
            errores = Coalesce(Sum(
                Case(
                    When(cuenta__cod_cta='51109501', then=Value(1)),
                        default=Value(0),
                    output_field=IntegerField()
                )
            ), 0)
        )
        resultados = []
        xnum = 0
        for row in registros:
            cantidad_reg_pe = PE_CARTE_HIS.objects.filter(fecha=row['hecho_econo__fecha']) \
                .aggregate(cantidad=Coalesce(Count('id'), 0))['cantidad']
            xnum = row['hecho_econo__numero']
            resultados.append({
                'numero': row['hecho_econo__numero'],
                'fecha': row['hecho_econo__fecha'],
                'RegModelo' : cantidad_reg_pe,
                'generado' : '✔️',
                'protegido': '✔️' if row['hecho_econo__protegido'] == 'S' else '❌',
                'cantidad': row['cantidad'],
                'cuadrado': '✔️' if row['difer'] == 0 else '❌',
                'errores': row['errores'],
            })
        for mes in range(xnum + 1, 13):
            ultimo_dia = calendar.monthrange(per_con, mes)[1]
            fecha_final = datetime(per_con, mes, ultimo_dia).date()
            cantidad_reg_pe = PE_CARTE_HIS.objects.filter(fecha = fecha_final) \
                .aggregate(cantidad=Coalesce(Count('id'), 0))['cantidad']
            resultados.append({
                'numero': mes, 
                'fecha': fecha_final,
                'RegModelo' : cantidad_reg_pe,
                'generado' : '❌',
                'protegido': '❌',
                'cantidad': 0,
                'cuadrado': '✔️' if row['difer'] == 0 else '❌',
                'errores': 0,
            })
        return resultados
    
def liquidar_asiento(request, numero):
    per_con = request.session.get('per_con', datetime.now().year)
    ultimo_dia = calendar.monthrange(per_con, numero)[1]
    fecha_final = datetime(per_con, numero, ultimo_dia).date()
    previous_url = request.META.get('HTTP_REFERER', '/')
    contexto = {
        'numero': numero,
        'fecha_final': fecha_final,
        'previous_url': previous_url,
    }
    return render(request, 'liquidar_asiento.html', contexto)

def ejecutar_modelo(request):
    try:
        # Recibir fecha como cadena desde GET
        fecha_str = request.GET.get('fecha')
        if not fecha_str:
            raise ValueError("No se recibió la fecha.")
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        oficina = OFICINAS.objects.filter(codigo='A0001').first()
        if not oficina:
            raise ValueError("Oficina no encontrada.")
        start_time = time.time()  # Marcamos el tiempo de inicio
        print('Inicia calculo_z_punt', datetime.now())
        pe_mes = perdida_esperada(oficina, fecha)
        pe_mes.calculo_z_punt(fecha)

        print('Finaliza Modelo', datetime.now())
        elapsed_time = time.time() - start_time  # Tiempo transcurrido en segundos

        return JsonResponse({
            'status': 'success',
            'message': 'Modelo ejecutado correctamente.',
            'elapsed_time': round(elapsed_time,0)  # Tiempo en minutos
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error en la ejecución: {str(e)}',
            'elapsed_time': 0
        })

def ejecutar_rec_pe(request):
    try:
        # Recibir fecha como cadena desde GET
        fecha_str = request.GET.get('fecha')
        if not fecha_str:
            raise ValueError("No se recibió la fecha.")
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        oficina = OFICINAS.objects.filter(codigo='A0001').first()
        if not oficina:
            raise ValueError("Oficina no encontrada.")
        start_time = time.time()  # Marcamos el tiempo de inicio
        pe_mes = perdida_esperada(oficina, fecha)
        rec_mes = Reclasificacion(oficina,fecha)
        print('Inicia calculos base      ',datetime.now())
        rec_mes.calculos_base()
        print('Inicia asignar valores pe ',datetime.now())
        pe_mes.asignar_valores_pe(fecha)
        print('Inicia  .ReclaProv_capital',datetime.now())
        rec_mes.ReclaProv_capital(fecha)
        print('Inicia  ReclaIntCor       ',datetime.now())
        rec_mes.ReclaIntCor(oficina)
        print('Inicia calculos PE        ',datetime.now())
        rec_mes.calculos_pe(oficina)
        print('Inicia Contabilizar       ',datetime.now())
        rec_mes.AsiConRPE.netear_valores()
        docto = DOCTO_CONTA.objects.filter(oficina = oficina,per_con = fecha.year,codigo = 34).first()
        numero = fecha.month
        rec_mes.AsiConRPE.contabilizar(docto,numero,fecha)
        rec_mes.AsiConRPE.guardar_en_excel('rkipe_01.xlsx','c:/aaa/')
        print('Final Recla_kap_int_PE    ',datetime.now())
        elapsed_time = time.time() - start_time  # Tiempo transcurrido en segundos
        return JsonResponse({
            'status': 'success',
            'message': 'Modelo ejecutado correctamente.',
            'elapsed_time': round(elapsed_time,0)  # Tiempo en minutos
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error en la ejecución: {str(e)}',
            'elapsed_time': 0
        })

def iniciar_modelo(request):
    fecha_str = request.GET.get('fecha')
    print(f"Fecha recibida: {fecha_str}") 
    if not fecha_str:
        return JsonResponse({'error': 'No se recibió fecha'}, status=400)

    # Enviar tarea a Celery
    task = ejecutar_modelo_task.delay(fecha_str)  # 👈 ESTA LÍNEA ES CLAVE

    return JsonResponse({'task_id': task.id})

# Consulta el progreso
def progreso_modelo(request, task_id):
    r = redis.StrictRedis(host='localhost', port=6379, db=0)
    progreso = r.get(f'progreso:{task_id}')

    if progreso is None:
        porcentaje = 0
    else:
        porcentaje = int(progreso)

    return JsonResponse({'progreso': porcentaje})

def ejecutar_tarea(request):
    tarea_lenta.delay()
    return JsonResponse({'status': 'ok', 'message': 'Tarea lanzada en segundo plano'})

def ver_resumen_pe(request,fecha):

    def formatear_numero(valor, ancho=12):
        texto = f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return texto.rjust(ancho)

    def encabezado(pdf):
        print('Entra a encabezado ')
        pdf.drawString(100, height - 40, "COOPERATIVA ESPECIALIZADA DE AHORRO Y CREDITO DE LA ORINOQUIA")
        pdf.drawString(220, height - 56, "C O O R I N O Q U I A")
        pdf.drawString(80, height - 72, "ASIENTO CONTABLE DE RECLASIFICACION Y PERDIDA ESPERADA DE CARTERA")
        pdf.drawString(120, height - 84, "    REPORTE RESUMEN DE ASIENTO CONTABLE DEL  "+fecha.strftime('%Y-%m-%d') )
        titulos = ["Asiento","Cuenta Nivel 1","Debito","Credito","Total"]
        for col_num, col_name in enumerate(titulos):
            pdf.drawString(x + col_num * col_width,height - 110 , col_name)
            pdf.drawString(x + col_num * col_width,height - 120, '============')

    oficina_id = request.session.get('oficina_id', datetime.now().year)
    print('Fecha ',fecha,'  type of ',type(fecha))
    if fecha:
        try:
            fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            fecha = None
    else:
        fecha = None

    resultado = DETALLE_ECONO.objects.filter(
        hecho_econo__numero = fecha.month,hecho_econo__docto_conta__oficina_id = oficina_id,
        hecho_econo__docto_conta__per_con = fecha.year,hecho_econo__docto_conta__codigo = 34,cuenta__per_con = fecha.year,
        ).annotate(
            cta=Substr('cuenta__cod_cta', 1, 1)
        ).values(
            'item_concepto', 'cta'
        ).annotate(
            debito=ExpressionWrapper(
                Coalesce(Sum('debito'), Value(0)),
                output_field=DecimalField(max_digits = 12, decimal_places=2)
            ),
            credito=ExpressionWrapper(
                Coalesce(Sum('credito'), Value(0)),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            ),
        ).annotate(
            saldo=ExpressionWrapper(
                F('debito') - F('credito'),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
    ).order_by('item_concepto', 'cta')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="documento.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Courier", 10)
    width, height = letter  # 612 x 792 puntos
    row_height = 20  # 20 puntos entre filas
    col_width = 110  # 100 puntos entre columnas
    x = 25 # Empezar 100 puntos desde el borde izquierdo
    y = height - 25  # Empezar 100 puntos desde el borde superior
    encabezado(p)
    y_actual = y - 115
    sum_deb = sum_cre = 0
    for fila in resultado:
        p.drawString(x + 0 * col_width+10, y_actual, str(fila['item_concepto']))
        p.drawString(x + 1 * col_width+25, y_actual, str(fila['cta']))
        p.drawRightString(x + 2 * col_width + 80, y_actual, formatear_numero(fila['debito']))
        p.drawRightString(x + 3 * col_width + 80, y_actual, formatear_numero(fila['credito']))
        p.drawRightString(x + 4 * col_width + 80, y_actual, formatear_numero(fila['saldo']))
        y_actual -= 18  # Bajás a la siguiente línea (ajustá el espaciado según lo que necesites)
        sum_deb = sum_deb + fila['debito']
        sum_cre = sum_cre + fila['credito']
    p.drawString(x + 1 * col_width, y_actual, 'Sumas Iguales')
    p.drawRightString(x + 2 * col_width + 80, y_actual, formatear_numero(sum_deb))
    p.drawRightString(x + 3 * col_width + 80, y_actual, formatear_numero(sum_cre))
    p.drawRightString(x + 4 * col_width + 80, y_actual, formatear_numero(sum_deb-sum_cre))
        
    p.showPage()
    p.save()

    return response

def exportar_asiento_pe(request, fecha):
    print('Entra a exportar asiento ')
    oficina_id = request.session.get('oficina_id', datetime.now().year)
    try:
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)
    per_con = fecha.year
    numero = fecha.month
    mi_doc = DOCTO_CONTA.objects.filter(oficina_id = oficina_id,per_con = per_con,codigo = 34).first()
    if mi_doc == None:
        return HttpResponse("No Existe documento con codigo 34", status=400)
    mi_hec_eco = HECHO_ECONO.objects.filter(docto_conta = mi_doc,numero = numero).first()
    if mi_hec_eco == None:
        return HttpResponse("No Existe cOMprobante de perdiDa espErada", status=400)
    mis_det = DETALLE_ECONO.objects.filter(hecho_econo = mi_hec_eco)
    wb = Workbook()
    ws = wb.active
    ws.title = "ComprobantE_pe "
    campos = [
        field.name
        for field in DETALLE_ECONO._meta.fields
        if not field.is_relation or field.many_to_one is False
    ]

    ws.append(campos)
    for obj in mis_det:
        fila = [getattr(obj, campo) for campo in campos]
        ws.append(fila)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"comprobante_pe_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

def exportar_catego_pe(request, fecha):
    print('Entra a exportar catego ')
    oficina_id = request.session.get('oficina_id', datetime.now().year)
    try:
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)
    cars_his = CARTE_CAT_HIS.objects.filter(oficina_id=oficina_id, fecha=fecha)
    wb = Workbook()
    ws = wb.active
    ws.title = "CARTE_CAT_HIS"
    # campos = [
    #     field.name
    #     for field in CARTE_CAT_HIS._meta.fields
    #     if not field.is_relation or field.many_to_one is False
    # ]
    campos = ['fecha', 'cod_cre', 'nit', 'cod_lin_cre',	'cod_imp_con',	'for_pag',	'plazo',	'dias_mor',	'cap_ini',
              'sal_cap_pe',	'sal_cap_dia', 'sal_int_dia',	'int_cau_res_per',	'cat_mor',	'cat_arr',	'aporte',	
              'pro_ind_kap',	'pro_ind_int',	'saldo_1',	'saldo_2',	'val_gar_hip',	'int_pag_per',	'cat_int_mes',	
              'sal_cat_int',	'castigo',	'int_conkas_per',	'gas_pro_gen',	'sal_int_pe',	'int_cor_per',	
              'sal_int_contin',	'zeta',	'puntaje',	'cla_gar',	'cat_mod',	'cat_ree',	'cat_eva',	'cat_sel',	
              'categoria',	'int_pag_per',	'pro_inc',	'pdi',	'vea',	'per_esp',	'conta_ali',	'ali_acu',	
              'gas_pro_ind_acu'
    ]
    print('campos ordenados--->', campos)
    ws.append(campos)
    for obj in cars_his:
        fila = [getattr(obj, campo) for campo in campos]
        ws.append(fila)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"catego_pe_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

def exportar_rpki_pe(request, fecha):
    print('Entra a exportar Rpki ')
    oficina_id = request.session.get('oficina_id', datetime.now().year)
    try:
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)
    mis_rpki = RPKI.objects.filter(oficina_id=oficina_id, fecha=fecha)
    wb = Workbook()
    ws = wb.active
    ws.title = "RpKi"
    campos = [
        field.name
        for field in RPKI._meta.fields
        if not field.is_relation or field.many_to_one is False
    ]

    ws.append(campos)
    for obj in mis_rpki:
        fila = [getattr(obj, campo) for campo in campos]
        ws.append(fila)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"rPki_pe_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response

def indicador_de_cartera(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')

    if request.method == 'GET':
        return render(request, 'cartera_a_una_fecha.html')  # tu formulario

    if request.method == 'POST':
        fecha_corte = request.POST.get('fecha_corte')
        min_dias = int(request.POST.get('min_dias', 0))
        max_dias = int(request.POST.get('max_dias', 0))

        saldos = report_cartera_a_una_fecha(id_cli, id_ofi, fecha_corte, min_dias, max_dias)
        if not saldos:
            return HttpResponse("No se encontraron datos", status=404)

        # Inicializa tabla
        categorias = ['A', 'B', 'C', 'D', 'E', 'F']
        tabla = {
            cat: {
                'categoria': cat,
                'num_cre_mor': 0, 'por_cre_mor': 0.0,
                'num_cre_arr': 0, 'por_cre_arr': 0.0,
                'num_cre_pe': 0,  'por_cre_pe': 0.0,
            } for cat in categorias
        }

        total_saldo_capital = sum(s['saldo_capital'] for s in saldos)

        for saldo in saldos:
            if saldo['cat_mor'] != 'X':
                xCat = saldo['cat_mor']
                tabla[xCat]['num_cre_mor'] += 1
                tabla[xCat]['por_cre_mor'] += saldo['saldo_capital']
            if saldo['cat_arr'] != 'X':
                xCat = saldo['cat_arr']
                tabla[xCat]['num_cre_arr'] += 1
                tabla[xCat]['por_cre_arr'] += saldo['saldo_capital']
            if saldo['categoria'] != 'X':
                xCat = saldo['categoria']
                tabla[xCat]['num_cre_pe'] += 1
                tabla[xCat]['por_cre_pe'] += saldo['saldo_capital']

        # Convertir acumuladores a porcentaje
        for datos in tabla.values():
            if total_saldo_capital > 0:
                datos['por_cre_mor'] = round((datos['por_cre_mor'] / total_saldo_capital) * 100, 2)
                datos['por_cre_arr'] = round((datos['por_cre_arr'] / total_saldo_capital) * 100, 2)
                datos['por_cre_pe']  = round((datos['por_cre_pe']  / total_saldo_capital) * 100, 2)

        # Crear PDF en memoria
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        margin_x, margin_y = 50, 60
        # Título
        p.setFont("Helvetica-Bold", 16)
        titulo = f"Resumen Indicador de Cartera - Fecha: {fecha_corte}"
        x0 = margin_x
        y0 = height - margin_y - 10  # ajusta verticalmente
        ancho = width - 2 * margin_x
        alto = 25
        p.setFillColor(colors.lightblue)  
        p.rect(x0, y0, ancho, alto, fill=1, stroke=0)
        p.setFillColor(colors.black)
        text_x = x0 + ancho / 2
        text_y = y0 + alto / 2 - 6  # ajuste vertical para centrar texto según tamaño de fuente
        p.drawCentredString(text_x, text_y, titulo)
        # Encabezado
        y = height - margin_y - 30
        p.setFont("Helvetica-Bold", 10)
        titulo_y = height - margin_y - 20
        y = titulo_y - 25  # posición inicial de la tabla debajo del título
        headers = ["CATEG.", "N° CAT_MOR", "% CAT_MOR", "N° CAT_ARR", "% CAT_ARR", "N° CAT_PE", "% CAT_PE"]
        col_width = (width - 2 * margin_x) // len(headers)
        x_positions = [margin_x + i * col_width for i in range(len(headers))]

        # Fondo gris claro para el encabezado
        p.setFillColor(colors.lightgrey)
        p.rect(margin_x, y, width - 2 * margin_x, 18, fill=1, stroke=0)

        # Escribir los encabezados
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(colors.black)
        for i, header in enumerate(headers):
            p.drawString(x_positions[i] + 2, y + 4, header)
        
        y -= 25
        p.setFont("Helvetica", 10)

# Filas de datos por categoría
        for cat in sorted(tabla.keys()):
            fila = tabla[cat]
            valores = [
                fila['categoria'],
                str(fila['num_cre_mor']), f"{fila['por_cre_mor']}%",
                str(fila['num_cre_arr']), f"{fila['por_cre_arr']}%",
                str(fila['num_cre_pe']),  f"{fila['por_cre_pe']}%",
            ]
            for i, val in enumerate(valores):
                if i == 0:
                    p.drawString(x_positions[i], y, val)
                else:
                    p.drawRightString(x_positions[i] + col_width - 5, y, val)
            y -= 18

            if y < margin_y + 50:
                p.showPage()
                y = height - margin_y - 30
                p.setFont("Helvetica-Bold", 10)
                for i, header in enumerate(headers):
                    p.drawString(x_positions[i], y, header)
                y -= 20
                p.setFont("Helvetica", 10)

        total_num_cre_mor = sum(f['num_cre_mor'] for f in tabla.values())
        total_num_cre_arr = sum(f['num_cre_arr'] for f in tabla.values())
        total_num_cre_pe  = sum(f['num_cre_pe']  for f in tabla.values())

        total_valores = [
            "TOTAL",
            f"{total_num_cre_mor:,}", "100.00%",
            f"{total_num_cre_arr:,}", "100.00%",
            f"{total_num_cre_pe:,}",  "100.00%",
        ]

        p.setFillColorRGB(0.85, 0.85, 0.85)
        p.rect(margin_x - 5, y - 4, 512, 18, fill=1, stroke=0)

        p.setFont("Helvetica-Bold", 10)
        p.setFillColorRGB(0, 0, 0)
        for i, val in enumerate(total_valores):
            if i == 0:
                p.drawString(x_positions[i], y, val)
            else:
                p.drawRightString(x_positions[i] + col_width - 5, y, val)
        y -= 25  # dejar espacio tras la tabla

        qs = (
            CARTE_CAT_HIS.objects.annotate(year=ExtractYear('fecha'))
                .filter(year__gt=2015)
                .values('fecha')
                .annotate(
                    total=Sum('sal_cap_pe'),
                    total_a=Sum(Case(
                        When(cat_arr='A', then=F('sal_cap_pe')),
                        default=Value(0),
                        output_field=FloatField()
                    )),
                    porcentaje=ExpressionWrapper(
                    (F('total') - F('total_a')) * 100.0 / F('total'),
                    output_field=FloatField()
                )
            )
            .order_by('fecha')
        )

        fechas = [registro['fecha'] for registro in qs]
        valores = [registro['porcentaje'] for registro in qs]  # Asegúrate que 'porcentaje' es el nombre correcto
        p.setFont("Helvetica-Bold", 12)
        p.drawString(200, 515, "Evolución del Porcentaje de Morosidad")
        x0, y0 = 60, 350  # origen del gráfico
        ancho, alto = 500, 150
        if valores:
            min_valor = min(valores)
            max_valor = max(valores)
        else:
            min_valor = 0
            max_valor = 1

        margen = 5
        escala_min = max(min_valor - margen, 0)
        escala_max = min(max_valor + margen, 100)
        if escala_max == escala_min:
            escala_max += 1
        escala_y = alto / (escala_max - escala_min)

        p.setStrokeColorRGB(0, 0, 0)
        p.setLineWidth(0.5)
        p.line(x0, y0, x0, y0 + alto)     # eje Y
        p.line(x0, y0, x0 + ancho, y0)    # eje X

        p.setFont("Helvetica", 7)
        for i in range(int(escala_min), int(escala_max) + 1, 2):
            y = y0 + (i - escala_min) * escala_y
            p.setStrokeColorRGB(0.85, 0.85, 0.85)
            p.line(x0, y, x0 + ancho, y)
            p.setFillColorRGB(0, 0, 0)
            p.drawRightString(x0 - 5, y - 3, f"{i}%")

        num_puntos = len(valores)
        espacio_x = ancho / max(num_puntos - 1, 1)
        p.setStrokeColorRGB(0.2, 0.4, 0.6)
        p.setLineWidth(0.8)
        for i in range(1, num_puntos):
            x1 = x0 + (i - 1) * espacio_x
            y1 = y0 + (valores[i - 1] - escala_min) * escala_y
            x2 = x0 + i * espacio_x
            y2 = y0 + (valores[i] - escala_min) * escala_y
            p.line(x1, y1, x2, y2)

        for i in range(num_puntos):
            x = x0 + i * espacio_x
            y = y0 + (valores[i] - escala_min) * escala_y
            p.setFillColorRGB(0.1, 0.3, 0.6)
            p.circle(x, y, 1.5, fill=1)

        p.setFont("Helvetica", 6)
        etiquetas_max = 14  # Máximo número de etiquetas visibles
        salto = max(1, num_puntos // etiquetas_max)
        for i, fecha in enumerate(fechas):
            if i % salto == 0 or i == num_puntos - 1:
                x = x0 + i * espacio_x
                label = fecha.strftime('%Y-%m')
                if x + 20 < x0 + ancho:  # Evita que se monten al borde
                    p.drawCentredString(x, y0 - 10, label)
        p.showPage()
        p.save()
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

def riesgo_de_liquidez_cartera(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')

    if request.method == 'GET':
        return render(request, 'Riesgo_liq_cartera.html')  # tu formulario

    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_corte')
        try:
            fecha_base = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha inválida", status=400)
        accion = request.POST.get('accion')
        print('Accion ----- ',accion)
        por_semanas = request.POST.get('por_semanas') == '1'  # True si está marcado
        fechas = []
        if por_semanas:
            for i in range(7):
                fechas.append(fecha_base + timedelta(days=7 * i))
        else:
            meses = [1, 2, 3, 6, 9, 12, 72]
            for m in meses:
                nueva_fecha = fecha_base + relativedelta(months=+m)
                fechas.append(nueva_fecha.date() if isinstance(nueva_fecha, datetime) else nueva_fecha)
        
        resultados = []
        Creditos = CREDITOS.objects.filter(oficina_id = id_ofi, fec_des__lte = fecha_base
                ).exclude(
                    # Excluir créditos con estado 'C' y fec_ult_pag < self.fecha_ant_rec
                    Q(estado='C') & Q(fec_ult_pag__lt = fecha_base)
                ).exclude(
                    # Excluir créditos con estado 'H'
                    estado='H'
                ).exclude(
                    # Excluir créditos con estado 'H'
                    fec_des__gt = fecha_base
                )
        
        resumen_bandas = {}
        banda = 1
        num_fechas = len(fechas)

        fecha_cor = fecha_base
        for i in range(num_fechas):
            fecha_ini = fecha_cor
            if i < num_fechas - 1:
                fecha_fin = fechas[i] 
            else:
                fecha_fin = date(2999, 12, 31) 
            resumen_bandas[banda] = {
                'fecha_ini': fecha_ini,
                'fecha_fin': fecha_fin,
                'capital': 0,
                'interes': 0
            }
            fecha_cor = fechas[i] + timedelta(days=1)
            banda += 1
        coleccion = {}
        for credito in Creditos:
            liq_cre = Liquida_cre(credito.cod_cre,fecha_base)
            if liq_cre.lista_mov == None:
                credito.estado = 'H'
                credito.save()
                continue 
            liq_cre.liq_al_dia(fecha_base)
            if liq_cre.sal_cap_tot <= 0 :
                continue 
            liq_cre.calculo_periodo()
            xdias_mor = (liq_cre.fecha_focal-liq_cre.fec_al_dia).days
            xdias_mor = xdias_mor if xdias_mor > 0 else 0
            xCat  = 'A' if xdias_mor < 31 else 'B' if xdias_mor < 61 else 'C' if xdias_mor < 91 else 'D' if xdias_mor < 181 else 'E' if xdias_mor < 361 else 'F'
            bandas = [{'Cap': 0, 'Int': 0} for _ in range(7)]
            cod_cre = credito.cod_cre
            coleccion[cod_cre] = {
                'Nit' : credito.socio.tercero.doc_ide,
                'Cat': xCat,
                'CapTot' : liq_cre.sal_cap_tot,
                'Bandas': bandas
            }
            if xdias_mor < 31:
                xSaldo = liq_cre.sal_cap_tot
                xSCAnt = liq_cre.sal_cap_dia
                xSIAnt = liq_cre.sal_int_dia
                xIPT = liq_cre.int_pag_tot
                xacu_cap = 0
                xacu_int_cor = 0
                xacu_int_mor = 0
                banda = 0
                #print('Banda ',banda,'  saldo ',liq_cre.sal_cap_tot)
                for fecha_banda in fechas:
                    banda = banda + 1
                    liq_cre = Liquida_cre(credito.cod_cre,fecha_banda)
                    liq_cre.liq_al_dia(fecha_banda)
                    #print('Banda ',banda,'  saldo ',liq_cre.sal_cap_tot)
                    #print('Fecha anda ',fecha_banda,'  cap ',liq_cre.capital_a_pag,'  int ',liq_cre.int_cor_a_pag)
                    if ((xSaldo > 0 and banda == 1) or (liq_cre.sal_cap_dia >= xSCAnt and banda > 1)) :
                        xap_cap = liq_cre.capital_a_pag - xacu_cap
                        if banda == 1:
                            print('Bada 1 ',liq_cre.sal_cap_dia,xSaldo,liq_cre.sal_cap_tot)
                            xCapBan = (liq_cre.sal_cap_dia if liq_cre.sal_cap_dia > 0 else 0) + (xSaldo-liq_cre.sal_cap_tot if xSaldo-liq_cre.sal_cap_tot>0 else 0)
                            xIntBan = (liq_cre.sal_int_dia if liq_cre.sal_int_dia>0 else 0) + (liq_cre.int_pag_tot-xIPT if liq_cre.int_pag_tot-xIPT>0 else 0)
                        else:
                            if liq_cre.sal_cap_dia > 0:
                                xCapBan = liq_cre.sal_cap_dia - (xSCAnt if xSCAnt > 0 else 0)
                            else:
                                xCapBan = 0
                            if liq_cre.sal_int_dia > 0 and xSaldo > 0:
                                xIntBan = liq_cre.sal_int_dia - (xSIAnt if xSIAnt > 0 else 0)
                            else:
                                xIntBan = 0
                        xSaldo -= xCapBan
                        coleccion[cod_cre]['Bandas'][banda-1]['Cap'] = xCapBan + (xSaldo if banda == 7 else 0) 
                        coleccion[cod_cre]['Bandas'][banda-1]['Int'] = xIntBan
                        resumen_bandas[banda]['capital'] += xCapBan
                        resumen_bandas[banda]['interes'] += xIntBan
                        xSCAnt=liq_cre.sal_cap_dia
                        xSIAnt=liq_cre.sal_int_dia

                        if banda == 7 and xacu_cap < liq_cre.sal_cap_tot:
                            resumen_bandas[banda]['capital'] += (liq_cre.sal_cap_tot - xacu_cap)
                    else:
                        continue
            else:
                coleccion[cod_cre]['Bandas'][6]['Cap'] = liq_cre.sal_cap_tot
                coleccion[cod_cre]['Bandas'][6]['Int'] = 0
                resumen_bandas[7]['capital'] += liq_cre.sal_cap_tot
                resumen_bandas[7]['interes'] += 0

        if accion == 'excel':
            print('Generando riesgo de liquidez detallado')
# exportamos el detalle a excel 
            wb = Workbook()
            ws = wb.active
            ws.title = "Cartera"
            headers = [
                'Cod Crédito', 'Nit', 'Categoría','CapTot',
                'Banda 1 Cap', 'Banda 1 Int',
                'Banda 2 Cap', 'Banda 2 Int',
                'Banda 3 Cap', 'Banda 3 Int',
                'Banda 4 Cap', 'Banda 4 Int',
                'Banda 5 Cap', 'Banda 5 Int',
                'Banda 6 Cap', 'Banda 6 Int',
                'Banda 7 Cap', 'Banda 7 Int',
            ]
            ws.append(headers)
            for cod_cre, datos in coleccion.items():
                fila = [
                    cod_cre,
                    datos['Nit'],
                    datos['Cat'],
                    datos['CapTot'],
                ]
                for banda in datos['Bandas']:  # 7 bandas
                    fila.append(banda['Cap'])
                    fila.append(banda['Int'])
                ws.append(fila)
            ws.append([])          
            ws.append(['TOTALES', '', ''] + ['' for _ in range(14)])
            fila_totales = ['TOTAL', '', '']
            for i in range(1, 8):  # Bandas 1 a 7
                fila_totales.append(resumen_bandas[i]['capital'])
                fila_totales.append(resumen_bandas[i]['interes'])
            ws.append(fila_totales)
        # Preparar respuesta HTTP para descarga
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            nombre_archivo = f"riesgo_liq_cartera_{fecha_base}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            wb.save(response)
            return response

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        margin_x, margin_y = 50, 60
        p.setFont("Helvetica-Bold", 16)
        titulo = f"Resumen Riesgo de Liquidez de Cartera - Fecha: {fecha_base.strftime('%d/%m/%Y')}"
        x0 = margin_x
        y0 = height - margin_y - 10
        ancho = width - 2 * margin_x
        alto = 25
        p.setFillColor(colors.lightblue)
        p.rect(x0, y0, ancho, alto, fill=1, stroke=0)
        p.setFillColor(colors.black)
        p.drawCentredString(x0 + ancho / 2, y0 + alto / 2 - 6, titulo)

        y = y0 - 30
        p.setFont("Helvetica-Bold", 10)
        headers = ["BANDA", "FECHA INICIAL", "FECHA FINAL", "CAPITAL", "INTERÉS"]
        col_width = (width - 2 * margin_x) // len(headers)
        x_positions = [margin_x + i * col_width for i in range(len(headers))]

        p.setFillColor(colors.lightgrey)
        p.rect(margin_x, y, width - 2 * margin_x, 18, fill=1, stroke=0)
        p.setFillColor(colors.black)

        for i, header in enumerate(headers):
            p.drawString(x_positions[i] + 2, y + 4, header)

        # Filas de datos
        p.setFont("Helvetica", 10)
        y -= 20
        for banda, datos in resumen_bandas.items():
            if y < margin_y + 40:  # Salto de página si estamos muy abajo
                p.showPage()
                y = height - margin_y
                p.setFont("Helvetica", 10)

            fila = [
                str(banda),
                datos['fecha_ini'].strftime("%d/%m/%Y"),
                datos['fecha_fin'].strftime("%d/%m/%Y"),
                f"${datos['capital']:,.2f}",
                f"${datos['interes']:,.2f}",
            ]
            for i, valor in enumerate(fila):
                p.drawString(x_positions[i] + 2, y, valor)
            y -= 18
        total_capital = sum(b['capital'] for b in resumen_bandas.values())
        total_interes = sum(b['interes'] for b in resumen_bandas.values())
        p.setFont("Helvetica-Bold", 10)
        fila = [
            "TOTALES", "", "", 
            f"${total_capital:,.2f}", 
            f"${total_interes:,.2f}"
        ]
        for i, valor in enumerate(fila):
            p.drawString(x_positions[i] + 2, y, valor)
        #y -= 18

        # Grafica
        # Datos base
        bandas_dibujar = list(resumen_bandas.items())[:6]  # solo las 6 primeras
        total_dias = sum((b['fecha_fin'] - b['fecha_ini']).days for _, b in bandas_dibujar)
        max_densidad = max(
            (b['capital'] + b['interes']) / ((b['fecha_fin'] - b['fecha_ini']).days or 1)
            for _, b in bandas_dibujar
        )
        d = Drawing(450, 200)
        x0 = 30
        y0 = 30
        height_max = 150
        width_total = 390
        x_cursor = x0

        for banda, datos in bandas_dibujar:
            dias = (datos['fecha_fin'] - datos['fecha_ini']).days or 1
            base_px = width_total * dias / total_dias
            area_valor = datos['capital'] + datos['interes']
            altura = (area_valor / dias) / max_densidad * height_max

    # Dibujar rectángulo
            rect = Rect(x_cursor, y0, base_px, altura)
            rect.fillColor = colors.lightblue
            rect.strokeColor = colors.black
            d.add(rect)

    # Agregar etiqueta debajo
            label = Label()
            label.setOrigin(x_cursor + base_px / 2, y0 - 10)
            label.boxAnchor = 'n'
            label.fontSize = 7
            label.setText(f"Banda {banda}")
            d.add(label)
            x_cursor += base_px  # avanzar a la derecha
            renderPDF.draw(d, p, margin_x, y - 250)
        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Resumen_Liquidez_{fecha_base.strftime("%Y-%m-%d")}.pdf"'
        return response

    # ⚠️ Si no es POST, devolver algo
    return HttpResponse("Método no permitido o sin datos", status=400)

def listado_cifin(fec_corte: date, tipo_reporte: str):
    # Fecha focal según tipo de reporte
    if tipo_reporte == 1:
        f_focal = fec_corte
    else:
        f_focal = fec_corte + timedelta(days=20)
    # f_focal = fec_corte
    # Lista de líneas del archivo
    registros = []

    # print('Fecha', f_focal)
    encabezado = f"121050181          01{f_focal.strftime('%Y%m%d')}\n"
    registros.append(''.join(encabezado))

    # Créditos base
    creditos = CREDITOS.objects.filter(
        # ~Q(est_jur='P'),
        # ~Q(com_des_id=''),
        estado = 'A',
        rep_cen_rie='S',
        fec_des__lte=f_focal
    ).order_by('cod_cre')
    # print('estos creditos',creditos)
    for credito in creditos:
        cod_cre = credito.cod_cre
        # Busca la categoría del crédito a la fecha
        categoria = CARTE_CAT_HIS.objects.filter(
            cod_cre=cod_cre,
            fecha__lte=f_focal
        ).order_by('-fecha').first()
        # print('este crédito', cod_cre)
        # Validación si continúa
        if not categoria and not (
            credito.fec_des <= f_focal.replace(day=1) - timedelta(days=1)
            and credito.fec_ult_pag.year == f_focal.year
            and credito.fec_ult_pag.month == f_focal.month
        ) and credito.est_jur != "K":
            continue

        # # Lógica de cálculo del saldo, mora, cuotas, fechas, estado, etc.
        # # Aquí deberías llamar a tus funciones CalculoEdadCr y LiqCreAlDia adaptadas a Python
        # saldo = calcular_saldo_credito(credito, f_focal)
        # if saldo <= 0 and credito.est_jur != "K":
        #     continue

        # Información del titular (Socio y Tercero)
        asociado = ASOCIADOS.objects.filter(id=credito.socio_id).first()
        tercero = asociado.tercero if asociado else None
        pagador = PAGADORES.objects.filter(id=asociado.id_pag_id).first() if asociado else None

        # Información del codeudor
        codeudores = GAR_NO_IDONEA.objects.filter(credito_id=credito.id)
        codeudor_data = []
        for codeudor in codeudores:
            tercero_codeudor = TERCEROS.objects.filter(doc_ide=codeudor.doc_ide).first()
            if tercero_codeudor:
                codeudor_data.append(obtener_linea_codeudor(credito, tercero_codeudor, f_focal, categoria))

        # Construcción de la línea del titular
        # linea_titular = construir_linea_titular(credito, tercero, categoria, saldo, fec_corte, f_focal, asociado, pagador)
        linea_titular = construir_linea_titular(credito, tercero, f_focal, categoria)
        registros.append(linea_titular)

        # Agregar las líneas de codeudores
        registros.extend(codeudor_data)

    # Pie con cantidad de registros
    pie = f"9{str(len(registros)).zfill(8)}\n"
    registros.append(''.join(pie))

    # for linea in registros:
    # # Guardar el archivo
    # with open(nom_archivo, 'w', encoding='utf-8') as f:
    #     f.writelines(registros)
    # print('estos registros cifin', registros)
    return registros #nom_archivo
    # return "".join(registros)

# Funciones auxiliares básicas
def calcular_saldo_credito(credito, f_focal):
    # Adaptar la lógica de CalculoEdadCr y LiqCreAlDia
    # Ejemplo simplificado:
    return credito.cap_ini

# def construir_linea_titular(credito, tercero, categoria, saldo, fec_cor, f_focal, socio, pagador):
def construir_linea_titular(credito, tercero, fecha_corte, categoria):  
    def str_safe(valor, longitud):
        return str(valor).strip()[:longitud].ljust(longitud) if valor else ''.ljust(longitud)

    def fecha_safe(fecha):
        return fecha.strftime('%Y%m%d') if fecha else '00000000'

    def num_safe(valor, longitud):
        return f"{int(valor):0{longitud}}" if valor is not None else '0'.rjust(longitud, '0')
    
    # Aquí ensamblas la línea tipo 2, usando los campos que has traído.
    # Por simplicidad solo mostramos una parte
    # 🔹 Diccionarios de mapeo
    mapeo_tip_doc = {'C': '01', 'N': '02', 'E': '03', 'T': '04', 'P': '05','R': '09'}
    mapeo_tip_ter = {'N': '000', 'J': '002'} 
    mapeo_categoria = {'A': '01', 'B': '02', 'C': '03', 'D': '04', 'E': '05'}
    xCuoMor = 0
    xValMor = "000000000000"
    if categoria.dias_mor < 30:
        xEdadMora = "00"
        xValMor = "000000000000"
    elif 30 <= categoria.dias_mor < 60:
        if credito.not_mor and credito.fec_not_mor and credito.fec_not_mor < fecha_corte:
            xEdadMora = "01"
        else:
            xEdadMora = "00"
            xCuoMor = 0
    elif 60 <= categoria.dias_mor < 90:
        xEdadMora = "02"
    elif 90 <= categoria.dias_mor < 120:
        xEdadMora = "03"
    elif 120 <= categoria.dias_mor < 150:
        xEdadMora = "04"
    elif 150 <= categoria.dias_mor < 180:
        xEdadMora = "05"
    elif 180 <= categoria.dias_mor < 210:
        xEdadMora = "06"
    elif 210 <= categoria.dias_mor < 240:
        xEdadMora = "07"
    elif 240 <= categoria.dias_mor < 270:
        xEdadMora = "08"
    elif 270 <= categoria.dias_mor < 300:
        xEdadMora = "09"
    elif 300 <= categoria.dias_mor < 330:
        xEdadMora = "10"
    elif 330 <= categoria.dias_mor < 360:
        xEdadMora = "11"
    elif 360 <= categoria.dias_mor < 540:
        xEdadMora = "12"
    elif 540 <= categoria.dias_mor < 730:
        xEdadMora = "13"
    elif categoria.dias_mor >= 730:
        xEdadMora = "14"
        
    # xAnosMora = categoria.dias_mor // 360
        
    tip_doc = mapeo_tip_doc.get(tercero.cla_doc) if tercero else ''.ljust(2)
    doc_ide = str_safe(getattr(tercero, 'doc_ide', ''), 15)
    pri_ape = str_safe(getattr(tercero, 'pri_ape', ''), 15)
    seg_ape = str_safe(getattr(tercero, 'seg_ape', ''), 15)
    pri_nom = tercero.pri_nom.strip() if tercero else ''
    seg_nom = tercero.seg_nom.strip() if tercero else ''
    nombre_completo = pri_nom+' '+seg_nom
    nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
    # nombre = str_safe((getattr(tercero, 'pri_nom', '') + ' ' + getattr(tercero, 'seg_nom', '')), 30)
    # nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
    # espacio2 = ''.ljust(2)
    fec_lim_pag = fecha_safe(fecha_corte)
    cod_cre = str_safe(getattr(credito, 'cod_cre', ''), 20)
    cod_age = '000001'
    deu_pri = 'P'
    calificacion = mapeo_categoria.get(getattr(categoria, 'cat_arr', ''), '05')   #mapeo_categoria.get(categoria.cat_arr, '05') if tercero else ''.ljust(2)
    sit_est_tit = '05' # xSitEstTit+ ;		&& Situacion o Estado del Titular (Tabla 24 Cifin)
    estado = '01' # xEstado+ ;		&& Estado (Tabla 10 Cifin) xEstado = IIF(S24SALCAP<=1,"07","01") && 07=SALDADO, 01=VIGENTE
    edad_mora = xEdadMora #		carte_cat_his__dias_mor     && Edad de Mora (Tabla 8 Cifin)
    anos_mora = str(getattr(categoria, 'dias_mor', 0) // 360).rjust(2, '0')  #10   #xAnosMora # xAnosMora+ ;		&& Anos En Mora
    fec_act = fecha_safe(fecha_corte) #fecha_corte.strftime('%Y%m%d') # xFecAct+ ;		fec_cor                     && Fecha de Corte  (AAAAMMDD)
    fec_des = fecha_safe(getattr(credito, 'fec_des', None))  #credito.fec_des.strftime('%Y%m%d')#		creditos__fec_des           && Fecha Inicial o de Expedicion (AAAAMMDD)
    fec_ven = fecha_safe(getattr(credito, 'fec_ven', None))  #credito.fec_ven.strftime('%Y%m%d')# xFecVen+ ;creditos__fec_ven       && Fecha Terminacion (AAAAMMDD)                 
    fec_exi = fec_ven # xFecExi+ ;		&& Fecha Exigibilidad (AAAAMMDD)
    # espacio8 = ''.ljust(8)# space(8)+ ;		&& Fecha Prescripcion (AAAAMMDD)
    fec_ult_pag = fecha_safe(getattr(credito, 'fec_ult_pag', None)) # credito.fec_ult_pag.strftime('%Y%m%d') # xFecPag+ ;		creditos__fec_ult_pag       && Fecha Pago (AAAAMMDD)
    mod_ext = '  '# xModExt+ ;		&& Modo Extincion (Tabla 9 Cifin)
    tip_pago = '01' # "01" + ;			&& Tipo Pago (Tabla 31 Ci"01" --> Volunt"02" -->No voluntario
    per_pag = str_safe(getattr(credito, 'per_ano', ''), 2) #credito.per_ano # xPerPag + ;		creditos__per_ano           && Periocidad de Pago (Tabla 6 Cifin)
    pro_no_pago = '999' # xProbNoPag	+ ;	&& Probabilidad de No Pago (Valor entre 0 y 100)
    # xCuoCan+ ;		&& Numero Cuotas Pagadas
    tot_cuo = str(getattr(credito, 'num_cuo_ini', 0)).rjust(3, '0')  #credito.num_cuo_ini # xTotCuo+ ;		creditos__num_cuo_ini       && Numero Cuotas Pactadas
    cuo_mor = xCuoMor # xCuoMor+ ;		&& Cuotas en Mora
    cap_ini = num_safe(getattr(categoria, 'cap_ini', 0) // 1000, 12) #'{:012}'.format(int(categoria.cap_ini / 1000)) # xValIni+ ;		carte_cat_his__cap_ini      && Valor o Cupo (En Miles)
    val_mor = xValMor # xValMor+ ;		&& Valor de Mora (En Miles)
    val_saldo = num_safe(getattr(categoria, 'saldo_1', 0) // 1000, 12) # '{:012}'.format(int(categoria.saldo_1 / 1000))# xValSaldo+ ;		carte_cat_his__sal_cap      && Valor Saldo (En Miles)
    val_cuo = num_safe(getattr(credito, 'val_cuo_ini', 0) // 1000, 12) # '{:012}'.format(int(credito.val_cuo_ini / 1000)) # xValCuo+ ;		creditos__val_cuo_ini       && Valor de la Cuota (En Miles)
    # espacio12 = ''.ljust(12) # space(12)+ ;		&& Valor de Cargo Fijo ### (No Aplica) ###
    lin_cre = '008' # "008"+ ;		carte_cat_his__cod_lin_cre  && Linea Credito (Tabla 3 Cifin)
    espacio3 = ''.ljust(3)# space(3)+ ;		&& Clausula de Permanencia ### (No Aplica) ###
    # "001"+ ;			&& Tipo de Contrato (Tabla 25 Cifin)
    # "001"+ ;			&& Estado del Contrato (Tabla 26 Cifin)
    # espacio5 = ''.ljust(5)
    # space(3)+ ;		&& Termino o Vigencia del Contrato ### (No Aplica) ###
    # space(2)+ ;		&& Numero Meses Contrato ### (No Aplica) ###
    naturaleza = mapeo_tip_ter.get(getattr(tercero, 'tip_ter', ''), '000') # mapeo_tip_ter.get(tercero.tip_ter) if tercero else ''.ljust(3)# xNaturaleza+ ;	&& Naturaleza Juridica (Tabla 19 Cifin)
    mod_cre = '02' # "02"+ ;			&& Modalidad de Credito (Tabla 2 Cifin) -- consumo
    # xTipMon+ ;		&& Tipo de Moneda (Tabla 5 Cifin)
    # xTipGar+ ;		&& Tipo Garantia (Tabla 4 Cifin)
    # "000000000000"+ ;	&& Valor Garantia (En Miles)
    # "02"+ ;			&& Obligacion Reestructurada (Tabla 32 Cifin)
    # espacio39 = ''.ljust(39)
    # space(2)+ ;		&& Naturaleza Reestructuracion (Tabla 11 Cifin)
    # space(3)+ ;		&& Numero Reestructuraciones
    # space(3)+ ;		&& Clase Tarjeta ### (No Aplica) ### 
    # space(4)+ ;		&& Numero de Cheques Devueltos ### (No Aplica) ###		
    # space(2)+ ;		&& Categoria Servicios ### (No Aplica) ###
    # space(2)+ ;		&& Plazo ### (No Aplica) ### 
    # space(6)+ ;		&& Dias Cartera ### (No Aplica) ###
    # space(2)+ ;		&& Tipo Cuenta ### (No Aplica) ###
    # space(12)+ ;		&& Cupo Sobregiro ### (No Aplica) ###
    # space(3)+ ;		&& Dias Autorizados ### (No Aplica) ###
    direccion = str_safe(getattr(tercero, 'direccion', ''), 60) # tercero.direccion # xDirRes + ; 		tercero__direccion          && Direccion Titular
    tel_res = str_safe(getattr(tercero, 'celular1', ''), 10) # tercero.celular1 # xTelRes + ;		tercero__celular1           && Telefono Casa Titular
    cod_ciu =  str_safe(getattr(getattr(tercero, 'cod_ciu_res', None), 'codigo', ''), 5) # tercero.cod_ciu_res.codigo # "000001" + ;		tercero__localidades_codigo && Codigo Ciudad Casa
    ciu_res = str_safe(getattr(getattr(tercero, 'cod_ciu_res', None), 'nombre', ''), 30)  # tercero.cod_ciu_res.nombre.strip()  # xCiuRes + ;		tercero__localidades_nombre && Ciudad Casa Titular
    cod_dep = str_safe(getattr(getattr(tercero, 'cod_ciu_res', None), 'cod_pos', '')[:2], 3).ljust(3, '0') #tercero.cod_ciu_res.cod_pos[:2]  #.ljust(3, '0') # "050" + ;	tercero__localidades_cod_pos[:2]    && Codigo Departamento Casa
    departamento = str_safe(getattr(getattr(tercero, 'cod_ciu_res', None), 'departamento', ''), 30) # tercero.cod_ciu_res.departamento.strip() # "META" + ; tercero__localidades_departamento   && Departamento Casa Titular
    # espacio293 = ''.ljust(293)
    # space(60)+ ;		&& Nombre Empresa
    # space(60)+ ;		&& Direccion Empresa
    # space(20)+ ;		&& Telefono Empresa
    # space(6)+ ;		&& Codigo Ciudad Empresa
    # space(6)+ ;		&& Codigo Departamento Empresa
    # space(20)+ ;		&& Departamento Empresa Titular
    # space(8)+ ;		&& Fecha Inicio Excension GMF
    # space(8)+ ;		&& Fecha Terminacion Escension GMF
    # space(2)+ ;		&& Numero Renovacion CDT
    # space(2)+ ;		&& Cuenta Ahorro Excenta GMF
    # space(2)+ ;		&& Tipo Identificacion Originaria
    # space(14)+ ;		&& Numero Identificacion Originaria
    # space(3)+ ;		&& Tipo Entidad Originaria
    # space(3)+ ;		&& Codigo Entidad Originaria
    # space(2)+ ;		&& Tipo Fideicomiso
    # space(3)+ ;		&& Numero Fideicomiso
    # space(60)+ ;		&& Nombre Fideicomiso
    # space(4)+ ;		&& Tipo Deuda Cartera
    # space(4)+ ;		&& Tipo de Poliza
    # space(6) +;		&& Codigo de Ramo
    
    espacio2 = '  '
    espacio3 = ' ' * 3
    espacio5 = ' ' * 5
    espacio8 = ' ' * 8
    espacio12 = ' ' * 12
    espacio39 = ' ' * 39
    espacio293 = ' ' * 293
    

    # Debes agregar todos los campos en el orden y tamaño exacto como en tu FoxPro
    datos_titular = f"2{tip_doc}{doc_ide}{pri_ape}{seg_ape}{nombre}{espacio2}{fec_lim_pag}{cod_cre}{cod_age}{deu_pri}{calificacion}{sit_est_tit}{estado}{edad_mora}{anos_mora}{fec_act}{fec_des}{fec_ven}{fec_exi}{espacio8}{fec_ult_pag}{mod_ext}{tip_pago}{fec_ult_pag}{per_pag}{pro_no_pago}{tot_cuo}{cuo_mor}{cap_ini}{val_saldo}{val_mor}{val_cuo}{espacio12}{lin_cre}{espacio3}{espacio5}{naturaleza}{mod_cre}{espacio39}{direccion}{tel_res}{cod_ciu}{ciu_res}{cod_dep}{departamento}{espacio293}\n"
    return (''.join(datos_titular))

def obtener_linea_codeudor(credito, tercero, fecha_corte, categoria):
    # Similar a construir_linea_titular pero con los datos del codeudor
    # nit = tercero.doc_ide.ljust(15)
    # nombre = tercero.nombre.ljust(30)
    # cod_cre = credito.cod_cre.ljust(20)
    
    # 🔹 Diccionarios de mapeo
    mapeo_tip_doc = {'C': '01', 'N': '02', 'E': '03', 'T': '04', 'P': '05','R': '09'}
    mapeo_tip_ter = {'N': '000', 'J': '002'} 
    mapeo_categoria = {'A': '01', 'B': '02', 'C': '03', 'D': '04', 'E': '05'}
    
    if categoria.dias_mor < 30:
        xEdadMora = "00"
        xValMor = "000000000000"
    elif 30 <= categoria.dias_mor < 60:
        # if credito.not_mor and credito.fec_not_mor < fec_cor:
            # xEdadMora = "01"
        # else:
            xEdadMora = "00"
            # xCuoMor = 0
    elif 60 <= categoria.dias_mor < 90:
        xEdadMora = "02"
    elif 90 <= categoria.dias_mor < 120:
        xEdadMora = "03"
    elif 120 <= categoria.dias_mor < 150:
        xEdadMora = "04"
    elif 150 <= categoria.dias_mor < 180:
        xEdadMora = "05"
    elif 180 <= categoria.dias_mor < 210:
        xEdadMora = "06"
    elif 210 <= categoria.dias_mor < 240:
        xEdadMora = "07"
    elif 240 <= categoria.dias_mor < 270:
        xEdadMora = "08"
    elif 270 <= categoria.dias_mor < 300:
        xEdadMora = "09"
    elif 300 <= categoria.dias_mor < 330:
        xEdadMora = "10"
    elif 330 <= categoria.dias_mor < 360:
        xEdadMora = "11"
    elif 360 <= categoria.dias_mor < 540:
        xEdadMora = "12"
    elif 540 <= categoria.dias_mor < 730:
        xEdadMora = "13"
    elif categoria.dias_mor >= 730:
        xEdadMora = "14"
        
    xAnosMora = 10  #categoria.dias_mor / 360
        
    tip_doc =  mapeo_tip_doc.get(tercero.cla_doc) if tercero else ''.ljust(2)
    doc_ide = tercero.doc_ide.ljust(15) if tercero else ''.ljust(15)
    pri_ape = tercero.pri_ape.ljust(15) if tercero else ''.ljust(15)
    seg_ape = tercero.seg_ape.ljust(15) if tercero else ''.ljust(15)
    pri_nom = tercero.pri_nom.strip() if tercero else ''
    seg_nom = tercero.seg_nom.strip() if tercero else ''
    nombre_completo = pri_nom+' '+seg_nom
    nombre = nombre_completo.ljust(30) if tercero else ''.ljust(30)
    espacio2 = ''.ljust(2)
    fec_lim_pag = fecha_corte.strftime('%Y%m%d')
    cod_cre = credito.cod_cre.ljust(20)
    cod_age = '000001'
    deu_pri = 'C'
    calificacion = mapeo_categoria.get(categoria.cat_arr, '05') if tercero else ''.ljust(2)
    sit_est_tit = '05' # xSitEstTit+ ;		&& Situacion o Estado del Titular (Tabla 24 Cifin)
    estado = '01' # xEstado+ ;		&& Estado (Tabla 10 Cifin) xEstado = IIF(S24SALCAP<=1,"07","01") && 07=SALDADO, 01=VIGENTE
    edad_mora = xEdadMora #		carte_cat_his__dias_mor     && Edad de Mora (Tabla 8 Cifin)
    anos_mora = xAnosMora # xAnosMora+ ;		&& Anos En Mora
    fec_act = fecha_corte.strftime('%Y%m%d') # xFecAct+ ;		fec_cor                     && Fecha de Corte  (AAAAMMDD)
    fec_des = credito.fec_des.strftime('%Y%m%d')#		creditos__fec_des           && Fecha Inicial o de Expedicion (AAAAMMDD)
    fec_ven = credito.fec_ven.strftime('%Y%m%d')# xFecVen+ ;creditos__fec_ven       && Fecha Terminacion (AAAAMMDD)                 
    fec_exi = fec_ven # xFecExi+ ;		&& Fecha Exigibilidad (AAAAMMDD)
    espacio8 = ''.ljust(8)# space(8)+ ;		&& Fecha Prescripcion (AAAAMMDD)
    fec_ult_pag = credito.fec_ult_pag.strftime('%Y%m%d') # xFecPag+ ;		creditos__fec_ult_pag       && Fecha Pago (AAAAMMDD)
    mod_ext = '  '# xModExt+ ;		&& Modo Extincion (Tabla 9 Cifin)
    tip_pago = '01' # "01" + ;			&& Tipo Pago (Tabla 31 Ci"01" --> Volunt"02" -->No voluntario
    per_pag = credito.per_ano # xPerPag + ;		creditos__per_ano           && Periocidad de Pago (Tabla 6 Cifin)
    # xProbNoPag	+ ;	&& Probabilidad de No Pago (Valor entre 0 y 100)
    # xCuoCan+ ;		&& Numero Cuotas Pagadas
    tot_cuo = credito.num_cuo_ini # xTotCuo+ ;		creditos__num_cuo_ini       && Numero Cuotas Pactadas
    # xCuoMor+ ;		&& Cuotas en Mora
    cap_ini = '{:012}'.format(int(categoria.cap_ini / 1000)) # xValIni+ ;		carte_cat_his__cap_ini      && Valor o Cupo (En Miles)
    # xValMor+ ;		&& Valor de Mora (En Miles)
    val_saldo = '{:012}'.format(int(categoria.saldo_1 / 1000))# xValSaldo+ ;		carte_cat_his__sal_cap      && Valor Saldo (En Miles)
    val_cuo = '{:012}'.format(int(credito.val_cuo_ini / 1000)) # xValCuo+ ;		creditos__val_cuo_ini       && Valor de la Cuota (En Miles)
    espacio12 = ''.ljust(12) # space(12)+ ;		&& Valor de Cargo Fijo ### (No Aplica) ###
    # "008"+ ;			carte_cat_his__cod_lin_cre  && Linea Credito (Tabla 3 Cifin)
    espacio3 = ''.ljust(3)# space(3)+ ;		&& Clausula de Permanencia ### (No Aplica) ###
    # "001"+ ;			&& Tipo de Contrato (Tabla 25 Cifin)
    # "001"+ ;			&& Estado del Contrato (Tabla 26 Cifin)
    # space(3)+ ;		&& Termino o Vigencia del Contrato ### (No Aplica) ###
    # space(2)+ ;		&& Numero Meses Contrato ### (No Aplica) ###
    naturaleza = mapeo_tip_ter.get(tercero.tip_ter) if tercero else ''.ljust(3)# xNaturaleza+ ;	&& Naturaleza Juridica (Tabla 19 Cifin)
    mod_cre = '02' # "02"+ ;			&& Modalidad de Credito (Tabla 2 Cifin) -- consumo
    # xTipMon+ ;		&& Tipo de Moneda (Tabla 5 Cifin)
    # xTipGar+ ;		&& Tipo Garantia (Tabla 4 Cifin)
    # "000000000000"+ ;	&& Valor Garantia (En Miles)
    # "02"+ ;			&& Obligacion Reestructurada (Tabla 32 Cifin)
    # space(2)+ ;		&& Naturaleza Reestructuracion (Tabla 11 Cifin)
    # space(3)+ ;		&& Numero Reestructuraciones
    # space(3)+ ;		&& Clase Tarjeta ### (No Aplica) ### 
    # space(4)+ ;		&& Numero de Cheques Devueltos ### (No Aplica) ###		
    # space(2)+ ;		&& Categoria Servicios ### (No Aplica) ###
    # space(2)+ ;		&& Plazo ### (No Aplica) ### 
    # space(6)+ ;		&& Dias Cartera ### (No Aplica) ###
    # space(2)+ ;		&& Tipo Cuenta ### (No Aplica) ###
    # space(12)+ ;		&& Cupo Sobregiro ### (No Aplica) ###
    # space(3)+ ;		&& Dias Autorizados ### (No Aplica) ###
    direccion = tercero.direccion # xDirRes + ; 		tercero__direccion          && Direccion Titular
    tel_res = tercero.celular1 # xTelRes + ;		tercero__celular1           && Telefono Casa Titular
    # "000001" + ;		tercero__localidades_codigo && Codigo Ciudad Casa
    # xCiuRes + ;		tercero__localidades_nombre && Ciudad Casa Titular
    # "050" + ;			tercero__localidades_cod_pos[:2]    && Codigo Departamento Casa
    # "META" + ;		tercero__localidades_departamento   && Departamento Casa Titular
    # space(60)+ ;		&& Nombre Empresa
    # space(60)+ ;		&& Direccion Empresa
    # space(20)+ ;		&& Telefono Empresa
    # space(6)+ ;		&& Codigo Ciudad Empresa
    # space(6)+ ;		&& Codigo Departamento Empresa
    # space(20)+ ;		&& Departamento Empresa Titular
    # space(8)+ ;		&& Fecha Inicio Excension GMF
    # space(8)+ ;		&& Fecha Terminacion Escension GMF
    # space(2)+ ;		&& Numero Renovacion CDT
    # space(2)+ ;		&& Cuenta Ahorro Excenta GMF
    # space(2)+ ;		&& Tipo Identificacion Originaria
    # space(14)+ ;		&& Numero Identificacion Originaria
    # space(3)+ ;		&& Tipo Entidad Originaria
    # space(3)+ ;		&& Codigo Entidad Originaria
    # space(2)+ ;		&& Tipo Fideicomiso
    # space(3)+ ;		&& Numero Fideicomiso
    # space(60)+ ;		&& Nombre Fideicomiso
    # space(4)+ ;		&& Tipo Deuda Cartera
    # space(4)+ ;		&& Tipo de Poliza
    # space(6) +;		&& Codigo de Ramo    
    datos_codeudor = f"2{tip_doc}{doc_ide}{pri_ape}{seg_ape}{nombre}{espacio2}{fec_lim_pag}{cod_cre}{cod_age}{deu_pri}{calificacion}{sit_est_tit}{estado}{edad_mora}{anos_mora}{fec_act}{fec_des}{fec_ven}{fec_exi}{espacio8}{fec_ult_pag}{mod_ext}{tip_pago}{fec_ult_pag}{per_pag}{tot_cuo}{cap_ini}{val_saldo}{val_cuo}{espacio12}{espacio3}{naturaleza}{mod_cre}{direccion}{tel_res}\n"
    return (''.join(datos_codeudor))

def deterioro_de_cartera(request, fecha):
    print('Entra a exportar deterioro de cartera ')
    oficina_id = request.session.get('oficina_id', datetime.now().year)

    try:
        fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Fecha inválida", status=400)

    per_con = fecha.year
    numero = fecha.month

    mi_doc = DOCTO_CONTA.objects.filter(oficina_id=oficina_id, per_con=per_con, codigo=34).first()
    if mi_doc is None:
        return HttpResponse("No existe documento con código 34", status=400)

    mi_hec_eco = HECHO_ECONO.objects.filter(docto_conta=mi_doc, numero=numero).first()
    if mi_hec_eco is None:
        return HttpResponse("No existe comprobante de pérdida esperada", status=400)

    mis_det = DETALLE_ECONO.objects.filter(hecho_econo=mi_hec_eco)

    wb = Workbook()
    ws = wb.active
    ws.title = "Comprobante_PE"

    # 🟢 Encabezados personalizados
    encabezado_1 = "Superintendencia de la Economía Solidaria"
    encabezado_2 = "COOPERATIVA ESPECIALIZADA DE AHORRO Y CRÉDITO DE LA ORINOQUÍA"
    encabezado_3 = f"INFORME INDIVIDUAL CÁLCULO DE DETERIORO DE CARTERA DE CRÉDITO {fecha.strftime('%d/%m/%Y')}"

    # Escribir encabezados en las primeras tres filas
    ws.append([encabezado_1])
    ws.append([encabezado_2])
    ws.append([encabezado_3])

    # Centrar el texto en las primeras tres filas (en columnas A a X)
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=24):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # 🟢 Nombres de columnas personalizados
    columnas_legibles = [
        "TipoIden", "NUMIDENT", "NUMCREDITO", "CodigoContable", "SaldoCapital", "SaldoIntereses", 
        "PagoxAsociados", "APORTES", "AHORROPERMANENTE", "VEA", "VALORGARANTIA", "TIPOGARANTIA",
        "PORCGARANTIA", "MODELO", "PI", "PDI", "PE", "DETERIOROCAPITAL", "DETERIOROINTERESES",
        "DETERIOROPAGOASOCIADOS", "SALDOPENDIENTE", "PLAZOALICUOTA", "VALORALICUOTA", "DETERIOROACUM"
    ]
    ws.append(columnas_legibles)
    registros = []
    mis_rpki = RPKI.objects.filter(oficina_id = oficina_id,fecha = fecha)
    for mi_rpki in mis_rpki:
        if mi_rpki.sal_cap_fin == 0:
            continue
        mi_cred = CREDITOS.objects.filter(oficina_id = oficina_id,cod_cre = mi_rpki.cod_cre).first()
        mi_car_his = CARTE_CAT_HIS.objects.filter(oficina_id = oficina_id,cod_cre = mi_rpki.cod_cre,fecha = fecha).first()
        regis = {
            'TipoIden': mi_cred.socio.tercero.cla_doc,
            'NUMIDENT': mi_cred.socio.tercero.doc_ide,
            'NUMCREDITO' : mi_cred.cod_cre,
            'CodigoContable' : 0,
            'SaldoCapital' : mi_rpki.sal_cap_fin,
            'SaldoIntereses' : mi_rpki.final,	
            'PagoxAsociados' : 0,
            'APORTES' : mi_car_his.aporte if mi_car_his else 0,
            'AHORROPERMANENTE'  : 0,	
            'VEA' : mi_car_his.vea if mi_car_his else 0,
            'VALORGARANTIA' : mi_car_his.val_gar_hip if mi_car_his else 0,
            'TIPOGARANTIA' : mi_car_his.cla_gar if mi_car_his else 0,
            'PORCGARANTIA' : 0,
            'MODELO' : 0,
            'PI' : mi_car_his.pro_inc if mi_car_his else 0,
            'PDI' : mi_car_his.pdi if mi_car_his else 0,	
            'PE' : mi_car_his.per_esp if mi_car_his else 0,
            'DETERIOROCAPITAL' : mi_car_his.sal_cap_pe if mi_car_his else 0,	
            'DETERIOROINTERESES' : mi_car_his.sal_int_pe if mi_car_his else 0,
            'DETERIOROPAGOASOCIADOS' : 0,
            'SALDOPENDIENTE' : 0,
            'PLAZOALICUOTA'	: 0,
            'VALORALICUOTA'	: 0,
            'DETERIOROACUM' : mi_car_his.gas_pro_ind_acu if mi_car_his else 0,
        }
        registros.append(regis)
    campos = list(registros[0].keys())  # Obtener encabezados dinámicos
    ws.append(campos)
    for fila in registros:
        ws.append([fila[campo] for campo in campos])
    
    # 🟢 Preparar archivo para descarga
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"deterioro_de_cartera_{fecha}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response
