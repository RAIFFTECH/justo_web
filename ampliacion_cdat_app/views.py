import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import CTA_CDAT_AMP

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = CTA_CDAT_AMP
    form = CrearForm
    template_name = 'lista_ampliacion_cdats.html'
    # ordering = ['cliente','per_con','cod_cta']

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CTA_CDAT_AMP
    form = CrearForm
    template_name = 'detalles_ampliacion_cdat.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CTA_CDAT_AMP
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_ampliacion_cdat.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_ampliacion_cdat')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CTA_CDAT_AMP
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_ampliacion_cdat.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_ampliacion_cdat')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = CTA_CDAT_AMP
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_ampliacion_cdat')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        ampliacion_cdats = CTA_CDAT_AMP.objects.all().order_by('cliente','per_con','cod_cta')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="ampliacion_cdats.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in ampliacion_cdats:
            # p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
            # p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
            # p.drawString(80, 760, f"Fecha: {dato.fecha}")
            # p.drawString(80, 740, f"Número Liquidación: {dato.num_liq}")
            # p.drawString(80, 720, f"Valor: {dato.valor}")
            # p.drawString(80, 700, f"Cuenta de Ahorro Afectada: {dato.cta_aho_afe}")
            # p.drawString(80, 680, f"Número Documento: {dato.docto}")
            # p.drawString(80, 660, f"Clase: {dato.clase}")
            # p.drawString(80, 640, f"Documento: {dato.documento}")
            # p.drawString(80, 620, f"Aplicado?: {dato.aplicado}")
        

        datos_tabla = [["Cuenta de Ahorro", "Cuenta Ampliación","Fecha","Número Liquidación","Valor","Cuenta de Ahorro Afectada","Número Documento","Clase","Documento","Aplicado?"]]

        for dato in ampliacion_cdats:
            datos_tabla.append([dato.cta_aho,dato.cta_amp,dato.fecha,dato.num_liq,dato.valor,dato.cta_aho_afe,dato.docto,dato.clase,dato.documento,dato.aplicado])
 
        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CTA_CDAT_AMP.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="ampliacion_cdats.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
        p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
        p.drawString(80, 760, f"Fecha: {dato.fecha}")
        p.drawString(80, 740, f"Número Liquidación: {dato.num_liq}")
        p.drawString(80, 720, f"Valor: {dato.valor}")
        p.drawString(80, 700, f"Cuenta de Ahorro Afectada: {dato.cta_aho_afe}")
        p.drawString(80, 680, f"Número Documento: {dato.docto}")
        p.drawString(80, 660, f"Clase: {dato.clase}")
        p.drawString(80, 640, f"Documento: {dato.documento}")
        p.drawString(80, 620, f"Aplicado?: {dato.aplicado}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        ampliacion_cdats = CTA_CDAT_AMP.objects.all()

        p = canvas.Canvas(response)

        for dato in ampliacion_cdats:
            p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
            p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
            p.drawString(80, 760, f"Fecha: {dato.fecha}")
            p.drawString(80, 740, f"Número Liquidación: {dato.num_liq}")
            p.drawString(80, 720, f"Valor: {dato.valor}")
            p.drawString(80, 700, f"Cuenta de Ahorro Afectada: {dato.cta_aho_afe}")
            p.drawString(80, 680, f"Número Documento: {dato.docto}")
            p.drawString(80, 660, f"Clase: {dato.clase}")
            p.drawString(80, 640, f"Documento: {dato.documento}")
            p.drawString(80, 620, f"Aplicado?: {dato.aplicado}")
        
            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CTA_CDAT_AMP._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        ampliacion_cdats = CTA_CDAT_AMP.objects.all()
        for row_num, data in enumerate(ampliacion_cdats, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        ampliacion_cdats = CTA_CDAT_AMP.objects.all()
        for data in ampliacion_cdats:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
