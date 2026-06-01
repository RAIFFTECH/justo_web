import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django import forms
from .forms import CrearForm
from .models import UserProfile


from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import User
from .models import UserProfile

def get_user_photo(request, user_id):
    user_profile = get_object_or_404(UserProfile, user__id=user_id)
    photo_url = user_profile.photo.url if user_profile.photo else None
    return JsonResponse({'photo_url': photo_url})


class Lista(LoginRequiredMixin, ListView):
    model = UserProfile
    form = CrearForm
    template_name = 'lista_usuarios.html'

class Detalles(LoginRequiredMixin, DetailView):
    model = UserProfile
    form = CrearForm
    template_name = 'detalles_usuario.html'

class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = UserProfile
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_usuario.html'
    success_message = 'Registro añadido correctamente.'
    def get_success_url(self):
        return reverse('listar_usuarios')

class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = UserProfile
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_usuario.html'
    success_message = 'Registro actualizado correctamente.'
    def get_success_url(self):
        return reverse('listar_usuarios')


class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = UserProfile
    form = CrearForm
    fields = "__all__"
    def get_success_url(self):
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_usuarios')


class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        usuarios = UserProfile.objects.all().order_by('oficina')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="usuarios.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)
        datos_tabla = [["Oficina", "Nombre Usuario",
                        "Identificación", "Nombre Completo", "Fecha Ingreso", "Es Cajero?", "Código Cajero", "Fecha Saldo", "Cuenta Contable", "Está Activo?"]]

        for dato in usuarios:
            datos_tabla.append([dato.oficina, dato.login, dato.nit, dato.nombre, dato.fec_ing, dato.es_cajero, dato.cod_caj, dato.fec_sal, dato.cta_con_acr, dato.activo])

            # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = UserProfile.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="usuarios.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Línea de Ahorro: {dato.lin_aho}")
        p.drawString(80, 780, f"Fecha Inicial: {dato.fecha_inicial}")
        p.drawString(80, 760, f"Fecha Final: {dato.fecha_final}")
        p.drawString(80, 740, f"Tasa Interés Anual Efectiva: {dato.tiae}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        usuarios = UserProfile.objects.all()

        p = canvas.Canvas(response)

        for dato in usuarios:
            p.drawString(80, 800, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 780, f"Fecha Inicial: {dato.fecha_inicial}")
            p.drawString(80, 760, f"Fecha Final: {dato.fecha_final}")
            p.drawString(80, 740, f"Tasa Interés Anual Efectiva: {dato.tiae}")
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in UserProfile._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        usuarios = UserProfile.objects.all()
        for row_num, data in enumerate(usuarios, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        usuarios = UserProfile.objects.all()
        for data in usuarios:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])
        return response    

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from .models import UserProfile
from .forms import UserProfileForm

# Listar Perfiles de Usuario
@login_required
def profile_list(request):
    profiles = UserProfile.objects.all()
    return render(request, 'profile_list.html', {'profiles': profiles})

# Crear Perfil de Usuario
@login_required
def profile_create(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('profile_list')
    else:
        form = UserProfileForm()
    return render(request, 'profile_form.html', {'form': form})

# Editar Perfil de Usuario
@login_required
def profile_update(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            return redirect('profile_list')
    else:
        form = UserProfileForm(instance=profile)
    return render(request, 'profile_form.html', {'form': form})

# Eliminar Perfil de Usuario
@login_required
def profile_delete(request, pk):
    profile = get_object_or_404(UserProfile, pk=pk)
    if request.method == 'POST':
        profile.delete()
        return redirect('profile_list')
    return render(request, 'profile_confirm_delete.html', {'profile': profile})

