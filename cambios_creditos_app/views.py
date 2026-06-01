import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import CAMBIOS_CRE

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = CAMBIOS_CRE
    form = CrearForm
    template_name = 'lista_cambios_creditos.html'
    ordering = ['det_pro','tip_cam']


# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CAMBIOS_CRE
    form = CrearForm
    template_name = 'detalles_cambios_credito.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CAMBIOS_CRE
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_cambios_credito.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_cambios_creditos')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CAMBIOS_CRE
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_cambios_credito.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_cambios_creditos')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = CAMBIOS_CRE
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_cambios_creditos')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cambios_creditos = CAMBIOS_CRE.objects.all().order_by(
            'det_pro', 'tip_cam')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cambios_creditos.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in cambios_creditos:
        # p.drawString(80, 800, f"Detalle Producto: {dato.det_pro}")
        # p.drawString(80, 780, f"Tipo de Cambio?: {dato.tip_cam}")
        # p.drawString(80, 760, f"Capital: {dato.capital}")
        # p.drawString(80, 740, f"Interés Corriente: {dato.int_cor}")
        # p.drawString(80, 720, f"Interés de Mora: {dato.int_mor}")
        # p.drawString(80, 700, f"Póliza: {dato.pol_seg}")
        # p.drawString(80, 680, f"Descuento Pronto Pago: {dato.des_pp}")
        # p.drawString(80, 660, f"Acreedor: {dato.acreedor}")

        datos_tabla = [["Detalle Producto", "Tipo de Cambio?", "Capital", "Interés Corriente", "Interés de Mora","Póliza", "Descuento Pronto Pago", "Acreedor"]]

        for dato in cambios_creditos:
            datos_tabla.append([dato.det_pro, dato.tip_cam, dato.capital, dato.int_cor, dato.int_mor,
                               dato.pol_seg, dato.des_pp, dato.acreedor])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CAMBIOS_CRE.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cambios_creditos.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cambios_creditos = CAMBIOS_CRE.objects.all()

        p = canvas.Canvas(response)

        for dato in cambios_creditos:
            p.drawString(80, 800, f"Detalle Producto: {dato.det_pro}")
            p.drawString(80, 780, f"Tipo de Cambio?: {dato.tip_cam}")
            p.drawString(80, 760, f"Capital: {dato.capital}")
            p.drawString(80, 740, f"Interés Corriente: {dato.int_cor}")
            p.drawString(80, 720, f"Interés de Mora: {dato.int_mor}")
            p.drawString(80, 700, f"Póliza: {dato.pol_seg}")
            p.drawString(80, 680, f"Descuento Pronto Pago: {dato.des_pp}")
            p.drawString(80, 660, f"Acreedor: {dato.acreedor}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CAMBIOS_CRE._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cambios_creditos = CAMBIOS_CRE.objects.all()
        for row_num, data in enumerate(cambios_creditos, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cambios_creditos = CAMBIOS_CRE.objects.all()
        for data in cambios_creditos:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
