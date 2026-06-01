import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django import forms
from .forms import CrearForm
from .models import CLIENTES
from creditos_app.models import CREDITOS


# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = CLIENTES
    form_class = CrearForm
    template_name = 'lista_clientes.html'
    ordering = ['codigo']


# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CLIENTES
    form_class = CrearForm
    template_name = 'detalles_cliente.html'


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CLIENTES
    form_class = CrearForm
    fields = '__all__'
    template_name = 'crear_cliente.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_clientes')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CLIENTES
    form_class = CrearForm
    fields = '__all__'
    template_name = 'actualizar_cliente.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_clientes')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = CLIENTES
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_clientes')

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cliente = CLIENTES.objects.all().order_by('codigo')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cliente.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in cliente:
        # p.drawString(80, 800, f"Código: {dato.codigo}")
        # p.drawString(80, 780, f"Nit: {dato.doc_ide}")
        # p.drawString(80, 760, f"DV: {dato.dv}")
        # p.drawString(80, 740, f"Sigla: {dato.sigla}")
        # p.drawString(80, 720, f"Nombre: {dato.nombre}") 
        # p.drawString(80, 700, f"Clase Cooperativa: {dato.clase_coop}")
        # p.drawString(80, 680, f"Dirección: {dato.direccion}")
        # p.drawString(80, 660, f"Teléfono: {dato.telefono}")
        # p.drawString(80, 640, f"Celular: {dato.celular}")
        # p.drawString(80, 620, f"Ciudad: {dato.ciudad}")
        # p.drawString(80, 600, f"E-mail: {dato.email}")
        # p.drawString(80, 580, f"Dominio: {dato.dominio}")
        # p.drawString(80, 560, f"Documento Gerente: {dato.nit_ger}")
        # p.drawString(80, 540, f"Nombre Gerente: {dato.nom_ger}")
        # p.drawString(80, 520, f"Documento Contador: {dato.nit_con}")
        # p.drawString(80, 500, f"Nombre Contador: {dato.nom_con}")
        # p.drawString(80, 480, f"Tar. Prof. Contador: {dato.tp_con}")
        # p.drawString(80, 460, f"Documento Revisor Fiscal: {dato.nit_rev_fis}")
        # p.drawString(80, 440, f"Nombre Revisor Fiscal: {dato.nom_rev_fis}")
        # p.drawString(80, 420, f"Tar. Prof. Revisor: {dato.tp_rev_fis}")
        # p.drawString(80, 400, f"Agente Retención?: {dato.age_ret}")
        # p.drawString(80, 380, f"Retiene Iva?: {dato.ret_iva}")
        # p.drawString(80, 360, f"Autorretenedor?: {dato.aut_ret}")
        # p.drawString(80, 340, f"Logo: {dato.logo}")
        # p.drawString(80, 320, f"Número Licencia: {dato.num_lic}")
        # p.drawString(80, 300, f"Licencia Activa?: {dato.lic_act}")
        # p.drawString(80, 280, f"Fecha Incio Licencia: {dato.ini_lic}")
        # p.drawString(80, 260, f"Fecha Fin Licencia: {dato.fin_lic}")

        datos_tabla = [["codigo","doc_ide","dv","sigla","nombre","clase_coop","direccion","telefono","celular","ciudad","email","dominio","nit_ger","nom_ger","nit_con","nom_con","tp_con","nit_rev_fis","nom_rev_fis","tp_rev_fis","age_ret","ret_iva","aut_ret","logo","num_lic","lic_act","ini_lic","fin_lic"]]

        for dato in cliente:
            datos_tabla.append([dato.codigo, dato.doc_ide, dato.dv, dato.sigla, dato.nombre, dato.clase_coop, dato.direccion, dato.telefono, dato.celular, dato.ciudad, dato.email, dato.dominio, dato.nit_ger, dato.nom_ger, dato.nit_con, dato.nom_con, dato.tp_con, dato.nit_rev_fis, dato.nom_rev_fis, dato.tp_rev_fis, dato.age_ret, dato.ret_iva, dato.aut_ret, dato.logo, dato.num_lic, dato.lic_act, dato.ini_lic, dato.fin_lic])   
                                                 
        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CLIENTES.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cliente.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Código: {dato.codigo}")
        p.drawString(80, 780, f"Nit: {dato.doc_ide}")
        p.drawString(80, 760, f"DV: {dato.dv}")
        p.drawString(80, 740, f"Sigla: {dato.sigla}")
        p.drawString(80, 720, f"Nombre: {dato.nombre}") 
        p.drawString(80, 700, f"Clase Cooperativa: {dato.clase_coop}")
        p.drawString(80, 680, f"Dirección: {dato.direccion}")
        p.drawString(80, 660, f"Teléfono: {dato.telefono}")
        p.drawString(80, 640, f"Celular: {dato.celular}")
        p.drawString(80, 620, f"Ciudad: {dato.ciudad}")
        p.drawString(80, 600, f"E-mail: {dato.email}")
        p.drawString(80, 580, f"Dominio: {dato.dominio}")
        p.drawString(80, 560, f"Documento Gerente: {dato.nit_ger}")
        p.drawString(80, 540, f"Nombre Gerente: {dato.nom_ger}")
        p.drawString(80, 520, f"Documento Contador: {dato.nit_con}")
        p.drawString(80, 500, f"Nombre Contador: {dato.nom_con}")
        p.drawString(80, 480, f"Tar. Prof. Contador: {dato.tp_con}")
        p.drawString(80, 460, f"Documento Revisor Fiscal: {dato.nit_rev_fis}")
        p.drawString(80, 440, f"Nombre Revisor Fiscal: {dato.nom_rev_fis}")
        p.drawString(80, 420, f"Tar. Prof. Revisor: {dato.tp_rev_fis}")
        p.drawString(80, 400, f"Agente Retención?: {dato.age_ret}")
        p.drawString(80, 380, f"Retiene Iva?: {dato.ret_iva}")
        p.drawString(80, 360, f"Autorretenedor?: {dato.aut_ret}")
        p.drawString(80, 340, f"Logo: {dato.logo}")
        p.drawString(80, 320, f"Número Licencia: {dato.num_lic}")
        p.drawString(80, 300, f"Licencia Activa?: {dato.lic_act}")
        p.drawString(80, 280, f"Fecha Incio Licencia: {dato.ini_lic}")
        p.drawString(80, 260, f"Fecha Fin Licencia: {dato.fin_lic}")

         # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cliente = CLIENTES.objects.all()

        p = canvas.Canvas(response)

        for dato in cliente:
            p.drawString(80, 800, f"Código: {dato.codigo}")
            p.drawString(80, 780, f"Nit: {dato.doc_ide}")
            p.drawString(80, 760, f"DV: {dato.dv}")
            p.drawString(80, 740, f"Sigla: {dato.sigla}")
            p.drawString(80, 720, f"Nombre: {dato.nombre}") 
            p.drawString(80, 700, f"Clase Cooperativa: {dato.clase_coop}")
            p.drawString(80, 680, f"Dirección: {dato.direccion}")
            p.drawString(80, 660, f"Teléfono: {dato.telefono}")
            p.drawString(80, 640, f"Celular: {dato.celular}")
            p.drawString(80, 620, f"Ciudad: {dato.ciudad}")
            p.drawString(80, 600, f"E-mail: {dato.email}")
            p.drawString(80, 580, f"Dominio: {dato.dominio}")
            p.drawString(80, 560, f"Documento Gerente: {dato.nit_ger}")
            p.drawString(80, 540, f"Nombre Gerente: {dato.nom_ger}")
            p.drawString(80, 520, f"Documento Contador: {dato.nit_con}")
            p.drawString(80, 500, f"Nombre Contador: {dato.nom_con}")
            p.drawString(80, 480, f"Tar. Prof. Contador: {dato.tp_con}")
            p.drawString(80, 460, f"Documento Revisor Fiscal: {dato.nit_rev_fis}")
            p.drawString(80, 440, f"Nombre Revisor Fiscal: {dato.nom_rev_fis}")
            p.drawString(80, 420, f"Tar. Prof. Revisor: {dato.tp_rev_fis}")
            p.drawString(80, 400, f"Agente Retención?: {dato.age_ret}")
            p.drawString(80, 380, f"Retiene Iva?: {dato.ret_iva}")
            p.drawString(80, 360, f"Autorretenedor?: {dato.aut_ret}")
            p.drawString(80, 340, f"Logo: {dato.logo}")
            p.drawString(80, 320, f"Número Licencia: {dato.num_lic}")
            p.drawString(80, 300, f"Licencia Activa?: {dato.lic_act}")
            p.drawString(80, 280, f"Fecha Incio Licencia: {dato.ini_lic}")
            p.drawString(80, 260, f"Fecha Fin Licencia: {dato.fin_lic}")


            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CLIENTES._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cliente = CLIENTES.objects.all()
        for row_num, data in enumerate(cliente, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cliente = CLIENTES.objects.all()
        for data in cliente:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
