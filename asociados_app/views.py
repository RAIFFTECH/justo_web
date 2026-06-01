import csv, re
from io import BytesIO
from django.db import connection
from math import ceil
from datetime import timedelta
from openpyxl import Workbook
from datetime import datetime, date
from django.views import View
from django.db.models import Q, Max, F, Case, When, Value, IntegerField
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.db.models.functions import ExtractMonth
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django_xhtml2pdf.utils import generate_pdf
from django_xhtml2pdf.views import PdfMixin
from django.views.generic import DetailView
from django import forms
from django.core.mail import EmailMessage
from django.conf import settings

from .forms import CrearForm, BeneficiarioForm, ReferenciaForm
from .models import ASOCIADOS, ASO_BENEF, ASO_REFERENCIAS
from .models import TERCEROS
from justo_app.funciones_principales import formato_fecha, formatear_cod_aso
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from creditos_app.models import CREDITOS
from ctas_ahorros_app.models import CTAS_AHORRO
from recla_carte_app.models import CARTE_CAT_HIS
from creditos_app.views import obtener_fecha_desembolso, obtener_fecha_movimiento, lista_creditos_asociado, lista_deudor_solidario
from creditos_app.view_liq_cre import liquidacion_justo
from ctas_ahorros_app.views import obtener_fecha_ctas_ahorros, lista_ahorros_asociado
from aportes_app.models import PLAN_APORTES
from aportes_app.views import saldo_aporte_socio_fecha, calcular_edad, saldo_aportes_fecha, aporte_mensual, saldo_aporte_extra_socio_fecha, saldo_aporte_voluntario_socio_fecha, saldo_revalorizacion_aportes_socio_fecha
from cxc_app.views import lista_cxp
from justo_app.justo_ahorros import obtener_saldos_ctas_ahorros
from justo_app.justo_creditos import Liquida_cre

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = ASOCIADOS
    form = CrearForm
    queryset = ASOCIADOS.objects.all()
    context_object_name = "resultados"
    paginate_by = 10
    template_name = 'lista_asociados.html'
    # ordering = ['cliente','per_con','cod_cta']

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return ASOCIADOS.objects.filter(Q(cod_aso__icontains=query) | Q(tercero__nombre__icontains=query) |
                                            Q(tercero__pri_nom__icontains=query) | Q(tercero__pri_ape__icontains=query)
                                            )
        return ASOCIADOS.objects.all().order_by('estado', 'tercero__nombre')

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = ASOCIADOS
    form = CrearForm
    template_name = 'detalles_asociados.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = ASOCIADOS
    form_class = CrearForm
    template_name = 'crear_asociado.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'
    
    def get_initial(self):
        initial = super().get_initial()
        initial['oficina'] = 1  # Reemplaza '001' con el valor por defecto deseado
        return initial

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_asociados')

class CrearBeneficiarios(LoginRequiredMixin, View):
    template_name = 'beneficiarios.html'

    def get(self, request, pk, *args, **kwargs):
        formset = BeneficiarioForm()
        data = ASO_BENEF.objects.filter(asociado__id=pk)
        return render(request, self.template_name, {'form': formset, 'pk': pk, 'data': data})

    def post(self, request, *args, **kwargs):
        pk = request.POST.get('pk')
        asociado = ASOCIADOS.objects.get(pk=pk)
        i = 0
        errors = []

        while True:
            cla_doc = request.POST.get(f'cla_doc_{i}') if i > 0 else request.POST.get('cla_doc')
            doc_ide = request.POST.get(f'doc_ide_{i}') if i > 0 else request.POST.get('doc_ide')
            nombre = request.POST.get(f'nombre_{i}') if i > 0 else request.POST.get('nombre')
            agno_nac = request.POST.get(f'agno_nac_{i}') if i > 0 else request.POST.get('agno_nac')
            parentesco = request.POST.get(f'parentesco_{i}') if i > 0 else request.POST.get('parentesco')
            porcentaje = request.POST.get(f'porcentaje_{i}') if i > 0 else request.POST.get('porcentaje')

            if not cla_doc:
                break

            try:
                beneficiario = ASO_BENEF(
                    asociado=asociado,
                    cla_doc=cla_doc,
                    doc_ide=doc_ide,
                    nombre=nombre,
                    agno_nac=agno_nac,
                    parentesco=parentesco,
                    porcentaje=porcentaje
                )
                beneficiario.save()

            except Exception as e:
                errors.append(f"Error en el beneficiario {i}: {str(e)}")

            i += 1
        messages.success(request, 'Beneficiarios añadidos correctamente.')
        return redirect('listar_asociados')

class CrearReferencia(LoginRequiredMixin, View):
    template_name = 'referencias.html'

    def get(self, request, pk, *args, **kwargs):
        formset = ReferenciaForm()
        data = ASO_REFERENCIAS.objects.filter(asociado__id=pk)
        return render(request, self.template_name, {'form': formset, 'pk': pk, 'data': data})

    def post(self, request, *args, **kwargs):
        pk = request.POST.get('pk')
        asociado = ASOCIADOS.objects.get(pk=pk)
        i = 0
        errors = []

        while True:
            tipo_ref = request.POST.get(f'tipo_ref_{i}') if i > 0 else request.POST.get('tipo_ref')
            parentesco = request.POST.get(f'parentesco_{i}') if i > 0 else request.POST.get('parentesco')
            nombre = request.POST.get(f'nombre_{i}') if i > 0 else request.POST.get('nombre')
            ocupacion = request.POST.get(f'ocupacion_{i}') if i > 0 else request.POST.get('ocupacion')
            empresa = request.POST.get(f'empresa_{i}') if i > 0 else request.POST.get('empresa')
            direccion = request.POST.get(f'direccion_{i}') if i > 0 else request.POST.get('direccion')
            tel_fijo = request.POST.get(f'tel_fijo_{i}') if i > 0 else request.POST.get('tel_fijo')
            tel_cel = request.POST.get(f'tel_cel_{i}') if i > 0 else request.POST.get('tel_cel')
            tel_emp = request.POST.get(f'tel_emp_{i}') if i > 0 else request.POST.get('tel_emp')
            es_fam_dir_cli = request.POST.get(f'es_fam_dir_cli_{i}') if i > 0 else request.POST.get('es_fam_dir_cli')

            if not tipo_ref:
                break

            try:
                beneficiario = ASO_REFERENCIAS(
                    asociado=asociado,
                    tipo_ref=tipo_ref,
                    parentesco=parentesco,
                    nombre=nombre,
                    ocupacion=ocupacion,
                    empresa=empresa,
                    direccion=direccion,
                    tel_fijo=tel_fijo,
                    tel_cel=tel_cel,
                    tel_emp=tel_emp,
                    es_fam_dir_cli=es_fam_dir_cli
                )
                beneficiario.save()

            except Exception as e:
                errors.append(f"Error en el beneficiario {i}: {str(e)}")

            i += 1
        messages.success(request, 'Referencias añadidas correctamente.')
        return redirect('listar_asociados')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ASOCIADOS
    form_class = CrearForm
    template_name = 'actualizar_asociado.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_asociados')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = ASOCIADOS
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_asociados')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Asociados"

        # Crear los encabezados en la primera fila
        sheet.append([
            "Oficina", "Código Asociado", "Tercero", "Estado en la Entidad", "Sexo",
            "Estado Civil", "Fecha Nacimiento", "Zona", "Profesión", "Ocupación",
            "Código Ocupación", "Estrato", "Nivel de Estudio", "Cabeza de Familia?",
            "Empresa Nominadora", "Fecha Afiliación", "Cargo en la Empresa",
            "Personas a Cargo", "Número Hijos Menores", "Número Hijos Mayores",
            "Tipo de Vivienda", "Tiempo en la Ciudad", "med_con", "Fecha Ingreso Trabajo",
            "Teléfono Trabajo", "Tipo de Salario", "Ciudad de Trabajo", "Actividad Económica",
            "Código CIIU", "Tipo Contrato", "Nombre Empresa", "Nit Empresa", "Dirección Empresa",
            "e-mail Empresa", "Sector Empresa", "Empresa Anterior", "Numero Empleados",
            "Negocio Propio?", "Nombre Negocio", "Teléfono Negocio", "Local Propio?",
            "Tiene Cámara Comercio?", "Antigüedad Negocio", "Entidad de Pensión",
            "Tiene Pensión", "Otra Pensión?", "Entidad Otra Pensión", "Tiene Familiar PEP?",
            "Parentesco Familiar PEP", "Nombre Familiar PEP", "Tiene Cargo Público PEP?",
            "Cargo PEP", "PEP Eje Pod?", "Administra Recursos del Estado PEP?",
            "Tiene Gre Car?", "Recibe Pagos del Extranjero?", "Recibe Extranjero +186?",
            "Recibe Ingresos Extranjeros?", "Estado Anteia", "Nombre Conyuge",
            "Doc_ide Conyuge", "Conyuge Trabaja?", "Ingresos por Honorarios",
            "Ciudad Conyuge", "Ocupacion Conyuge", "Telefono Conyuge",
            "Empresa Conyuge", "Dir Empresa Conyuge", "Cargo Conyuge",
            "Año ing empresas Conyuge", "Barrio Asociado"
        ])

        # Obtener los datos del modelo ASOCIADOS
        asociados = ASOCIADOS.objects.select_related('tercero', 'id_pag', 'ciu_tra').all()

        # Agregar los datos de cada asociado a la hoja de cálculo
        for asociado in asociados:
            sheet.append([
                asociado.oficina.codigo if asociado.oficina else '',
                asociado.cod_aso,
                asociado.tercero.nombre if asociado.tercero else '',
                asociado.estado,
                asociado.sexo,
                asociado.est_civ,
                asociado.fec_nac,
                asociado.zona,
                asociado.profesion,
                asociado.ocupacion,
                asociado.ocupacion_cod,
                asociado.estrato,
                asociado.niv_est,
                asociado.cab_fam,
                asociado.id_pag.nombre if asociado.id_pag else '',
                asociado.fec_afi,
                asociado.cargo_emp,
                asociado.per_a_cargo,
                asociado.num_hij_men,
                asociado.num_hij_may,
                asociado.tip_viv,
                asociado.tie_en_ciu,
                asociado.med_con,
                asociado.fec_ing_tra,
                asociado.tel_tra,
                asociado.tip_sal,
                asociado.ciu_tra.nombre if asociado.ciu_tra else '',
                asociado.act_eco,
                asociado.cod_ciiu,
                asociado.tip_con,
                asociado.nom_emp,
                asociado.nit_emp,
                asociado.dir_emp,
                asociado.email_emp,
                asociado.sector_emp,
                asociado.empresa_ant,
                asociado.emp_num_emp,
                asociado.negocio_pro,
                asociado.negocio_nom,
                asociado.negocio_tel,
                asociado.negocio_loc_pro,
                asociado.negocio_cam_com,
                asociado.negocio_ant,
                asociado.pension_ent,
                asociado.pension_tie,
                asociado.pension_otr,
                asociado.pension_ent_otr,
                asociado.pep_es_fam,
                asociado.pep_fam_par,
                asociado.pep_fam_nom,
                asociado.pep_car_pub,
                asociado.pep_cargo,
                asociado.pep_eje_pod,
                asociado.pep_adm_rec_est,
                asociado.tie_gre_car,
                asociado.recibe_pag_ext,
                asociado.recide_ext_mas_186,
                asociado.recibe_ing_ext,
                asociado.estado_anteia,
                asociado.conyuge_nombre,
                asociado.conyuge_doc_ide,
                asociado.conyuge_trabaja,
                asociado.conyuge_ingresos,
                asociado.conyuge_ciudad,
                asociado.conyuge_ocupacion,
                asociado.conyuge_telefono,
                asociado.conyuge_empresa,
                asociado.conyuge_dir_empresa,
                asociado.conyuge_cargo,
                asociado.conyuge_fec_ing_emp,
                asociado.negocio_pro_local,
                asociado.barrio_aso
            ])

        # Preparar la respuesta HTTP para devolver el archivo Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=asociados.xlsx"

        # Guardar el libro de trabajo en la respuesta
        workbook.save(response)

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = ASOCIADOS.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="asociados.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)
        start = 800

        # Agregamos contenido al PDF utilizando datos de la base de datos
        for field in dato._meta.fields:
            p.drawString(60, start, f"{field.verbose_name}: {getattr(dato, field.name)}")
            start -= 20

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

class ExportPDF(PdfMixin, DetailView):
    model = ASOCIADOS
    template_name = "pdf-asociados.html"
    context_object_name = "asociado"
   
def generar_pdf(request, pk, destino):
    # Crear un buffer en memoria
    buffer = BytesIO()
    
    # Configurar el canvas de reportlab
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Aquí puedes agregar el contenido que deseas al PDF
    p.drawString(100, 750, "Estado de Cuenta")
    p.drawString(100, 730, f"Usuario: {pk}")
    p.drawString(100, 710, "Detalles del saldo y pagos")
    
    # Terminamos de crear la página y guardar el PDF
    p.showPage()
    p.save()
    
    # Mover el puntero de lectura al principio del buffer
    buffer.seek(0)
    
    # Aquí verificamos si se quiere devolver el PDF como una respuesta HTTP
    como_http_response = request.GET.get('como_http_response', 'true').lower() == 'true'
    
    if como_http_response:
        # Si se quiere ver el PDF en el navegador
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="estado_cuenta_{pk}.pdf"'
        return response
    else:
        # Si se quiere enviar el PDF por correo
        email = EmailMessage(
            subject="Estado de Cuenta",
            body="Adjunto encontrarás el estado de cuenta solicitado.",
            from_email=settings.EMAIL_HOST_USER,
            to=[destino],  # Cambiar con el destinatario real
        )
        email.attach('estado_cuenta.pdf', buffer.read(), 'application/pdf')
        email.send()
        
        # Regresamos una respuesta indicando que el correo fue enviado
        return HttpResponse("Correo enviado con éxito.", content_type="text/plain")
    
def estado_cuenta(request,pk):
    como_http_response = request.GET.get('como_http_response', 'true').lower() == 'true'

    if request.method == 'GET':
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        accion = request.GET.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fecha_str = request.GET.get('fecha_corte')
        if not fecha_str:
            return HttpResponse("Debe seleccionar una fecha de corte", status=400)
        try:
            fec_cor = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Fecha de corte inválida", status=400)
    
        asociado = ASOCIADOS.objects.get(id=pk)#  .first()
        cod_aso = asociado.cod_aso
        
        ahorros_encontrados = lista_ahorros_asociado(pk)
        creditos_encontrados = lista_creditos_asociado(pk, fec_cor)
        deudores_encontrados = lista_deudor_solidario(pk, fec_cor)
        cxp_encontradas = lista_cxp(pk, fec_cor)
        
        fecha_formateada = formato_fecha(fec_cor)
    
        # if accion == "imprimir":
        # Lógica para imprimir o generar vista previa  Configuración del PDF
        entidad = CLIENTES.objects.filter(id=id_cli).first()
        oficina = OFICINAS.objects.filter(id=id_ofi).first()
                
        #if como_http_response == True:
        nombre_archivo = f"estado_cuenta_{cod_aso}_{fec_cor}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'

        width, height = letter
        margin_x = 50 
        margin_y = 60        

        # Función para dibujar encabezado
        def dibujar_encabezado():
            p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 20, height - 70, 60, 60)           
            empresa = f"{entidad.nombre.strip()}"
            texto_empresa = stringWidth(empresa, "Times-Roman", 12)
            p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
            p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

            reporte = f"ESTADO DE CUENTA"
            texto_reporte = stringWidth(reporte, "Times-Roman", 12)
            p.drawString((width - texto_reporte) / 2, height - 60, reporte)

            p.setFont("Helvetica", 10)
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de Corte: {fecha_formateada}"
            texto_detalles = stringWidth(detalles, "Helvetica", 10)
            p.drawString((width - texto_detalles) / 2, height - 75, detalles)
            
            datos = "Datos de la Asociada" if asociado.sexo == "F" else "Datos del Asociado"
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, height - 115, datos)
            p.setFillColor(colors.black)
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = height - 115
            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)
            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, datos)
             # Restaurar color por defecto
            p.setFillColor(colors.black)
            
            fecha_afiliacion = formato_fecha(asociado.fec_afi) 
            p.setFont("Helvetica", 11)                      
            filtro_cuentas = f"{asociado.cod_aso.strip()} - {asociado.tercero.nombre.strip()[:64]}"
            p.drawString(margin_x + 20, height - 130, filtro_cuentas)
            afiliado = "Afiliada desde: " if asociado.sexo == "F" else "Afiliado desde: "
            email = f"E-mail: {asociado.tercero.email.strip()}    {afiliado}{fecha_afiliacion}"
            p.drawString(margin_x + 20, height - 145, email)
           
        
        # Función para dibujar pie de página
        def dibujar_pie():
            line_y = 30  # Ajusta para que esté justo encima del texto
            p.line(margin_x-40, line_y, margin_x+552, line_y)  # Dibuja la línea

            p.setFont("Courier", 9)
            texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
            p.drawString(margin_x-20, line_y -10, texto_pie)

            # texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
            # p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)
        
        # === APORTES ===
        def dibujar_aportes(titulo):
            margin_y = height - 170
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += alto_texto
            
        def dibujar_tabla_aportes():
            margin_y = height - 180
            encabezados = ["Tipo Aporte", "            Saldo Total", "Aporte Mensual" ]
            col_widths = [125, 90, 90]
            p.setFont("Helvetica-Bold", 7)
            x = margin_x
            for i, h in enumerate(encabezados):
                p.drawString(x, margin_y, h)
                x += col_widths[i]
            margin_y -= 10
            p.setFont("Helvetica", 7)
            
            extra = saldo_aporte_extra_socio_fecha(id_ofi, cod_aso, fec_cor)
            voluntario = saldo_aporte_voluntario_socio_fecha(id_ofi, cod_aso, fec_cor)
            revalorizacion = saldo_revalorizacion_aportes_socio_fecha(id_ofi, cod_aso, fec_cor)
            ordinario = saldo_aporte_socio_fecha(id_ofi, cod_aso, fec_cor)-revalorizacion
            aportes_mensuales = aporte_mensual(id_ofi, cod_aso, fec_cor)
            # total_aportes = saldo_aporte_socio_fecha(id_ofi, cod_aso, fec_cor)-revalorizacion
            # # for i, item in enumerate(total_aportes):
            # p.drawString(margin_x, margin_y, "APORTE ORDINARIO")
            # p.drawString(170, margin_y, f"{aportes_mensuales:>15,.0f}              {total_aportes:>15,.0f}")
            
            line_height = 10  # altura entre líneas
            if ordinario > 0:
                p.drawString(margin_x, margin_y, "APORTE ORDINARIO")
                # p.drawString(170, margin_y, f"{ordinario:>15,.0f}              {aportes_mensuales:>15,.0f}")
                p.drawRightString(300, margin_y, f"{aportes_mensuales:,.0f}")
                p.drawRightString(235, margin_y, f"{ordinario:,.0f}")
                margin_y -= line_height  # solo baja si se imprimió algo

            if extra > 0:
                p.drawString(margin_x, margin_y, "APORTE EXTRAORDINARIO")
                p.drawString(170, margin_y, f"{extra:>15,.0f}")
                margin_y -= line_height

            if voluntario > 0:
                p.drawString(margin_x, margin_y, "APORTE VOLUNTARIO")
                p.drawString(170, margin_y, f"{voluntario:>15,.0f}")
                margin_y -= line_height

            if revalorizacion > 0:
                p.drawString(margin_x, margin_y, "REVALORIZACIÓN DE APORTES")
                # p.drawString(170, margin_y, f"{revalorizacion:>15,.0f}")
                p.drawRightString(235, margin_y, f"{revalorizacion:,.0f}")
                margin_y -= line_height
            
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual
            
            x = margin_x
            total_aportes = ordinario + extra + voluntario + revalorizacion
            fila_total = [
                "Total Aportes", f"${total_aportes:>15,.0f}"
            ]
            for i, item in enumerate(fila_total):
                p.setFont("Helvetica-Bold", 10) if i == 0 else p.setFont("Helvetica", 10)
                p.drawString(x, margin_y, str(item))
                x += col_widths[i]

            margin_y -= 10  # Si quieres seguir escribiendo más abajo
           
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual
            return total_aportes

        dibujar_aportes("Aportes")
        dibujar_tabla_aportes()

        margin_y -= 10

        # === AHORROS ===
        def dibujar_ahorros(titulo):
            margin_y = height - 270
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += alto_texto

        def dibujar_tabla_ahorros(ahorros_encontrados):
            margin_y = height - 280
            encabezados = ["Línea de Ahorro", "Núm. Cuenta", "Apertura", "    Saldo Total"]
            col_widths = [125, 60, 50, 70, 50]
            p.setFont("Helvetica-Bold", 7)
            x = margin_x
            for i, h in enumerate(encabezados):
                p.drawString(x, margin_y, h)
                x += col_widths[i]
            margin_y -= 10
            p.setFont("Helvetica", 7)

            total_ahorros = 0
            for ahorro in ahorros_encontrados:
                x = margin_x
                resultado = obtener_saldos_ctas_ahorros(id_ofi, fec_cor, ahorro["num_cta"])
                valor = resultado.get(ahorro["num_cta"], {}).get("total_valor", 0)
                if valor is None:
                    valor = 0
                ahorro['saldo'] = valor # if valor is not None else 0
                total_ahorros += ahorro["saldo"]
                valor_formateado = "{:>15,.0f}".format(ahorro["saldo"])
                fila = [
                    ahorro["lin_aho"][:28], ahorro["num_cta"], ahorro["fec_ape"], valor_formateado 
                ]
                #f"{ahorro["saldo"]:>15,.2f}" #, ahorro['fec_ult_mov']
                for i, item in enumerate(fila):
                    p.drawString(x, margin_y, str(item))
                    x += col_widths[i]
                margin_y -= 10
            
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual

            # # Fila de totales
            x = margin_x
            fila_total = [
                "Total Ahorros", f"${total_ahorros:>15,.0f}"
            ]
            for i, item in enumerate(fila_total):
                p.setFont("Helvetica-Bold", 10) if i == 0 else p.setFont("Helvetica", 10)
                p.drawString(x, margin_y, str(item))
                x += col_widths[i]

            margin_y -= 10  # Si quieres seguir escribiendo más 
            
            return total_ahorros

        dibujar_ahorros("Cuentas de Ahorro")
        dibujar_tabla_ahorros(ahorros_encontrados)
        
        margin_y -= 10

        # === CRÉDITOS ===
        def dibujar_creditos(titulo):
            margin_y = height - 370
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += alto_texto     

        # === Tabla de Créditos ===
        def dibujar_tabla_creditos(creditos_encontrados):
            margin_y = height - 380
            encabezados = ["Línea de Crédito", "Número", "Fec. Des.","    Cap. Inicial", "     Vr. Cuota", "  Cuotas", "  Altura", " Pagadas", "       Vr. Al día", "  Saldo Total", "Estado"]
            col_widths = [130, 40, 40, 40, 40, 30, 30, 30, 50, 50, 40]
            p.setFont("Helvetica-Bold", 7)
            x = margin_x
            for i, h in enumerate(encabezados):
                p.drawString(x, margin_y, h)
                x += col_widths[i]
            margin_y -= 10
            p.setFont("Helvetica", 7)
            total_cartera = 0
            total_cuotas_dia = 0
            for cred in creditos_encontrados:
                x = margin_x
                total_cartera += cred["saldo"]
                total_cuotas_dia += cred["val_cuo_dia"]
                x = margin_x
                fila = [
                    cred["lin_cre"][:28], cred["cod_cre"], cred["fec_des"], f"{cred['cap_ini']:>15,.0f}", f"{cred['val_cuo_ini']:>15,.0f}", cred["num_cuo_ini"], cred["altura"],  cred["cuotas_pagadas"], f"{cred['val_cuo_dia']:>15,.0f}",  f"{cred['saldo']:>15,.0f}",cred["estado"]
                ]                    
                for i, item in enumerate(fila):
                    col_width = col_widths[i]
                    valor = str(item)
                    # Vamos a centrar solo si es una de las columnas que tú quieres
                    if i in [4, 5, 6, 7]:  # posiciones de "num_cuo_ini" en tu fila
                        ancho_texto = p.stringWidth(valor, "Helvetica", 7)
                        x_centrado = x + (col_width - ancho_texto) / 2
                        p.drawString(x_centrado, margin_y, valor)
                    else:
                        p.drawString(x, margin_y, valor)
                    x += col_width
            
                margin_y -= 10
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual

            # Fila de totales
            x = margin_x
            fila_total = [
                "Total Créditos", f"${total_cartera:>15,.0f}", "                 Total Cuotas al día ", f"                                  ${total_cuotas_dia:>15,.0f}"
            ]                
            for i, item in enumerate(fila_total):
                p.setFont("Helvetica-Bold", 10) if i == 0 else p.setFont("Helvetica", 10)
                p.drawString(x, margin_y, str(item))
                x += col_widths[i]

            margin_y -= 10  # Si quieres seguir escribiendo más abajo
            
            return total_cartera

        dibujar_creditos("Cartera de Créditos")
        dibujar_tabla_creditos(creditos_encontrados)
       
        # === CUENTAS POR PAGAR ===
        def dibujar_cxp(titulo):
            margin_y = height - 490
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += alto_texto   
        
        def dibujar_tabla_cxp(cxp_encontradas):
            margin_y = height - 500
            encabezados = ["Concepto", "Valor Inicial", "Fec. Des.", "Fec. Ven.", "  Saldo Total"]
            col_widths = [130, 50, 50, 40, 50]
            p.setFont("Helvetica-Bold", 7)
            x = margin_x
            for i, h in enumerate(encabezados):
                p.drawString(x, margin_y, h)
                x += col_widths[i]
            margin_y -= 10
            p.setFont("Helvetica", 7)
            total_cxp = 0
            for cta in cxp_encontradas:
                x = margin_x
                total_cxp += cta["saldo"]
                x = margin_x                
                fila = [
                    cta["concepto"][:28], f"{cta['val_ini']:>15,.0f}", cta["fec_des"], cta["fec_ven"], f"{cta['saldo']:>15,.0f}"
                ]                    
                for i, item in enumerate(fila):
                    col_width = col_widths[i]
                    valor = str(item)
                    # Vamos a centrar solo si es una de las columnas que tú quieres
                    if i in [4, 5, 6, 7]:  # posiciones de "num_cuo_ini" en tu fila
                        ancho_texto = p.stringWidth(valor, "Helvetica", 7)
                        x_centrado = x + (col_width - ancho_texto) / 2
                        p.drawString(x_centrado, margin_y, valor)
                    else:
                        p.drawString(x, margin_y, valor)
                    x += col_width
            
                margin_y -= 10
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual

            # Fila de totales
            x = margin_x
            fila_total = [
                "Total Cuentas por Pagar", f"${total_cxp:>15,.0f}"
            ]                
            for i, item in enumerate(fila_total):
                p.setFont("Helvetica-Bold", 10) if i == 0 else p.setFont("Helvetica", 10)
                p.drawString(x, margin_y, str(item))
                x += col_widths[i]

            margin_y -= 10  # Si quieres seguir escribiendo más abajo
        
            return total_cxp

        dibujar_cxp("Cuentas por Pagar")
        dibujar_tabla_cxp(cxp_encontradas)
       
        margin_y -= 10

        # === DEUDOR SOLIDARIO ===
        def dibujar_deudor(titulo):
            margin_y = height - 610
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += 10
            
        def dibujar_tabla_deudor(deudores_encontrados):
            margin_y = height - 620
            encabezados = ["Nombre Deudor", "Número", "Línea de Crédito", "Fec. Des.", "    Cap. Inicial", "  Cuotas", "    Saldo Total", "Estado"]
            col_widths = [125, 30, 125, 40, 60, 30, 50, 40]
            p.setFont("Helvetica-Bold", 7)
            x = margin_x
            for i, h in enumerate(encabezados):
                p.drawString(x, margin_y, h)
                x += col_widths[i]
            margin_y -= 10
            p.setFont("Helvetica", 7)
            for deudor in deudores_encontrados:                
                x = margin_x
                fila = [
                    deudor["nom_deu"][:28], deudor["cod_cre"], deudor["lin_cre"][:28], deudor["fec_des"], f"{deudor['cap_ini']:>15,.0f}",
                    deudor["num_cuo_ini"], f"{deudor['saldo']:>15,.0f}", deudor["estado"]
                ]
                for i, item in enumerate(fila):
                    col_width = col_widths[i]
                    valor = str(item)
                    # Vamos a centrar solo si es una de las columnas que tú quieres
                    if i in [5, 6]:  # posiciones de "num_cuo_ini" en tu fila
                        ancho_texto = p.stringWidth(valor, "Helvetica", 7)
                        x_centrado = x + (col_width - ancho_texto) / 2
                        p.drawString(x_centrado, margin_y, valor)
                    else:
                        p.drawString(x, margin_y, valor)
                    x += col_width
            
                margin_y -= 10
            # Línea en blanco
            margin_y -= 10  # Espacio extra visual   
        
        dibujar_deudor("Deudor Solidario")
        dibujar_tabla_deudor(deudores_encontrados)
       
        margin_y -= 10

        # === Resumen ===
        p.setFillColor(colors.lightgrey)
        p.rect(margin_x - 2, height - 725, 520, 60, fill=1, stroke=0)
        p.setFillColor(colors.black)
        total_aportes = dibujar_tabla_aportes()
        total_ahorros = dibujar_tabla_ahorros(ahorros_encontrados)
        total_creditos = dibujar_tabla_creditos(creditos_encontrados)
        total_cxp = dibujar_tabla_cxp(cxp_encontradas)

        total_activos = total_aportes + total_ahorros
        total_pasivos = total_creditos + total_cxp
        diferencia = total_activos - total_pasivos
        datos = " de la Asociada" if asociado.sexo == "F" else " del Asociado"
        mensaje = "Saldo a favor" + datos if diferencia > 0 else "Saldo por pagar" + datos
        color_mensaje = colors.green if diferencia > 0 else colors.red

        # Dibuja resumen
        p.setFont("Helvetica-Bold", 11)
        p.drawString(margin_x, height - 680, f"Total Aportes + Ahorros = ${total_activos:,.2f}")
        margin_y -= 15
        p.drawString(margin_x, height - 700, f"Total Créditos + Cuentas por Pagar = ${total_pasivos:,.2f}")
        margin_y -= 15
        texto_diferencia = f"Diferencia = ${abs(diferencia):,.2f} "
        p.drawString(margin_x, height - 720, texto_diferencia)
        ancho_texto = stringWidth(texto_diferencia, "Helvetica-Bold", 11)
        # p.drawString(margin_x, margin_y+430, f"Diferencia: ${abs(diferencia):,.2f}")
            
        p.setFillColor(color_mensaje)
        p.drawString(margin_x + ancho_texto, height - 720, mensaje)
        p.setFillColor(colors.black)
        
        margin_y += 20
       
        dibujar_encabezado()
        dibujar_pie()

        p.showPage()
        p.save()

        if como_http_response == True:
            response.write(buffer.getvalue()) 
            
            return response
        else:
            buffer.seek(0)
            return buffer.read() 
        
    # else:
        # return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)
        
# ---------------------------- End Points ------------------------------

def obtener_socio(request,cod_aso):
    oficina_id = request.session.get('oficina_id')
    if cod_aso:
        try:
            persona = ASOCIADOS.objects.get(cod_aso = cod_aso)
            print('Nombre ',persona)
            return JsonResponse({'nombre': persona.tercero.nombre})
        except ASOCIADOS.DoesNotExist:
            return JsonResponse({'nombre': 'No Existe'})
    return JsonResponse({'nombre': 'formato de dato errado'})

def listado_mensual_asociados(request):
    if request.method == 'GET':
        return render(request, 'reporte_mensual_asociados.html')
    if request.method == 'POST':
        accion = request.POST.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
        # cta_ini = request.POST['cta_ini'] or '1'
        # cta_fin = request.POST['cta_fin'] or '9999999999'
        estado_aso = request.POST['estado_aso']
        estado_aso_nom = {
            'N': 'Nuevos',
            'R': 'Retirados',
            'T': 'Todos'
            }
        estado_nombre = estado_aso_nom.get(estado_aso)
        
        resultado = listado_asociados(fec_ini, fec_fin, estado_aso)
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte_Mensual_Asociados" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = listado_asociados(fec_ini, fec_fin, estado_aso)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "REPORTE MENSUAL DE ASOCIADOS"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"reporte_asociados_{estado_nombre}_{fec_ini}_{fec_fin}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = listado_asociados(fec_ini, fec_fin, estado_aso)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"reporte_asociados_{estado_nombre}_{fec_ini}_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            nombre_archivo = f"reporte_asociados_{estado_nombre}_{fec_ini}_{fec_fin}.pdf"
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="{nombre_archivo}"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # # Función para dibujar subtotales
            # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
            #     x = margin_x - 30
            #     p.setFont("Helvetica-Bold", 9)
            #     p.drawString(x, y, "Subtotal: "+cod_cta)
            #     x += 640  # Posicionar en la columna de débito
            #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
            #     x += 80  # Posicionar en la columna de crédito
            #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "REPORTE MENSUAL DE ASOCIADOS"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Cod_Aso", 60),
                    ("Tip_Ide", 40), 
                    ("Nombre", 210), 
                    ("Estado", 40), 
                    ("Profesión", 150),  
                    ("Fec Afi", 60), 
                    ("Fec Ret", 50), 
                    ("   Saldo Aportes", 70), 
                    ("   Saldo Ahorros", 70), 
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_aso'], 60), 
                    (row['tercero__cla_doc'], 40), 
                    (row['tercero__nombre'][:36], 210), 
                    (row['estado'], 40), 
                    (row['profesion'][:30], 150), 
                    (row['fec_afi'], 60), 
                    (row['fec_ret'], 50),  
                    (f"{row['saldo_aportes']:,.2f}",70,'right'), 
                    (f"{row['saldo_ahorros']:,.2f}", 70,'right'), 
                    ] 

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)

def listado_asociados(fecha_inicial, fecha_final, estado_aso):

    filtros = Q()

    if estado_aso == 'N':  # Asociados Nuevos (Estados 'A' o 'I')
        filtros &= Q(fec_afi__range=(fecha_inicial, fecha_final), estado__in=['A', 'I'])

    elif estado_aso == 'R':  # Asociados Retirados (Estados 'R', 'Z', 'F')
        filtros &= Q(fec_ret__range=(fecha_inicial, fecha_final), estado__in=['R', 'Z', 'F'])

    elif estado_aso == 'T':  # Todos (Activos y Retirados en sus respectivos rangos de fecha)
        filtros &= (
            Q(fec_afi__range=(fecha_inicial, fecha_final), estado__in=['A', 'I']) |
            Q(fec_ret__range=(fecha_inicial, fecha_final), estado__in=['R', 'Z', 'F'])
        )

    # Consultar en la base de datos con los filtros aplicados
    asociados = ASOCIADOS.objects.filter(filtros).select_related('tercero').order_by('estado', 'tercero__nombre')

    # Convertir el QuerySet a una lista de diccionarios
    resultados = list(asociados.values(
        'id',
        'oficina_id',
        'cod_aso',
        'tercero__cla_doc',
        'tercero__nombre',
        'estado',
        'profesion',
        'fec_afi',
        'fec_ret'
    ))

    # saldo_ahorros = obtener_saldos_ctas_ahorros(resultados['oficina_id'], fecha_final) 

    for resultado in resultados:
        resultado['fec_afi'] = resultado['fec_afi'].strftime('%d-%m-%Y') if resultado['fec_afi'] else ''
        resultado['fec_ret'] = resultado['fec_ret'].strftime('%d-%m-%Y') if resultado['fec_ret'] else ''
        saldo_aporte = saldo_aporte_socio_fecha(resultado['oficina_id'], resultado['cod_aso'], fecha_final)  # Llama la función con el código del asociado
        resultado['saldo_aportes'] = saldo_aporte
        # resultado['saldo_aportes'] = 0

        ctas_aho_aso = CTAS_AHORRO.objects.filter(oficina_id = resultado['oficina_id'], asociado_id = resultado['id'])
        saldo_ahorros = obtener_saldos_ctas_ahorros(resultado['oficina_id'], fecha_final)  
        total_saldo = 0
        for cta_aho in ctas_aho_aso:
            saldo = saldo_ahorros.get(cta_aho.num_cta, {}).get('total_valor', 0)
            total_saldo = total_saldo + saldo 

        resultado['saldo_ahorros'] = total_saldo
        # resultado['saldo_ahorros'] = 0

    return resultados

def asociados_super(request):
    if request.method == 'GET':
        return render(request, 'asociados_supersolidaria.html')
    if request.method == 'POST':
        accion = request.POST.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
            
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte_Asociados_Supersolidaria" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = reporte_asociados(id_ofi, fec_fin)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "REPORTE ASOCIADOS SUPERSOLIDARIA"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de corte: {fec_fin}"
            # filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                # [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"reporte_asociados_supersolidaria_{fec_fin}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = reporte_asociados(id_ofi, fec_fin)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"reporte_asociados_supersolidaria_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            nombre_archivo = f"reporte_asociados_supersolidaria_{fec_fin}.pdf"
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="{nombre_archivo}"'
            
            resultado = reporte_asociados(id_ofi, fec_fin)

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # # Función para dibujar subtotales
            # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
            #     x = margin_x - 30
            #     p.setFont("Helvetica-Bold", 9)
            #     p.drawString(x, y, "Subtotal: "+cod_cta)
            #     x += 640  # Posicionar en la columna de débito
            #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
            #     x += 80  # Posicionar en la columna de crédito
            #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "REPORTE ASOCIADOS SUPERSOLIDARIA"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de Corte: {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                # filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
                # texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                # p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Cód_Créd.", 50), 
                    ("Cod_Aso", 50), 
                    ("Nombre", 100), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        
        if accion == "exportar_actividad":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte_Actividad_Asociados" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = reporte_actividad_asociados(id_ofi, fec_fin, id_cli)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "REPORTE ACTIVIDAD DE ASOCIADOS"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de corte: {fec_fin}"
            # filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                # [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"reporte_actividad_asociados_{fec_fin}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv_actividad":
            # Llama a la función para obtener los datos
            resultado = reporte_actividad_asociados(id_ofi, fec_fin, id_cli)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"reporte_actividad_asociados_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir_actividad":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            resultado = reporte_actividad_asociados(id_ofi, fec_fin, id_cli)
            
            nombre_archivo = f"reporte_actividad_asociados_{fec_fin}.pdf"
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="{nombre_archivo}"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # # Función para dibujar subtotales
            # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
            #     x = margin_x - 30
            #     p.setFont("Helvetica-Bold", 9)
            #     p.drawString(x, y, "Subtotal: "+cod_cta)
            #     x += 640  # Posicionar en la columna de débito
            #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
            #     x += 80  # Posicionar en la columna de crédito
            #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "REPORTE ACTIVIDAD DE ASOCIADOS"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de corte: {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                # filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
                # texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                # p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("TD", 10),
                    ("Cod_Aso", 60), 
                    ("Pri_Ape", 80), 
                    ("Seg_Ape", 80), 
                    ("Pri_Nom", 80), 
                    ("Seg_Nom", 80),  
                    ("Fec_Afi", 60),
                    ("Fec_Ult_Apo", 60),
                    ("Fec_Ult_Aho", 60),
                    ("Fec_Ult_Des", 60),
                    ("Fec_Ult_Mov", 60),
                    ("Activo", 30)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cla_doc'], 10), 
                    (row['cod_aso'], 60), 
                    (row['pri_ape'], 80), 
                    (row['seg_ape'], 80),
                    (row['pri_nom'], 80), 
                    (row['seg_nom'], 80), 
                    (row['fec_afi'], 60), 
                    (row['fec_ult_apo'], 60),
                    (row['fec_ult_aho'], 60),
                    (row['fec_ult_des'], 60),
                    (row['fec_ult_mov'], 60),
                    (row['activo'], 30)
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        
        if accion == "exportar_edades":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Reporte_Edades_Asociados" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = reporte_edades_asociados(id_ofi, fec_fin)
           
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "REPORTE EDADES DE ASOCIADOS"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de corte: {fec_fin}"
            # filtro_cuentas = f"Estado de los Asociados: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                # [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"reporte_edades_asociados_{fec_fin}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv_edades":
            # Llama a la función para obtener los datos
            resultado = reporte_edades_asociados(id_ofi, fec_fin)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"reporte_edades_asociados_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir_edades":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            resultado = reporte_edades_asociados(id_ofi, fec_fin)

            nombre_archivo = f"reporte_edades_asociados_{fec_fin}.pdf"
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="{nombre_archivo}"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "REPORTE EDADES DE ASOCIADOS"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de corte: {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Cod_Aso", 60), 
                    ("Pri_Ape", 100), 
                    ("Seg_Ape", 100), 
                    ("Pri_Nom", 100), 
                    ("Seg_Nom", 100),  
                    ("Fec_Nac", 60), 
                    ("Fec_Afi", 60),
                    ("Edad", 60)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_aso'], 60), 
                    (row['pri_ape'], 100), 
                    (row['seg_ape'], 100),
                    (row['pri_nom'], 100), 
                    (row['seg_nom'], 100), 
                    (row['fec_nac'], 60),
                    (row['fec_afi'], 60), 
                    (row['edad'], 60)
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)

def reporte_actividad_asociados(id_oficina, fecha_corte, id_cliente):
    # Consultar en la base de datos con los filtros aplicados
    asociados = ASOCIADOS.objects.filter(
        oficina_id = id_oficina, 
        cod_aso__in=TERCEROS.objects.values_list("doc_ide", flat=True)
        ).filter(
            Q(estado__in=["A", "I"], fec_afi__lte=fecha_corte) |
            Q(~Q(estado="A"), fec_ret__gt=fecha_corte - timedelta(days=31))
            ).select_related('tercero')#.order_by('tercero__pri_ape')
        
    resultados = []
    for asociado in asociados:
        cod_aso = str(asociado.cod_aso) if asociado.cod_aso else ""
        # Formatear cod_aso si cla_doc es "N"
        if asociado.tercero and asociado.tercero.cla_doc == "N":
            cod_aso_formateado = formatear_cod_aso(cod_aso)
            if asociado.tercero.dig_ver:  # Agregar el DV si existe
                cod_aso_formateado += f"-{asociado.tercero.dig_ver}"
        else:
            cod_aso_formateado = cod_aso
                
        # Obtener el resultado de la consulta de fecha último aporte
        fecha_apor = saldo_aportes_fecha(id_oficina, asociado.cod_aso, fecha_corte) 
        
        fecha_aporte = fecha_apor["Fecha"] if fecha_apor else date(2000, 1, 1)
        
        # Obtener el resultado de la consulta de fecha último ahorro
        fecha_ahorro = obtener_fecha_ctas_ahorros(id_oficina, fecha_corte, asociado.cod_aso,id_cliente)
        
        # Obtener el resultado de la consulta de fecha último desembolso
        fecha_desembolso = obtener_fecha_desembolso(id_oficina, fecha_corte, asociado.cod_aso,id_cliente)
        
        # Obtener el resultado de la consulta de fecha último movimiento
        fecha_movimiento = obtener_fecha_movimiento(id_oficina, fecha_corte, asociado.cod_aso,id_cliente) 
        
        activo = (
            ("N" if fecha_aporte and fecha_aporte + timedelta(days=365) < fecha_corte else "S") +
            ("N" if fecha_ahorro and fecha_ahorro + timedelta(days=365) < fecha_corte else "S") +
            ("N" if fecha_desembolso and fecha_desembolso + timedelta(days=365) < fecha_corte else "S")
        )     
                           
        # Construir el diccionario con los datos
        resultados.append({
            "cla_doc": "I" if asociado.tercero and asociado.tercero.cla_doc == "T" else (asociado.tercero.cla_doc if asociado.tercero else ""),
            "cod_aso": cod_aso_formateado,
            "pri_ape": asociado.tercero.pri_ape if asociado.tercero else "",
            "seg_ape": asociado.tercero.seg_ape if asociado.tercero else "",
            "pri_nom": asociado.tercero.pri_nom if asociado.tercero else "",
            "seg_nom": asociado.tercero.seg_nom if asociado.tercero else "",
            "fec_afi": asociado.fec_afi,
            "fec_ult_apo": fecha_aporte,
            "fec_ult_aho": fecha_ahorro,
            "fec_ult_des": fecha_desembolso,
            "fec_ult_mov": fecha_movimiento,
            "activo": activo + ("=Inactivo" if activo == "NNN" else "=Activo")
        })
    return resultados
    
def reporte_edades_asociados(id_oficina, fecha_corte):
    # Consultar en la base de datos con los filtros aplicados
    asociados = ASOCIADOS.objects.filter(oficina_id = id_oficina, cod_aso__in=TERCEROS.objects.exclude(tip_ter="J").values_list("doc_ide", flat=True),fec_nac__month=ExtractMonth(Value(fecha_corte))).select_related('tercero')
    # Filtrar socios con edad 14 o 18
    resultados = [
    {
        "cod_aso": asociado.cod_aso,
        "pri_ape": asociado.tercero.pri_ape if asociado.tercero else "",
        "seg_ape": asociado.tercero.seg_ape if asociado.tercero else "",
        "pri_nom": asociado.tercero.pri_nom if asociado.tercero else "",
        "seg_nom": asociado.tercero.seg_nom if asociado.tercero else "",
        "fec_nac": asociado.fec_nac,
        "fec_afi": asociado.fec_afi,
        "edad": calcular_edad(asociado.fec_nac, fecha_corte)
    }
    for asociado in asociados
    if calcular_edad(asociado.fec_nac, fecha_corte) in [14, 18]
    ]
    return resultados

def reporte_asociados(id_oficina, fecha_corte):
    filtros = Q(
            Q(fec_afi__range=(fecha_corte), estado__in=['A', 'I']) |
            Q(fec_ret__range=(fecha_corte), estado__in=['R', 'Z', 'F'])
        )
    # Consultar en la base de datos con los filtros aplicados
    asociados = ASOCIADOS.objects.filter(oficina_id=id_oficina).select_related('tercero', 'tercero__localidad', 'tercero__estado_fin').order_by('tercero__nombre').annotate(
            fecha_maxima=Max('tercero__estados_fin__fec_inf'),  # Fecha más reciente por tercero
        ).filter(tercero__estados_fin__fec_inf=F('fecha_maxima'))

    # Convertir el QuerySet a una lista de diccionarios
    resultados = list(asociados.values(
        'tercero__cla_doc',
        'cod_aso',
        'tercero__pri_ape',
        'tercero__seg_ape',
        'tercero__pri_nom',
        'tercero__seg_nom',
        'fec_afi',
        'tercero__tel_res',
        'tercero__celular1',
        'tercero__direccion',
        'estado',
        'act_eco',
        'tercero__cod_ciu_res_id__codigo',
        'tercero__email',
        'sexo',
        'emp_ent',
        'tip_con',
        'niv_est',
        'estrato',
        
        'tercero__estados_fin__ing_sal_fij',
        
        # 'niv_ing',
        'fec_nac',
        'est_civ',
        'cab_fam',
        'ocupacion',
        'sector_emp',
        # 'jor_lab',
        'fec_ret',
        # 'asi_ult_asa'
    ))
    
    # 🔹 Diccionarios de mapeo
    mapeo_est_civ = {'S': 1, 'C': 2, 'U': 3, 'E': 4, 'D': 5, 'V': 6}  # Soltero, Casado, Viudo, Divorciado
    mapeo_sexo = {'M': 1, 'F': 2}  # Masculino, Femenino
    mapeo_cab_fam = {'S': 1, 'N': 0}  # Sí, No
    mapeo_emp_ent = {'S': 1, 'N': 0}  # Sí, No   
        
    resultado_final = []

    for a in resultados:
        plan_aporte = PLAN_APORTES.objects.filter(agno=fecha_corte.year).first()
        sal_min = plan_aporte.sal_min if plan_aporte else None
        ingresos = a.get('tercero__estados_fin__ing_sal_fij')
       
        NivIng = ''
        if ingresos > 0 and ingresos < sal_min:
            NivIng = 1
        elif ingresos >= sal_min and ingresos < sal_min * 2:
            NivIng = 2
        elif ingresos >= sal_min * 2 and ingresos < sal_min * 3:
            NivIng = 3
        elif ingresos >= sal_min * 3 and ingresos < sal_min * 4:
            NivIng = 4
        elif ingresos >= sal_min * 4 and ingresos < sal_min * 6:
            NivIng = 5
        elif ingresos >= sal_min * 6 and ingresos < sal_min * 8:
            NivIng = 6
        elif ingresos >= sal_min * 8 and ingresos < sal_min * 11:
            NivIng = 7
        elif ingresos >= sal_min * 11 and ingresos < sal_min * 17:
            NivIng = 8
        elif ingresos >= sal_min * 17 and ingresos < sal_min * 24:
            NivIng = 9
        elif ingresos >= sal_min * 24 and ingresos < sal_min * 48:
            NivIng = 10
        elif ingresos >= sal_min * 48:
            NivIng = 11
       
        nuevo = {
            'Tipo de identificación': a.get('tercero__cla_doc'),
            'Número de identificación': a.get('cod_aso'),
            'Primer apellido': f"{a.get('tercero__pri_ape', '').strip()}",
            'Segundo apellido': f"{a.get('tercero__seg_ape', '').strip()}",           
            'Nombres': f"{a.get('tercero__pri_nom', '').strip()} {a.get('tercero__seg_nom', '').strip()}".strip(),
            'Fecha de ingreso': a.get('fec_afi').strftime('%d/%m/%Y') if a.get('fec_afi') else '',
            'Teléfono': a.get('tercero__tel_res'),
            'Direccion': a.get('tercero__direccion'),
            'Asociado': 1,
            'Activo': a.get('estado'),
            'Actividad económica': a.get('act_eco'),
            'Código municipio': a.get('tercero__cod_ciu_res_id__codigo'),
            'Email': a.get('tercero__email'),
            'Genero': mapeo_sexo.get(a.get('sexo'), 3),
            'Empleado': mapeo_emp_ent.get(a.get('emp_ent'), 0),
            'TipoContrato': a.get('tip_con'),
            'NivelEscolaridad': a.get('niv_est'),
            'Estrato': a.get('estrato'),
            'NivelIngresos': NivIng,
            'FechaNacimiento': a.get('fec_nac').strftime('%d/%m/%Y') if a.get('fec_nac') else '',
            'EstadoCivil': mapeo_est_civ.get(a.get('est_civ'), 0),
            'MujerCabezaFamilia': mapeo_cab_fam.get(a.get('cab_fam'), 0),
            'Ocupacion': a.get('ocupacion'),
            'Sector Económico': a.get('sector_emp'),
            'Jornada Laboral': 1,
            'Fecha de Retiro (ExAsociado)': a.get('fec_ret').strftime('%d/%m/%Y') if a.get('fec_ret') else '',
            'AsistioUltAsamblea': 1,
            'Celular': a.get('tercero__celular1')
        }
        resultado_final.append(nuevo)

    return resultado_final

def enviar_email(request, pk):
   if request.method == 'GET':
        asociado = ASOCIADOS.objects.get(id=pk)#  .first()
        email = asociado.tercero.email
        print('email--->', email)
        enviar_correo_con_pdf(request, pk, email, "prueba", "prueba dos")
        return HttpResponse("Estado de cuenta enviado al correo "+email+" exitosamente.", status=405)

def enviar_correo_con_pdf(request, pk, destinatario, asunto, mensaje):
    print('Entro a enviar email', pk)

    # Obligatorio para controlar el comportamiento interno de estado_cuenta
    request.GET = request.GET.copy()
    request.GET['como_http_response'] = 'false'
    pdf_data = estado_cuenta(request, pk)
    email = EmailMessage(
        subject="Estado de cuenta a la fecha ",
        body="Adjunto PDF generado",
        from_email=settings.EMAIL_HOST_USER,
        to=[destinatario],
    )
    email.attach("estado_cuenta.pdf", pdf_data, "application/pdf")
    email.send()
    return HttpResponse("Correo enviado correctamente.")

from django.db.models import Count, Sum, Max
from django.db.models.functions import ExtractYear, ExtractMonth

def calificacion_socio(request,pk):
    como_http_response = request.GET.get('como_http_response', 'true').lower() == 'true'

    if request.method == 'GET':
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        accion = request.GET.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fecha_str = request.GET.get('fecha_corte')
        if not fecha_str:
            return HttpResponse("Debe seleccionar una fecha de corte", status=400)
        try:
            fec_cor = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse("Fecha de corte inválida", status=400)
    
        asociado = ASOCIADOS.objects.get(id=pk)#  .first()
        cod_aso = asociado.cod_aso
        
        
        fecha_formateada = formato_fecha(fec_cor)
    
        # if accion == "imprimir":
        # Lógica para imprimir o generar vista previa  Configuración del PDF
        entidad = CLIENTES.objects.filter(id=id_cli).first()
        oficina = OFICINAS.objects.filter(id=id_ofi).first()
                
        #if como_http_response == True:
        nombre_archivo = f"estado_cuenta_{cod_aso}_{fec_cor}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'

        width, height = letter
        margin_x = 50 
        margin_y = 60        

        # Función para dibujar encabezado
        def dibujar_encabezado():
            p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 20, height - 70, 60, 60)           
            empresa = f"{entidad.nombre.strip()}"
            texto_empresa = stringWidth(empresa, "Times-Roman", 12)
            p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
            p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

            reporte = f"CALIFICACION DEL ASOCIADO"
            texto_reporte = stringWidth(reporte, "Times-Roman", 12)
            p.drawString((width - texto_reporte) / 2, height - 60, reporte)

            p.setFont("Helvetica", 10)
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Fecha de Corte: {fecha_formateada}"
            texto_detalles = stringWidth(detalles, "Helvetica", 10)
            p.drawString((width - texto_detalles) / 2, height - 75, detalles)
            
            datos = "Datos de la Asociada" if asociado.sexo == "F" else "Datos del Asociado"
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, height - 115, datos)
            p.setFillColor(colors.black)
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = height - 115
            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)
            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, datos)
             # Restaurar color por defecto
            p.setFillColor(colors.black)
            
            fecha_afiliacion = formato_fecha(asociado.fec_afi) 
            p.setFont("Helvetica", 11)                      
            filtro_cuentas = f"{asociado.cod_aso.strip()} - {asociado.tercero.nombre.strip()[:64]}"
            p.drawString(margin_x + 20, height - 130, filtro_cuentas)
            afiliado = "Afiliada desde: " if asociado.sexo == "F" else "Afiliado desde: "

            total_creditos = CREDITOS.objects.filter(
                oficina_id=oficina.id,
                socio__tercero__doc_ide=asociado.tercero.doc_ide
            ).aggregate(total=Count('id'))['total']

            hoy = date.today()
            antiguedad_meses = hoy.year * 12 + hoy.month - asociado.fec_afi.year * 12 - asociado.fec_afi.month
            antiguedad = f"Afiliado desde : {fecha_afiliacion}   Antiguedad en Meses : {str(antiguedad_meses)}     Nro Creditos : {str(total_creditos)} "
            p.drawString(margin_x + 20, height - 145, antiguedad)
        

        def dibujar_calificacion(titulo):
            margin_y = height - 170
            alto_texto = 15
            ancho_texto = 520  # Ajusta según lo que necesites
            y_pos = margin_y

            # Dibujar fondo (puedes ajustar color y tamaño)
            p.setFillColor(colors.HexColor("#cce5ff"))  # Color de fondo azul claro
            p.rect(margin_x - 2, y_pos - 3, ancho_texto, alto_texto, fill=1, stroke=0)

            # Dibujar texto del título encima del fondo
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(colors.HexColor("#003366"))
            p.drawString(margin_x, y_pos, titulo)

            # Restaurar color por defecto
            p.setFillColor(colors.black)
            margin_y += alto_texto    

            hoy = date.today()
            resultados = CARTE_CAT_HIS.objects.filter(
                oficina_id=oficina.id,
                nit=asociado.tercero.doc_ide,
                fecha__year__gt = hoy.year - 3,
            ).annotate(
                anio=ExtractYear('fecha'),
                mes=ExtractMonth('fecha')
            ).values('anio', 'mes').annotate(
                total=Count('id'),
                suma_sal_cap_pe=Sum('sal_cap_pe'),
                max_cat_arr=Max('cat_arr')
            ).order_by('anio', 'mes')

            arreglo_arr = {}

            for fila in resultados:
                anio = fila['anio']
                mes = fila['mes']
                max_cat = fila['max_cat_arr']
                if anio not in arreglo_arr:
                    arreglo_arr[anio] = [None] * 12
                arreglo_arr[anio][mes - 1] = max_cat

            y_pos -= 15
            p.setFont("Courier", 10)  # Fuente monoespaciada

            for anio in sorted(arreglo_arr.keys()):
                valores_mes = arreglo_arr[anio]
                valores_str = '  '.join(
                    f"{v if v is not None else '-':^3}"  # centra cada valor en espacio de 3 caracteres
                    for v in valores_mes
                )
                linea = f"              {anio}:  {valores_str}"
                p.drawString(margin_x, y_pos, linea)
                y_pos -= 15

        
        dibujar_encabezado()
        dibujar_calificacion('CALIFICACION  AÑO     ENE  FEB  MAR  ABR  MAY  JUN  JUL  AGO   SEP  OCT  NOV   DIC')
        
        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response


    return HttpResponse("Método no permitido", status=405)
