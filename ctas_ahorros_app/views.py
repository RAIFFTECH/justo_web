import json, re, os, csv
import pandas as pd
from math import ceil
from openpyxl import Workbook
from django.views import View
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, View, CreateView, UpdateView, DeleteView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.pagesizes import landscape, letter, legal
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from django import forms
from django.db.models import Max, IntegerField, Func, F, Q, Sum, Min, Value, FloatField, Case, When
from django.db.models.functions import Cast, Coalesce
from django.views.decorators.http import require_GET
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from datetime import datetime, date, timedelta
from django.core.paginator import Paginator
from django.db.models import Prefetch

from justo_app.funciones_principales import formato_fecha, formatear_cod_aso
from detalle_economico_app.models import DETALLE_PROD
from terceros_app.models import TERCEROS
from asociados_app.models import ASOCIADOS
from ctas_ahorros_app.models import CTAS_AHORRO,INT_DIA_AHO
from cdat_app.models import CTA_CDAT
from ampliacion_cdat_app.models import CTA_CDAT_AMP
from liquidacion_cdat_app.models import CTA_CDAT_LIQ
from lineas_ahorro_app.models import LINEAS_AHORRO
from tasas_lin_aho_app.models import TAS_LIN_AHO
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from hecho_economico_app.models import HECHO_ECONO
from detalle_producto_app.models import DETALLE_PROD
from detalle_economico_app.models import DETALLE_ECONO
from documentos_app.models import DOCTO_CONTA
from .models import CTAS_AHORRO
from contabilizacion_lineas_ahorros_app.models import IMP_CON_LIN_AHO
from .forms import CtasAhorroForm, CrearForm
from justo_app.justo_ahorros import obtener_saldos_ctas_ahorros

# Asegúrate de tener la función get_max_consecutive_for_line en el mismo archivo o importarla
def get_max_consecutive_for_line(id_lin):   
    lin_aho = LINEAS_AHORRO.objects.filter(id=id_lin).first()
    queryset = CTAS_AHORRO.objects.filter(lin_aho=lin_aho)
    queryset = queryset.annotate(
        consecutive_number=Cast(
            Func(
                F('num_cta'),
                function='SUBSTRING',
                template="SUBSTRING(%(expressions)s FROM 4 FOR 6)",  # Desde el carácter 4 por 6 posiciones
                output_field=IntegerField()
            ),
            output_field=IntegerField()
        )
    )
    max_consecutive = queryset.aggregate(max_consecutive=Max('consecutive_number'))['max_consecutive']+1
    if max_consecutive is None:
        max_consecutive = 0
    max_consecutive_str = str(max_consecutive).zfill(6)
    max_value = f"{lin_aho.cod_lin_aho}-{max_consecutive_str}"
    return max_value

@require_GET
def max_consecutivo_view(request):
    line_code = request.GET.get('lin_aho_id')  # Recibimos lin_aho_id desde la URL
    if not line_code:
        return JsonResponse({'error': 'No se recibió ningún código de línea.'}, status=400)
    
    max_consecutivo = get_max_consecutive_for_line(line_code)
    
    if max_consecutivo is not None:
        response = {
            'id_lin_aho': line_code,
            'max_consecutivo': max_consecutivo
        }
    else:
        response = {
            'id_lin_aho': line_code,
            'error': 'No se encontraron registros para esta línea.'
        }
    
    return JsonResponse(response)

class BuscarCtaAhoView(View):
    def get(self, request):
        num_cta = request.GET.get('num_cta', '')
        nombre = request.GET.get('nombre', '')
        est_cta = request.GET.get('est_cta', None)  # Permitir que sea None
        doc_ide = request.GET.get('doc_ide',None)
        resultados = CTAS_AHORRO.objects.exclude(est_cta='C')
        if num_cta:
            resultados = resultados.filter(num_cta__icontains=num_cta)
        if nombre:
            resultados = resultados.filter(asociado__tercero__nombre__icontains=nombre)
        if doc_ide:
            resultados = resultados.filter(asociado__tercero__doc_ide__icontains=doc_ide)    
        if est_cta:  # Solo filtrar si est_cta no es None
            resultados = resultados.filter(est_cta=est_cta)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = list(resultados.values("num_cta", "asociado__tercero__nombre", "est_cta"))
            return JsonResponse(data, safe=False)

        # Si es una petición normal, devolver HTML
        context = {
            'resultados': resultados,
            'num_cta': num_cta,
            'nombre': nombre,
            'est_cta': est_cta,
        }
        return render(request, 'lista_cta_aho.html', context)

class CtasAhorroListView(ListView):
    model = CTAS_AHORRO
    template_name = 'lista_cta_aho.html'
    context_object_name = 'cuentas'
    paginate_by = 10  # Puedes ajustar el número de resultados por página

    def get_queryset(self):
        query = self.request.GET.get('q')
        object_list = CTAS_AHORRO.objects.all()

        if query:
            object_list = object_list.filter(
                Q(num_cta__icontains=query) | Q(asociado__nombre__icontains=query)
            )

        return object_list
    
class CtasAhorroBaseView(View):
    template_name = 'lista_cta_aho.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cta_aho = self.get_object() if 'pk' in self.kwargs else None
        if cta_aho:  # Si existe un objeto, estamos en modo actualización
            form = self.form_class(instance=cta_aho)
            context['cod_soc'] = cta_aho.asociado.cod_aso if cta_aho.asociado else 'No disponible'
            context['asociado_nombre'] = cta_aho.asociado.tercero.nombre if cta_aho.asociado and cta_aho.asociado.tercero else 'Desconocido'
            context['operation'] = 'update'
            context['button_text'] = 'Modificar'
        else:  # Si no hay objeto, estamos en modo creación
            form = self.form_class()
            context['operation'] = 'create'
            context['button_text'] = 'Guardar'
        context['form'] = form  # Asigna el formulario al contexto
        return context

 #    def get_object(self):
 #        return get_object_or_404(self.model, id=self.kwargs['pk']) if 'pk' in self.kwargs else None


    def validate_models(self, form):
        errors = {}
        fec_apertura = form.cleaned_data.get('fec_apertura')
        if not self.is_valid_date_format(fec_apertura):
            errors['fec_apertura'] = 'La fecha debe estar en formato dd/mm/yyyy.'
        est_cta = form.cleaned_data.get('est_cta')
        cod_aso = form.cleaned_data.get('cod_aso')
        if not cod_aso :
            errors['cod_aso'] = 'Debe Existir el documento del asociado Titular de la Cuenta.'
        if est_cta == 'C':
            fec_cancela = form.cleaned_data.get('fec_cancela')
            if not self.is_valid_date_format(fec_cancela):
                errors['fecha cancelacion'] = 'cuando el estado es Cancelada, debe diligenciarse este campo.'
            else:
                if fec_cancela < fec_apertura:
                    errors['fecha cancelacion'] = 'La fecha de Cancelacion debe ser posterior a la fecha de Apertura.'
        exc_tas_mil = form.cleaned_data.get('exc_tas_mil')
        fec_ini_exc = form.cleaned_data.get('fec_ini_exc')
        if exc_tas_mil == 'S':
            if not self.is_valid_date_format(fec_ini_exc):
                errors['fec_ini_exc'] = 'cuando la cuenta esta excenta debe indicarse la Fecha Exención.'
            else:
                if fec_ini_exc < fec_apertura:
                    errors['fecha cancelacion1'] = 'La fecha inicial de excención debe ser igual o posterior a la fecha de Apertura.'
        lin_aho = form.cleaned_data.get('lin_aho')
        if not lin_aho:
            errors['lin_aho'] = 'Debe seleccionar una linea de ahorro.'
        else:
            # Verifica que el valor seleccionado sea un objeto válido en la base de datos
            try:
                get_object_or_404(LINEAS_AHORRO, pk=lin_aho.id)
            except ValidationError:
                errors['lin_aho'] = 'la Línea de Ahorro no es valida.'
        
        return errors

    def is_valid_date_format(self,date_value):
        # Verificar si date_value es una cadena
        if isinstance(date_value, str):
            try:
                datetime.strptime(date_value, '%d/%m/%Y')
                return True
            except ValueError:
                return False
        # Verificar si date_value es un objeto datetime
        elif isinstance(date_value, (datetime, date)):
            return True
        else:
            return False

    def save_models(self, form):
        print('Importante saber si pasa por aqui')
        try:
            with transaction.atomic():
                # Guarda el modelo principal HechoEcono
                CTAS_AHORRO = form.save(commit=False)
                CTAS_AHORRO.save()
                self.save_or_update_details(CTAS_AHORRO.id, form)
                self.save_or_update_accounting_details(CTAS_AHORRO, form)
        except Exception as e:
            # Maneja cualquier error que pueda ocurrir durante el guardado
            raise e

    def handle_errors(self, errors):
        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        return None
                           
class CtasAhorroCreateView(CtasAhorroBaseView,CreateView):
    model = CTAS_AHORRO
    form_class = CtasAhorroForm
    template_name = 'ctas_ahorro_form.html'
    success_url = reverse_lazy('ctas_ahorro_list')

    def form_valid(self, form):
        lin_aho = form.cleaned_data['lin_aho']
        ultimo_num = CTAS_AHORRO.objects.filter(num_cta__startswith=f"{lin_aho.cod_lin_aho}-").count()
    #    form.instance.num_cta = f"{lin_aho.cod_lin_aho}-{ultimo_num + 1}"
        return super().form_valid(form)
    
    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)  # Esto llama a get_context_data
        return response

    def post(self, request):
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        per_con = self.request.session.get('per_con')
        form = CtasAhorroForm(request.POST)
        cod_aso = request.POST.get('cod_aso')
        asociado = ASOCIADOS.objects.filter(oficina_id = oficina_id,cod_aso = cod_aso).first()
        if asociado:
            form.instance.asociado = asociado  
        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
                else:
                    messages.error(request, 'Errores en el formulario.')
                    return render(request, 'hecho_econo_form.html', {'form': form})
            cta_ahorro = form.save(commit=False)
            cta_ahorro.oficina_id = oficina_id
            cta_ahorro.asociado = asociado 
            cta_ahorro.save()
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'num_cta': cta_ahorro.num_cta,
                    'mensaje': 'La cuenta se ha grabado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
           
            else:
                messages.success(request, 'Cuenta de ahorro guardada correctamente. Num_cta -->'+cta_ahorro.num_cta)
                return redirect('nombre_de_la_vista')  # Redirige a otra vista si no es una solicitud AJAX
        else:
            return JsonResponse({'success': False, 'errors': ['Método no permitido']}, status=405)
        
class CtasAhorroUpdateView(CtasAhorroBaseView,UpdateView):
    model = CTAS_AHORRO
    form_class = CtasAhorroForm
    template_name = 'ctas_ahorro_form.html'
    success_url = reverse_lazy('ctas_ahorro_list')
    
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        return obj

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()  # Asegúrate de asignar el objeto
        print("Valor de fec_apertura:", self.object.fec_apertura)
        context = self.get_context_data(object=self.object)
        context['pk'] = self.kwargs.get('pk')
        if self.object.asociado:
            context['cod_soc'] = self.object.asociado.cod_aso if self.object.asociado else 'No disponible'
            context['asociado_nombre'] = self.object.asociado.tercero.nombre if self.object.asociado.tercero else 'Desconocido'
        else:
            context['cod_soc'] = 'No disponible'
            context['asociado_nombre'] = 'Desconocido'
        return render(request, self.template_name, context)

    def form_valid(self, form):
        lin_aho = form.cleaned_data['lin_aho']
 #        ultimo_num = CTAS_AHORRO.objects.filter(num_cta__startswith=f"{lin_aho.cod_lin_aho}-").count()
 #        form.instance.num_cta = f"{lin_aho.cod_lin_aho}-{ultimo_num + 1}"
        return super().form_valid(form)

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        pk = self.kwargs.get('pk')
        print('Entra a post Update ---->')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')
        per_con = self.request.session.get('per_con')
        form = CtasAhorroForm(request.POST)
        cod_aso = request.POST.get('cod_aso')
        asociado = ASOCIADOS.objects.filter(oficina_id = oficina_id,cod_aso = cod_aso).first()
        if asociado:
            form.instance.asociado = asociado  
        if form.is_valid():
            errors = self.validate_models(form)
            print('errores --->',errors)
            if errors:
                error_messages = [f"{key}: {value}" for key, value in errors.items()]       
                if is_ajax:
                    return JsonResponse({'success': False, 'errors': error_messages}, status=200)
                else:
                    messages.error(request, 'Errores en el formulario.')
                    return render(request, 'hecho_econo_form.html', {'form': form})
            
            cta_ahorro = form.save(commit=False)
            cta_ahorro.id = pk 
            print('cta_ahorro ---> ',cta_ahorro.id)
            print('cta_ahorro ---> ',cta_ahorro.num_cta)
            print('cta_ahorro ---> ',cta_ahorro.fec_ini_exc)

            cta_ahorro.oficina_id = oficina_id
            cta_ahorro.asociado = asociado 
            cta_ahorro.save()
            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'num_cta': cta_ahorro.num_cta,
                    'mensaje': 'La cuenta se ha grabado exitosamente.',
                    'mostrar_boton_imprimier': True, 
                }, status=200)
           
            else:
                messages.success(request, 'Cuenta de ahorro guardada correctamente. Num_cta -->'+cta_ahorro.num_cta)
                return redirect('nombre_de_la_vista')  # Redirige a otra vista si no es una solicitud AJAX

        else:
            return JsonResponse({'success': False, 'errors': ['Método no permitido']}, status=405)

class CtasAhorroDeleteView(View):
    def post(self, request, pk):
        try:
            ctaaho = get_object_or_404(CTAS_AHORRO, pk=pk)
            movs_cta = DETALLE_PROD.objects.filter(hecho_econo__docto_conta__oficina_id=ctaaho.oficina_id,
                producto = 'AH',subcuenta = ctaaho.num_cta).exclude(valor=0)
            if movs_cta.exists():
                messages.error(request, "Error Existen Movimientos en esta cuenta por lo tanto no se puede eliminar ")
                return redirect('listar_ctas_ahorros')
            if ctaaho.est_cta == 'C':
                messages.error(request, "❌ No se puede eliminar una cuenta cancelada.")
                return redirect('listar_ctas_ahorros')
            ctaaho.delete()
            messages.success(request, "✅ La cuenta fue eliminada exitosamente.")
        except Exception as e:
            messages.error(request, f"❌ Error al eliminar: {str(e)}")
        return redirect('listar_ctas_ahorros')

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CTAS_AHORRO
    form = CrearForm
    template_name = 'detalles_cta_aho.html'

class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        cta_ahos = CTAS_AHORRO.objects.all()

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cta_ahos.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        for dato in cta_ahos:
            p.drawString(80, 800, f"Oficina: {dato.oficina}")
            p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
            p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
            p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
            p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
            p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
            p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
            p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")
            
            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CTAS_AHORRO.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cta_ahos.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Oficina: {dato.oficina}")
        p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
        p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
        p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
        p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
        p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
        p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
        p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
        p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cta_ahos = CTAS_AHORRO.objects.all()

        p = canvas.Canvas(response)

        for dato in cta_ahos:
            p.drawString(80, 800, f"Oficina: {dato.oficina}")
            p.drawString(80, 780, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 760, f"Nombre Asociado: {dato.asociado}")
            p.drawString(80, 740, f"Número Cuenta: {dato.num_cta}")
            p.drawString(80, 720, f"Estado Cuenta: {dato.est_cta}")
            p.drawString(80, 700, f"Fecha Apertura: {dato.fec_apertura}")
            p.drawString(80, 680, f"Fecha Cancelación: {dato.fec_cancela}")
            p.drawString(80, 660, f"Exenta 4x1000?: {dato.exc_tas_mil}")
            p.drawString(80, 640, f"Fecha Exención: {dato.fec_ini_exc}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CTAS_AHORRO._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cta_ahos = CTAS_AHORRO.objects.all()
        for row_num, data in enumerate(cta_ahos, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cta_ahos = CTAS_AHORRO.objects.all()
        for data in cta_ahos:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response

def liquidar_ctas_ahorro(request):
    if request.method == 'POST':
        fecha_corte_str = request.POST.get('fecha_corte')
        nombre_archivo = request.POST.get('nombre_archivo')
        liquidar_ahorros_process(fecha_corte_str, nombre_archivo, request)
        try:
            return JsonResponse({'message': 'Cálculos de liquidación iniciados.'})
        except ValueError as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return render(request, 'saldos_cuentas.html')
    
def liquidar_ahorros_process(fecha_corte_str, nombre_archivo, request):
    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    if not os.path.exists(os.path.dirname(nombre_archivo)):
        raise ValueError("El directorio especificado no existe.")
    
    resultados = []
    print('Hora de Inicio ',datetime.now(),'  archivo  ',nombre_archivo)
    result= (
        DETALLE_PROD.objects.filter(
            producto = 'AH',
            hecho_econo__fecha__lte= fecha_corte,
            hecho_econo__docto_conta__oficina_id = 1
        )
        .values('subcuenta')  # Agrupamos por 'subcuenta'
        .annotate(
            cod_cta = F('subcuenta'),
            saldo_fecha = -Sum(F('valor')),  # Negamos la suma como en la SQL
            fec_ult_mov = Max('hecho_econo__fecha'),
            fec_pri_mov = Min('hecho_econo__fecha')
        )
    )
    result_list = list(result)
    nun_ctas = 0
    for item in result_list:
        if item['cod_cta'] is not None or item['fec_ult_mov'] is not None:
            cta_aho = CTAS_AHORRO.objects.filter(num_cta = item['cod_cta'],oficina_id = 1).first()
            asociado = ASOCIADOS.objects.filter(oficina_id = 1,id = cta_aho.asociado_id).first()
            if asociado == None:
                continue
            if asociado.tercero_id == None:
                print('Error de Integridad cod_aso ',asociado.cod_aso)
                continue
            tercero = TERCEROS.objects.filter(id = asociado.tercero_id).first()
            if tercero != None:
                nun_ctas = nun_ctas + 1
                resultados.append({
                    'cod_cta' : item['cod_cta'],
                    'nombre' : tercero.nombre,
                    'doc_ide' : tercero.doc_ide,
                    'saldo_fecha' : item['saldo_fecha'],
                    'fec_pri_mov' : item['fec_pri_mov'],
                    'fec_ult_mov' : item['fec_ult_mov']
                })

    print('Hora de Final  ',datetime.now())    
    df = pd.DataFrame(resultados)
    print('df ya    ',datetime.now())    
    # df.to_excel(nombre_archivo, index=False)
    #df.to_excel(nombre_archivo, index=False, engine='xlsxwriter')
    df.to_excel(nombre_archivo, index=False, engine='openpyxl')
    print(df.head())
    print('a excel  ',datetime.now())    
    return(nun_ctas)

#  ------------------------------------------------------------------------------------------------ 

def ejecutar_consulta_orm(oficina_id,subcuenta,fecha_ini,fecha_fin):
    # Primera parte de la consulta
    xmovs_ant = DETALLE_PROD.objects.filter(
        producto = 'AH', subcuenta = subcuenta,
        hecho_econo__fecha__lt = fecha_ini,
        hecho_econo__docto_conta__oficina_id=oficina_id
    )
    xsal_cor = 0
    xsal_canje = 0
    for xmov_ant  in xmovs_ant:
        if xmov_ant.concepto == 'AHOCH' or xmov_ant.concepto == 'CHCON' or xmov_ant.concepto == 'CHDEV':
            xaho_canje = CANJE_AHORROS.objects.filter(oficina = oficina_id,num_cta = subcuenta,hec_eco_1_id = xmov_ant.hecho_econo_id).first()
            if xmov_ant.concepto == 'AHOCH':
                xsal_canje = xsal_canje - xaho_canje.valor_1
                if xaho_canje == 'P':
                    xsal_cor = xsal_cor - xaho_canje.valor_1   
                else:
                    xsal_cor = xsal_cor - xmov_ant.valor
            elif xmov_ant.concepto == 'CHCON':
                xsal_canje = xsal_canje + xaho_canje.valor_1
            elif xmov_ant.concepto == 'CHDEV':
                xsal_canje = xsal_canje + xaho_canje.valor_1
                xsal_cor = xsal_cor + xaho_canje.valor_1   
        else:
            xsal_cor = xsal_cor - xmov_ant.valor
    resultado_saldo_anterior = {
        'Fecha' : fecha_ini - timedelta(days = 1),
        'Comprobante' : 'Saldo Anterior',
        'Numero' : 0,
        'Concepto': '',
        'Consignacion' : 0,
        'Retiro' : 0,
        'Saldo' : xsal_cor
    }
    
    movimientos = DETALLE_PROD.objects.filter(
        producto = 'AH', subcuenta = subcuenta,
        hecho_econo__fecha__range = (fecha_ini, fecha_fin),
        hecho_econo__docto_conta__oficina_id=oficina_id).order_by('hecho_econo__fecha')
    res_mov = [] 
    for movi in movimientos:
        if movi.concepto == 'AHOCH' or movi.concepto == 'CHCON' or movi.concepto == 'CHDEV':
            if movi.concepto == 'AHOCH':
                xaho_canje = CANJE_AHORROS.objects.filter(oficina = oficina_id,num_cta = subcuenta,hec_eco_1_id = movi.hecho_econo_id).first()
                xsal_cor = xsal_cor - xaho_canje.valor_1
                xsal_canje = xsal_canje - xaho_canje.valor_1
                val_mov = xaho_canje.valor_1
            elif movi.concepto == 'CHCON':
                xaho_canje = CANJE_AHORROS.objects.filter(oficina = oficina_id,num_cta = subcuenta,hec_eco_2_id = movi.hecho_econo_id).first()
                xsal_canje = xsal_canje + xaho_canje.valor_1
                val_mov = 0
            elif movi.concepto == 'CHDEV':
                xaho_canje = CANJE_AHORROS.objects.filter(oficina = oficina_id,num_cta = subcuenta,hec_eco_2_id = movi.hecho_econo_id).first()
                xsal_canje = xsal_canje + xaho_canje.valor_1
                xsal_cor = xsal_cor + xaho_canje.valor_1
                val_mov = xaho_canje.valor_1   
        else:
            xsal_cor = xsal_cor - movi.valor
            val_mov = movi.valor 
        print('Fecha ',movi.hecho_econo.fecha,'  concepto ',movi.concepto,'  saldo canje ',xsal_canje,' saldo ',xsal_cor,'  valmov ',val_mov)
    
        res_mov.append({
            'Fecha' : movi.hecho_econo.fecha,
            'Comprobante' : movi.hecho_econo.docto_conta.nombre,
            'Numero' : movi.hecho_econo.numero,
            'Concepto': movi.concepto,
            'Consignacion' : -val_mov if val_mov < 0 else 0,
            'Retiro' : val_mov if val_mov >= 0 else 0,
            'Saldo' : xsal_cor
        })

    result_int_mes = INT_DIA_AHO.objects.filter(aplicado='N', num_cta=subcuenta,fecha__lte = fecha_fin).aggregate(
        total_int_dia=Coalesce(Sum('int_dia'), Value(0)),
        total_ret_fue=Coalesce(Sum('ret_fue'), Value(0))
    )
    xint_x_abonar = result_int_mes['total_int_dia'] - result_int_mes['total_ret_fue']
    res_int_mes = [] 
    if result_int_mes['total_int_dia'] > 0:
        xsal_cor = xsal_cor + result_int_mes['total_int_dia'] 
        res_int_mes.append({
            'Fecha' : fecha_fin,
            'Comprobante' : 'Interés del Mes',
            'Numero' : 0,
            'Concepto': 'IntMes',
            'Consignacion' : result_int_mes['total_int_dia'],
            'Retiro' : 0,
            'Saldo' : xsal_cor
        } )

    if result_int_mes['total_ret_fue']:
        xsal_cor = xsal_cor + result_int_mes['total_ret_fue']
        res_int_mes.append({
            'Fecha' : fecha_fin,
            'Comprobante' : 'Retencion del Mes',
            'Numero' : 0,
            'Concepto': 'RetFueMes',
            'Consignacion' : 0,
            'Retiro' : result_int_mes['total_ret_fue'],
            'Saldo' : xsal_cor
        }) 
    # Unir las dos partes
    # print('Resultados -anterior -->> ',[resultado_saldo_anterior])
    # print('Resultados -res_mov  -->> ',res_mov)
    # print('Resultados -res_int_mes>> ',res_int_mes)
    if len(res_int_mes) > 0:
        resultados = [resultado_saldo_anterior] + res_mov + res_int_mes
    else:
        resultados = [resultado_saldo_anterior] + res_mov
    return {'resultados': resultados,'saldo_canje': xsal_canje,'int_x_abonar': xint_x_abonar}

def obtener_titular_cta_aho(request, num_cta):
    CtaAho = CTAS_AHORRO.objects.filter(oficina_id = 1,num_cta = num_cta).first()
    if CtaAho:
        nom_titular = CtaAho.asociado.tercero.nombre
        return JsonResponse({"nom_titular": nom_titular})
    return JsonResponse({"nom_titular": ""})  # Devuelve vacío si no encuentra la cuenta

def listar_movtos_cta_ahorro(request, cliente_id=None, oficina_id=None):
    cta_aho = request.GET.get('num_cta', '').strip()
    fecha_ini = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_final', '')
    accion = request.GET.get('accion', '') 
    # Si aún no se ha enviado el formulario con num_cta, solo renderiza vacío con fechas por defecto
    if not cta_aho:
        hoy = datetime.today().date()
        fecha_inicio = datetime(hoy.year - 1, 1, 1).date()
        fecha_final = hoy

        return render(request, 'lista_movtos_cta_aho.html', {
            'num_cta': '',
            'nom_titular': '',
            'context': [],
            'page_obj': [],
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_final': fecha_final.strftime('%Y-%m-%d'),
        })
    
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year,1,1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()

    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))  # Obtiene el valor de oficina_id de los parámetros GET
    resultados = []
    nom_titular = ''
    if len(cta_aho) > 3: 
        CtaAho = CTAS_AHORRO.objects.filter(oficina_id = oficina_id,num_cta = cta_aho).first()
        if CtaAho == None:
            return render(request, 'lista_movtos_cta_aho.html', {
                'page_obj': page_obj,
                'num_cta': cta_aho,
                'fecha_inicio': fecha_inicio_str,
                'fecha_final': fecha_final_str,
                'nom_titular': nom_titular,  # Se agrega al contexto
            })
        else:
            nom_titular = CtaAho.asociado.tercero.nombre
            respuesta = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)
            resultados = respuesta['resultados']
            saldo_canje = respuesta['saldo_canje']
            saldo_x_abonar = respuesta['int_x_abonar']
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        upd_row = row
        sal_acu = row['Saldo']
        rows.append(upd_row)
        print('consigna ',float(row['Consignacion']),'   Retiro ',)

    paginator = Paginator(rows, 10)  # 10 resultados por página
    page_number = request.GET.get('page',1)
    page_obj = paginator.get_page(page_number)

    saldo_formateado = "${:,.2f}".format(sal_acu)
    en_canje_formateado = "${:,.2f}".format(saldo_canje)
    int_por_abonar_formateado = "${:,.2f}".format(saldo_x_abonar)
    print('- Saldo Por Abonar ',int_por_abonar_formateado)

    disponible_formateado = "${:,.2f}".format(sal_acu-saldo_canje-saldo_x_abonar)

    if accion == "exportar":
        return exportar(request, rows)
    elif accion == "csv":
        return exportar_csv(request, rows)
    elif accion == 'imprimir':
        return imprimir(request, rows)

    return render(request, 'lista_movtos_cta_aho.html', {
        'context': page_obj,
        'page_obj': page_obj,
        'num_cta': cta_aho,
        'nom_titular': nom_titular,
        'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
        'fecha_final': fecha_final.strftime('%Y-%m-%d'),
        'saldo': saldo_formateado, 
        'en_canje': en_canje_formateado, 
        'int_x_abonar': int_por_abonar_formateado,
        'disponible': disponible_formateado, 
    })
    
def exportar(request, rows):
    cta_aho = request.GET.get('num_cta', '').strip()
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    
    CtaAho = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = cta_aho).first()
    if CtaAho:
        filtro_socio = CtaAho.asociado.tercero.doc_ide 
    
    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))
  
    # Lógica para exportar Crear el libro de Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimiento Ahorros "+cta_aho 

    entidad = CLIENTES.objects.filter(id=id_cli).first()
    oficina = OFICINAS.objects.filter(id=id_ofi).first()
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()

    # Llama a la función para obtener los datos
    resultados = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)

    if len(filtro_socio) > 3: 
        resultados = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        sal_acu = sal_acu + float(row['Consignacion']) - float(row['Retiro']) + float(row['Saldo'])
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)
                    
    if not resultados:
        return HttpResponse("No se encontraron datos para exportar", status=404)
    
    fecha_inicio_formateada = formato_fecha(fecha_inicio)
    fecha_final_formateada = formato_fecha(fecha_final)
                        
    # Añadir datos adicionales encima de los encabezados
    empresa = f"{entidad.nombre.strip()}"
    nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
    reporte = "MOVIMIENTO CUENTA DE AHORRO "+cta_aho
    asociado = f"Asociado: {filtro_socio} {tercero.nombre}"
    detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fecha_inicio_formateada} a {fecha_final_formateada}"
                        
    datos_adicionales = [
        [empresa],
        [nit_empresa],
        [reporte],
        [asociado],
        [detalles]                
    ]

    # Insertar los datos adicionales
    for row_num, fila in enumerate(datos_adicionales, start=1):
        for col_num, valor in enumerate(fila, start=1):
            sheet.cell(row=row_num, column=col_num, value=valor)

    # Determinar la fila donde empiezan los encabezados
    header_row = len(datos_adicionales) + 1

    # Añadir los encabezados
    headers = resultados[0].keys()
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col_num, value=header)

    # Añadir los datos
    for row_num, data in enumerate(resultados, start=header_row + 1):
        for col_num, field in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=data[field])

    nombre_archivo = f"mov_cta_aho_{cta_aho}_{tercero.nombre.strip()}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    workbook.save(response)
    return response
            
def exportar_csv(request, rows):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    cta_aho = request.GET.get('num_cta', '').strip()
    CtaAho = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = cta_aho).first()
    if CtaAho:
        filtro_socio = CtaAho.asociado.tercero.doc_ide

    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    print('Filtro  ',filtro_socio,'    fecha_ini ',fecha_ini,'   fecha_fin ',fecha_fin)
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))
        
    # Llama a la función para obtener los datos
    resultados = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)

    if len(filtro_socio) > 3: 
        resultados = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)
    else:
        resultados = []
    rows = []
    sal_acu = 0
    for row in resultados:
        sal_acu = sal_acu + float(row['Consignacion']) - float(row['Retiro']) + float(row['Saldo'])
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)

    if not resultados:
        return HttpResponse("No se encontraron datos para exportar", status=404)
    
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()         
    # Configurar la respuesta HTTP para un archivo CSV
    nombre_archivo = f"mov_cta_aho_{cta_aho}_{tercero.nombre.strip()}.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                
    # Crear el escritor CSV
    writer = csv.writer(response)
                
    # Añadir encabezados de las columnas
    headers = resultados[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
    writer.writerow(headers)

    # Añadir los datos
    for fila in resultados:
        writer.writerow([fila[col] for col in headers])
    return response
            
def imprimir(request, rows):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    cta_aho = request.GET.get('num_cta', '').strip()
    CtaAho = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = cta_aho).first()
    if CtaAho:
        filtro_socio = CtaAho.asociado.tercero.doc_ide

    fecha_ini = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_final', None)
    
    oficina_id = request.session.get('oficina_id')
    date_format = "%Y-%m-%d"
    fecha_actual = datetime.now()
    if fecha_ini == None:
        fecha_inicio = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_inicio = datetime.strptime(fecha_ini, date_format).date()
    if fecha_fin == None:
        fecha_final = datetime(fecha_actual.year - 1, 1, 1).date()
    else:
        fecha_final = datetime.strptime(fecha_fin, date_format).date()
    fecha_final = fecha_actual.date()
    
    fecha_inicio_formateada = formato_fecha(fecha_inicio)
    fecha_final_formateada = formato_fecha(fecha_final)    
    fecha_inicio_str = fecha_inicio.strftime('%Y-%m-%d')
    fecha_final_str = fecha_final.strftime('%Y-%m-%d')
    oficina_id = int(request.GET.get('oficina_id', 1))
    
    resultados = ejecutar_consulta_orm(oficina_id,cta_aho,fecha_inicio,fecha_final)
    rows = []
    sal_acu = 0
    rows_resul = resultados['resultados']
    for row in rows_resul:
        print('Con ',row)  
        sal_acu = sal_acu + row['Consignacion'] - row['Retiro'] + row['Saldo']
        upd_row = row
        upd_row['Saldo'] = sal_acu
        rows.append(upd_row)
        
    # Lógica para imprimir o generar vista previa  Configuración del PDF
    entidad = CLIENTES.objects.filter(id=id_cli).first()
    oficina = OFICINAS.objects.filter(id=id_ofi).first()
    tercero = TERCEROS.objects.filter(doc_ide=filtro_socio).first()

    nombre_archivo = f"mov_cta_aho_{cta_aho.strip()}_{tercero.nombre.strip()}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    margin_x, margin_y = 50, 60

    # Configuración inicial
    filas_por_pagina = 64 # Número máximo de filas por página
    total_filas = len(resultados)
    total_paginas = ceil(total_filas / filas_por_pagina)

    # Función para dibujar encabezado
    def dibujar_encabezado():
        empresa = f"{entidad.nombre.strip()}"
        texto_empresa = stringWidth(empresa, "Times-Roman", 12)
        p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                    
        nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
        texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
        p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

        reporte = "MOVIMIENTO CUENTA DE AHORRO "+cta_aho.strip()
        texto_reporte = stringWidth(reporte, "Times-Roman", 12)
        p.drawString((width - texto_reporte) / 2, height - 60, reporte)

        p.setFont("Helvetica", 10)
        detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fecha_inicio_formateada} a {fecha_final_formateada}"
        texto_detalles = stringWidth(detalles, "Helvetica", 10)
        p.drawString((width - texto_detalles) / 2, height - 75, detalles)

        asociado = f"Asociado: {filtro_socio} {tercero.nombre.strip()}"
        texto_filtro = stringWidth(asociado, "Helvetica", 12)
        p.drawString((width - texto_filtro) / 2, height - 90, asociado)

        p.setFont("Helvetica", 8)
        texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
        p.drawRightString(width - margin_x, height - 90, texto_paginas)

        p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 20, height - margin_y - 10, 60, 60)

    # Función para dibujar pie de página
    def dibujar_pie(pagina_actual, total_paginas):
        line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
        p.line(margin_x-40, line_y, margin_x+552, line_y)  # Dibuja la línea

        p.setFont("Courier", 9)
        texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
        p.drawString(margin_x-20, margin_y - 40, texto_pie)

    # Función para dibujar encabezado de tabla
    def dibujar_encabezado_tabla(y):
        line_y = height - 115  # Ajusta para que esté justo encima del texto
        p.line(margin_x-40, line_y, margin_x+552, line_y)
        p.line(margin_x-40, height - 100, margin_x+552, height - 100)
        columnas = [
            ("Fecha", 50), 
            ("Comprobante", 180), 
            ("Número", 50), 
            ("Concepto", 60), 
            ("        Consignación", 75), 
            ("        Retiro", 75), 
            ("        Saldo", 80)
            ]
        x = margin_x-30
        p.setFont("Helvetica-Bold", 9)
        for col, ancho in columnas:
            p.drawString(x, y, col)
            x += ancho
                        
    # Dibujar contenido
    pagina_actual = 1
    dibujar_encabezado()
    dibujar_pie(pagina_actual, total_paginas)
    dibujar_encabezado_tabla(height - margin_y - 50)
    y = height - margin_y - 65

    def dibujar_fila(y, row):
        x = margin_x-30
        columnas = [
            (row['Fecha'], 50), 
            (row['Comprobante'],180), 
            (row['Numero'], 50),
            (row['Concepto'], 50), 
            (f"{row['Consignacion']:,.2f}",70,'right'), 
            (f"{row['Retiro']:,.2f}", 70,'right'), 
            (f"{row['Saldo']:,.2f}", 80, 'right')
            ]

        p.setFont("Helvetica", 9)
        for col in columnas:
            if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                texto, ancho, _ = col
                p.drawRightString(x + ancho, y, texto)
            else:  # Alineación izquierda
                texto, ancho = col[:2]
                p.drawString(x, y, str(texto))
            x += ancho

    for idx, row in enumerate(rows_resul):
        if y < margin_y - 30:
            p.showPage()
            pagina_actual += 1
            dibujar_encabezado()
            dibujar_encabezado_tabla(height - margin_y - 50)
            dibujar_pie(pagina_actual, total_paginas)
            y = height - margin_y - 65    

        dibujar_fila(y, row)
        y -= 10

    p.showPage()
    p.save()
    return response

def obtener_fecha_ctas_ahorros(oficina_id,fecha,subcuenta=None,cliente_id=None):
    fecha_ahorro = DETALLE_PROD.objects.filter(
            producto='AH',
            concepto='AHO',
            hecho_econo__fecha__lte=fecha,
            subcuenta__in=CTAS_AHORRO.objects.filter(
                oficina_id=oficina_id,
                asociado__tercero__doc_ide=subcuenta,
                asociado__tercero__cliente_id=cliente_id
            ).values_list('num_cta', flat=True)
        ).aggregate(max_fecha=Max('hecho_econo__fecha'))['max_fecha']
    
    return fecha_ahorro if fecha_ahorro else date(2000, 1, 1)
    
def lista_ahorros_asociado(id_socio):
    ahorros = CTAS_AHORRO.objects.filter(
        est_cta = "A",
        asociado_id = id_socio
    )
    resultados = [
        {
            'num_cta': ah.num_cta,
            'lin_aho': ah.lin_aho.nombre,
            'fec_ape': ah.fec_apertura
        }
        for ah in ahorros
    ]
    return resultados

def ahorros_super(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    if request.method == 'GET':
        return render(request, 'ahorros_supersolidaria.html') 
    if request.method == 'POST':
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        accion = request.POST.get("accion")   
      
        fecha_corte = request.POST.get('fecha_corte')
    
        saldos = reporte_ahorros(id_cli, id_ofi, fecha_corte)

        if accion == "exportar":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
        
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "AHORROS SUPERSOLIDARIA" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos        
            saldos = reporte_ahorros(id_cli, id_ofi, fecha_corte)                            
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            fecha_corte_formateada = formato_fecha(fecha_corte)
                      
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = f"AHORROS SUPERSOLIDARIA A LA FECHA {fecha_corte_formateada}"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}"
                                
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = saldos[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(saldos, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field])

            nombre_archivo = f"ahorros_supersolidaria_{fecha_corte_formateada}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
           
        elif accion == "csv":
            id_cli = request.session.get('cliente_id')
            id_ofi = request.session.get('oficina_id')
                            
            # Llama a la función para obtener los datos
            saldos = reporte_ahorros(id_cli, id_ofi, fecha_corte)
           
            if not saldos:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"ahorros_supersolidaria_{fecha_corte}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                        
            # Crear el escritor CSV
            writer = csv.writer(response)
                        
            # Añadir encabezados de las columnas
            headers = saldos[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in saldos:
                writer.writerow([fila[col] for col in headers])
            return response
        else:
            return HttpResponse("Acción no permitida", status=405)
                
    return render(request, 'ahorros_supersolidaria.html')

def reporte_ahorros(cliente_id, oficina_id, fecha_corte_str):

    fecha_corte = datetime.strptime(fecha_corte_str, '%Y-%m-%d').date()
    
    ahorros = CTAS_AHORRO.objects.filter(
        oficina_id=oficina_id,
        est_cta__in=["A", "I", "C"],
        fec_apertura__lte=fecha_corte,
        fec_cancela__lte=fecha_corte        
    ).select_related(
        'lin_aho',
        'asociado__tercero'
    ).prefetch_related(
        Prefetch(
            'lin_aho__tas_lin_aho_set',
            queryset=TAS_LIN_AHO.objects.filter(
                fecha_inicial__lte=fecha_corte
            ).filter(
                fecha_final__gte=fecha_corte
            ) | TAS_LIN_AHO.objects.filter(
                fecha_final__isnull=True,
                fecha_inicial__lte=fecha_corte
            ),
            to_attr='tasas_vigentes'
        )
    )
    resultados = []
    resultado = obtener_saldos_ctas_ahorros(oficina_id, fecha_corte)
    
    for ahorro in ahorros:
        asociado = ahorro.asociado
        tercero = asociado.tercero
        linea = ahorro.lin_aho
        num_cta = ahorro.num_cta
        
        consigna = ejecutar_consulta_orm(oficina_id, num_cta, ahorro.fec_apertura, fecha_corte)
        consignaciones = consigna['resultados']
        # Buscar la primera consignación sin modificar la función
        primera_consignacion = next((r['Consignacion'] for r in consignaciones if r['Consignacion'] > 0), None)
        # Tasa vigente más reciente
        tasas = sorted(linea.tasas_vigentes, key=lambda t: t.fecha_inicial, reverse=True)
        tasa_vigente = tasas[0].tiae if tasas else 0
        
        saldo = resultado.get(num_cta, {}).get("total_valor", 0)
        if saldo is None:
            saldo = 0     
            
        if saldo == 0:
            continue
        
        # Formatear cod_aso si cla_doc es "N"
        if tercero.cla_doc == "N":
            cod_aso_formateado = formatear_cod_aso(asociado.cod_aso) #tercero.doc_ide)
            if tercero.dig_ver:  # Agregar el DV si existe
                cod_aso_formateado += f"-{tercero.dig_ver}"
        else:
            cod_aso_formateado = asociado.cod_aso
            
        # Buscar IMP_CON_LIN_AHO por línea y cod_imp
        imp_con = IMP_CON_LIN_AHO.objects.filter(
            linea_ahorro=linea,
            cod_imp=ahorro.cod_imp
        ).first()
            
        # Selección de cuenta contable según estado
        cuenta_contable = ""
        if imp_con:
            if ahorro.est_cta == "A":
                cuenta_contable = imp_con.ctaafeact
            elif ahorro.est_cta == "I":
                cuenta_contable = imp_con.ctaafeina
            elif ahorro.est_cta == "C":
                cuenta_contable = imp_con.ctaafeina
                
        cdat = CTA_CDAT.objects.filter(
            cta_aho=ahorro,
        ).order_by('-fecha').first()

        cdat_amp = None
        tiea = tasa_vigente
        fec_ven = None
        plazo = 0
        int_cau = 0

        if cdat:
            tiea = cdat.tiae or tiea

            cdat_amp = CTA_CDAT_AMP.objects.filter(
                cta_amp=cdat,
            ).order_by('-fecha').first()

            if cdat_amp and cdat_amp.fecha:
                fec_ven = cdat_amp.fecha.strftime('%d/%m/%Y')
                plazo = (cdat_amp.fecha - ahorro.fec_apertura).days
                
            # if cdat_amp:
            #     fec_ven = cdat_amp.fecha
            #     plazo = (fec_ven - ahorro.fec_apertura).days
            
                int_cau = CTA_CDAT_LIQ.objects.filter(
                    cta_amp=cdat_amp,
                    tip_liq__in=['C', 'D'],
                    fecha__year=fecha_corte.year,
                    fecha__month=fecha_corte.month
                ).aggregate(total_int=Sum('val_int'))['total_int'] or 0

        resultados.append({
            'tip_doc': "I" if tercero.cla_doc == "T" else (tercero.cla_doc if tercero.cla_doc else ""),
            'doc_ide': cod_aso_formateado, 
            'cod_con': cuenta_contable[:6],
            'nombre': linea.nombre,
            'termino': linea.termino,
            'amortizacion': 30, 
            'fec_apertura': ahorro.fec_apertura.strftime('%d/%m/%Y'),
            'plazo': plazo,
            'fec_ven': fec_ven,
            'modalidad': 2,
            'tas_int_nom': round(tasa_nominal_anual(tiea, 12),3),
            'tas_int_efe': tiea,
            'int_cau': int_cau,
            'saldo': saldo,
            'dep_ini': 40000 if primera_consignacion is not None and primera_consignacion < 40000 else primera_consignacion or 0, 
            'num_cta': ahorro.num_cta,
            'cta_exc_gmf': 1 if ahorro.exc_tas_mil == "S" else 0,
            'fec_exc': ahorro.fec_ini_exc.strftime('%d/%m/%Y') if ahorro.fec_ini_exc else "",
            'est_cta': 1 if ahorro.est_cta == "A" else 0,
            'cta_bajo_monto': 1,
            'cotitular': 0,
            'conjunta': 0
        })

    return resultados

def tasa_nominal_anual(tea, m):
    if tea < 0 or m <= 0:
        raise ValueError("La Tasa Efectiva Anual debe ser >= 0 y el periodo m debe ser un entero positivo.")
    return (m * ((1 + tea/100) ** (1 / m) - 1))*100


# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from .models import CANJE_AHORROS
from .forms import CanjeAhorrosFiltroForm

def listar_canje_ahorros(request):
    form = CanjeAhorrosFiltroForm(request.GET or None)

    registros = CANJE_AHORROS.objects.select_related(
        'cta_aho__asociado__tercero'
    )

    historico = request.GET.get('historico')
    estado_filtro = request.GET.get('estado', '')

    # Si no es histórico ni hay filtro explícito de estado, mostrar solo 'P'
    if not historico and not estado_filtro:
        registros = registros.filter(estado='P')

    if form.is_valid():
        num_cta = form.cleaned_data.get('num_cta')
        nombre = form.cleaned_data.get('nombre')
        estado = form.cleaned_data.get('estado')

        if num_cta:
            registros = registros.filter(num_cta__icontains=num_cta)

        if nombre:
            registros = registros.filter(
                cta_aho__asociado__tercero__nombre__icontains=nombre
            )

        if estado:
            registros = registros.filter(estado=estado)

    # Paginación
    paginator = Paginator(registros, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    data = []
    for registro in page_obj:
        nombre_usuario = getattr(registro.cta_aho.asociado.tercero, 'nombre', '')
        data.append({
            'registro': registro,
            'nombre': nombre_usuario,
        })

    return render(request, 'movtos_canje.html', {
        'form': form,
        'data': data,
        'page_obj': page_obj,
    })

from django.shortcuts import render, redirect, get_object_or_404
from .forms import CambioEstadoCanjeForm

@csrf_exempt
def canje_editar(request, pk):
    registro = get_object_or_404(CANJE_AHORROS, pk=pk)
    if request.method == 'POST':
        form = CambioEstadoCanjeForm(request.POST)
        if form.is_valid():
            nuevo_estado = form.cleaned_data['nuevo_estado']
            fecha_nueva = form.cleaned_data['fecha_nueva']
            print('Nuevo Estado ',nuevo_estado)
            print('Nueva Fecha  ',fecha_nueva,'  tipo ',type(fecha_nueva))
            if fecha_nueva < registro.fecha_1:
                form.add_error('fecha_nueva', 'La nueva fecha no puede ser menor a la fecha de canje original.')
                return render(request, 'cambiar_estado_canje.html', {'registro': registro,'form': form
            })
            xdoc = DOCTO_CONTA.objects.filter(oficina_id = registro.oficina_id,per_con = fecha_nueva.year,codigo = 7).first()
            hec_eco = HECHO_ECONO.objects.filter(docto_conta = xdoc,numero = registro.hec_eco_1.numero).first()
            if hec_eco == None:
                hec_eco = HECHO_ECONO.objects.create(docto_conta = xdoc,numero = registro.hec_eco_1.numero)
            hec_eco.fecha = fecha_nueva
            hec_eco.descripcion = 'Ajuste de Canje'
            det_pro = DETALLE_PROD.objects.filter(hecho_econo = hec_eco,producto = 'AH',subcuenta = registro.num_cta,valor = 0).first()
            if nuevo_estado == 'C':
                if det_pro == None:
                    det_pro = DETALLE_PROD.objects.create(hecho_econo = hec_eco,producto = 'AH',concepto = 'CHCON',subcuenta = registro.num_cta,valor = 0)
                det_pro.concepto = 'CHCON'
                det_pro.save()
                det_ini = DETALLE_PROD.objects.filter(hecho_econo_id = registro.hec_eco_1_id,producto = 'AH',concepto = 'AHOCH',subcuenta = registro.num_cta).first()
                det_ini.valor = registro.valor_1
                det_ini.save()
                registro.hec_eco_2 = hec_eco
                registro.fecha_2 = fecha_nueva
                registro.valor_2 = registro.valor_1
                registro.estado = 'C'
                registro.aplicado = 'S'
                registro.save()
            if nuevo_estado == 'D':
                if det_pro == None:
                    det_pro = DETALLE_PROD.objects.create(hecho_econo = hec_eco,producto = 'AH',concepto = 'CHDEV',subcuenta = registro.num_cta,valor = 0)
                det_pro.concepto = 'CHDEV'
                det_pro.save()
                det_ini = DETALLE_PROD.objects.filter(hecho_econo_id = registro.hec_eco_1_id,producto = 'AH',subcuenta = registro.num_cta).first()
                det_ini.valor = 0
                det_ini.save()
                registro.hec_eco_2 = hec_eco
                registro.fecha_2 = fecha_nueva
                registro.valor_2 = registro.valor_1
                registro.estado = 'D'
                registro.aplicado = 'S'
                registro.save() 
                dets_eco = DETALLE_ECONO.objects.filter(hecho_econo_id = registro.hec_eco_1_id)
                for det_eco in dets_eco:
                    deco = DETALLE_ECONO.objects.filter(hecho_econo_id = hec_eco.id,cuenta_id = det_eco.cuenta_id,tercero_id = det_eco.tercero_id).first()
                    if deco == None:
                        deco = DETALLE_ECONO.objects.create(hecho_econo_id = hec_eco.id,cuenta_id = det_eco.cuenta_id,tercero_id = det_eco.tercero_id)
                    deco.item_concepto = det_eco.item_concepto
                    deco.descripcion = 'Reversion asiento canje'
                    deco.debito = det_eco.credito
                    deco.credito = det_eco.debito
                    deco.save()
            return redirect('movtos_en_canje')
    else:
        print("🚀 GET: se está cargando la vista inicialmente")
        form = CambioEstadoCanjeForm()  # Aquí sí aplica el initial de fecha_nueva

    return render(request, 'cambiar_estado_canje.html', {
        'registro': registro,
        'form': form
    })

from datetime import date, timedelta
from dateutil.relativedelta import relativedelta 

def riesgo_de_liquidez_ahorros(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    print('id_ofi  ',id_ofi,'liquidez_ahorros')
    if request.method == 'GET':
        return render(request, 'Riesgo_liq_ahorros.html')  # tu formulario
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_corte')
        try:
            fecha_base = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha inválida", status=400)
        accion = request.POST.get('accion')
        fecha_inicio = (fecha_base - relativedelta(months=12)) + timedelta(days=1)
        ahorro = []  # Lista de dicts con claves: 'DIA', 'M01', ..., 'M12'
        for dia in range(1, 32):
            fila = {'DIA': dia}
            for mes in range(1, 13):
                clave_mes = f"M{mes:02}"  # M01, M02, ..., M12
                fila[clave_mes] = 0
            ahorro.append(fila)
        MESES = {}
        MESES = [[0 for _ in range(32)] for _ in range(13)]

        dets_aho = DETALLE_PROD.objects.filter(
            producto='AH',
            hecho_econo__anulado='N',
            hecho_econo__fecha__year__gt=2014,
            hecho_econo__docto_conta__oficina_id=id_ofi,
        ).values(
            'hecho_econo__fecha','subcuenta','concepto','valor'
        )
        xAntMov = 0
        for det_aho in dets_aho:
            fecha = det_aho['hecho_econo__fecha']
            concepto = det_aho['concepto']
            valor = det_aho['valor']
            num_cta = det_aho['subcuenta']
            xvalor = 0
            if num_cta[:2] != '04' and fecha <= fecha_base:
                xvalor = valor
                if concepto == 'AHOCH':
                    canje = CANJE_AHORROS.objects.filter(oficina_id = id_ofi,num_cta=num_cta,fecha_1 = fecha).first()
                    if canje != None:
                        if canje.estado != 'D' or (canje.estado == 'D' and canje.fecha_2 > fecha_base):
                            xvalor = canje.valor_1 if canje.valor_1 != None else 0
                        else:
                            if canje.fecha_2 <= fecha_base:
                                MESES[canje.fecha_2.month][canje.fecha_2.day]+= canje.valor_1
                                xvalor = canje.valor_1 if canje.valor_1 != None else 0
                else:
                    if concepto == 'INTCA':
                        xvalor = 0
                if fecha < fecha_inicio:
                    xAntMov = xAntMov + xvalor
                else:
                    MESES[fecha.month][fecha.day] = (MESES[fecha.month][fecha.day] or 0.0) - xvalor
                    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")
        ws.append(['Accion Riesgo de Aportes ... '+accion+'   de la fecha '+fecha_base.strftime("%Y-%m-%d")])
        ws.append([])
        encabezado = ['Día'] + [f"Mes {m+1:02}" for m in range(12)]
        ws.append(encabezado)
        for dia in range(1, 32):  # Días del 1 al 31
            fila = [f"Día {dia:02}"]
            for mes in range(1, 13):  # Meses del 1 al 12
                fila.append(MESES[mes][dia])
            ws.append(fila)
        ws.append([])
        ws.append([])
        ws.append(['Aporte Anterior ',xAntMov])
        nombre_archivo = f"rl_ahorros_{fecha_base}.xlsx"
        print('Ahora Graba ',nombre_archivo )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(response)
        return response
    

def saldos_diarios_ahorros(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    print('id_ofi  ',id_ofi,'liquidez_ahorros')
    if request.method == 'GET':
        return render(request, 'saldos_diarios.html')  # tu formulario
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_inicial')
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Inicial inválida", status=400)
        fecha_str = request.POST.get('fecha_final')
        try:
            fecha_final = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Final inválida", status=400)
        num_dias = (fecha_final - fecha_inicio).days + 1
        DIAMES = [0] * num_dias
        for i in range(num_dias):
            fecha = fecha_inicio + timedelta(days=i)
        movs_aho = DETALLE_PROD.objects.select_related(
            'hecho_econo__docto_conta'
            ).filter(
            producto='AH',
            hecho_econo__fecha__lte=fecha_final,
            hecho_econo__docto_conta__oficina_id=id_ofi
            ).exclude(
                concepto='INTCA'
            ).values(
                fecha=F('hecho_econo__fecha'),
                num_cta=F('subcuenta'),
                tipo=F('concepto'),
                val_mov=F('valor')
            )
        xAntMov = 0
        for mov_aho in movs_aho:
            xValMov = 0
            if mov_aho['tipo'] == 'CHCON':
                canje = CANJE_AHORROS.objects.filter(oficina_id = id_ofi,num_cta=mov_aho.subcuenta ,fecha_1 = mov_aho.fecha).first()
                if canje != None:
                    xvalmov = -canje.valor_1
            else:
                xvalmov = mov_aho['val_mov']
            if mov_aho['fecha'] < fecha_inicio:
                xAntMov -= mov_aho['val_mov']
            else:
                indice = (mov_aho['fecha'] - fecha_inicio).days
                DIAMES[indice] -= xvalmov 
        SalDia = []

        for i in range(num_dias):
            fecha_actual = fecha_inicio + timedelta(days=i)
            saldo = DIAMES[i]
            SalDia.append({
                'DIA': i + 1,              # Día 1, 2, ..., n
                'fecha': fecha_actual,     # La fecha correspondiente
                'Saldo': saldo             # El valor de DIAMES[i]
            })
        wb = Workbook()
        ws = wb.active
        ws.title = "SalDia"
        ws.append(['DIA', 'Fecha', 'Saldo'])
        for fila in SalDia:
            ws.append([fila['DIA'], fila['fecha'], fila['Saldo']])
        ws.append([])
        ws.append(['Aporte Anterior ',xAntMov])
        nombre_archivo = f"saldos_dia_{fecha_inicio}_{fecha_final}.xlsx"
        print('Ahora Graba ',nombre_archivo )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(response)
        return response

from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce
from datetime import date
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def interes_y_ret_fuente(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    print('id_ofi  ',id_ofi,'liquidez_ahorros')
    if request.method == 'GET':
        return render(request, 'intereses_retenftes.html')  # tu formulario
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_inicial')
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Inicial inválida", status=400)
        fecha_str = request.POST.get('fecha_final')
        try:
            fecha_final = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Final inválida", status=400)
        
        resultado = (
            DETALLE_ECONO.objects
            .filter(
                hecho_econo__protegido='N',
                hecho_econo__fecha__range=[fecha_inicio, fecha_final],
                hecho_econo__docto_conta__oficina_id=id_ofi,
                detalle_prod__hecho_econo=F('hecho_econo'),  # Este filtro simula el AND he.id = dp.hecho_econo_id
                detalle_prod__producto = 'AH'
            )
            .values(        
                'detalle_prod__subcuenta',
                'item_concepto'
            )
            .annotate(
                total=Sum(
                    ExpressionWrapper(
                        F('debito') - F('credito'),
                        output_field=DecimalField()
                    )
                )
            )
        )

        #   cuentas deahooros no cdat
        registros = {}
        for fila in resultado:
            num_cta = fila['detalle_prod__subcuenta']
            concepto = fila['item_concepto']
            total = -fila['total']
            cta_aho = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = num_cta).first()
            clave = (num_cta)
            if clave in registros:
                registros[clave]['PagActInt'] += (total if concepto[:3] == 'Int' else 0) 
                registros[clave]['ret_fue'] += (total if concepto[:3] != 'Int' else 0)
            else:
                registros[clave] = {
                    'NitTer': cta_aho.asociado.tercero.doc_ide,
                    'NomTer': cta_aho.asociado.tercero.nombre,
                    'PagAntInt': 0.0,
                    'PagActInt': total if concepto[:3] == 'Int' else 0.0 ,
                    'CauResMesInt': 0.0,
                    'ret_fue': total if concepto[:3] != 'Int' else 0.0 
                }

        #   Ahora cdat
        resul_cdat = (CTA_CDAT_LIQ.objects.filter(
            cta_aho__oficina_id = id_ofi,
            fecha__range=[fecha_inicio,fecha_final],
            tip_liq='P'
        )
        .annotate(num_cta=F('cta_aho__num_cta'))
        .values('num_cta', 'fecha')
        .annotate(
            interes=Sum('val_int'),
            val_ret=Sum('val_ret')
        )
        .order_by('num_cta', 'fecha')
        )
        for regis in resul_cdat:
            cdat = (CTA_CDAT_LIQ.objects.filter(
                        fecha__range=[fecha_inicio,fecha_final],
                        cta_aho__num_cta=regis['num_cta']
                    )
                    .values('cta_aho__num_cta')
                    .annotate(
                        intcaufin=Sum(Case(When(tip_liq='D', then='val_int'),default=Value(0),output_field=FloatField())),
                        intpagact=Sum(Case(When(tip_liq='C', then='val_int'),default=Value(0),output_field=FloatField()))
                    )   
            )
            regis1 = list(cdat)[0] if cdat else None
            int_pag_act = -regis['interes']
            int_pag_ant = 0
            int_cau_mes = 0
            if regis1:
                int_pag_ant = int_pag_act - regis1['intpagact']
                int_pag_act = int_pag_act - int_pag_ant
                int_cau_mes = regis1['intcaufin'] 
            num_cta_int = CTA_CDAT.objects.filter(cta_aho__num_cta=regis['num_cta'],cta_aho__oficina_id=id_ofi,ampliacion=0).values_list('cta_int_ret', flat=True).first()
            cta_cda = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = regis['num_cta']).first()
            clave = (num_cta_int)
            if clave in registros:
                registros[clave]['PagActInt'] = float(registros[clave]['PagActInt']) + float(regis['interes']) - float(regis['val_ret'])
                #registros[clave]['ret_fue'] += (total if concepto[:3] != 'Int' else 0)
            clave = (cta_cda.num_cta)
            print(F"cta_cda.num_cta  {cta_cda.num_cta}")
            if clave in registros:
                registros[clave]['PagActInt'] -= (int_pag_ant+int_pag_act)
                registros[clave]['ret_fue'] += regis['val_ret'] 
            else:
                registros[clave] = {
                    'num_cta': cta_cda.num_cta,
                    'NitTer': cta_cda.asociado.tercero.doc_ide,
                    'NomTer': cta_cda.asociado.tercero.nombre,
                    'PagAntInt': int_pag_ant,
                    'PagActInt': int_pag_act,
                    'CauResMesInt': int_cau_mes,
                    'ret_fue': -regis['val_ret'] 
                }
        registros = dict(sorted(registros.items(), key=lambda item: item[1]['NitTer']))
        wb = Workbook()
        ws = wb.active
        ws.title = "Inter y Reten"
        encabezado = ['NumCta', 'NitTer', 'NomTer', 'PagAntInt', 'PagActInt', 'CauResMesInt', 'ret_fue']
        ws.append(encabezado)
        for col_num, col_name in enumerate(encabezado, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
        suma_pag_ant_int = 0
        suma_pag_act_int = 0
        suma_cau_res_mes_int = 0
        suma_ret_fue = 0

        for num_cta, datos in registros.items():
            pag_ant = float(datos['PagAntInt'])
            pag_act = float(datos['PagActInt'])
            cau_res = float(datos['CauResMesInt'])
            retfue = float(datos['ret_fue'])
            fila = [
                num_cta,
                datos['NitTer'],
                datos['NomTer'],
                pag_ant,
                pag_act,
                cau_res,
                retfue
            ]
            ws.append(fila)
            suma_pag_ant_int += pag_ant
            suma_pag_act_int += pag_act
            suma_cau_res_mes_int += cau_res
            suma_ret_fue += retfue
        total_fila = [
            'TOTAL', '', '',  # columnas de texto
            suma_pag_ant_int,
            suma_pag_act_int,
            suma_cau_res_mes_int,
            suma_ret_fue
        ]
        ws.append(total_fila)

        total_row_index = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row=total_row_index, column=col)
            cell.font = Font(bold=True)

        for i in range(1, len(encabezado) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        ws.append([])
        nombre_archivo = f"inter_y_ret_{fecha_inicio}_{fecha_final}.xlsx"
        print('Ahora Graba ',nombre_archivo )
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        wb.save(response)
        return response

from openpyxl.styles import Font, Alignment    

def ahorros_x_terceros(request):
    id_cli = request.session.get('cliente_id')
    id_ofi = request.session.get('oficina_id')
    print('id_ofi  ',id_ofi,'liquidez_ahorros')
    if request.method == 'GET':
        return render(request, 'ahorros_x_terceros.html')  # tu formulario
    if request.method == 'POST':
        fecha_str = request.POST.get('fecha_inicial')
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Inicial inválida", status=400)
        fecha_str = request.POST.get('fecha_final')
        try:
            fecha_final = datetime.strptime(fecha_str, '%Y-%m-%d').date()  # ✅ Conversión obligatoria
        except (ValueError, TypeError):
            return HttpResponse("Fecha Final inválida", status=400)
        accion = request.POST.get('accion')
        print(f"accion {accion}")
        if accion == 'terceros':
            movimientos = (
                DETALLE_PROD.objects
                .filter(
                    producto='AH',
                    concepto='AHO',
                    hecho_econo__fecha__range=(fecha_inicio,fecha_final),
                    hecho_econo__docto_conta__oficina_id=1
                )
                .values('subcuenta')
                .annotate(
                    Consignaciones=Sum(
                        Case(When(valor__lt=0, then=F('valor') * -1),default=Value(0), output_field=FloatField())),
                    Retiros=Sum(
                        Case(When(valor__gt=0, then=F('valor')),default=Value(0),output_field=FloatField()))
                )
            )

            registros = {}
            total_consignaciones = 0
            total_retiros = 0
            for movto in movimientos:
                num_cta = movto['subcuenta']
                Consignaciones = movto['Consignaciones']
                Retiros = movto['Retiros']
                cta_aho = CTAS_AHORRO.objects.filter(oficina_id = id_ofi,num_cta = num_cta).first()
                clave = (cta_aho.asociado.tercero.doc_ide,num_cta)         
                registros[clave] = {
                    'NitTer': cta_aho.asociado.tercero.doc_ide,
                    'NomTer': cta_aho.asociado.tercero.nombre,
                    'num_cta': num_cta,
                    'Excenta': cta_aho.exc_tas_mil,
                    'Consignaciones': Consignaciones,
                    'Retiros':  Retiros
                }
                total_consignaciones += Consignaciones
                total_retiros += Retiros
            registros = dict(sorted(registros.items(), key=lambda item: item[1]['NitTer']))
            wb = Workbook()
            ws = wb.active
            ws.title = "Ahorros por Tercero"
            titulo = f"Reporte de Ahorros por Tercero entre las fechas {fecha_inicio} y {fecha_final}"
            ws.merge_cells('A1:F1')
            ws['A1'] = titulo
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            columnas = ['NitTer', 'NomTer', 'num_cta', 'Excenta', 'Consignaciones', 'Retiros']
            ws.append(columnas)
            for registro in registros.values():
                fila = [
                    registro['NitTer'],
                    registro['NomTer'],
                    registro['num_cta'],
                    "Sí" if registro['Excenta'] else "No",
                    registro['Consignaciones'],
                    registro['Retiros'],
                ]
                ws.append(fila)
            fila_total = [
                '', '', 'TOTAL GENERAL:', '',
                total_consignaciones,
                total_retiros,
            ]
            ws.append([])  # Línea en blanco antes de totales
            ws.append(fila_total)
            for i, col in enumerate(columnas, 1):
                ws.column_dimensions[get_column_letter(i)].width = 18
            nombre_archivo = f"ahorros_tercero_{fecha_inicio}_{fecha_final}.xlsx"    
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            wb.save(response)
            return response
        elif accion == 'canceladas':
            canceladas = canceladas = (
                CTAS_AHORRO.objects
                .filter(
                    oficina_id=id_ofi,
                    fec_cancela__range=(fecha_inicio, fecha_final)
                )
                .annotate(
                    doc_ide=F('asociado__tercero__doc_ide'),
                    nombre=F('asociado__tercero__nombre')
                )
                .values(
                    'num_cta',
                    'doc_ide',
                    'nombre',
                    'fec_apertura',
                    'fec_cancela'
                )
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Cuentas_canceladas"
            titulo = f"Reporte de Cuentas de Ahorros Canceadas fechas {fecha_inicio} y {fecha_final}"
            ws.merge_cells('A1:F1')
            ws['A1'] = titulo
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:E1')
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            headers = ['Número de Cuenta', 'Documento', 'Nombre', 'Fecha Apertura', 'Fecha Cancelación']
            ws.append([])  # Fila 2 vacía
            ws.append(headers)
            for row in canceladas:
                ws.append([
                    row['num_cta'],
                    row['doc_ide'],
                    row['nombre'],
                    row['fec_apertura'],
                    row['fec_cancela'],
                ])
            for i, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
            nombre_archivo = f"ctas_canceladas_{fecha_inicio}_{fecha_final}.xlsx"    
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            wb.save(response)
            return response
        elif accion == 'exentas':
            canceladas = canceladas = (
                CTAS_AHORRO.objects
                .filter(
                    oficina_id=id_ofi,
                    exc_tas_mil = 'S'
                )
                .exclude(est_cta='C')
                .annotate(
                    doc_ide=F('asociado__tercero__doc_ide'),
                    nombre=F('asociado__tercero__nombre')
                )
                .values(
                    'num_cta',
                    'doc_ide',
                    'nombre',
                    'fec_apertura',
                    'fec_ini_exc'
                )
            )
            wb = Workbook()
            ws = wb.active
            ws.title = "Cuentas_Excentas de impuesto mil"
            titulo = f"Reporte de Cuentas de Ahorros xcentas del impuesto por mil fechas {fecha_inicio} y {fecha_final}"
            ws.merge_cells('A1:F1')
            ws['A1'] = titulo
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:E1')
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            headers = ['Número de Cuenta', 'Documento', 'Nombre', 'Fecha Apertura', 'Fecha Cancelación']
            ws.append([])  # Fila 2 vacía
            ws.append(headers)
            for row in canceladas:
                ws.append([
                    row['num_cta'],
                    row['doc_ide'],
                    row['nombre'],
                    row['fec_apertura'],
                    row['fec_ini_exc'],
                ])
            for i, header in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = 20
            nombre_archivo = f"ctas_exentas_{fecha_inicio}_{fecha_final}.xlsx"    
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            wb.save(response)
            return response