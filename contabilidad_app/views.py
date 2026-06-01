import datetime, csv
from math import ceil
from reportlab.lib.pagesizes import landscape, letter, legal
from django.http import HttpResponse, JsonResponse
import time, locale, re, json, os, django
from io import BytesIO
from num2words import num2words
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from django_xhtml2pdf.utils import generate_pdf
from django_xhtml2pdf.views import PdfMixin
from django.db.models import Sum, F, Q,OuterRef, Value, IntegerField,Subquery,QuerySet
from django.db.models.functions import Coalesce
import pandas as pd
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from django.views import View
from django.views.generic import CreateView, UpdateView, DeleteView, DetailView
from django.views.decorators.http import require_GET
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.utils.dateformat import format
from django.utils.decorators import method_decorator
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.db import transaction, connection, IntegrityError
from django.utils.dateparse import parse_date

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'Justo_proy.settings')
django.setup()

from hecho_economico_app.models import HECHO_ECONO
from hecho_economico_app.forms import HechoEconoForm
from documentos_app.models import DOCTO_CONTA
from detalle_producto_app.models import DETALLE_PROD
from localidades_app.models import LOCALIDADES
from conceptos_app.models import CONCEPTOS
from creditos_app.models import CREDITOS
from cambios_creditos_app.models import CAMBIOS_CRE
from causacion_creditos_app.models import CREDITOS_CAUSA
from asociados_app.models import ASOCIADOS
from detalle_economico_app.models import DETALLE_ECONO
from cuentas_app.models import PLAN_CTAS
from ctas_ahorros_app.models import CTAS_AHORRO
from contabilizacion_lineas_ahorros_app.models import IMP_CON_LIN_AHO
from contabilizacion_capital_creditos_app.models import IMP_CON_CRE
from terceros_app.models import TERCEROS
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from justo_app.justo_creditos import Liquida_cre
from justo_app.opciones import OPC_CANALES


def Buscar_Cuenta_Contable(request, template_name): 
    fecha_ini = request.GET.get('fec_ini')  # Fecha inicial ingresada por el usuario
    if fecha_ini:   # Si el usuario proporciona una fecha, filtramos por el período contable (año)
        fecha_ini_parsed = parse_date(fecha_ini)  # Convertir la fecha en objeto datetime.date
        if fecha_ini_parsed:
            periodo_contable = fecha_ini_parsed.year  # Extraer el año de la fecha
            cuentas = PLAN_CTAS.objects.filter(per_con__gte=periodo_contable).order_by('cod_cta')  # Filtrar cuentas con per_con >= año
        else:
            cuentas = PLAN_CTAS.objects.all().order_by('-per_con', 'cod_cta')  # Si la fecha no es válida, obtener todas las cuentas
    else:
        cuentas = PLAN_CTAS.objects.all().order_by('-per_con', 'cod_cta')  # Si no se proporciona fecha, mostrar todas las cuentas
    return render(request, template_name, {'cuentas': cuentas})  # Pasar las cuentas filtradas al template


def get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre):
        with connection.cursor() as cursor:
            query = '''
            SELECT pc.cod_cta AS cod_cta,
                pc.nom_cta AS nom_cta,
                '' AS doc_ide,
                '' AS raz_soc,
                0 AS docto,
                0 AS numero,
                DATE_SUB(%s, INTERVAL 1 DAY) AS fecha,
                'Saldo Anterior' AS detalle,
                '0' AS id_dc,
                0 AS debito,
                0 AS credito,
                SUM(de.debito - de.credito) AS sal_acu
            FROM detalle_econo de
            INNER JOIN (
                SELECT he.ID AS id,
                    he.fecha AS fecha,
                    dc.codigo AS cod_doc,
                    dc.nom_cto AS nom_doc,
                    he.numero AS numero
                FROM hecho_econo he
                INNER JOIN oficinas fi ON fi.id = %s
                INNER JOIN docto_conta dc ON dc.id = he.docto_conta_id
                                        AND dc.oficina_id = fi.id
                                        AND YEAR(he.fecha) = dc.per_con
                                        AND (%s = 'S' OR (%s = 'N' AND dc.codigo != 255))
                WHERE ((he.fecha = %s AND YEAR(he.fecha) >= YEAR(%s))
                    OR (dc.codigo = 0 AND YEAR(he.fecha) = YEAR(%s)))
                AND he.anulado = 'N'
            ) AS co ON co.id = de.hecho_econo_id
            INNER JOIN plan_ctas AS pc ON pc.cliente_id = %s
                                        AND pc.id = de.cuenta_id
                                        AND pc.cod_cta >= %s
                                        AND pc.cod_cta <= %s
            GROUP BY cod_cta, nom_cta
            UNION ALL
            SELECT pc.cod_cta AS cod_cta,
                pc.nom_cta AS nom_cta,
                te.doc_ide AS doc_ide,
                te.nombre AS raz_soc,
                co.cod_doc AS docto,
                co.numero AS numero,
                co.fecha AS fecha,
                de.detalle AS detalle,
                de.id_ds AS id_dc,
                de.debito AS debit0,
                de.credito AS credito,
                0.0 AS sal_acu
            FROM detalle_econo de
            INNER JOIN (
                SELECT he.ID AS id,
                    he.fecha AS fecha,
                    dc.codigo AS cod_doc,
                    dc.nom_cto AS nom_doc,
                    he.numero AS numero
                FROM hecho_econo he
                INNER JOIN oficinas fi ON fi.id = %s
                INNER JOIN docto_conta dc ON dc.id = he.docto_conta_id
                                        AND dc.oficina_id = fi.id
                                        AND YEAR(he.fecha) = dc.per_con
                                        AND (%s = 'S' OR (%s = 'N' AND dc.codigo != 255))
                WHERE he.fecha >= %s
                AND he.fecha <= %s
                AND dc.codigo != 0
                AND he.anulado = 'N'
            ) AS co ON co.id = de.hecho_econo_id
            INNER JOIN plan_ctas AS pc ON pc.cliente_id = %s
                                        AND pc.id = de.cuenta_id
                                        AND pc.cod_cta >= %s
                                        AND pc.cod_cta <= %s
            INNER JOIN terceros te ON te.id = de.tercero_id
            ORDER BY 1, 2, 7, 5, 6
        '''
            params = [
                fec_ini,  # %s para DATE_SUB
                id_ofi,   # %s para INNER JOIN oficinas
                con_cierre,  # %s para el primer AND en docto_conta
                con_cierre,  # %s para el segundo AND en docto_conta
                fec_ini,  # %s para el primer DATE en WHERE
                fec_ini,  # %s para el segundo DATE en WHERE
                fec_ini,  # %s para YEAR en el tercer DATE en WHERE
                id_cli,   # %s para INNER JOIN plan_ctas
                cta_ini,  # %s para AND pc.cod_cta >=
                cta_fin,  # %s para AND pc.cod_cta <=
                id_ofi,   # %s para INNER JOIN oficinas en UNION ALL
                con_cierre,  # %s para el primer AND en docto_conta en UNION ALL
                con_cierre,  # %s para el segundo AND en docto_conta en UNION ALL
                fec_ini,  # %s para el primer DATE en WHERE en UNION ALL
                fec_fin,  # %s para el segundo DATE en WHERE en UNION ALL
                id_cli,   # %s para INNER JOIN plan_ctas en UNION ALL
                cta_ini,  # %s para AND pc.cod_cta >= en UNION ALL
                cta_fin,  # %s para AND pc.cod_cta <= en UNION ALL
            ]

            cursor.execute(query, params)
            results = cursor.fetchall()
            # Define los nombres de los campos como una lista
            columns = ['cod_cta', 'nom_cta', 'doc_ide', 'raz_soc', 'docto',
                'numero', 'fecha', 'detalle', 'id_dc', 'debito', 'credito', 'sal_acu']
            results_dict = [dict(zip(columns, row)) for row in results]

        return results_dict


def saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide):
        with connection.cursor() as cursor:
            query = '''
            SELECT pc.cod_cta AS cod_cta,
                pc.nom_cta AS nom_cta,
                '' AS doc_ide,
                '' AS raz_soc,
                0 AS docto,
                0 AS numero,
                DATE_SUB(%s, INTERVAL 1 DAY) AS fecha,
                'Saldo Anterior' AS detalle,
                '0' AS id_dc,
                0 AS debito,
                0 AS credito,
                SUM(de.debito - de.credito) AS sal_acu
            FROM detalle_econo de
            INNER JOIN (
                SELECT he.ID AS id,
                    he.fecha AS fecha,
                    dc.codigo AS cod_doc,
                    dc.nom_cto AS nom_doc,
                    he.numero AS numero
                FROM hecho_econo he
                INNER JOIN oficinas fi ON fi.id = %s
                INNER JOIN docto_conta dc ON dc.id = he.docto_conta_id
                                        AND dc.oficina_id = fi.id
                                        AND YEAR(he.fecha) = dc.per_con
                                        AND (%s = 'S' OR (%s = 'N' AND dc.codigo != 255))
                WHERE ((he.fecha = %s AND YEAR(he.fecha) >= YEAR(%s))
                    OR (dc.codigo = 0 AND YEAR(he.fecha) = YEAR(%s)))
                AND he.anulado = 'N' 
                AND (%s IS NULL OR te.doc_ide = %s)
            ) AS co ON co.id = de.hecho_econo_id
            INNER JOIN plan_ctas AS pc ON pc.cliente_id = %s
                                        AND pc.id = de.cuenta_id
                                        AND pc.cod_cta >= %s
                                        AND pc.cod_cta <= %s
            GROUP BY cod_cta, nom_cta
            UNION ALL
            SELECT pc.cod_cta AS cod_cta,
                pc.nom_cta AS nom_cta,
                te.doc_ide AS doc_ide,
                te.nombre AS raz_soc,
                co.cod_doc AS docto,
                co.numero AS numero,
                co.fecha AS fecha,
                de.detalle AS detalle,
                de.id_ds AS id_dc,
                de.debito AS debit0,
                de.credito AS credito,
                0.0 AS sal_acu
            FROM detalle_econo de
            INNER JOIN (
                SELECT he.ID AS id,
                    he.fecha AS fecha,
                    dc.codigo AS cod_doc,
                    dc.nom_cto AS nom_doc,
                    he.numero AS numero
                FROM hecho_econo he
                INNER JOIN oficinas fi ON fi.id = %s
                INNER JOIN docto_conta dc ON dc.id = he.docto_conta_id
                                        AND dc.oficina_id = fi.id
                                        AND YEAR(he.fecha) = dc.per_con
                                        AND (%s = 'S' OR (%s = 'N' AND dc.codigo != 255))
                WHERE he.fecha >= %s
                AND he.fecha <= %s
                AND dc.codigo != 0
                AND he.anulado = 'N'
                AND (%s IS NULL OR te.doc_ide = %s)
            ) AS co ON co.id = de.hecho_econo_id
            INNER JOIN plan_ctas AS pc ON pc.cliente_id = %s
                                        AND pc.id = de.cuenta_id
                                        AND pc.cod_cta >= %s
                                        AND pc.cod_cta <= %s
            INNER JOIN terceros te ON te.id = de.tercero_id
            AND (%s IS NULL OR te.doc_ide = %s)
            ORDER BY 1, 2, 7, 5, 6
        '''
            params = [
                fec_ini,  # %s para DATE_SUB
                id_ofi,   # %s para INNER JOIN oficinas
                con_cierre,  # %s para el primer AND en docto_conta
                con_cierre,  # %s para el segundo AND en docto_conta
                fec_ini,  # %s para el primer DATE en WHERE
                fec_ini,  # %s para el segundo DATE en WHERE
                fec_ini,  # %s para YEAR en el tercer DATE en WHERE
                id_cli,   # %s para INNER JOIN plan_ctas
                cta_ini,  # %s para AND pc.cod_cta >=
                cta_fin,  # %s para AND pc.cod_cta <=
                id_ofi,   # %s para INNER JOIN oficinas en UNION ALL
                con_cierre,  # %s para el primer AND en docto_conta en UNION ALL
                con_cierre,  # %s para el segundo AND en docto_conta en UNION ALL
                fec_ini,  # %s para el primer DATE en WHERE en UNION ALL
                fec_fin,  # %s para el segundo DATE en WHERE en UNION ALL
                id_cli,   # %s para INNER JOIN plan_ctas en UNION ALL
                cta_ini,  # %s para AND pc.cod_cta >= en UNION ALL
                cta_fin,  # %s para AND pc.cod_cta <= en UNION ALL
                doc_ide,  # %s para el filtro por te.doc_ide en ambas consultas
                doc_ide,  # %s para el filtro por te.doc_ide en UNION ALL
            ]

            cursor.execute(query, params)
            results = cursor.fetchall()
            # Define los nombres de los campos como una lista
            columns = ['cod_cta', 'nom_cta', 'doc_ide', 'raz_soc', 'docto',
                'numero', 'fecha', 'detalle', 'id_dc', 'debito', 'credito', 'sal_acu']
            results_dict = [dict(zip(columns, row)) for row in results]

        return results_dict


def auxiliar_cuenta(request):
    if request.method == 'GET':
        return render(request, 'auxiliar_x_cuenta.html')
    if request.method == 'POST':
        accion = request.POST.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'
        con_cierre = request.POST['con_cierre']
        
        resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
        
        cta_act = None
        saldo_cta = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Maneja el primer caso
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta == cta_act:
                saldo_cta = saldo_cta + debito - credito
                row['sal_acu'] = saldo_cta
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Auxiliar_Cuenta" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "Reporte Auxiliar por Cuenta"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_cuenta.xlsx"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_cuenta.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="auxiliar_por_cuenta.pdf"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "Reporte Auxiliar por Cuenta"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Página {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 100), 
                    ("doc_ide", 60), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)


def auxiliar_tercero(request):
    if request.method == 'GET':
        return render(request, 'auxiliar_x_tercero.html')

    if request.method == 'POST':

        accion = request.POST.get("accion")  # Obtener la acción

        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
        
        # doc_ide = request.POST['tercero']
        doc_ide = request.POST.get('tercero', None)

        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'
        con_cierre = request.POST['con_cierre']

        print("Parámetros enviados a saldo_anterior:", id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)

        resultado = saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)

        cta_act = None
        saldo_cta = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Maneja el primer caso
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta == cta_act:
                saldo_cta = saldo_cta + debito - credito
                row['sal_acu'] = saldo_cta
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Auxiliar_Tercero" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "Reporte Auxiliar por Tercero"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_tercero.xlsx"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_tercero.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response

        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="auxiliar_por_tercero.pdf"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "Reporte Auxiliar por Tercero"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Página {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 100), 
                    ("doc_ide", 60), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)


def auxiliar_comprobante(request):
    if request.method == 'GET':
        return render(request, 'auxiliar_x_comprobante.html')

    if request.method == 'POST':

        accion = request.POST.get("accion")  # Obtener la acción

        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
        
        # doc_ide = request.POST['tercero']
        doc_ide = request.POST.get('tercero', None)

        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'
        con_cierre = request.POST['con_cierre']

        print("Parámetros enviados a saldo_anterior:", id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)

        resultado = saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)

        cta_act = None
        saldo_cta = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Maneja el primer caso
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta == cta_act:
                saldo_cta = saldo_cta + debito - credito
                row['sal_acu'] = saldo_cta
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Auxiliar_Comprobante" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre, doc_ide)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "Reporte Auxiliar por Comprobante"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_comprobante.xlsx"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="auxiliar_comprobante.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response

        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="auxiliar_por_comprobante.pdf"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "Reporte Auxiliar por Comprobante"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Página {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 100), 
                    ("doc_ide", 60), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)


def balance_prueba(request):
    if request.method == 'GET':
        return render(request, 'balance_prueba.html')

    if request.method == 'POST':

        accion = request.POST.get("accion")  # Obtener la acción

        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_cor'], '%Y-%m-%d').date()

        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'

        con_cierre = request.POST['con_cierre']

        resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)

        print(resultado)

        # cta_act = None
        # saldo_cta = 0
        # for row in resultado:
        #     cod_cta = row['cod_cta']
        #     debito = row['debito']
        #     credito = row['credito']
        #     sal_acu = row['sal_acu']
        #     if cta_act is None:  # Maneja el primer caso
        #         saldo_cta = sal_acu
        #         cta_act = cod_cta
        #     elif cod_cta == cta_act:
        #         saldo_cta = saldo_cta + debito - credito
        #         row['sal_acu'] = saldo_cta
        #     else:
        #         saldo_cta = sal_acu
        #         cta_act = cod_cta
       
        cta_act = None
        saldo_cta = 0
        resultado_con_subtotales = []
        # Variables para acumular subtotales
        subtotal_debito = 0
        subtotal_credito = 0
        subtotal_saldo = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Primer registro
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta != cta_act:  # Cambio de cuenta -> Insertar subtotal
                # Insertar fila de subtotal antes de cambiar de cuenta
                resultado_con_subtotales.append({
                    'cod_cta': cta_act,
                    'descripcion': 'Subtotal',
                    'debito': subtotal_debito,
                    'credito': subtotal_credito,
                    'sal_acu': subtotal_saldo
                })
                # Reiniciar subtotales
                subtotal_debito = 0
                subtotal_credito = 0
                subtotal_saldo = 0

                saldo_cta = sal_acu
                cta_act = cod_cta

            elif cod_cta == cta_act:
                # Actualizar saldo y agregar fila al resultado final
                saldo_cta += debito - credito
                row['sal_acu'] = saldo_cta
                resultado_con_subtotales.append(row)

                # Acumular subtotales
                subtotal_debito += debito
                subtotal_credito += credito
                subtotal_saldo = saldo_cta

                # Agregar el último subtotal después del bucle
                if cta_act is not None:
                    resultado_con_subtotales.append({
                        'cod_cta': cta_act,
                        'descripcion': 'Subtotal',
                        'debito': subtotal_debito,
                        'credito': subtotal_credito,
                        'sal_acu': subtotal_saldo
                    })
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Balance_Prueba" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # cta_act = None
            # saldo_cta = 0
            # for row in resultado:
            #     cod_cta = row['cod_cta']
            #     debito = row['debito']
            #     credito = row['credito']
            #     sal_acu = row['sal_acu']
            #     if cta_act is None:  # Maneja el primer caso
            #         saldo_cta = sal_acu
            #         cta_act = cod_cta
            #     elif cod_cta == cta_act:
            #         saldo_cta = saldo_cta + debito - credito
            #         row['sal_acu'] = saldo_cta
            #     else:
            #         saldo_cta = sal_acu
            #         cta_act = cod_cta

            cta_act = None
            saldo_cta = 0
            resultado_con_subtotales = []
            # Variables para acumular subtotales
            subtotal_debito = 0
            subtotal_credito = 0
            subtotal_saldo = 0
            for row in resultado:
                cod_cta = row['cod_cta']
                debito = row['debito']
                credito = row['credito']
                sal_acu = row['sal_acu']
                if cta_act is None:  # Primer registro
                    saldo_cta = sal_acu
                    cta_act = cod_cta
                elif cod_cta != cta_act:  # Cambio de cuenta -> Insertar subtotal
                    # Insertar fila de subtotal antes de cambiar de cuenta
                    resultado_con_subtotales.append({
                        'cod_cta': cta_act,
                        'descripcion': 'Subtotal',
                        'debito': subtotal_debito,
                        'credito': subtotal_credito,
                        'sal_acu': subtotal_saldo
                    })
                    # Reiniciar subtotales
                    subtotal_debito = 0
                    subtotal_credito = 0
                    subtotal_saldo = 0

                    saldo_cta = sal_acu
                    cta_act = cod_cta

                elif cod_cta == cta_act:
                    # Actualizar saldo y agregar fila al resultado final
                    saldo_cta += debito - credito
                    row['sal_acu'] = saldo_cta
                    resultado_con_subtotales.append(row)

                    # Acumular subtotales
                    subtotal_debito += debito
                    subtotal_credito += credito
                    subtotal_saldo = saldo_cta

                    # Agregar el último subtotal después del bucle
                    if cta_act is not None:
                        resultado_con_subtotales.append({
                            'cod_cta': cta_act,
                            'descripcion': 'Subtotal',
                            'debito': subtotal_debito,
                            'credito': subtotal_credito,
                            'sal_acu': subtotal_saldo
                        })
                else:
                    saldo_cta = sal_acu
                    cta_act = cod_cta
                    
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "BALANCE DE PRUEBA"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="balance_prueba.xlsx"'
            workbook.save(response)
            return response

        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            cta_act = None
            saldo_cta = 0
            for row in resultado:
                cod_cta = row['cod_cta']
                debito = row['debito']
                credito = row['credito']
                sal_acu = row['sal_acu']
                if cta_act is None:  # Maneja el primer caso
                    saldo_cta = sal_acu
                    cta_act = cod_cta
                elif cod_cta == cta_act:
                    saldo_cta = saldo_cta + debito - credito
                    row['sal_acu'] = saldo_cta
                else:
                    saldo_cta = sal_acu
                    cta_act = cod_cta
                
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="balance_prueba.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response

        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="balance_prueba.pdf"'

            p = canvas.Canvas(response, pagesize=letter)
            width, height = letter
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 64 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "BALANCE DE PRUEBA"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.setFont("Helvetica", 8)
                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, height - 90, texto_paginas)
                # p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 20, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x-20, margin_y - 40, texto_pie)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+552, line_y)
                p.line(margin_x-40, height - 100, margin_x+552, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 200), 
                    ("Saldo Anterior", 60), 
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],200),  
                    # (f"{row['sal_ant']:,.2f}",80,'right'),
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida en Balance de Prueba", status=400)
    return HttpResponse("Método no permitido", status=405)


def balance_general(request):
    if request.method == 'GET':
        return render(request, 'balance_general.html')

    if request.method == 'POST':

        accion = request.POST.get("accion")  # Obtener la acción

        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()

        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'

        con_cierre = request.POST['con_cierre']

        resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)

        cta_act = None
        saldo_cta = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Maneja el primer caso
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta == cta_act:
                saldo_cta = saldo_cta + debito - credito
                row['sal_acu'] = saldo_cta
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Balance_General" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "Balance de Prueba"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="balance_general.xlsx"'
            workbook.save(response)
            return response

        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="balance_general.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response

        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="balance_general.pdf"'

            p = canvas.Canvas(response, pagesize=letter)
            width, height = letter
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "Balance de General"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Página {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 100), 
                    ("doc_ide", 60), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)


def estado_resultados(request):
    if request.method == 'GET':
        return render(request, 'estado_resultados.html')

    if request.method == 'POST':

        accion = request.POST.get("accion")  # Obtener la acción

        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()

        cta_ini = request.POST['cta_ini'] or '1'
        cta_fin = request.POST['cta_fin'] or '9999999999'

        con_cierre = request.POST['con_cierre']

        resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)

        cta_act = None
        saldo_cta = 0
        for row in resultado:
            cod_cta = row['cod_cta']
            debito = row['debito']
            credito = row['credito']
            sal_acu = row['sal_acu']
            if cta_act is None:  # Maneja el primer caso
                saldo_cta = sal_acu
                cta_act = cod_cta
            elif cod_cta == cta_act:
                saldo_cta = saldo_cta + debito - credito
                row['sal_acu'] = saldo_cta
            else:
                saldo_cta = sal_acu
                cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Estado_Resultados" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "Estado de Resultados"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="estado_resultados.xlsx"'
            workbook.save(response)
            return response

        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = get_saldo_anterior(id_cli, id_ofi, fec_ini, fec_fin, cta_ini, cta_fin, con_cierre)
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="estado_resultados.csv"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response

        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="estado_resultados.pdf"'

            p = canvas.Canvas(response, pagesize=letter)
            width, height = letter
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
                x = margin_x - 30
                p.setFont("Helvetica-Bold", 9)
                p.drawString(x, y, "Subtotal: "+cod_cta)
                x += 640  # Posicionar en la columna de débito
                p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
                x += 80  # Posicionar en la columna de crédito
                p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "Balance de Prueba"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Desde la Cuenta: {cta_ini} Hasta la Cuenta: {cta_fin}      Con Cierre: {con_cierre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Página {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Código", 50), 
                    ("Cuenta", 100), 
                    ("doc_ide", 60), 
                    ("raz_soc", 95), 
                    ("Dto", 20), 
                    ("Número", 40), 
                    ("Fecha", 50), 
                    ("Detalle", 125),
                    ("        Débito", 70), 
                    ("        Crédito", 70), 
                    ("        Saldo", 80)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cta'], 50), 
                    (row['nom_cta'][:17],100), 
                    (row['doc_ide'], 60),
                    (row['raz_soc'][:15], 95), 
                    (row['docto'],20), 
                    (row['numero'], 40),
                    (row['fecha'], 50), 
                    (f"{row['detalle'][:24]}", 125), 
                    (f"{row['debito']:,.2f}",70,'right'), 
                    (f"{row['credito']:,.2f}", 70,'right'), 
                    (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)


def conciliacion_bancaria(request):
    messages('En construcción')
    pass


def activos_fijos(request):
    messages('En construcción')
    pass

