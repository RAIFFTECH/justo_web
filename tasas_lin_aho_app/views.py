import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django import forms
from .forms import CrearForm
from .models import TAS_LIN_AHO,LINEAS_AHORRO

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = TAS_LIN_AHO
    form = CrearForm
    template_name = 'lista_tasa_linea_ahorros.html'

# Para obtener todos los detalles de un registro

class Detalles(LoginRequiredMixin, DetailView):
    model = TAS_LIN_AHO
    template_name = 'detalles_tasa_linea_ahorro.html'
    context_object_name = 'object'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        obj = self.get_object()
        fecha_inicial = self.request.GET.get('fecha_inicial')
        fecha_final = self.request.GET.get('fecha_final')
        lin_aho_id = self.request.GET.get('lin_aho')  # supondremos que se pasa el ID
        tasas_relacionadas = TAS_LIN_AHO.objects.all()
        if lin_aho_id:
            tasas_relacionadas = tasas_relacionadas.filter(lin_aho_id=lin_aho_id)
        else:
            tasas_relacionadas = tasas_relacionadas.filter(lin_aho=obj.lin_aho)
        if fecha_inicial:
            tasas_relacionadas = tasas_relacionadas.filter(fecha_inicial__gte=fecha_inicial)
        if fecha_final:
            tasas_relacionadas = tasas_relacionadas.filter(fecha_final__lte=fecha_final)
        context['tasas_relacionadas'] = tasas_relacionadas
        context['fecha_inicial_filtro'] = fecha_inicial
        context['fecha_final_filtro'] = fecha_final
        context['lin_aho_filtro'] = int(lin_aho_id) if lin_aho_id else obj.lin_aho.id
        context['lineas_ahorro'] = LINEAS_AHORRO.objects.all()  # para el <select> de líneas
        return context


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TAS_LIN_AHO
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_tasa_linea_ahorro.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_tasa_linea_ahorro')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TAS_LIN_AHO
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_tasa_linea_ahorro.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_tasa_linea_ahorro')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = TAS_LIN_AHO
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_tasa_linea_ahorro')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        tasa_linea_ahorros = TAS_LIN_AHO.objects.all().order_by('lin_aho')
           
        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="tasa_linea_ahorros.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)
       
        datos_tabla = [["Línea de Ahorro", "Fecha Inicial","Fecha Final","Tasa Interés Anual Efectiva"]]

        for dato in tasa_linea_ahorros:
            datos_tabla.append([dato.lin_aho,dato.fecha_inicial,dato.fecha_final,dato.tiae])
 
            # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = TAS_LIN_AHO.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="tasa_linea_ahorros.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Línea de Ahorro: {dato.lin_aho}")
        p.drawString(80, 780, f"Fecha Inicial: {dato.fecha_inicial}")
        p.drawString(80, 760, f"Fecha Final: {dato.fecha_final}")
        p.drawString(80, 740, f"Tasa Interés Anual Efectiva: {dato.tiae}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        tasa_linea_ahorros = TAS_LIN_AHO.objects.all()

        p = canvas.Canvas(response)

        for dato in tasa_linea_ahorros:
            p.drawString(80, 800, f"Línea de Ahorro: {dato.lin_aho}")
            p.drawString(80, 780, f"Fecha Inicial: {dato.fecha_inicial}")
            p.drawString(80, 760, f"Fecha Final: {dato.fecha_final}")
            p.drawString(80, 740, f"Tasa Interés Anual Efectiva: {dato.tiae}")
            p.showPage()
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in TAS_LIN_AHO._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        tasa_linea_ahorros = TAS_LIN_AHO.objects.all()
        for row_num, data in enumerate(tasa_linea_ahorros, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        tasa_linea_ahorros = TAS_LIN_AHO.objects.all()
        for data in tasa_linea_ahorros:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
