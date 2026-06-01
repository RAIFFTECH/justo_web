import csv, re, json, time
from math import ceil
from django.db import connection
from openpyxl import Workbook
from django.db import IntegrityError, transaction
from django.db.models import IntegerField, Max, Q
from django.db.models.functions import Cast
from django.views import View
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.pdfbase.pdfmetrics import stringWidth
from dateutil.relativedelta import relativedelta
from django.views.generic import UpdateView, CreateView, ListView
from django.views.generic.edit import CreateView, UpdateView
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse, QueryDict
from django.urls import reverse_lazy, reverse
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.contrib import messages
from datetime import datetime, date
from django.shortcuts import render, redirect, get_object_or_404
from .forms import FechaForm, CreditoForm, CreditosFilterForm, CodeudorForm
from asociados_app.models import ASOCIADOS
from cambios_creditos_app.models import CAMBIOS_CRE
from justo_app.justo_creditos import calculo_cuota, Liquida_cre
from terceros_app.models import TERCEROS
from .models import CODEUDORES, CREDITOS, GAR_NO_IDONEA
from clientes_app.models import CLIENTES
from oficinas_app.models import OFICINAS
from detalle_producto_app.models import DETALLE_PROD

def prueba_fecha(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
    # if request.method == 'POST' and request.is_ajax():
        form = FechaForm(request.POST)
        if form.is_valid():
            fecha = form.cleaned_data['fec_des']
            print('fecha ',fecha,'  type ',type(fecha))
            return JsonResponse({'fecha': str(fecha)})
        else:
            return JsonResponse({'error': 'Datos inválidos'}, status=400)
    else:
        form = FechaForm()
    return render(request, 'prueba.html', {'form': form})

class CreditosListView(ListView):
    model = CREDITOS
    template_name = 'creditos_list.html'
    context_object_name = 'creditos'
    paginate_by = 10  # Opcional: número de créditos por página

class CreditosListView(ListView):
    model = CREDITOS
    template_name = 'creditos_list.html'
    context_object_name = 'creditos'
    paginate_by = 10

    def get_queryset(self):
        queryset = CREDITOS.objects.all()
        form = CreditosFilterForm(self.request.GET or None)
        
        if form.is_valid():
            print("GET:", self.request.GET)
            print("Form válido:", form.is_valid())
            print("Form cleaned_data:", form.cleaned_data if form.is_valid() else "No válido")

            estado = form.cleaned_data.get('estado') or 'activo'
            cod_cre = form.cleaned_data.get('cod_cre')
            nombre_deudor = form.cleaned_data.get('nombre_deudor')
            cod_aso = form.cleaned_data.get('cod_aso')

            pk = self.kwargs.get('pk')

            if estado:
                queryset = queryset.filter(estado=estado)
            if cod_cre:
                queryset = queryset.filter(cod_cre__icontains=cod_cre)
            if nombre_deudor:
                queryset = queryset.filter(socio__tercero__nombre__icontains=nombre_deudor)
            if cod_aso:
                queryset = queryset.filter(socio__cod_aso__icontains=cod_aso)
            if pk:
                queryset = queryset.filter(socio_id=pk)
        else:
            queryset = queryset.filter(estado='activo')

        return queryset


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        get_copy = self.request.GET.copy()
        if not get_copy.get('estado'):
            get_copy['estado'] = 'A'
        form = CreditosFilterForm(get_copy)
        context['filter_form'] = form
        return context


DATE_PATTERN = re.compile(r'^\d{4}-(0[1-9]|1[0-2])-(0[1-9]|[12][0-9]|3[01])$')

class BaseCreditoView(View):
    model = CREDITOS
    template_name = 'creditos_form.html'

    def validate_models(self, form):
        print('Entra a Validacion ')
        errors = {}
        cod_soc = form.cleaned_data.get('doc_ide')
        print('Cod Aso ',cod_soc)
        if not isinstance(cod_soc, str):
            errors[cod_soc] = ' Invalido'
            return errors
        socio = ASOCIADOS.objects.filter(oficina_id = 1,cod_aso = cod_soc).first()
        if socio == None:
             print('cos_aso   no Existe ')
             errors[cod_soc] = 'No Existe este Código de Asociado'
        fec_des = form.cleaned_data.get('fec_des')
        if not isinstance(fec_des, str):
            return errors
        if not self.is_valid_date_format(fec_des):
            errors['fec_des'] = 'La fecha debe estar en formato dd/mm/yyyy'
        if not DATE_PATTERN.match(fec_des):
            errors['fec_des'] = 'La fecha debe estar en formato dd/mm/yyyy 1'
        fec_pag_ini = form.cleaned_data.get('fec_pag_ini')
        if not isinstance(fec_pag_ini, str):
            return errors
        if not self.is_valid_date_format(fec_pag_ini):
            errors['fec_pag_ini'] = 'La fecha debe estar en formato dd/mm/yyyy'
        if not DATE_PATTERN.match(fec_des):
            errors['fec_pag_ini'] = 'La fecha debe estar en formato dd/mm/yyyy 1'
        val_cuo_ini = form.cleaned_data.get('val_cuo_ini')
        if val_cuo_ini <= 0:
            errors['val_cuo_ini'] = 'Valor de Cuota Invalido'
        return errors

    def is_valid_date_format(date_value):
        # Verificar si date_value es una cadena
        if isinstance(date_value, str):
            try:
                datetime.strptime(date_value, '%d/%m/%Y')
                return True
            except ValueError:
                return False
        # Verificar si date_value es un objeto datetime
        elif isinstance(date_value, (datetime, date)):
            return True
        else:
            return False

def validate_codeudores(form,cliente_id,crud):   # Pre Validacion
    errors = {}
    codeudores = form.data.get('codeudores', [])
    for codeudor in codeudores:
        documento = codeudor.get('doc_ide')
        tercero = TERCEROS.objects.filter(cliente_id=cliente_id, doc_ide = documento).first()
        if tercero == None:
            errors[documento] = ' No existe Tercero Registrado '
    return errors

class CreditoCreateView(BaseCreditoView, CreateView):
    model = CREDITOS
    form_class = CreditoForm
    template_name = 'creditos_form.html'
    success_url = reverse_lazy('crear_credito_justo')

    MAX_RETRIES = 3

    def get_form(self, form_class=None):
        oficina_id = self.request.session.get('oficina_id')
        form = super().get_form(form_class)
        form.fields['cod_cre'].widget.attrs['readonly'] = 'readonly'
        form.fields['cod_cre'].widget.attrs['class'] = 'form-control bg-light'  # Clase para resaltado
        form.fields['fec_des'].widget.attrs['class'] = 'form-control custom-background'
        return form
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')

        max_cod_cre = (
            CREDITOS.objects.filter(cod_cre__regex=r'^\d+$',oficina_id = oficina_id)  # Filtra solo los valores numéricos
                .annotate(cod_cre_int=Cast('cod_cre', IntegerField()))  # Convierte `cod_cre` a entero
                .aggregate(max_value=Max('cod_cre_int'))['max_value']
            )
        new_cod_cre = str(max_cod_cre + 1).zfill(8) if max_cod_cre is not None else '00000001'
        kwargs['request'] = self.request
        kwargs['initial'] = {
            'cliente_id': cliente_id,
            'cod_cre': new_cod_cre  # Asigna `new_cod_cre` como valor inicial para `cod_cre`
        }
        return kwargs


    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        print('GET Response Context:', response.context_data)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['operation'] = 'create'  # Definir la operación como 'create'
        return context

    def post(self, request, *args, **kwargs):
        print('entra a post')
        # ✅ 1. Acceder a la sesión (esto no se toca)
        cliente_id = self.request.session.get('cliente_id')
        oficina_id = self.request.session.get('oficina_id')

        # ✅ 2. Validar que hay datos en el cuerpo de la petición
        if not request.body:
            return JsonResponse({'success': False, 'message': 'No data received'}, status=400)

        raw_data = json.loads(request.body)
        data = QueryDict('', mutable=True)
        for key, value in raw_data.items():
            data[key] = value

        form = CreditoForm(data, request=request)
        if form.is_valid():
            errors = self.validate_models(form)
            if errors: 
                for field, error_list in errors.items():
                    if not isinstance(error_list, list):
                        error_list = [error_list]
                        form.errors[field] = form.errors.get(field, []) + error_list      

                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            hay_error = validate_codeudores(form,cliente_id,'C')
            if hay_error:
                for field, error_list in hay_error.items():
                    form.errors[field] = form.errors.get(field, []) + error_list
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)

            cod_cre = form.cleaned_data.get('cod_cre')
            doc_ide = form.cleaned_data.get('doc_ide')
            cap_ini = form.cleaned_data.get('cap_ini')
            num_cuo_ini = form.cleaned_data.get('num_cuo_ini')
            num_cuo_gra = form.cleaned_data.get('num_cuo_gra')
            #tian_ic_ini = form.cleaned_data.get('tian_ic_ini')
            cod_lin_cre = form.cleaned_data.get('cod_lin_cre')
            fec_des = form.cleaned_data.get('fec_des')
            #decreciente = form.cleaned_data.get('decreciente')
            tian_pol_seg = form.cleaned_data.get('tian_pol_seg')
            per_ano = form.cleaned_data.get('per_ano')
            fec_pag_ini = form.cleaned_data.get('fec_pag_ini')
            val_cuo_ini = form.cleaned_data.get('val_cuo_ini')
            print('val_cuo_ini  ',val_cuo_ini)
            val_cuo_act = form.cleaned_data.get('val_cuo_act')
            print('val_cuo_act  ',val_cuo_act)
            tian_ic_act = form.cleaned_data.get('tian_ic_act')
            tian_im = form.cleaned_data.get('tian_im')
            imputacion = form.cleaned_data.get('imputacion')
            libranza = form.cleaned_data.get('libranza')
            pagare = form.cleaned_data.get('pagare')
            termino = form.cleaned_data.get('termino')
            for_pag = form.cleaned_data.get('for_pag')
            tip_gar = form.cleaned_data.get('tip_gar')
            por_des_pro_pag = form.cleaned_data.get('por_des_pro_pag')
            estado = form.cleaned_data.get('estado')
            est_jur = form.cleaned_data.get('est_jur')
            rep_cen_rie = form.cleaned_data.get('rep_cen_rie')
            figarantias = form.cleaned_data.get('figarantias')
            fec_ult_pag = form.cleaned_data.get('fec_ult_pag')
            num_cuo_act = form.cleaned_data.get('num_cuo_act')
            val_gar_hip = form.cleaned_data.get('val_gar_hip')
            mat_inm_gar = form.cleaned_data.get('mat_inm_gar')
            num_pol_gar_hip = form.cleaned_data.get('num_pol_gar_hip')
            cat_nue = form.cleaned_data.get('cat_nue')
            prima = form.cleaned_data.get('prima')
            porcentaje_prima = form.cleaned_data.get('porcentaje_prima')
            retries = 0
            while retries < self.MAX_RETRIES:
                try:
                    with transaction.atomic():
                        if CREDITOS.objects.filter(oficina_id= oficina_id,cod_cre = cod_cre).exists():
                            cod_cre_numeric = int(cod_cre) + 1
                            cod_cre = f'{cod_cre_numeric:08d}'
                        else:
                            socio = ASOCIADOS.objects.filter(oficina_id = oficina_id,cod_aso = doc_ide).first()
                            credito = CREDITOS.objects.create(oficina_id= oficina_id,cod_cre = cod_cre)
                            credito.socio = socio
                            credito.cap_ini = cap_ini
                            credito.num_cuo_ini = num_cuo_ini
                            credito.num_cuo_act = num_cuo_ini
                            credito.num_cuo_gra = num_cuo_gra
                            #credito.tian_ic_ini = tian_ic_ini
                            credito.tian_ic_act = tian_ic_act  # revisar si es la variable
                            credito.cod_lin_cre = cod_lin_cre
                            credito.fec_des = fec_des
                            credito.decreciente = 'N'
                            credito.tian_pol_seg = tian_pol_seg
                            credito.per_ano = per_ano
                            credito.fec_pag_ini = fec_pag_ini
                            credito.val_cuo_ini = val_cuo_ini
                            credito.val_cuo_act = val_cuo_ini
                            credito.tian_im = tian_im
                            credito.imputacion = imputacion
                            credito.libranza = libranza
                            credito.pagare = pagare
                            credito.termino = termino
                            credito.for_pag = for_pag
                            credito.tip_gar = tip_gar
                            credito.por_des_pro_pag = por_des_pro_pag
                            credito.estado = estado
                            credito.est_jur = est_jur
                            credito.rep_cen_rie = rep_cen_rie
                            credito.figarantias = figarantias
                            credito.fec_ult_pag = fec_ult_pag
                            credito.num_cuo_act = num_cuo_act
                            credito.val_gar_hip = val_gar_hip
                            credito.mat_inm_gar = mat_inm_gar
                            credito.num_pol_gar_hip = num_pol_gar_hip
                            credito.cat_nue = cat_nue
                            credito.prima = prima
                            credito.porcentaje_prima = porcentaje_prima
                            credito.save()
                            lista_codeudores = data.get('codeudores', [])
                            if lista_codeudores:
                                for codeudo in lista_codeudores:
                                    documento = codeudo.get('doc_ide')
                                    tercero = TERCEROS.objects.filter(cliente_id=cliente_id, doc_ide = documento).first()
                                    codeudor = CODEUDORES.objects.create(oficina_id = oficina_id,credito_id=credito.id,tercero_id=tercero.id)
                                    codeudor.save()
                            print('Grabacion Perfecta')
                            return JsonResponse({'success': True, 'message': 'Se grabó correctamente.','Numero' : credito.cod_cre})                 
                except IntegrityError as e:
                    return JsonResponse({'success': False, 'error': 'Fallos en el Registro del Credito'}, status=400)
            if retries >= self.MAX_RETRIES:
                return JsonResponse({'success': False, 'error': 'Fallos en el Registro del Credito'}, status=400)
            else:
                retries += retries
                time.sleep(1)  # Esperar un segundo antes de reintentar
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
         
    def form_valid(self, form):
        oficina_id = self.request.session.get('oficina_id')
        if not oficina_id:
            return JsonResponse({'success': False, 'errors': {'oficina_id': 'ID de oficina no encontrado en la sesión.'}}, status=400)
        form.instance.oficina_id = oficina_id
        errors = self.validate_models(form)
        if errors:       
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        socio = ASOCIADOS.objects.filter(oficina_id=oficina_id, cod_aso=form.cleaned_data.get('doc_ide')).first()
        if socio is None:
            return JsonResponse({'success': False, 'errors': {'doc_ide': 'No hay registro de un asociado con ese documento de identidad'}}, status=400)
        form.instance.socio = socio
        self.object = form.save()  # Guarda el crédito aquí
        return JsonResponse({'success': True, 'message': 'Registro guardado exitosamente.', 'cod_cre': form.cleaned_data.get('cod_cre')}, status=200)

    def form_invalid(self, form):
        print('Datos recibidos Invalidos en POST:', self.request.POST)
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list
            print('Errrores ',error_list)
        return JsonResponse({'success': True, 'errors': errors}, status=200)

class UpdateViewCredito(BaseCreditoView, UpdateView):
    model = CREDITOS
    form_class = CreditoForm
    template_name = 'creditos_form.html'
    success_url = reverse_lazy('crear_credito_justo')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request  # ✅ importante: para que el form use request
        return kwargs

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        credito = self.object
        codeudores = CODEUDORES.objects.filter(credito_id=credito.id)

        # ✅ Usamos self.get_form() para que incluya 'request' y demás kwargs
        form = self.get_form()
        form.initial.update({
            'doc_ide': credito.socio.tercero.doc_ide if credito.socio and credito.socio.tercero else '',
            'nom_soc': credito.socio.tercero.nombre if credito.socio and credito.socio.tercero else '',
        })

        context = {
            'form': form,
            'operation': 'update',
            'codeudores': codeudores,
            'credito': credito,
        }
        return render(request, self.template_name, context)

    def get_form(self, form_class=None):
        print('get_form de update')
        form = super().get_form(form_class)

        # Personalización de los widgets
        form.fields['cod_cre'].widget.attrs.update({
            'readonly': 'readonly',
            'class': 'form-control bg-light'
        })
        form.fields['fec_des'].widget.attrs.update({
            'class': 'form-control custom-background'
        })
        return form

    @method_decorator(csrf_exempt)
    def post(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
        cliente_id = request.session.get('cliente_id')
        oficina_id = request.session.get('oficina_id')
        data = json.loads(request.body)

        credito = get_object_or_404(CREDITOS, id=pk)

        # ✅ Pasamos 'request' al form también en POST
        form = CreditoForm(data, instance=credito, request=request)

        if form.is_valid():
            errors = self.validate_models(form)
            if errors:
                return JsonResponse({'success': False, 'errors': errors}, status=400)

            hay_error = validate_codeudores(form, cliente_id, 'U')
            if hay_error:
                hay_error = {key: [value] if isinstance(value, str) else value for key, value in hay_error.items()}
                return JsonResponse({'success': False, 'error': hay_error})

            socio = ASOCIADOS.objects.filter(oficina_id=oficina_id, cod_aso=form.cleaned_data['doc_ide']).first()
            credito.socio = socio
            # 🛠 Asignas todos los campos del form a la instancia
            for field in [
                'cap_ini', 'num_cuo_ini', 'num_cuo_gra', 'tian_ic_act',
                'cod_lin_cre', 'fec_des', 'tian_pol_seg', 'per_ano',
                'fec_pag_ini', 'val_cuo_ini', 'tian_im', 'imputacion',
                'libranza', 'pagare', 'termino', 'for_pag', 'tip_gar',
                'por_des_pro_pag', 'estado', 'est_jur', 'rep_cen_rie',
                'figarantias', 'fec_ult_pag', 'num_cuo_act', 'val_cuo_act',
                'val_gar_hip', 'mat_inm_gar', 'num_pol_gar_hip', 'cat_nue',
                'prima', 'porcentaje_prima'
            ]:
                setattr(credito, field, form.cleaned_data.get(field))

            try:
                with transaction.atomic():
                    credito.save()
                    # Manejo de codeudores (igual que antes)
                    codeudors_bac = data.get('codeudores', [])
                    codeudors_mod = CODEUDORES.objects.filter(credito_id=credito.id)
                    for codeudor_mod in codeudors_mod:
                        if not any(c['doc_ide'] == codeudor_mod.tercero.doc_ide for c in codeudors_bac):
                            codeudor_mod.delete()
                    for codeudo in codeudors_bac:
                        documento = codeudo.get('doc_ide')
                        tercero = TERCEROS.objects.filter(cliente_id=cliente_id, doc_ide=documento).first()
                        if tercero and not CODEUDORES.objects.filter(oficina_id=oficina_id, credito_id=credito.id, tercero_id=tercero.id).exists():
                            CODEUDORES.objects.create(oficina_id=oficina_id, credito_id=credito.id, tercero_id=tercero.id)

                    if is_ajax:
                        return JsonResponse({'success': True, 'message': 'Se grabó correctamente.', 'Numero': credito.cod_cre})
                    else:
                        messages.success(request, 'Se grabó correctamente.')
                        return JsonResponse({'success': True})

            except IntegrityError:
                return JsonResponse({'success': False, 'errors': form.errors})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})

# -----------------  end point ---------------------

@csrf_exempt  
def calculo_cuota_view(request):
    print('Entrada a Calculo Cuota.........')
    if request.method == 'POST':
        ikapital = float(request.POST.get('cap_ini', 0))
        per_ano = int(request.POST.get('per_ano', 0))
        iTIDIC = float(request.POST.get('tian_im', 0))
        iTIEA = float(request.POST.get('tiae_ic_ini', 0)) 
        iTIAN = xTasIntAnuNom = round(((1+iTIEA/100)**(1/per_ano)-1)*per_ano*100,3)
        xTasIntPer = round((iTIEA/100+1)**(1/per_ano)-1,6)           
        iTIDIC = round(xTasIntPer*per_ano*100/36525,6)                   
        iNumCuo = int(request.POST.get('num_cuo_ini', 0))
        iFecDes = request.POST.get('fec_des')
        iFecPagIni = request.POST.get('fec_pag_ini')
        iNumCuoGra = int(request.POST.get('num_cuo_gra', 0))
        periodos = {12: 'M', 6: 'B', 4: 'T', 3: 'C', 2: 'B', 24: 'Q'}
        iPerio = periodos.get(per_ano, None)
        if not iPerio:
            return JsonResponse({'cuota': 0}, status=400)
        if iNumCuo <= 0:
            return JsonResponse({'cuota': 0})
        FecDes = datetime.strptime(iFecDes, '%d-%m-%Y').date()
        FecPagIni = datetime.strptime(iFecPagIni, '%d-%m-%Y').date()
        valor_cuota = calculo_cuota(
                    ikapital = ikapital,
                    iTIEA = iTIEA/100,
                    iTIDIC = iTIDIC,
                    iPerio = iPerio,
                    iNumCuo = iNumCuo,
                    iFecDes = FecDes,
                    iFecPagIni = FecPagIni,
                    iNumCuoGra = iNumCuoGra,
                    iTIDPS = 0,
                    iIntCorAnt = 0
                )
        return JsonResponse({'cuota': valor_cuota})
    return JsonResponse({'cuota': 0}, status=405)

def gomonth(fecha, meses):
    return fecha + relativedelta(months=meses)

@csrf_exempt
def ImprimePlanAmortizacionPDF(request):

    def formatear_numero(numero):
        return f'{numero:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

    def encabezado(pdf):
        print('Entra a encabezado ')
        pdf.setFont("Helvetica", 12)
        pdf.drawString(100, height - 40, "COOPERATIVA ESPECIALIZADA DE AHORRO Y CREDITO DE LA ORINOQUIA")
        pdf.drawString(160, height - 56, "               O R I N O Q U I A")
        pdf.drawString(120, height - 72, "        PLAN DE AMORTIZACION DE CREDITOS")
        titulos = ["Cuota","Fecha","Capital","Int. Corriente","Poliza","Saldo"]
        pdf.setFont("Helvetica", 9)
        pdf.drawString(x, y-80, 'Credito ---> ')
        pdf.drawString(x+80, y-80, cod_cre)
        pdf.drawString(x+200, y-80, 'Deudor --> ')
        pdf.drawString(x+320, y-80,nom_deu)

        pdf.drawString(x, y-96, 'Capital Inicial --> ')
        pdf.drawString(x+80, y-96,formatear_numero(ikapital))
        pdf.drawString(x+200, y-96, 'Tas Int. Efectiva Anual --> ')
        pdf.drawString(x+320, y-96,format(iTIEA, '.2f'))
        pdf.drawString(x+400, y-96, 'Numero de Cuotas --->')
        pdf.drawString(x+520, y-96,formatear_numero(iNumCuo))

        pdf.drawString(x, y-112, 'Forma de Pago --> ')
        pdf.drawString(x+80, y-112,for_pag)
        pdf.drawString(x+200, y-112, 'Tas Int. Nominal Anual --> ')
        pdf.drawString(x+320, y-112,format(iTIAN, '.2f'))
        pdf.drawString(x+400, y-112, 'Tas Int. Mora Anual --> ')
        pdf.drawString(x+520, y-112,formatear_numero(tian_im))

        pdf.drawString(x, y-128,'Modalidad --> ')
        pdf.drawString(x+80, y-128,cod_lin_cre)
        pdf.drawString(x+200, y-128, 'Fecha Desembolso ')
        pdf.drawString(x+320, y-128,FecDes.strftime('%d/%m/%Y'))
        pdf.drawString(x+400, y-128, 'Fecha Pago Inicial --> ')
        pdf.drawString(x+520, y-128,FecPagIni.strftime('%d/%m/%Y'))

        pdf.drawString(x, y-144,'Valor Cuota --> ')
        pdf.drawString(x+80, y-144,formatear_numero(valor_cuota))
        pdf.drawString(x+200, y-144, 'Valor total ')
        pdf.drawString(x+320, y-144,formatear_numero(valor_cuota*iNumCuo))
        pdf.drawString(x+400, y-144, ' Fecha Venc. --> ')
        pdf.drawString(x+520, y-144,FecPagIni.strftime('%d/%m/%Y'))
        
        for col_num, col_name in enumerate(titulos):
            pdf.drawString(x + col_num * col_width, y-170, col_name)
            pdf.drawString(x + col_num * col_width, y-186, '============')
        

    def pie_de_pagina(pdf):
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(colors.grey)
        pdf.drawString(100, 50, "Este es el pie de página - Página 1")


    if request.method == 'POST':
        try:
            cod_cre = request.POST.get('cod_cre', '')
            nom_deu = request.POST.get('nom_deu', '')
            ikapital = float(request.POST.get('cap_ini', 0))
            per_ano = int(request.POST.get('per_ano', 0))
            iTIAN = float(request.POST.get('tian_ic_act', 0))
            print('por que recibe aqui cero   ',request.POST.get('tiae_ic_ini', 0))
            iTIEA = float(request.POST.get('tiae_ic_ini', 0))
            print('iTIEA  ',iTIAN, iTIEA)

            xTasIntPer = round((iTIEA / 100 + 1) ** (1 / per_ano) - 1, 6)
            iTIDIC = round(xTasIntPer * per_ano * 100 / 36525, 6)
            iNumCuo = int(request.POST.get('num_cuo_ini', 0))
            iFecDes = request.POST.get('fec_des')
            iFecPagIni = request.POST.get('fec_pag_ini')
            iNumCuoGra = int(request.POST.get('num_cuo_gra', 0))
            periodos = {12: 'M', 6: 'B', 4: 'T', 3: 'C', 2: 'B', 24: 'Q'}
            iPerio = periodos.get(per_ano, None)
            FecDes = datetime.strptime(iFecDes, '%d-%m-%Y').date()
            FecPagIni = datetime.strptime(iFecPagIni, '%d-%m-%Y').date()
            tian_im =  float(request.POST.get('tian_im', 0))
            for_pag = request.POST.get('for_pag')
            cod_lin_cre = request.POST.get('cod_lin_cre')
            valor_cuota = calculo_cuota(
                ikapital=ikapital,
                iTIEA=iTIEA / 100,
                iTIDIC=iTIDIC,
                iPerio=iPerio,
                iNumCuo=iNumCuo,
                iFecDes=FecDes,
                iFecPagIni=FecPagIni,
                iNumCuoGra=iNumCuoGra,
                iTIDPS=0,
                iIntCorAnt=0
            )
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="plan_amortizacion.pdf"'
            p = canvas.Canvas(response, pagesize=letter)
            width, height = letter  # 612 x 792 puntos
            row_height = 20  # 20 puntos entre filas
            col_width = 100  # 100 puntos entre columnas
            x = 25 # Empezar 100 puntos desde el borde izquierdo
            y = height - 25  # Empezar 100 puntos desde el borde superior
            encabezado(p)

            col = 560
            xFecAnt = FecDes
            xSaldo = ikapital
            xFecCuo = FecPagIni
            iTIDPS = 0
            xIntPorApl = 0
            xIntApl = 0
            xMeses = 1
            xTotCap = 0
            xTotIntCor = 0
            xTotPS = 0
            flujo_caja = []
            flujo_caja.append([0,-xSaldo])
            for xPer in range(1,iNumCuo+iNumCuoGra+1):
                xDifDias = (xFecCuo-xFecAnt).days
                xIntIC = round(xSaldo * iTIDIC * 1000,0)
                xIntIC = round(xIntIC/1000,0)*xDifDias
                xIntPS = round(xSaldo * iTIDPS * 1000,0)*(xDifDias)
                xIntPS = round(xIntPS/1000,0)*(xDifDias)
                xIntPer = xIntIC + xIntPS + xIntPorApl
                xCapPer = round(valor_cuota-xIntApl-xIntPer if xPer > iNumCuoGra and valor_cuota-xIntApl-xIntPer > 0 else 0) if xPer < iNumCuo+iNumCuoGra else xSaldo

                xNueCapPer = xCapPer 
                xIntApl = xIntApl + (xIntPer if xPer < iNumCuoGra else 0)
                xIntApl = xIntApl + (valor_cuota - xNueCapPer - xIntPer if xIntPorApl > 0 else 0)
                xIntApl = xIntApl if xIntApl > 0 else 0
                xFecAnt = xFecCuo
                xSaldo = xSaldo - xCapPer 
                p.drawString(60, col,str(xPer) )
                p.drawString(100, col,xFecCuo.strftime("%d/%m/%Y"))
                p.drawString(200+50 - p.stringWidth("{:,.0f}".format(xCapPer)), col,"{:,.0f}".format(xCapPer))
                p.drawString(300+50 - p.stringWidth("{:,.0f}".format(xIntPer)), col,"{:,.0f}".format(xIntPer))
                p.drawString(400+50 - p.stringWidth("{:,.0f}".format(xIntPS)), col,"{:,.0f}".format(xIntPS))
                p.drawString(500+50 - p.stringWidth("{:,.0f}".format(xSaldo)), col,"{:,.0f}".format(xSaldo))
                flujo_caja.append([xPer,-xCapPer])
                if col < 50:  # Suponiendo que 50 es el límite inferior de la página
                    p.showPage()  # Crear una nueva página
                    encabezado(p)
                    col = 576  # Reiniciar la posición vertical
                xTotCap += xCapPer
                xTotIntCor += xIntPer
                xTotPS += xIntPS
                col -= 20
                if xMeses > 0 :
                    xFecCuo = gomonth(FecPagIni,xMeses*xPer)
                elif iPerio == 'E':
                    xFecCuo = xFecCuo + 7
                elif  xPer % 2 == 1:
                    xFecCuo = xFecCuo + 15
                else:
                    xFecCuo = gomonth(xFecCuo-15,1)
            p.drawString(60, col,'Total')
            p.drawString(200+50 - p.stringWidth("{:,.0f}".format(xTotCap)), col,"{:,.0f}".format(xTotCap))
            p.drawString(300+50 - p.stringWidth("{:,.0f}".format(xTotIntCor)), col,"{:,.0f}".format(xTotIntCor))
            p.drawString(400+50 - p.stringWidth("{:,.0f}".format(xTotPS)), col,"{:,.0f}".format(xTotPS))
            width, height = letter
            y_position = height - 120
            
            p.showPage()
            p.save()

            return response
        except Exception as e:
            print(f'Error: {e}')
            return HttpResponse(status=500)

    return HttpResponse(status=405)

def confirmar_eliminar_credito(request, pk):
    credito = get_object_or_404(CREDITOS, pk=pk)    
    if request.method == "POST":
        credito.delete()
        return redirect('creditos_list')
    return render(request, 'confirmar_eliminar_credito.html', {'credito': credito})

def detalle_credito(request, credito_id):
    credito = get_object_or_404(CREDITOS, id=credito_id)
    codeudores = GAR_NO_IDONEA.objects.filter(credito=credito)
    return render(request, 'detalle_credito.html', {'credito': credito, 'codeudores': codeudores})

def agregar_codeudor(request, credito_id):
    credito = get_object_or_404(CREDITOS, id=credito_id)
    if request.method == 'POST':
        form = CodeudorForm(request.POST)
        if form.is_valid():
            codeudor = form.save(commit=False)
            codeudor.credito = credito
            codeudor.save()
            return redirect('detalle_credito', credito_id=credito.id)  # Redirigir a la vista de detalle
    else:
        form = CodeudorForm()
    return render(request, 'agregar_codeudor.html', {'form': form, 'credito': credito})

def editar_codeudor(request, codeudor_id):
    codeudor = get_object_or_404(GAR_NO_IDONEA, id=codeudor_id)
    if request.method == 'POST':
        form = CodeudorForm(request.POST, instance=codeudor)
        if form.is_valid():
            form.save()
            return redirect('detalle_credito', credito_id=codeudor.credito.id)
    else:
        form = CodeudorForm(instance=codeudor)
    return render(request, 'editar_codeudor.html', {'form': form, 'codeudor': codeudor})

def eliminar_codeudor(request, codeudor_id):
    codeudor = get_object_or_404(CODEUDORES, id=codeudor_id)
    if request.method == 'POST':
        codeudor.delete()
        return redirect('detalle_credito', credito_id=codeudor.credito.id)
    return render(request, 'confirmar_eliminacion.html', {'codeudor': codeudor})

def creditos_desembolsados(request):
    if request.method == 'GET':
        return render(request, 'creditos_desembolsados.html')
    if request.method == 'POST':
        accion = request.POST.get("accion")  # Obtener la acción
        # Recupera los parámetros del formulario
        id_cli = request.session.get('cliente_id')
        id_ofi = request.session.get('oficina_id')
        fec_ini = datetime.strptime(request.POST['fec_ini'], '%Y-%m-%d').date()
        fec_fin = datetime.strptime(request.POST['fec_fin'], '%Y-%m-%d').date()
        estado_cre = request.POST['estado_cre']
        estado_cre_nom = {
            'A': 'Activos',
            'C': 'Cancelados',
            'T': 'Todos'
            }
        estado_nombre = estado_cre_nom.get(estado_cre)
        
        resultado = listado_desembolsos(fec_ini, fec_fin, estado_cre)
        
        # cta_act = None
        # saldo_cta = 0
        # for row in resultado:
        #     cod_cta = row['cod_cta']
        #     debito = row['debito']
        #     credito = row['credito']
        #     sal_acu = row['sal_acu']
        #     if cta_act is None:  # Maneja el primer caso
        #         saldo_cta = sal_acu
        #         cta_act = cod_cta
        #     elif cod_cta == cta_act:
        #         saldo_cta = saldo_cta + debito - credito
        #         row['sal_acu'] = saldo_cta
        #     else:
        #         saldo_cta = sal_acu
        #         cta_act = cod_cta
    
        if accion == "exportar":
            # Lógica para exportar Crear el libro de Excel
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Creditos_Desembolsados" 

            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            # Llama a la función para obtener los datos
            resultado = listado_desembolsos(fec_ini, fec_fin, estado_cre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
                  
            # Añadir datos adicionales encima de los encabezados
            empresa = f"{entidad.nombre.strip()}"
            nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
            reporte = "CRÉDITOS DESEMBOLSADOS"
            detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
            filtro_cuentas = f"Estado de los Créditos: {estado_nombre}"
            
            datos_adicionales = [
                [empresa],
                [nit_empresa],
                [reporte],
                [detalles],
                [filtro_cuentas]                
            ]

            # Insertar los datos adicionales
            for row_num, fila in enumerate(datos_adicionales, start=1):
                for col_num, valor in enumerate(fila, start=1):
                    sheet.cell(row=row_num, column=col_num, value=valor)

            # Determinar la fila donde empiezan los encabezados
            header_row = len(datos_adicionales) + 1

            # Añadir los encabezados
            headers = resultado[0].keys()
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=header_row, column=col_num, value=header)

            # Añadir los datos
            for row_num, data in enumerate(resultado, start=header_row + 1):
                for col_num, field in enumerate(headers, start=1):
                    sheet.cell(row=row_num, column=col_num, value=data[field]) 
            
            nombre_archivo = f"cred_desem_{estado_nombre}_{fec_ini}_{fec_fin}.xlsx"
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            workbook.save(response)
            return response
        
        elif accion == "csv":
            # Llama a la función para obtener los datos
            resultado = listado_desembolsos(fec_ini, fec_fin, estado_cre)
            
            if not resultado:
                return HttpResponse("No se encontraron datos para exportar", status=404)
            
            # Configurar la respuesta HTTP para un archivo CSV
            nombre_archivo = f"cred_desem_{estado_nombre}_{fec_ini}_{fec_fin}.csv"
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            
            # Crear el escritor CSV
            writer = csv.writer(response)
            
            # Añadir encabezados de las columnas
            headers = resultado[0].keys()  # Suponemos que todos los diccionarios tienen las mismas claves
            writer.writerow(headers)

            # Añadir los datos
            for fila in resultado:
                writer.writerow([fila[col] for col in headers])
            return response
        
        elif accion == "imprimir":
            # Lógica para imprimir o generar vista previa  Configuración del PDF
            entidad = CLIENTES.objects.filter(id=id_cli).first()
            oficina = OFICINAS.objects.filter(id=id_ofi).first()

            nombre_archivo = f"cred_desem_{estado_nombre}_{fec_ini}_{fec_fin}.pdf"
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="{nombre_archivo}"'

            p = canvas.Canvas(response, pagesize=landscape(letter))
            width, height = landscape(letter)
            margin_x, margin_y = 50, 60

            # Configuración inicial
            filas_por_pagina = 46 # Número máximo de filas por página
            total_filas = len(resultado)
            total_paginas = ceil(total_filas / filas_por_pagina)

            # Función para dibujar subtotales
            # def dibujar_subtotales(y, subtotal_debito, subtotal_credito):
            #     x = margin_x - 30
            #     p.setFont("Helvetica-Bold", 9)
            #     p.drawString(x, y, "Subtotal: "+cod_cta)
            #     x += 640  # Posicionar en la columna de débito
            #     p.drawRightString(x, y, f"{subtotal_debito:,.2f}")
            #     x += 80  # Posicionar en la columna de crédito
            #     p.drawRightString(x, y, f"{subtotal_credito:,.2f}")

            # Función para dibujar encabezado
            def dibujar_encabezado():
                empresa = f"{entidad.nombre.strip()}"
                texto_empresa = stringWidth(empresa, "Times-Roman", 12)
                p.drawString((width - texto_empresa) / 2, height - 30, empresa)
                
                nit_empresa = f"NIT. {entidad.doc_ide.strip()}-{entidad.dv.strip()}"
                texto_nit = stringWidth(nit_empresa, "Times-Roman", 12)
                p.drawString((width - texto_nit) / 2, height - 45, nit_empresa)

                reporte = "LISTADO DE CRÉDITOS DESEMBOLSADOS"
                texto_reporte = stringWidth(reporte, "Times-Roman", 12)
                p.drawString((width - texto_reporte) / 2, height - 60, reporte)

                p.setFont("Helvetica", 10)
                detalles = f"Oficina: {oficina.nombre_oficina.upper()}      Periodo: {fec_ini} a {fec_fin}"
                texto_detalles = stringWidth(detalles, "Helvetica", 10)
                p.drawString((width - texto_detalles) / 2, height - 75, detalles)

                filtro_cuentas = f"Estado de los Créditos: {estado_nombre}"
                texto_filtro = stringWidth(filtro_cuentas, "Helvetica", 10)
                p.drawString((width - texto_filtro) / 2, height - 90, filtro_cuentas)

                p.drawImage("hecho_economico_app\\Logo\\LOGO.jpg", margin_x - 10, height - margin_y - 10, 60, 60)

            # Función para dibujar pie de página
            def dibujar_pie(pagina_actual, total_paginas):
                line_y = margin_y - 30  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)  # Dibuja la línea

                p.setFont("Courier", 9)
                texto_pie = f"{oficina.direccion.strip()}     Tel.: {oficina.celular.strip()}     E-mail: {oficina.email.strip()}     {oficina.ciudad.nombre.strip()}-{oficina.ciudad.departamento.strip()}"
                p.drawString(margin_x, margin_y - 40, texto_pie)

                texto_paginas = f"Pág. {pagina_actual} de {total_paginas}"
                p.drawRightString(width - margin_x, margin_y - 40, texto_paginas)

            # Función para dibujar encabezado de tabla
            def dibujar_encabezado_tabla(y):
                line_y = height - 115  # Ajusta para que esté justo encima del texto
                p.line(margin_x-40, line_y, margin_x+732, line_y)
                p.line(margin_x-40, height - 100, margin_x+732, height - 100)
                columnas = [
                    ("Cód_Créd.", 40), 
                    ("Cod_Aso", 50), 
                    ("Nombre", 180), 
                    ("F_P", 20), 
                    ("Fec_Des", 50), 
                    ("Ult_Pag", 40), 
                    ("       Cap_Ini", 80), 
                    ("C_I", 20),
                    ("C_A", 20),
                    ("   Cuo_Ini", 60), 
                    ("   Cuo_Act", 60),
                    ("E_J", 15),
                    ("ic_ini", 30),
                    ("ic_act", 30),
                    ("Fig.", 20),
                    ("Estado", 20)
                    ]
                x = margin_x-30
                p.setFont("Helvetica-Bold", 9)
                for col, ancho in columnas:
                    p.drawString(x, y, col)
                    x += ancho
                    
            # Dibujar contenido
            pagina_actual = 1
            dibujar_encabezado()
            dibujar_pie(pagina_actual, total_paginas)
            dibujar_encabezado_tabla(height - margin_y - 50)
            y = height - margin_y - 65

            def dibujar_fila(y, row):
                x = margin_x-30
                columnas = [
                    (row['cod_cre'], 40), 
                    (row['cod_aso'], 50),
                    (row['nombre'], 180), 
                    (row['for_pag'], 20),
                    (row['fec_des'], 50), 
                    (row['fec_ult_pag'],40),
                    (f"{row['cap_ini']:,.2f}",70,'right'),  
                    (f"{row['num_cuo_ini']}", 20,'right'), 
                    (f"{row['num_cuo_act']}", 20,'right'), 
                    # (row['val_cuo_ini'], 70),
                    (f"{row['val_cuo_ini']:,.2f}",60,'right'), 
                    # (row['val_cuo_act'], 70), 
                    (f"{row['val_cuo_act']:,.2f}",60,'right'), 
                    (f"{row['est_jur']}", 20,'right'),
                    # (f"{row['tian_ic_ini']}", 30,'right'), 
                    (f"{row['tian_ic_act']}", 30,'right'),
                    (f"{row['figarantias']}", 20,'right'),
                    (f"{row['estado']}", 20,'right'),
                    # (f"{row['detalle'][:24]}", 125), 
                    # (f"{row['debito']:,.2f}",70,'right'), 
                    # (f"{row['credito']:,.2f}", 70,'right'), 
                    # (f"{row['sal_acu']:,.2f}", 80, 'right')
                    ]

                p.setFont("Helvetica", 9)
                for col in columnas:
                    if len(col) == 3 and col[2] == 'right':  # Alineación derecha
                        texto, ancho, _ = col
                        p.drawRightString(x + ancho, y, texto)
                    else:  # Alineación izquierda
                        texto, ancho = col[:2]
                        p.drawString(x, y, str(texto))
                    x += ancho

            for idx, row in enumerate(resultado):
                if y < margin_y - 30:
                    p.showPage()
                    pagina_actual += 1
                    dibujar_encabezado()
                    dibujar_encabezado_tabla(height - margin_y - 50)
                    dibujar_pie(pagina_actual, total_paginas)
                    y = height - margin_y - 65    

                dibujar_fila(y, row)
                y -= 10

            p.showPage()
            p.save()
            return response
        else:
            return HttpResponse("Acción no válida", status=400)
    return HttpResponse("Método no permitido", status=405)

def listado_desembolsos(fecha_inicial, fecha_final, estado_cre):
    if estado_cre == 'A':  # Créditos Activos
        filtros = Q(fec_des__range=(fecha_inicial, fecha_final), estado='A')
    elif estado_cre == 'C':  # Créditos Cancelados
        filtros = Q(fec_ult_pag__range=(fecha_inicial, fecha_final), estado='C')
    elif estado_cre == 'T':  # Todos (Activos y Cancelados)
        filtros = Q(estado='A', fec_des__range=(fecha_inicial, fecha_final)) | \
                  Q(estado='C', fec_ult_pag__range=(fecha_inicial, fecha_final))
    else:
        return []  # En caso de estado inválido, devuelve una lista vacía

    # Consulta con el ORM de Django
    creditos = CREDITOS.objects.filter(filtros).select_related('socio__tercero').order_by('estado', 'socio__tercero__nombre')

    # Convertir resultados en una lista de diccionarios
    resultados = [
        {
            'cod_cre': c.cod_cre,
            'cod_aso': c.socio.cod_aso,
            'nombre': c.socio.tercero.nombre,
            'for_pag': c.for_pag,
            'fec_des': c.fec_des,
            'fec_ult_pag': c.fec_ult_pag,
            'cap_ini': c.cap_ini,
            'num_cuo_ini': c.num_cuo_ini,
            'num_cuo_act': c.num_cuo_act,
            'val_cuo_ini': c.val_cuo_ini,
            'val_cuo_act': c.val_cuo_act,
            'est_jur': c.est_jur,
            'tian_ic_act': c.tian_ic_act,
            'figarantias': c.figarantias,
            'estado': c.estado
        }
        for c in creditos
    ]
    return resultados

def obtener_fecha_desembolso(oficina_id,fecha,subcuenta=None,cliente_id=None):    
    fecha_desembolso = DETALLE_PROD.objects.filter(
            producto='CR',
            concepto='DESEM',
            hecho_econo__fecha__lte=fecha,
            subcuenta__in=CREDITOS.objects.filter(
                oficina_id=oficina_id,
                socio__tercero__doc_ide=subcuenta,
                socio__tercero__cliente_id=cliente_id
            ).values_list('cod_cre', flat=True)
        ).aggregate(max_fecha=Max('hecho_econo__fecha'))['max_fecha']
    
    return fecha_desembolso if fecha_desembolso else date(2000, 1, 1)

def obtener_fecha_movimiento(oficina_id,fecha,subcuenta=None,cliente_id=None): 
    fecha_movimiento = DETALLE_PROD.objects.filter(
            producto='CR',
            hecho_econo__fecha__lte=fecha,
            subcuenta__in=CREDITOS.objects.filter(
                oficina_id=oficina_id,
                socio__tercero__doc_ide=subcuenta,
                socio__tercero__cliente_id=cliente_id
                ).values_list('cod_cre', flat=True)
            ).exclude(concepto='DESEM').aggregate(max_fecha=Max('hecho_econo__fecha'))['max_fecha']
    return fecha_movimiento if fecha_movimiento else date(2000, 1, 1)

def fecha_ultimo_movimiento(fecha, cod_cre): 
    fecha_movimiento = DETALLE_PROD.objects.filter(
            producto='CR',
            hecho_econo__fecha__lte=fecha,
            subcuenta = cod_cre
            ).exclude(concepto='DESEM').aggregate(max_fecha=Max('hecho_econo__fecha'))['max_fecha']
    print('fecha ,ovimiento',fecha_movimiento)
    return fecha_movimiento.strftime('%d/%m/%Y') if fecha_movimiento else ''

def lista_creditos_asociado(id_socio, fecha_corte):
    creditos = CREDITOS.objects.filter(
        estado = "A",
        socio_id = id_socio
    )
    resultados = []
    for credito in creditos:
        liq_cre = Liquida_cre(credito.cod_cre, fecha_corte)
        liq_cre.liq_al_dia(fecha_corte, recarga = True)
        saldo_total = (liq_cre.sal_cap_tot + liq_cre.sal_int_dia + liq_cre.int_cau_fra + liq_cre.int_mor_a_pag + liq_cre.pol_seg_a_pag + liq_cre.acreedor_a_pag)
        val_cuo_dia = (liq_cre.capital_a_pag + liq_cre.sal_int_dia + liq_cre.int_cau_fra + liq_cre.int_mor_a_pag + liq_cre.pol_seg_a_pag + liq_cre.acreedor_a_pag)
        liq_cre.calculo_periodo()
        xdias_mor = (liq_cre.fecha_focal-liq_cre.fec_al_dia).days
        xdias_mor = xdias_mor if xdias_mor > 0 else 0
        cuotas_pagadas = liq_cre.cuo_pag
        altura = liq_cre.altura
        if cuotas_pagadas == altura:
            estado = 'AL DÍA'
        elif cuotas_pagadas > altura:
            estado = 'ADELANTADO'
        else:
            estado = 'EN MORA '+str(xdias_mor)+' DÍAS'
            
        resultados.append(
            {
            'cod_cre': credito.cod_cre,
            'lin_cre': credito.cod_lin_cre.descripcion,
            'fec_des': credito.fec_des,
            'cap_ini': credito.cap_ini,
            'num_cuo_ini': credito.num_cuo_ini,
            'val_cuo_ini': credito.val_cuo_ini,
            'saldo': saldo_total,
            'val_cuo_dia': val_cuo_dia,
            'cuotas_pagadas': cuotas_pagadas,
            'altura': altura,
            'estado': estado
            }
        )
    return resultados

def lista_deudor_solidario(id_socio,fecha_corte):
    # Obtener el tercero relacionado al id_asociado
    asociado = ASOCIADOS.objects.select_related('tercero').get(id=id_socio)
    tercero = asociado.tercero.doc_ide
    # Buscar los créditos en los que es codeudor
    codeudores = GAR_NO_IDONEA.objects.filter(doc_ide=tercero, credito__estado='A').select_related('credito','credito__oficina')
    resultados = []
    # Mostrar la información
    for codeudor in codeudores:
        credito = codeudor.credito
        liq_cre = Liquida_cre(credito.cod_cre,fecha_corte)  # función propia
        liq_cre.liq_al_dia(fecha_corte, recarga = True)
        saldo_total = (liq_cre.sal_cap_tot + liq_cre.sal_int_dia + liq_cre.int_cau_fra + liq_cre.int_mor_a_pag + liq_cre.pol_seg_a_pag + liq_cre.acreedor_a_pag)
        liq_cre.calculo_periodo()
        xdias_mor = (liq_cre.fecha_focal-liq_cre.fec_al_dia).days
        xdias_mor = xdias_mor if xdias_mor > 0 else 0
        cuotas_pagadas = liq_cre.cuo_pag
        altura = liq_cre.altura
        if cuotas_pagadas == altura:
            estado = 'AL DÍA'
        elif cuotas_pagadas > altura:
            estado = 'ADELANTADO'
        else:
            estado = 'EN MORA '+str(xdias_mor)+' DÍAS'  
        datos = {
        'cod_cre': credito.cod_cre,
        'lin_cre': credito.cod_lin_cre.descripcion,
        'nom_deu': credito.socio.tercero.nombre,
        'fec_des': credito.fec_des,
        'cap_ini': credito.cap_ini,
        'num_cuo_ini': credito.num_cuo_ini,
        'saldo': saldo_total,
        'estado': estado
        }    
        resultados.append(datos)
    return resultados

def novacion_reestrucion(request, pk):

    return