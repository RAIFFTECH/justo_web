import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import CTA_CDAT_LIQ

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = CTA_CDAT_LIQ
    form = CrearForm
    template_name = 'lista_liq_cdat.html'
    # ordering = ['cliente', 'codigo']

    # def get_queryset(self):
    #     # Obtén el queryset base para el modelo LOCALIDADES
    #     queryset = super().get_queryset()

    #     # Puedes obtener el cliente desde la solicitud (request) de diferentes maneras
    #     # Por ejemplo, si el cliente está en los argumentos de la URL, puedes usar:
    #     # cliente_id = self.kwargs.get('cliente_id')

    #     # Si el cliente es parte de la sesión del usuario, puedes usar:
    #     # cliente_id = self.request.user.cliente_id
    #     cliente_id = 2

    #     # Por simplicidad, aquí se toma un cliente_id específico como ejemplo
    #     # Cambia este valor según tus necesidades
    #     # cliente_id = self.request.GET.get('cliente_id', None)

    #     # Filtra el queryset por cliente si se proporciona un cliente_id
    #     if cliente_id:
    #         queryset = queryset.filter(cliente=cliente_id)

    #         # Retorna el queryset filtrado
    #         return queryset

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = CTA_CDAT_LIQ
    form = CrearForm
    template_name = 'detalles_liq_cdat.html'

# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = CTA_CDAT_LIQ
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_liq_cdat.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_liq_cdat')

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = CTA_CDAT_LIQ
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_liq_cdat.html'

    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_liq_cdat')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = CTA_CDAT_LIQ
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_liq_cdat')

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        liq_cdat = CTA_CDAT_LIQ.objects.all().order_by('cliente', 'codigo')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="liq_cdat.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in liq_cdat:
        # p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
        # p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
        # p.drawString(80, 760, f"Fecha: {dato.fecha}")
        # p.drawString(80, 740, f"Tipo Liquidación: {dato.tip_liq}")
        # p.drawString(80, 720, f"Valor Intereses: {dato.val_int}")
        # p.drawString(80, 700, f"Valor Retefuente: {dato.val_ret}")
        # p.drawString(80, 680, f"Valor Retefuente Nueva: {dato.val_ret_nue}")
        # p.drawString(80, 660, f"Aplicado?: {dato.aplicado}")
        # p.drawString(80, 640, f"Número Documento: {dato.docto}")


        datos_tabla = [["cta_aho","cta_amp","fecha","tip_liq","val_int","val_ret","val_ret_nue","aplicado","docto"]]

        for dato in liq_cdat:
            datos_tabla.append([dato.cta_aho, dato.cta_amp, dato.fecha, dato.tip_liq, dato.val_int, dato.val_ret, dato.val_ret_nue, dato.aplicado, dato.docto])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = CTA_CDAT_LIQ.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="localidades.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
        p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
        p.drawString(80, 760, f"Fecha: {dato.fecha}")
        p.drawString(80, 740, f"Tipo Liquidación: {dato.tip_liq}")
        p.drawString(80, 720, f"Valor Intereses: {dato.val_int}")
        p.drawString(80, 700, f"Valor Retefuente: {dato.val_ret}")
        p.drawString(80, 680, f"Valor Retefuente Nueva: {dato.val_ret_nue}")
        p.drawString(80, 660, f"Aplicado?: {dato.aplicado}")
        p.drawString(80, 640, f"Número Documento: {dato.docto}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        liq_cdat = CTA_CDAT_LIQ.objects.all()

        p = canvas.Canvas(response)

        for dato in liq_cdat:
            p.drawString(80, 800, f"Cuenta de Ahorro: {dato.cta_aho}")
            p.drawString(80, 780, f"Cuenta Ampliación: {dato.cta_amp}")
            p.drawString(80, 760, f"Fecha: {dato.fecha}")
            p.drawString(80, 740, f"Tipo Liquidación: {dato.tip_liq}")
            p.drawString(80, 720, f"Valor Intereses: {dato.val_int}")
            p.drawString(80, 700, f"Valor Retefuente: {dato.val_ret}")
            p.drawString(80, 680, f"Valor Retefuente Nueva: {dato.val_ret_nue}")
            p.drawString(80, 660, f"Aplicado?: {dato.aplicado}")
            p.drawString(80, 640, f"Número Documento: {dato.docto}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in CTA_CDAT_LIQ._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        liq_cdat = CTA_CDAT_LIQ.objects.all()
        for row_num, data in enumerate(liq_cdat, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        liq_cdat = CTA_CDAT_LIQ.objects.all()
        for data in liq_cdat:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
