import csv, re, json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.db.models import Sum, F
from django.db import IntegrityError
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, TableStyle, PageBreak
from datetime import datetime, date, timedelta
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User, Group
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import get_object_or_404

# from django.contrib.auth import password_reset, password_reset_done,password_reset_confirm, password_reset_complete

from django.contrib.auth.decorators import login_required
from num2words import num2words
from .forms import CustomUserCreationForm
from django import forms

# from movimientos_app.models import REGISTROS
# from trabajadores_app.models import TRABAJADORES
# from trabajos_app.models import TRABAJOS    
# from localidades_app.models import LOCALIDADES
from backup_app.views import backup_database, restore_backup
# from indusanchez_app.funciones_principales import fecha_corta, fecha_larga, fecha_año_mes_dia, fecha_primero

def Inicio(request):
    return render(request, 'dashboard.html')


def Registrar_Usuario(request):
    if request.method == 'GET':
        return render(request, 'Registrar_Usuario1.html', {
            'form': CustomUserCreationForm
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                usuario = User.objects.create_user(
                    username=request.POST['username'], password=request.POST['password1'],
                    first_name=request.POST['first_name'], last_name=request.POST['last_name'], email=request.POST['email'])
                usuario.save()
                # Obtiene el grupo deseado
                grupo = Group.objects.get(name='Invitados')
                # Asigna el usuario al grupo
                usuario.groups.add(grupo)  
                login(request, usuario)
                return redirect('inicio')
            except IntegrityError:
                return render(request, 'Registrar_Usuario1.html', {
                    'form': CustomUserCreationForm,
                    'error': 'El Usuario ya existe'
                })
        return render(request, 'Registrar_Usuario1.html', {
            'form': CustomUserCreationForm,
            'error': 'Las Contraseñas no coinciden'
        }) 

@login_required
def Cerrar_Sesion(request):
    backup_database(request)
    logout(request)
    return redirect('iniciar_sesion')


def Iniciar_Sesion(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # request.session['cliente_id'] = 1  # Valor constante de ejemplo
            # request.session['oficina_id'] = 1  # Valor constante de ejemplo
            # request.session['per_con'] = 2026  # Valor constante de ejemplo
            # user_profile = UserProfile.objects.get(user=request.user)
            # foto_perfil = user_profile.photo.url #if user_profile.photo else None
            # request.session['foto_perfil'] = foto_perfil
            return redirect('inicio')
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
    return render(request, 'Iniciar_Sesion.html')       


def dashboard(request):
    return render(request, 'hola mundo')

# @login_required  # Asegúrate de que el usuario esté autenticado
def dashboard(request):
    fecha_inicio = request.GET.get('fecha_inicial')
    fecha_fin = request.GET.get('fecha_final')

    # registros = REGISTROS.objects.all()
    
    # if fecha_inicio and fecha_fin:
    #     registros = registros.filter(fecha__range=[fecha_inicio, fecha_fin])

    # datos = registros.values(
    #     'trabajo__nombre'
    # ).annotate(
    #     total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
    # )

    # labels = [d['trabajo__nombre'] for d in datos]
    # valores = [float(d['total']) for d in datos]

    return render(request, 'dashboard.html', {
        # 'labels': json.dumps(labels),
        # 'valores': json.dumps(valores),
        'inicio': fecha_inicio,
        'fin': fecha_fin
    })
    

def exportar_excel_resumen(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # 🔥 TITULO
    ws.merge_cells('A1:D1')
    ws['A1'] = "REPORTE DE PRODUCCIÓN"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 🔥 ENCABEZADOS
    headers = ["Trabajador", "Trabajo", "Cantidad", "Total"]
    ws.append(headers)

    for cell in ws[2]:
        cell.font = Font(bold=True)

    datos = REGISTROS.objects.values(
        'trabajador__pri_nom',
        'trabajador__pri_ape',
        'trabajo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad'),
        total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
    )

    total_general = 0

    for d in datos:
        total_general += d['total']

        ws.append([
            f"{d['trabajador__pri_nom']} {d['trabajador__pri_ape']}",
            d['trabajo__nombre'],
            d['total_cantidad'],
            d['total']
        ])

    # 🔥 TOTAL
    ws.append(["", "", "TOTAL", total_general])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte.xlsx"'

    wb.save(response)
    return response  


def exportar_pdf_resumen(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    elementos = []

    # 🔥 TITULO
    elementos.append(Paragraph("REPORTE DE PRODUCCIÓN", styles['Title']))

    datos = [["Trabajador", "Trabajo", "Cantidad", "Total"]]

    registros = REGISTROS.objects.values(
        'trabajador__pri_nom',
        'trabajador__pri_ape',
        'trabajo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad'),
        total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
    )

    total_general = 0

    for r in registros:
        total_general += r['total']

        datos.append([
            f"{r['trabajador__pri_nom']} {r['trabajador__pri_ape']}",
            r['trabajo__nombre'],
            r['total_cantidad'],
            r['total']
        ])

    datos.append(["", "", "TOTAL", total_general])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.black),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.grey)
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    return response



def Restablecer_Contraseña(request):
    return render(request, "hola mundo")




def Resumen(request):
    
    hoy = date.today()

    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    # 🔥 valores por defecto (semana actual)
    if not fecha_inicial or not fecha_final:
        fecha_inicial = hoy - timedelta(days=hoy.weekday())
        fecha_final = hoy
        
    # inicio = hoy - timedelta(days=hoy.weekday())

    resumen = REGISTROS.objects.filter(
        fecha__range=[fecha_inicial, fecha_final]
    ).values(
        'trabajador_id', 'trabajador__pri_nom', 'trabajador__pri_ape',
    ).annotate(
        total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
    )

    total_general = sum(r['total'] for r in resumen)
    
    return render(request, 'resumen.html', {
        'resumen': resumen,
        'fecha_inicial': fecha_inicial,
        'fecha_final': fecha_final,
        'inicio': fecha_inicial,
        'fin': fecha_final,
        'total_general': total_general
    })



def Cuentas_Por_Pagar(request, trabajador_id):
    
    hoy = date.today()

    fecha_inicial = request.GET.get('fecha_inicial') or ""
    fecha_final = request.GET.get('fecha_final') or ""

    # 🔥 Si vienen del formulario → convertir
    if fecha_inicial and fecha_final:
        fecha_inicial = datetime.strptime(fecha_inicial, "%Y-%m-%d").date()
        fecha_final = datetime.strptime(fecha_final, "%Y-%m-%d").date()
    else:
        # 🔥 valores por defecto (semana actual)
        fecha_inicial = hoy - timedelta(days=hoy.weekday())
        fecha_final = hoy

    registros = REGISTROS.objects.filter(
        trabajador_id=trabajador_id,
        fecha__range=[fecha_inicial, fecha_final]
    ).values(
        'trabajo__nombre', unidades= Sum(F('cantidad')),
    ).annotate(
        
        total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
    ).order_by('trabajo__nombre')

    # 🔥 obtener trabajador
    trabajador = REGISTROS.objects.filter(
        trabajador_id=trabajador_id
    ).first().trabajador

    total_general = sum(r['total'] for r in registros)

    # ================= PDF =================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="cuenta_cobro_{trabajador_id}.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elementos = []
    
    estilo_centrado = ParagraphStyle(
        name='Centrado',
        parent=styles['Normal'],
        alignment=TA_CENTER,
        fontSize=12,
        spaceAfter=5
    )

    # 🔥 TITULO
    elementos.append(Paragraph(f"Granada, {fecha_larga(hoy)}", styles['Normal']))
    elementos.append(Spacer(6, 30))
    elementos.append(Paragraph("CUENTA DE COBRO", styles['Title']))
    elementos.append(Spacer(10, 30))

    # 🔥 DATOS EMPRESA
    elementos.append(Paragraph("CAREN ELIANA SÁNCHEZ", estilo_centrado))
    elementos.append(Paragraph(f"NIT. 1.130.638.008-1", estilo_centrado))
    elementos.append(Spacer(8, 30))
    elementos.append(Paragraph(f"DEBE A:", estilo_centrado))
    elementos.append(Spacer(6, 20))
    
    
    # 🔥 TABLA
    datos = [["Labor", "Cantidad", "Total    "]]

    for r in registros:
        datos.append([
            r['trabajo__nombre'],
            r['unidades'],
            f"${int(r['total']):,}".replace(",", ".")
        ])

    # 🔥 TOTAL
    datos.append(["", "TOTAL", f"${int(total_general):,}".replace(",", ".")])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.black),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        # 🔹 Columna 0 → izquierda
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),

        # 🔹 Columna 1 → derecha
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),

        # 🔹 Columna 2 → derecha
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),

        ('GRID', (0,0), (-1,-1), 1, colors.grey),
    ]))
    pesos = ''
    if total_general == 0:
        pesos = 'PESOS M/CTE.'
    elif total_general % 1_000_000 == 0:
        pesos = 'DE PESOS M/CTE.'
    else:
        pesos = 'PESOS M/CTE.'
        
    numero_en_letras = num2words(total_general, lang='es').upper()
    valor_en_letras = f"{numero_en_letras} {pesos}" + f"(${total_general:,.0f})".replace(",", ".")+f" M/CTE."
    
    # 🔥 DATOS COLABORADOR
    nombre = f"{trabajador.pri_nom} {trabajador.seg_nom} {trabajador.pri_ape} {trabajador.seg_ape}"
    tipo_identificacion = trabajador.get_cla_doc_display().lower()
    numero_identificacion = f"{trabajador.doc_ide}"
    ciudad_expedicion = trabajador.cod_ciu_exp.nombre+" - "+trabajador.cod_ciu_exp.departamento

    identificacion = "identificada" if trabajador.sexo == 'F' else "identificado"
    elementos.append(Paragraph(f"{nombre} {identificacion} con {tipo_identificacion} {numero_identificacion} expedida en la ciudad de {ciudad_expedicion}, la suma de {valor_en_letras}, por concepto de servicios en las siguientes labores:", styles['Normal']))
    
    elementos.append(Spacer(2, 20))

    elementos.append(tabla)
        
    elementos.append(Spacer(4, 20))
    fec_ini = fecha_corta(fecha_inicial)
    fec_fin = fecha_corta(fecha_final)
    elementos.append(Paragraph(f"Desde: {fec_ini}   Hasta: {fec_fin}", styles['Normal']))
    print('cuenta',trabajador.tipo_cuenta)
    elementos.append(Spacer(4, 20))
    
    tipo_cuenta = " de Ahorros" if trabajador.tipo_cuenta == 'A' else "Corriente"
    print('tipo_cuenta', tipo_cuenta)
    elementos.append(Paragraph(f"Favor consignar en la Cuenta {tipo_cuenta} {trabajador.cuenta_bancaria} del Banco {trabajador.get_banco_display()}, de la cual soy titular.", styles['Normal']))
    
    elementos.append(Spacer(4, 20))
    elementos.append(Paragraph(f"Cordialmente,", styles['Normal']))
    
    elementos.append(Spacer(24, 80))
    fec_ini = fecha_corta(fecha_inicial)
    fec_fin = fecha_corta(fecha_final)
    elementos.append(Paragraph(f"{nombre}", styles['Normal']))
    elementos.append(Paragraph(f"Dirección: {trabajador.direccion}", styles['Normal']))
    elementos.append(Paragraph(f"Celular: {trabajador.celular1}", styles['Normal']))
    elementos.append(Paragraph(f"e-mail: {trabajador.email}", styles['Normal']))

    doc.build(elementos)

    return response



def cuentas_masivas(request):

    hoy = date.today()

    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    if fecha_inicial and fecha_final:
        fecha_inicial = datetime.strptime(fecha_inicial, "%Y-%m-%d").date()
        fecha_final = datetime.strptime(fecha_final, "%Y-%m-%d").date()
    else:
        fecha_inicial = hoy - timedelta(days=hoy.weekday())
        fecha_final = hoy

    # 🔥 obtener trabajadores únicos
    trabajadores = REGISTROS.objects.filter(
        fecha__range=[fecha_inicial, fecha_final]
    ).values('trabajador_id').distinct()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="cuentas_masivas.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elementos = []

    for t in trabajadores:

        trabajador_id = t['trabajador_id']

        registros = REGISTROS.objects.filter(
            trabajador_id=trabajador_id,
            fecha__range=[fecha_inicial, fecha_final]
        ).values(
            'trabajo__nombre'
        ).annotate(
            unidades=Sum('cantidad'),
            total=Sum(F('cantidad') * F('trabajo__valor_unitario'))
        )

        trabajador = REGISTROS.objects.filter(
            trabajador_id=trabajador_id
        ).first().trabajador

        total_general = sum(r['total'] for r in registros)

        # ================= PDF =================        
        estilo_centrado = ParagraphStyle(
            name='Centrado',
            parent=styles['Normal'],
            alignment=TA_CENTER,
            fontSize=12,
            spaceAfter=5
        )

        # 🔥 TITULO
        elementos.append(Paragraph(f"Granada, {fecha_larga(hoy)}", styles['Normal']))
        elementos.append(Spacer(6, 30))
        elementos.append(Paragraph("CUENTA DE COBRO", styles['Title']))
        elementos.append(Spacer(10, 30))

        # 🔥 DATOS EMPRESA
        elementos.append(Paragraph("CAREN ELIANA SÁNCHEZ", estilo_centrado))
        elementos.append(Paragraph(f"NIT. 1.130.638.008-1", estilo_centrado))
        elementos.append(Spacer(8, 30))
        elementos.append(Paragraph(f"DEBE A:", estilo_centrado))
        elementos.append(Spacer(6, 20))
        
        
        # 🔥 TABLA
        datos = [["Labor", "Cantidad", "Total    "]]

        for r in registros:
            datos.append([
                r['trabajo__nombre'],
                r['unidades'],
                f"${int(r['total']):,}".replace(",", ".")
            ])

        # 🔥 TOTAL
        datos.append(["", "TOTAL", f"${int(total_general):,}".replace(",", ".")])

        tabla = Table(datos)

        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.black),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            # 🔹 Columna 0 → izquierda
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),

            # 🔹 Columna 1 → derecha
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),

            # 🔹 Columna 2 → derecha
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),

            ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ]))
        
            
        numero_en_letras = num2words(total_general, lang='es').upper()
        valor_en_letras = f"{numero_en_letras} PESOS "+f"(${total_general:,.0f})".replace(",", ".")+f" M/CTE."
        
        # 🔥 DATOS COLABORADOR
        nombre = f"{trabajador.pri_nom} {trabajador.seg_nom} {trabajador.pri_ape} {trabajador.seg_ape}"
        tipo_identificacion = trabajador.get_cla_doc_display().lower()
        numero_identificacion = f"{trabajador.doc_ide}"
        ciudad_expedicion = trabajador.cod_ciu_exp.nombre+" - "+trabajador.cod_ciu_exp.departamento

        identificacion = "identificada" if trabajador.sexo == 'F' else "identificado"
        elementos.append(Paragraph(f"{nombre} {identificacion} con {tipo_identificacion} {numero_identificacion} expedida en la ciudad de {ciudad_expedicion}, la suma de {valor_en_letras}, por concepto de servicios en las siguientes labores:", styles['Normal']))
        
        elementos.append(Spacer(2, 20))

        elementos.append(tabla)
            
        elementos.append(Spacer(4, 20))
        fec_ini = fecha_corta(fecha_inicial)
        fec_fin = fecha_corta(fecha_final)
        elementos.append(Paragraph(f"Desde: {fec_ini}   Hasta: {fec_fin}", styles['Normal']))
        print('cuenta',trabajador.tipo_cuenta)
        elementos.append(Spacer(4, 20))
        
        tipo_cuenta = " de Ahorros" if trabajador.tipo_cuenta == 'A' else "Corriente"
        print('tipo_cuenta', tipo_cuenta)
        elementos.append(Paragraph(f"Favor consignar en la Cuenta {tipo_cuenta} {trabajador.cuenta_bancaria} del Banco {trabajador.get_banco_display()}, de la cual soy titular.", styles['Normal']))
        
        elementos.append(Spacer(4, 20))
        elementos.append(Paragraph(f"Cordialmente,", styles['Normal']))
        
        elementos.append(Spacer(24, 80))
        fec_ini = fecha_corta(fecha_inicial)
        fec_fin = fecha_corta(fecha_final)
        elementos.append(Paragraph(f"{nombre}", styles['Normal']))
        elementos.append(Paragraph(f"Dirección: {trabajador.direccion}", styles['Normal']))
        elementos.append(Paragraph(f"Celular: {trabajador.celular1}", styles['Normal']))
        elementos.append(Paragraph(f"e-mail: {trabajador.email}", styles['Normal']))

        # 🔥 SALTO DE PÁGINA
        elementos.append(PageBreak())

    doc.build(elementos)

    return response


def Archivo_Plano(request):
    return render(request, "hola mundo")


def keep_alive(request):
    return JsonResponse({'status': 'ok'})