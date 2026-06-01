import csv
from django.utils.dateparse import parse_date
from django.db.models import Q
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django_xhtml2pdf.utils import generate_pdf
from django_xhtml2pdf.views import PdfMixin
from django.views.generic import DetailView

from django import forms
from .forms import CrearForm
from .models import PLAN_CTAS


# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = PLAN_CTAS
    form = CrearForm
    paginate_by = 10
    template_name = 'lista_cuentas.html'
    ordering = ['cliente', 'per_con', 'cod_cta']

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return self.model.objects.filter(
                Q(cod_cta__icontains=query) | Q(nom_cta__icontains=query)
            )
        return self.model.objects.all()


def Buscar_Cuenta_Contable(request, template_name): 
    fecha_ini = request.GET.get('fec_ini')  # Fecha inicial ingresada por el usuario
    if fecha_ini:   # Si el usuario proporciona una fecha, filtramos por el período contable (año)
        fecha_ini_parsed = parse_date(fecha_ini)  # Convertir la fecha en objeto datetime.date
        if fecha_ini_parsed:
            periodo_contable = fecha_ini_parsed.year  # Extraer el año de la fecha
            cuentas = PLAN_CTAS.objects.filter(per_con__gte=periodo_contable).order_by('cod_cta')  # Filtrar cuentas con per_con >= año
        else:
            cuentas = PLAN_CTAS.objects.all().order_by('-per_con', 'cod_cta')  # Si la fecha no es válida, obtener todas las cuentas
    else:
        cuentas = PLAN_CTAS.objects.all().order_by('-per_con', 'cod_cta')  # Si no se proporciona fecha, mostrar todas las cuentas
    return render(request, template_name, {'cuentas': cuentas})  # Pasar las cuentas filtradas al template
    




# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = PLAN_CTAS
    form = CrearForm
    template_name = 'detalles_cuenta_contable.html'


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = PLAN_CTAS
    form_class = CrearForm
    # fields = '__all__'
    template_name = 'crear_cuenta_contable.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    def get_form_kwargs(self):
        kwargs = super(Crear, self).get_form_kwargs()
        kwargs.update({'request': self.request})
        return kwargs

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_cuenta')


# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = PLAN_CTAS
    form_class = CrearForm
    # fields = '__all__'
    template_name = 'actualizar_cuenta_contable.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    def get_form_kwargs(self):
        kwargs = super(Actualizar, self).get_form_kwargs()
        kwargs.update({'request': self.request})
        return kwargs

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_cuenta')


# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = PLAN_CTAS
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_cuenta')


# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Plan de Cuentas"
        # Crear los encabezados en la primera fila
        sheet.append([
            "Cliente", "Periodo Contable", "Código Cuenta", "Nombre Cuenta",
            "Tipo Cuenta", "Dinámica", "Naturaleza", "Cuenta Activa?",
            "Contabiliza por Tercero?", "Cuenta Activo Fijo?",
            "Cuenta Presupuesto?", "Cuenta de Balance?",
            "Cuenta de Resultados?", "Cuenta de Orden?",
            "Cuenta de Banco?", "Cuenta Ganancias?",
            "Cuenta Pérdidas?", "Cuenta Depreciación?",
            "Cuenta Ingresos y Retenciones?", "Cuenta Reteiva?",
            "Cuenta Recíproca?", "ID"
        ])
        # Obtener los datos del modelo PLAN_CTAS
        plan_ctas = PLAN_CTAS.objects.all()
        # Agregar los datos de cada cuenta a la hoja de cálculo
        for cuenta in plan_ctas:
            sheet.append([
                cuenta.cliente.nombre if cuenta.cliente else '',
                cuenta.per_con,
                cuenta.cod_cta,
                cuenta.nom_cta,
                cuenta.tip_cta,
                cuenta.dinamica,
                cuenta.naturaleza,
                cuenta.activa,
                cuenta.por_tercero,
                cuenta.cta_act_fij,
                cuenta.cta_pre,
                cuenta.cta_bal,
                cuenta.cta_res,
                cuenta.cta_ord,
                cuenta.cta_ban,
                cuenta.cta_gan_per,
                cuenta.cta_per_gan,
                cuenta.cta_dep,
                cuenta.cta_ing_ret,
                cuenta.cta_ret_iva,
                cuenta.cta_rec,
                cuenta.id_ds
            ])
        # Preparar la respuesta HTTP para devolver el archivo Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=plan_ctas.xlsx"
        # Guardar el libro de trabajo en la respuesta
        workbook.save(response)
        return response


# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = PLAN_CTAS.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="cuentas.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cliente: {dato.cliente}")
        p.drawString(80, 780, f"Periodo Contable: {dato.per_con}")
        p.drawString(80, 760, f"Código Cuenta: {dato.cod_cta}")
        p.drawString(80, 740, f"Nombre Cuenta: {dato.nom_cta}")
        p.drawString(80, 720, f"Tipo Cuenta: {dato.tip_cta}")
        p.drawString(80, 700, f"Dinámica: {dato.dinamica}")
        p.drawString(80, 680, f"Naturaleza: {dato.naturaleza}")
        p.drawString(80, 660, f"Cuenta Activa?: {dato.activa}")
        p.drawString(80, 640, f"Contabiliza por Tercero?: {dato.por_tercero}")
        p.drawString(80, 620, f"Cuenta Activo Fijo?: {dato.cta_act_fij}")
        p.drawString(80, 600, f"Cuenta Presupuesto?: {dato.cta_pre}")
        p.drawString(80, 580, f"Cuenta de Balance?: {dato.cta_bal}")
        p.drawString(80, 560, f"Cuenta de Resultados?: {dato.cta_res}")
        p.drawString(80, 540, f"Cuenta de Orden?: {dato.cta_ord}")
        p.drawString(80, 520, f"Cuenta de Banco?: {dato.cta_ban}")
        p.drawString(80, 500, f"Cuenta Ganancias?: {dato.cta_gan_per}")
        p.drawString(80, 480, f"Cuenta Pérdidas?: {dato.cta_per_gan}")
        p.drawString(80, 460, f"Cuenta Depreciación?: {dato.cta_dep}")
        p.drawString(80, 440, f"Cuenta Ingresos y Retenciones?: {dato.cta_ing_ret}")
        p.drawString(80, 420, f"Cuenta Reteiva?: {dato.cta_ret_iva}")
        p.drawString(80, 400, f"Cuenta Recíproca?: {dato.cta_rec}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        cuentas = PLAN_CTAS.objects.all()

        p = canvas.Canvas(response)

        for dato in cuentas:
            p.drawString(80, 800, f"Cliente: {dato.cliente}")
            p.drawString(80, 780, f"Periodo Contable: {dato.per_con}")
            p.drawString(80, 760, f"Código Cuenta: {dato.cod_cta}")
            p.drawString(80, 740, f"Nombre Cuenta: {dato.nom_cta}")
            p.drawString(80, 720, f"Tipo Cuenta: {dato.tip_cta}")
            p.drawString(80, 700, f"Dinámica: {dato.dinamica}")
            p.drawString(80, 680, f"Naturaleza: {dato.naturaleza}")
            p.drawString(80, 660, f"Cuenta Activa?: {dato.activa}")
            p.drawString(80, 640, f"Contabiliza por Tercero?: {dato.por_tercero}")
            p.drawString(80, 620, f"Cuenta Activo Fijo?: {dato.cta_act_fij}")
            p.drawString(80, 600, f"Cuenta Presupuesto?: {dato.cta_pre}")
            p.drawString(80, 580, f"Cuenta de Balance?: {dato.cta_bal}")
            p.drawString(80, 560, f"Cuenta de Resultados?: {dato.cta_res}")
            p.drawString(80, 540, f"Cuenta de Orden?: {dato.cta_ord}")
            p.drawString(80, 520, f"Cuenta de Banco?: {dato.cta_ban}")
            p.drawString(80, 500, f"Cuenta Ganancias?: {dato.cta_gan_per}")
            p.drawString(80, 480, f"Cuenta Pérdidas?: {dato.cta_per_gan}")
            p.drawString(80, 460, f"Cuenta Depreciación?: {dato.cta_dep}")
            p.drawString(80, 440, f"Cuenta Ingresos y Retenciones?: {dato.cta_ing_ret}")
            p.drawString(80, 420, f"Cuenta Reteiva?: {dato.cta_ret_iva}")
            p.drawString(80, 400, f"Cuenta Recíproca?: {dato.cta_rec}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in PLAN_CTAS._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        cuentas = PLAN_CTAS.objects.all()
        for row_num, data in enumerate(cuentas, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        cuentas = PLAN_CTAS.objects.all()
        for data in cuentas:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
class ExportPDF(PdfMixin, DetailView):
    model = PLAN_CTAS
    template_name = "pdf-cuenta-contable.html"
    context_object_name = "cuenta"