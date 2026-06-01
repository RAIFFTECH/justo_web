import csv
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from django import forms
from .forms import CrearForm
from .models import LINEAS_CREDITO
from django.http import JsonResponse

def obtener_detalles_lin_cre(request,pk):
    # print('Codigo de linea de Credito recibido --->',pk)
    if request.method == 'GET':
        lin_cre = get_object_or_404(LINEAS_CREDITO,id = pk)
        data = {
                'tas_int_cor': lin_cre.tas_int_anu,
                'tas_int_mor': lin_cre.tas_int_mor,
                'tas_por_pol': lin_cre.por_pol,
                'por_des_pp': lin_cre.por_des_pp,
                # 'dia_con_int_mor': lin_cre.dia_con_int_mor,
        }
        print(data)
        return JsonResponse(data)
    

def obtener_lin_cre(request,pk):
    print('Entro Lineas de Credito')
    cliente_id = request.session.get('cliente_id')
    if pk:
        try:
            lin_cre = LINEAS_CREDITO.objects.get(cliente_id = cliente_id,id = pk)
            context = {
                'cod_lin_cre': lin_cre.cod_lin_cre,
                'tas_int_cor': lin_cre.tas_int_anu,
                'tas_int_mor': lin_cre.tas_int_mor,
                'tas_por_pol': lin_cre.por_pol,
                'por_des_pp': lin_cre.por_des_pp,
                # 'dia_con_int_mor': lin_cre.dia_con_int_mor,
            }
            return JsonResponse({'nombre': context})
        except pk.DoesNotExist:
            return JsonResponse({'cod_lin_cre': 'No existe'})
    return JsonResponse({'cod_lin_cre': 'No existe'})

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = LINEAS_CREDITO
    form = CrearForm
    template_name = 'lista_linea_creditos.html'
    ordering = ['cliente', 'cod_lin_cre']

    # def get_queryset(self):
    #     # Obtén el queryset base para el modelo LOCALIDADES
    #     queryset = super().get_queryset()

    #     # Puedes obtener el cliente desde la solicitud (request) de diferentes maneras
    #     # Por ejemplo, si el cliente está en los argumentos de la URL, puedes usar:
    #     # cliente_id = self.kwargs.get('cliente_id')

    #     # Si el cliente es parte de la sesión del usuario, puedes usar:
    #     # cliente_id = self.request.user.cliente_id
    #     cliente_id = 2

    #     # Por simplicidad, aquí se toma un cliente_id específico como ejemplo
    #     # Cambia este valor según tus necesidades
    #     # cliente_id = self.request.GET.get('cliente_id', None)

    #     # Filtra el queryset por cliente si se proporciona un cliente_id
    #     if cliente_id:
    #         queryset = queryset.filter(cliente=cliente_id)

    #         # Retorna el queryset filtrado
    #         return queryset


# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = LINEAS_CREDITO
    form = CrearForm
    template_name = 'detalles_linea_credito.html'


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = LINEAS_CREDITO
    form = CrearForm
    fields = '__all__'
    template_name = 'crear_linea_credito.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_linea_credito')
    

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = LINEAS_CREDITO
    form = CrearForm
    fields = '__all__'
    template_name = 'actualizar_linea_credito.html'

    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_linea_credito')


# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = LINEAS_CREDITO
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_linea_credito')
    

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        linea_creditos = LINEAS_CREDITO.objects.all().order_by('cliente', 'codigo')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="linea_creditos.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in linea_creditos:
        # p.drawString(80, 800, f"Cliente: {dato.cliente}")
        # p.drawString(80, 780, f"Código Línea Crédito: {dato.cod_lin_cre}")
        # p.drawString(80, 760, f"Descripción: {dato.descripcion}")
        # p.drawString(80, 740, f"Tasa Interés Anual: {dato.tas_int_anu}")
        # p.drawString(80, 720, f"Tasa Interés Mora: {dato.tas_int_mor}")
        # p.drawString(80, 700, f"Porcentaje Póliza: {dato.por_pol}")
        # p.drawString(80, 680, f"Descuento Pronto Pago: {dato.por_des_pp}")
        # p.drawString(80, 660, f"Días Condonación Interés Mora: {dato.dia_con_int_mor}")
     
        datos_tabla = [["cliente","cod_lin_cre","descripcion","tas_int_anu","tas_int_mor","por_pol","por_des_pp","dia_con_int_mor"]]

        for dato in linea_creditos:
            datos_tabla.append([dato.cliente, dato.cod_lin_cre, dato.descripcion, dato.tas_int_anu,
                               dato.tas_int_mor, dato.por_pol, dato.por_des_pp, dato.dia_con_int_mor])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = LINEAS_CREDITO.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="localidades.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        p.drawString(80, 800, f"Cliente: {dato.cliente}")
        p.drawString(80, 780, f"Código Línea Crédito: {dato.cod_lin_cre}")
        p.drawString(80, 760, f"Descripción: {dato.descripcion}")
        p.drawString(80, 740, f"Tasa Interés Anual: {dato.tas_int_anu}")
        p.drawString(80, 720, f"Tasa Interés Mora: {dato.tas_int_mor}")
        p.drawString(80, 700, f"Porcentaje Póliza: {dato.por_pol}")
        p.drawString(80, 680, f"Descuento Pronto Pago: {dato.por_des_pp}")
        p.drawString(80, 660, f"Días Condonación Interés Mora: {dato.dia_con_int_mor}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        linea_creditos = LINEAS_CREDITO.objects.all()

        p = canvas.Canvas(response)

        for dato in linea_creditos:
            p.drawString(80, 800, f"Cliente: {dato.cliente}")
            p.drawString(80, 780, f"Código Línea Crédito: {dato.cod_lin_cre}")
            p.drawString(80, 760, f"Descripción: {dato.descripcion}")
            p.drawString(80, 740, f"Tasa Interés Anual: {dato.tas_int_anu}")
            p.drawString(80, 720, f"Tasa Interés Mora: {dato.tas_int_mor}")
            p.drawString(80, 700, f"Porcentaje Póliza: {dato.por_pol}")
            p.drawString(80, 680, f"Descuento Pronto Pago: {dato.por_des_pp}")
            p.drawString(80, 660, f"Días Condonación Interés Mora: {dato.dia_con_int_mor}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in LINEAS_CREDITO._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        linea_creditos = LINEAS_CREDITO.objects.all()
        for row_num, data in enumerate(linea_creditos, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        linea_creditos = LINEAS_CREDITO.objects.all()
        for data in linea_creditos:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
