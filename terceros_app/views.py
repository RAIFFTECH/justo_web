import csv

from django_xhtml2pdf.views import PdfMixin
from openpyxl import Workbook
from django.views import View
from django.http import HttpResponse
from django.db.models import Q
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, ListFlowable, ListItem
from django import forms
from django_xhtml2pdf.utils import generate_pdf
from django_xhtml2pdf.views import PdfMixin
from django.views.generic import DetailView

from .forms import CrearForm
from .models import TERCEROS
from django.template.loader import render_to_string
from django.http import JsonResponse

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q

def buscar_tercero(request):
    doc_ide = request.GET.get('doc_ide', None)
    if doc_ide:
        tercero = TERCEROS.objects.filter(cliente_id = 1,doc_ide=doc_ide).first()
        if tercero  != None:      
            return JsonResponse({'nombre': tercero.nombre, 'exists': True})
        else:
            return JsonResponse({'error': 'El tercero no existe.', 'exists': False})
    return JsonResponse({'error': 'No se proporcionó un doc_ide válido.'}, status=400)


@api_view(['GET'])
def buscar_terceros_query(request, query):
    queryset = TERCEROS.objects.filter(
        Q(doc_ide__icontains=query) |
        Q(nombre__icontains=query) |
        Q(raz_soc__icontains=query) |
        Q(celular1__icontains=query)
    )
    results = queryset.values('doc_ide', 'nombre')
    return Response(results, status=status.HTTP_200_OK)


class Lista(LoginRequiredMixin, ListView):
    model = TERCEROS
    form = CrearForm
    paginate_by = 10
    context_object_name = "resultados"
    template_name = 'lista_terceros.html'

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return TERCEROS.objects.filter(
                Q(pri_nom__icontains=query) | Q(pri_ape__icontains=query) | Q(doc_ide__icontains=query) |
                Q(celular1__icontains=query) | Q(email__icontains=query)
            ).order_by('doc_ide')
        return TERCEROS.objects.none()


# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = TERCEROS
    form = CrearForm
    template_name = 'detalles_tercero.html'


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = TERCEROS
    form_class = CrearForm
    template_name = 'Crear_Tercero.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'
    
    def form_valid(self, form):
        # Asignamos el usuario actual al campo asesor
        form.instance.usuario_asesor = self.request.user
        return super().form_valid(form)
    
    def get_initial(self):
        initial = super().get_initial()
        initial['usuario_asesor'] = self.request.user.id
        return initial

    def get_form_kwargs(self):
        kwargs = super(Crear, self).get_form_kwargs()
        kwargs.update({'request': self.request})
        kwargs['user'] = self.request.user
        return kwargs

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_terceros')


# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = TERCEROS
    form_class = CrearForm
    template_name = 'actualizar_tercero.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'
    
    def get_form_kwargs(self):
        kwargs = super(Actualizar, self).get_form_kwargs()
        kwargs.update({'request': self.request})
        return kwargs
  
    # def get_initial(self):
    #     initial = super().get_initial()
    #     instance = self.get_object()
    #     # Asegúrate de que esté en formato lista
    #     initial['grupos_especiales'] = instance.grupos_especiales.split(',') if instance.grupos_especiales else []
    #     return initial
    
    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_terceros')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = TERCEROS
    form = CrearForm
    fields = "__all__"

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_terceros')

# Para imprimir los registros
class ImprimirPDF(View):
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Terceros"

        # Crear los encabezados en la primera fila
        sheet.append([
            "Cliente", "Tipo Documento", "Número Documento", "DV", "Nit Rápido",
            "Ciudad Expedición", "Ciudad Residencia", "Tipo Régimen", "Fecha Exp. Documento",
            "Tipo Tercero", "Primer Apellido", "Segundo Apellido", "Primer Nombre",
            "Segundo Nombre", "Razón Social", "Dirección", "Código Postal", "Tel. Oficina",
            "Tel. Residencia", "Celular 1", "Celular 2", "Fax", "Email", "Nombre",
            "Fecha Actualización", "Observación", "Persona Expuesta PEP", "Nit Interno"
        ])

        # Obtener los datos del modelo TERCEROS
        terceros = TERCEROS.objects.all()

        # Agregar los datos de cada tercero a la hoja de cálculo
        for tercero in terceros:
            sheet.append([
                tercero.cliente.nombre if tercero.cliente else '',
                tercero.cla_doc,
                tercero.doc_ide,
                tercero.dig_ver,
                tercero.nit_rap,
                tercero.cod_ciu_exp.nombre if tercero.cod_ciu_exp else '',
                tercero.cod_ciu_res.nombre if tercero.cod_ciu_res else '',
                tercero.regimen,
                tercero.fec_exp_ced,
                tercero.tip_ter,
                tercero.pri_ape,
                tercero.seg_ape,
                tercero.pri_nom,
                tercero.seg_nom,
                tercero.raz_soc,
                tercero.direccion,
                tercero.cod_pos,
                tercero.tel_ofi,
                tercero.tel_res,
                tercero.celular1,
                tercero.celular2,
                tercero.fax,
                tercero.email,
                tercero.nombre,
                tercero.fec_act,
                tercero.observacion,
                tercero.per_pub_exp,
                tercero.nit_interno
            ])

        # Preparar la respuesta HTTP para devolver el archivo Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=terceros.xlsx"

        # Guardar el libro de trabajo en la respuesta
        workbook.save(response)

        return response


# Para imprimir un registro
class ImprimePDF(View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = TERCEROS.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="terceros.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        p.drawString(60, 800, f"Cliente: {dato.cliente}")
        p.drawString(60, 780, f"Tipo Documento: {dato.cla_doc}")
        p.drawString(60, 760, f"Número Documento: {dato.doc_ide}")
        p.drawString(60, 740, f"DV: {dato.dig_ver}")
        p.drawString(60, 720, f"Nit Rápido: {dato.nit_rap}")
        p.drawString(60, 700, f"Ciudad Expedición Documento: {dato.cod_ciu_exp}")
        p.drawString(60, 680, f"Ciudad Residencia: {dato.cod_ciu_res}")
        p.drawString(60, 660, f"Tipo Régimen: {dato.regimen}")
        p.drawString(60, 640, f"Fecha Expedición Documento: {dato.fec_exp_ced}")
        p.drawString(60, 620, f"Tipo Tercero: {dato.tip_ter}")
        p.drawString(60, 600, f"Primer Apellido: {dato.pri_ape}")
        p.drawString(60, 580, f"Segundo Apellido: {dato.seg_ape}")
        p.drawString(60, 560, f"Primer Nombre: {dato.pri_nom}")
        p.drawString(60, 540, f"Segundo Nombre: {dato.seg_nom}")
        p.drawString(60, 520, f"Razón Social: {dato.raz_soc}")
        p.drawString(60, 500, f"Dirección: {dato.direccion}")
        p.drawString(60, 480, f"Código Postal: {dato.cod_pos}")
        p.drawString(60, 460, f"Teléfono Oficina: {dato.tel_ofi}")
        p.drawString(60, 440, f"Teléfono Residencia: {dato.tel_res}")
        p.drawString(60, 400, f"Celular 1: {dato.celular1}")
        p.drawString(60, 380, f"Celular 2: {dato.celular2}")
        p.drawString(60, 360, f"Fax: {dato.fax}")
        p.drawString(60, 340, f"e-mail: {dato.email}")
        p.drawString(60, 320, f"Nombre: {dato.nombre}")
        p.drawString(60, 300, f"Fecha Actualización: {dato.fec_act}")
        p.drawString(60, 280, f"Observación: {dato.observacion}")
        p.drawString(60, 260, f"Persona Expuesta PEP: {dato.per_pub_exp}")
        p.drawString(60, 240, f"Nit Interno: {dato.nit_interno}")
        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportPDF(PdfMixin, DetailView):
    model = TERCEROS
    template_name = "pdf-terceros.html"
    context_object_name = "tercero"


class ExportarView(View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        terceros = TERCEROS.objects.all()

        p = canvas.Canvas(response)

        for dato in terceros:
            p.drawString(60, 800, f"Cliente: {dato.cliente}")
            p.drawString(60, 780, f"Tipo Documento: {dato.cla_doc}")
            p.drawString(60, 760, f"Número Documento: {dato.doc_ide}")
            p.drawString(60, 740, f"DV: {dato.dig_ver}")
            p.drawString(60, 720, f"Nit Rápido: {dato.nit_rap}")
            p.drawString(60, 700, f"Ciudad Expedición Documento: {dato.cod_ciu_exp}")
            p.drawString(60, 680, f"Ciudad Residencia: {dato.cod_ciu_res}")
            p.drawString(60, 660, f"Tipo Régimen: {dato.regimen}")
            p.drawString(60, 640, f"Fecha Expedición Documento: {dato.fec_exp_ced}")
            p.drawString(60, 620, f"Tipo Tercero: {dato.tip_ter}")
            p.drawString(60, 600, f"Primer Apellido: {dato.pri_ape}")
            p.drawString(60, 580, f"Segundo Apellido: {dato.seg_ape}")
            p.drawString(60, 560, f"Primer Nombre: {dato.pri_nom}")
            p.drawString(60, 540, f"Segundo Nombre: {dato.seg_nom}")
            p.drawString(60, 520, f"Razón Social: {dato.raz_soc}")
            p.drawString(60, 500, f"Dirección: {dato.direccion}")
            p.drawString(60, 480, f"Código Postal: {dato.cod_pos}")
            p.drawString(60, 460, f"Teléfono Oficina: {dato.tel_ofi}")
            p.drawString(60, 440, f"Teléfono Residencia: {dato.tel_res}")
            p.drawString(60, 400, f"Celular 1: {dato.celular1}")
            p.drawString(60, 380, f"Celular 2: {dato.celular2}")
            p.drawString(60, 360, f"Fax: {dato.fax}")
            p.drawString(60, 340, f"e-mail: {dato.email}")
            p.drawString(60, 320, f"Nombre: {dato.nombre}")
            p.drawString(60, 300, f"Fecha Actualización: {dato.fec_act}")
            p.drawString(60, 280, f"Observación: {dato.observacion}")
            p.drawString(60, 260, f"Persona Expuesta PEP: {dato.per_pub_exp}")
            p.drawString(60, 240, f"Nit Interno: {dato.nit_interno}")
            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in TERCEROS._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        terceros = TERCEROS.objects.all()
        for row_num, data in enumerate(terceros, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        terceros = TERCEROS.objects.all()
        for data in terceros:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response
