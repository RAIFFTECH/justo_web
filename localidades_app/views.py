import csv, json
from openpyxl import Workbook, load_workbook
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.generic import ListView, DetailView, View
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django import forms
from .forms import CrearForm
from .models import LOCALIDADES

# Para obtener todos los registros
class Lista(LoginRequiredMixin, ListView):
    model = LOCALIDADES
    form_class = CrearForm
    template_name = 'lista_localidades.html'
    ordering = ['codigo']
    

# Para obtener todos los detalles de un registro
class Detalles(LoginRequiredMixin, DetailView):
    model = LOCALIDADES
    form_class = CrearForm
    template_name = 'detalles_localidades.html'


# Para crear un nuevo registro
class Crear(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = LOCALIDADES
    form_class = CrearForm
    
    template_name = 'crear_localidades.html'

    # Mensaje que se mostrará cuando se inserte el registro
    success_message = 'Registro añadido correctamente.'

    # Redirigimos a la página principal tras insertar el registro
    def get_success_url(self):
        return reverse('listar_localidades')
    

# Para modificar un registro
class Actualizar(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = LOCALIDADES
    form_class = CrearForm
    
    template_name = 'actualizar_localidades.html'
    # Mensaje que se mostrará cuando se actualice el registro
    success_message = 'Registro actualizado correctamente.'

    # Redireccionamos a la página principal tras actualizar el registro
    def get_success_url(self):
        return reverse('listar_localidades')

# Para eliminar un registro
class Eliminar(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = LOCALIDADES
    form = CrearForm

    # Redireccionamos a la página principal tras de eliminar el registro
    def get_success_url(self):
        # Mensaje que se mostrará cuando se elimine el registro
        success_message = 'Registro eliminado correctamente.'
        messages.success(self.request, (success_message))
        return reverse('listar_localidades')

# Para imprimir los registros
class ImprimirPDF(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        localidades = LOCALIDADES.objects.all().order_by('cliente', 'codigo')

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="localidades.pdf"'

        # Creamos un objeto PDF con ReportLab
        # p = canvas.Canvas(response)
        p = canvas.Canvas(response, pagesize=letter)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # for dato in localidades:
        # p.drawString(80, 800, f"Cliente: {dato.cliente}")
        # p.drawString(80, 780, f"Código: {dato.codigo}")
        # p.drawString(80, 760, f"Ciudad: {dato.nombre}")
        # p.drawString(80, 740, f"Código Postal: {dato.cod_pos}")
        # p.drawString(80, 720, f"Departamento: {dato.departamento}")

        datos_tabla = [["Código", "nombre", "cod_pos", "departamento"]]

        for dato in localidades:
            datos_tabla.append([dato.codigo, dato.nombre, dato.cod_pos, dato.departamento])

        # Agrega más campos según tus necesidades

        tabla = Table(datos_tabla)

        # estilo_tabla = TableStyle([('BACKGROUND', (0, 0), (-1, 0), '#06153C'),
        #                            ('TEXTCOLOR', (0, 0), (-1, 0), (255, 255, 255)),
        #                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        #                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #                            ('BACKGROUND', (0, 1), (-1, -1), '#f7f7f7'),
        #                            ('GRID', (0, 0), (-1, -1), 1, '#06153C')])

        # tabla.setStyle(estilo_tabla)

        # Dibujamos la tabla en el PDF
        tabla.wrapOn(p, 0, 0)
        tabla.drawOn(p, 30, 600)

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response

# Para imprimir un registro
class ImprimePDF(LoginRequiredMixin, View):
    def get(self, request, pk, *args, **kwargs):
        # Recupera los datos de la base de datos
        # Asegúrate de adaptar esto a tu modelo y consulta específicos
        dato = LOCALIDADES.objects.get(pk=pk)

        # Creamos un objeto HttpResponse con el tipo de contenido PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="localidades.pdf"'

        # Creamos un objeto PDF con ReportLab
        p = canvas.Canvas(response)

        # Agregamos contenido al PDF utilizando datos de la base de datos
        # p.drawString(80, 800, f"Cliente: {dato.cliente}")
        p.drawString(80, 780, f"Código: {dato.codigo}")
        p.drawString(80, 760, f"Ciudad: {dato.nombre}")
        p.drawString(80, 740, f"Código Postal: {dato.cod_pos}")
        p.drawString(80, 720, f"Departamento: {dato.departamento}")

        # Agrega más campos según tus necesidades

        # Agrega un salto de página para el siguiente conjunto de datos
        p.showPage()

        # Cierra el objeto PDF y devuelve la respuesta
        p.save()

        return response


class ExportarView(LoginRequiredMixin, View):
    template_name = 'exportar.html'

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        export_type = request.POST.get('export_type')

        if export_type == 'pdf':
            return self.exportar_pdf()
        elif export_type == 'excel':
            return self.exportar_excel()
        elif export_type == 'csv':
            return self.exportar_csv()

    def exportar_pdf(self):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="exportacion.pdf"'
        localidades = LOCALIDADES.objects.all()

        p = canvas.Canvas(response)

        for dato in localidades:
            # p.drawString(80, 800, f"Cliente: {dato.cliente}")
            p.drawString(80, 780, f"Código: {dato.codigo}")
            p.drawString(80, 760, f"Ciudad: {dato.nombre}")
            p.drawString(80, 740, f"Código Postal: {dato.cod_pos}")
            p.drawString(80, 720, f"Departamento: {dato.departamento}")

            # Agrega más campos según tus necesidades

            # Agrega un salto de página para el siguiente conjunto de datos
            p.showPage()
        # Puedes agregar más contenido según tus necesidades.
        p.save()

        return response

    def exportar_excel(self):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exportacion.xlsx"'

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Datos"

        # Añade encabezados a la hoja de cálculo utilizando los nombres de campo del modelo
        headers = [field.name for field in LOCALIDADES._meta.fields]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Añade datos a la hoja de cálculo
        localidades = LOCALIDADES.objects.all()
        for row_num, data in enumerate(localidades, start=2):
            for col_num, field in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num,
                           value=getattr(data, field))

        workbook.save(response)
        return response

    def exportar_csv(self):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exportacion.csv"'

        writer = csv.writer(response)
        # Añade más encabezados según tus necesidades
        writer.writerow(['ID', 'Nombre'])

        localidades = LOCALIDADES.objects.all()
        for data in localidades:
            # Añade más campos según tus necesidades
            writer.writerow([data.pk, data.nombre])

        return response



class Vista_Previa(LoginRequiredMixin, View):
    template_name = 'importar_localidades.html'

    # ✅ PARA ABRIR LA PÁGINA (GET)
    def get(self, request):
        return render(request, self.template_name)

    # ✅ PROCESAR ARCHIVO (POST)
    def post(self, request):
        archivo = request.FILES.get('archivo')

        if not archivo:
            messages.error(request, 'Debe seleccionar un archivo')
            return redirect('vista_previa')

        extension = archivo.name.split('.')[-1].lower()
        data = []

        try:
            # ===== EXCEL =====
            if extension in ['xlsx', 'xlsm']:
                wb = load_workbook(archivo)
                sheet = wb.active

                for fila in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(fila):
                        continue

                    # 🔥 SOLUCIÓN ERROR columnas extra
                    codigo, nombre, cod_pos, departamento, *resto = fila

                    data.append({
                        'codigo': codigo,
                        'nombre': nombre,
                        'cod_pos': cod_pos,
                        'departamento': departamento,
                        'error': self.validar(codigo, nombre)
                    })

            # ===== CSV =====
            elif extension == 'csv':
                decoded = archivo.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded)

                next(reader)

                for fila in reader:
                    if not fila:
                        continue

                    codigo, nombre, cod_pos, departamento, *resto = fila

                    data.append({
                        'codigo': codigo,
                        'nombre': nombre,
                        'cod_pos': cod_pos,
                        'departamento': departamento,
                        'error': self.validar(codigo, nombre)
                    })

            # 🔥 CLAVE: convertir a JSON correcto
            data_json = json.dumps(data)

            return render(request, self.template_name, {
                'data': data,
                'data_json': data_json
            })

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            return redirect('vista_previa')

    def validar(self, codigo, nombre):
        if not codigo or not nombre:
            return "❌ Campos obligatorios"
        if LOCALIDADES.objects.filter(codigo=codigo).exists():
            return "⚠️ Código duplicado"
        return ""


class Confirmar_Importacion(LoginRequiredMixin, View):
    
    def post(self, request):
        import json

        data_raw = request.POST.get('data')

        if not data_raw:
            messages.error(request, "No hay datos para importar")
            return redirect('vista_previa')

        registros = json.loads(data_raw)

        creados = 0

        for r in registros:
            if r['error'] == "":
                LOCALIDADES.objects.create(
                    codigo=r['codigo'],
                    nombre=r['nombre'],
                    cod_pos=r['cod_pos'],
                    departamento=r['departamento']
                )
                creados += 1

        messages.success(request, f'{creados} Registros importados correctamente')
        return redirect('listar_localidades')